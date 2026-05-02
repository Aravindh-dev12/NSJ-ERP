import json
import os
from typing import Iterable, Optional


from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.db.models import Q

from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from nsj_backend.authentication import CsrfExemptSessionAuthentication


def jwt_login_required(view_func):
    """
    Drop-in replacement for @jwt_login_required that also accepts JWT Bearer tokens.
    Checks Authorization header first, falls back to Django session, returns 401 JSON if neither works.
    """
    from functools import wraps
    from rest_framework_simplejwt.authentication import JWTAuthentication
    from rest_framework_simplejwt.exceptions import InvalidToken, TokenError

    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            auth_header = request.META.get("HTTP_AUTHORIZATION", "")
            if auth_header.startswith("Bearer "):
                token_str = auth_header[7:]
                try:
                    jwt_auth = JWTAuthentication()
                    validated_token = jwt_auth.get_validated_token(token_str)
                    request.user = jwt_auth.get_user(validated_token)
                except (InvalidToken, TokenError):
                    return JsonResponse({"detail": "Authentication required"}, status=401)
            else:
                return JsonResponse({"detail": "Authentication required"}, status=401)
        return view_func(request, *args, **kwargs)

    return _wrapped


# Helper function to get authentication classes
def get_auth_classes():
    """Get authentication classes, with JWT if available"""
    auth_classes = [CsrfExemptSessionAuthentication]
    try:
        from rest_framework_simplejwt.authentication import JWTAuthentication

        auth_classes.insert(0, JWTAuthentication)
    except ImportError:
        pass
    return auth_classes


def parse_put_request(request):
    """
    Helper function to parse multipart/form-data for PUT/PATCH requests.
    Django doesn't parse these by default, so we do it manually.
    """
    if (
        request.method in ["PUT", "PATCH"]
        and request.content_type
        and "multipart" in request.content_type
    ):
        from django.http.multipartparser import MultiPartParser

        try:
            # Parse the multipart data
            parser = MultiPartParser(request.META, request, request.upload_handlers)
            post_data, files_data = parser.parse()

            # Replace request.POST and request.FILES with parsed data
            request._post = post_data
            request._files = files_data
        except Exception:
            # If parsing fails, continue with empty POST/FILES
            pass


from .models import Voucher, Sale, PurReturn, SalesReturn
from .models import Archives
from .models import PurchaseDiamond
from .models import Order
from .serializers import (
    VoucherSerializer,
    SaleSerializer,
    PurReturnSerializer,
    SalesReturnSerializer,
)
from .serializers import ArchivesSerializer
from .serializers import PurchaseDiamondSerializer
from .serializers import PurchaseMSerializer, PurchaseTagwiseMSerializer
from .models import ApprovalLooseM
from .serializers import ApprovalLooseMSerializer
from .models import ApprovalTagM
from .serializers import ApprovalTagMSerializer
from .models import PurAndApprovalM
from .serializers import PurAndApprovalMSerializer
from .models import Repair
from .serializers import RepairSerializer
from .serializers import PaymentEntrySerializer, JournalEntrySerializer
from django.utils import timezone
import traceback
from datetime import timedelta
from core.models import SeriesMaster, StampMaster, BaseMetalMaster
from django.db.models import Sum
from accounts.models import Account, SubAccount
from django.db.utils import OperationalError
from core.models import (
    ItemNameMaster,
    ClarityMaster,
    ShapeMaster,
    UnitMaster,
    SizeMaster,
    ColourMaster,
    LabMaster,
)
from openpyxl import Workbook
from django.utils.text import slugify
import decimal
import uuid

# Receive model/serializer
from .models import Receive
from .serializers import ReceiveSerializer

# Receipt model/serializer
from .models import Receipt, ReceiptDebitEntry
from .serializers import ReceiptSerializer


def _base_queryset(user) -> Iterable[Voucher]:
    queryset = Voucher.objects.select_related(
        "company", "account", "created_by", "updated_by"
    ).order_by("-created_at")
    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()
    return queryset


def _parse_json(body: bytes) -> Optional[dict]:
    try:
        return json.loads(body.decode("utf-8")) if body else {}
    except (json.JSONDecodeError, UnicodeDecodeError):
        return None


def _flatten_dict(d: dict, parent_key: str = "", sep: str = ".") -> dict:
    items = {}
    for k, v in (d or {}).items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            # collapse simple fk dicts to name
            if "name" in v and len(v.keys()) <= 4:
                items[new_key] = v.get("name")
            else:
                items.update(_flatten_dict(v, new_key, sep=sep))
        else:
            items[new_key] = v
    return items


def _excel_safe_value(v):
    if v is None:
        return ""
    if isinstance(v, uuid.UUID):
        return str(v)
    if isinstance(v, decimal.Decimal):
        try:
            return float(v)
        except Exception:
            return str(v)
    if isinstance(v, (list, tuple)):
        try:
            return json.dumps(v)
        except Exception:
            return str(v)
    if isinstance(v, dict):
        if "name" in v:
            return v.get("name")
        try:
            return json.dumps(v)
        except Exception:
            return str(v)
    if isinstance(v, bytes):
        try:
            return v.decode("utf-8")
        except Exception:
            return repr(v)
    if isinstance(v, (str, int, float, bool)):
        return v
    return str(v)


@api_view(["GET", "POST"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def vouchers_collection_view(request):
    user = request.user
    queryset = _base_queryset(user)

    if request.method == "GET":
        status = request.GET.get("status")
        if status:
            queryset = queryset.filter(status=status)

        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(
                Q(bill_no__icontains=search) | Q(item_name__icontains=search)
            )

        # Date filtering
        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        if date_from and date_to:
            if date_from > date_to:
                return JsonResponse({"detail": "From date cannot be after To date"}, status=400)
            queryset = queryset.filter(date__gte=date_from, date__lte=date_to)
        elif date_from:
            queryset = queryset.filter(date__gte=date_from)
        elif date_to:
            queryset = queryset.filter(date__lte=date_to)

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [
            VoucherSerializer(v, context={"request": request}).data for v in queryset[start:end]
        ]

        payload = {
            "count": total,
            "next": None,
            "previous": None,
            "results": results,
        }
        return JsonResponse(payload, status=200)

    # Accept either application/json or multipart/form-data
    if request.content_type and "multipart" in request.content_type:
        # request.POST is a QueryDict; convert lists to single values where applicable
        try:
            raw = {
                k: v[0] if isinstance(v, (list, tuple)) and len(v) == 1 else v
                for k, v in request.POST.lists()
            }
        except Exception:
            raw = dict(request.POST)

        # Try to parse JSON-encoded fields (common when frontend stringifies some objects)
        for k, v in list(raw.items()):
            if isinstance(v, str) and (v.startswith("{") or v.startswith("[")):
                try:
                    raw[k] = json.loads(v)
                except Exception:
                    pass

        payload = raw
        # include uploaded file object so serializer can accept it
        if request.FILES.get("upload_file"):
            payload["upload_file"] = request.FILES.get("upload_file")
    else:
        payload = _parse_json(request.body)
        if payload is None:
            return JsonResponse({"detail": "Invalid JSON"}, status=400)

    # Normalize FK aliases
    if isinstance(payload.get("account"), dict):
        payload["account"] = payload["account"].get("id") or payload["account"].get("pk")

    # Ensure the request user has a company assigned; we set company on save.
    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    serializer = VoucherSerializer(data=payload)
    # Route to Archives when advance_payment_received == "NO"
    advance_val = payload.get("advance_payment_received")
    if isinstance(advance_val, str) and advance_val.upper() == "NO":
        # create archive record instead of active order
        archive_serializer = ArchivesSerializer(data=payload)
        if archive_serializer.is_valid():
            archive = archive_serializer.save(company=company, created_by=user)
            return JsonResponse(
                {
                    "message": "Record archived for review",
                    "result": ArchivesSerializer(archive, context={"request": request}).data,
                },
                status=201,
            )
        return JsonResponse({"errors": archive_serializer.errors}, status=400)

    if serializer.is_valid():
        voucher = serializer.save(company=company, created_by=user)
        return JsonResponse(
            {
                "message": "Order successfully created",
                "result": VoucherSerializer(voucher, context={"request": request}).data,
            },
            status=201,
        )

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def sales_return_collection_view(request):
    """List or create SalesReturn entries for the current user's company."""
    user = request.user
    queryset = SalesReturn.objects.select_related(
        "company",
        "account",
        "item_name",
        "stamp",
        "unit",
        "shape",
        "clarity",
        "created_by",
        "updated_by",
    ).order_by("-created_at")
    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    if request.method == "GET":
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(
                Q(order_no__icontains=search)
                | Q(remark__icontains=search)
                | Q(tag_no__icontains=search)
            )

        # Date filtering
        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        if date_from and date_to:
            if date_from > date_to:
                return JsonResponse({"detail": "From date cannot be after To date"}, status=400)
            queryset = queryset.filter(date__gte=date_from, date__lte=date_to)
        elif date_from:
            queryset = queryset.filter(date__gte=date_from)
        elif date_to:
            queryset = queryset.filter(date__lte=date_to)

        # Ordering
        ordering = request.GET.get("ordering")
        if ordering:
            if ordering == "date_asc":
                queryset = queryset.order_by("date")
            elif ordering == "date_desc":
                queryset = queryset.order_by("-date")
            elif ordering == "party_asc":
                queryset = queryset.order_by("account__account_name")
            elif ordering == "party_desc":
                queryset = queryset.order_by("-account__account_name")

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [SalesReturnSerializer(v).data for v in queryset[start:end]]

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    # Map incoming FK aliases to write-only *_id fields
    if isinstance(payload.get("account"), dict):
        payload["account_id"] = payload["account"].get("id") or payload["account"].get("pk")
    elif payload.get("account"):
        payload["account_id"] = payload.get("account")

    if isinstance(payload.get("item_name"), dict):
        payload["item_name_id"] = payload["item_name"].get("id") or payload["item_name"].get("pk")
    elif payload.get("item_name"):
        payload["item_name_id"] = payload.get("item_name")

    if isinstance(payload.get("stamp"), dict):
        payload["stamp_id"] = payload["stamp"].get("id") or payload["stamp"].get("pk")
    elif payload.get("stamp"):
        payload["stamp_id"] = payload.get("stamp")

    if isinstance(payload.get("unit"), dict):
        payload["unit_id"] = payload["unit"].get("id") or payload["unit"].get("pk")
    elif payload.get("unit"):
        payload["unit_id"] = payload.get("unit")

    if isinstance(payload.get("shape"), dict):
        payload["shape_id"] = payload["shape"].get("id") or payload["shape"].get("pk")
    elif payload.get("shape"):
        payload["shape_id"] = payload.get("shape")

    if isinstance(payload.get("clarity"), dict):
        payload["clarity_id"] = payload["clarity"].get("id") or payload["clarity"].get("pk")
    elif payload.get("clarity"):
        payload["clarity_id"] = payload.get("clarity")

    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    serializer = SalesReturnSerializer(data=payload)
    if serializer.is_valid():
        obj = serializer.save(company=company, created_by=user)
        return JsonResponse(SalesReturnSerializer(obj).data, status=201)

    return JsonResponse({"errors": serializer.errors}, status=400)


@jwt_login_required
@require_http_methods(["GET"])
def sales_return_overview_view(request):
    """Return aggregates and recent entries for Sales.Return overview page."""
    user = request.user
    qs = SalesReturn.objects.select_related("company", "account").order_by("-created_at")
    if user.is_authenticated and getattr(user, "company_id", None):
        qs = qs.filter(company_id=user.company_id)
    else:
        qs = qs.none()

    total_count = qs.count()
    total_piece = 0
    total_value = 0
    for rec in qs.values("net_wt", "rate", "piece"):
        nw = rec.get("net_wt") or 0
        rt = rec.get("rate") or 0
        pc = rec.get("piece") or 0
        try:
            total_piece += int(pc)
        except Exception:
            try:
                total_piece += int(float(pc))
            except Exception:
                pass
        try:
            total_value += float(rt) * float(nw)
        except Exception:
            pass

    recent_qs = qs[:5]
    recent = [SalesReturnSerializer(r).data for r in recent_qs]

    payload = {
        "total_count": total_count,
        "total_piece": total_piece,
        "total_value": total_value,
        "recent": recent,
    }
    return JsonResponse(payload, status=200)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PUT", "PATCH", "DELETE"])
def sales_return_detail_view(request, pk):
    user = request.user
    try:
        obj = SalesReturn.objects.select_related(
            "company",
            "account",
            "item_name",
            "stamp",
            "unit",
            "shape",
            "clarity",
            "created_by",
            "updated_by",
        ).get(pk=pk)
    except SalesReturn.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    if getattr(user, "company_id", None) and obj.company_id != user.company_id:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(SalesReturnSerializer(obj).data, status=200)

    if request.method == "DELETE":
        obj.delete()
        return HttpResponse(status=204)

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    # Normalize FK aliases
    if isinstance(payload.get("account"), dict):
        payload["account_id"] = payload["account"].get("id") or payload["account"].get("pk")
    elif payload.get("account"):
        payload["account_id"] = payload.get("account")

    if isinstance(payload.get("item_name"), dict):
        payload["item_name_id"] = payload["item_name"].get("id") or payload["item_name"].get("pk")
    elif payload.get("item_name"):
        payload["item_name_id"] = payload.get("item_name")

    if isinstance(payload.get("stamp"), dict):
        payload["stamp_id"] = payload["stamp"].get("id") or payload["stamp"].get("pk")
    elif payload.get("stamp"):
        payload["stamp_id"] = payload.get("stamp")

    if isinstance(payload.get("unit"), dict):
        payload["unit_id"] = payload["unit"].get("id") or payload["unit"].get("pk")
    elif payload.get("unit"):
        payload["unit_id"] = payload.get("unit")

    if isinstance(payload.get("shape"), dict):
        payload["shape_id"] = payload["shape"].get("id") or payload["shape"].get("pk")
    elif payload.get("shape"):
        payload["shape_id"] = payload.get("shape")

    if isinstance(payload.get("clarity"), dict):
        payload["clarity_id"] = payload["clarity"].get("id") or payload["clarity"].get("pk")
    elif payload.get("clarity"):
        payload["clarity_id"] = payload.get("clarity")

    serializer = SalesReturnSerializer(obj, data=payload, partial=True)
    if serializer.is_valid():
        updated = serializer.save()
        return JsonResponse(SalesReturnSerializer(updated).data, status=200)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@api_view(["GET", "PATCH", "DELETE"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def voucher_detail_view(request, pk):
    user = request.user
    queryset = _base_queryset(user)
    try:
        voucher = queryset.get(pk=pk)
    except Voucher.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return Response(VoucherSerializer(voucher, context={"request": request}).data, status=200)

    if request.method == "PATCH":
        # Accept multipart/form-data for PATCH as well (file updates)
        if request.content_type and "multipart" in request.content_type:
            try:
                raw = {
                    k: v[0] if isinstance(v, (list, tuple)) and len(v) == 1 else v
                    for k, v in request.POST.lists()
                }
            except Exception:
                raw = dict(request.POST)

            for k, v in list(raw.items()):
                if isinstance(v, str) and (v.startswith("{") or v.startswith("[")):
                    try:
                        raw[k] = json.loads(v)
                    except Exception:
                        pass

            payload = raw
            if request.FILES.get("upload_file"):
                payload["upload_file"] = request.FILES.get("upload_file")
        else:
            payload = _parse_json(request.body)
            if payload is None:
                return Response({"detail": "Invalid JSON"}, status=400)

        if isinstance(payload.get("account"), dict):
            payload["account"] = payload["account"].get("id") or payload["account"].get("pk")

        serializer = VoucherSerializer(voucher, data=payload, partial=True)
        if serializer.is_valid():
            updated = serializer.save()
            return Response(
                VoucherSerializer(updated, context={"request": request}).data, status=200
            )

        return Response({"errors": serializer.errors}, status=400)

    if request.method == "DELETE":
        # Only admin/founder can delete orders
        if not user.is_staff and not user.is_superuser:
            # Check if user has founder role
            if not hasattr(user, "department") or user.department != "FOUNDER":
                return Response({"error": "Only admin or founder can delete orders"}, status=403)

        voucher.delete()
        return Response(status=204)


@jwt_login_required
@require_http_methods(["GET"])
def voucher_history_view(request, pk):
    """Return order history timeline for the given voucher/order."""
    user = request.user
    queryset = _base_queryset(user)
    try:
        voucher = queryset.get(pk=pk)
    except Voucher.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    # Build timeline events
    timeline = []

    # Event 1: Order Created
    if voucher.created_at:
        timeline.append(
            {
                "date": voucher.created_at.date().isoformat(),
                "event": "Order Created",
                "description": f"Order {voucher.bill_no or voucher.id} was created in the system",
                "status": "created",
            }
        )

    # Event 2: Order Date (if different from created_at)
    if voucher.date and voucher.date != voucher.created_at.date():
        timeline.append(
            {
                "date": voucher.date.isoformat(),
                "event": "Order Date",
                "description": "Official order date recorded",
                "status": "order_date",
            }
        )

    # Event 3: Advance Payment
    if voucher.advance_payment_received == "YES":
        # Use created_at as proxy for advance payment date
        # In a real system, you'd have a separate field for this
        timeline.append(
            {
                "date": voucher.created_at.date().isoformat(),
                "event": "Advance Payment Received",
                "description": "Advance payment confirmed and recorded",
                "status": "advance_paid",
            }
        )

    # Event 4: Last Update (if different from created)
    if voucher.updated_at and voucher.updated_at.date() != voucher.created_at.date():
        timeline.append(
            {
                "date": voucher.updated_at.date().isoformat(),
                "event": "Order Updated",
                "description": "Order details were last modified",
                "status": "updated",
            }
        )

    # Sort timeline by date
    timeline.sort(key=lambda x: x["date"])

    # Calculate days
    from datetime import date

    today = date.today()
    days_since_order = (today - voucher.created_at.date()).days
    days_since_advance = None
    if voucher.advance_payment_received == "YES":
        days_since_advance = (today - voucher.created_at.date()).days

    # Get customer name
    customer_name = "Unknown"
    if voucher.account:
        customer_name = voucher.account.account_name or voucher.account.name or "Unknown"

    # Build response
    response_data = {
        "order_id": str(voucher.id),
        "bill_no": voucher.bill_no or str(voucher.id),
        "timeline": timeline,
        "current_status": "Active"
        if voucher.advance_payment_received == "YES"
        else "Pending Advance",
        "customer_name": customer_name,
        "item": voucher.item_name or "â€”",
        "advance_payment": voucher.advance_payment_received or "NO",
        "days_since_order": days_since_order,
        "days_since_advance": days_since_advance,
    }

    return JsonResponse(response_data, status=200)


@jwt_login_required
@require_http_methods(["GET"])
def voucher_export_view(request, pk):
    """Export a full voucher/order as an Excel file (field | value)."""
    user = request.user
    queryset = _base_queryset(user)
    try:
        voucher = queryset.get(pk=pk)
    except Voucher.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    data = VoucherSerializer(voucher, context={"request": request}).data

    wb = Workbook()
    ws = wb.active
    ws.title = "Order Details"
    row = 1

    def w(label, value):
        nonlocal row
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value if value is not None else "")
        row += 1

    # Pick common voucher fields; serializer contains all saved fields
    w("Order Number", data.get("bill_no") or data.get("voucher_number"))
    w("Date", data.get("date"))
    w(
        "Item Name",
        (data.get("item_name") or {}).get("name")
        if isinstance(data.get("item_name"), dict)
        else data.get("item_name"),
    )
    w(
        "Account",
        (data.get("account") or {}).get("name")
        if isinstance(data.get("account"), dict)
        else data.get("account"),
    )
    # include a selection of numeric/text fields if present
    for key in (
        "rate",
        "gross_wt",
        "net_wt",
        "unit",
        "tunch",
        "remark",
        "amount",
        "advance_payment_received",
    ):
        val = data.get(key)
        if isinstance(val, dict) and "name" in val:
            val = val.get("name")
        w(key.replace("_", " ").title(), val)

    # footer
    from datetime import datetime

    row += 1
    ws.cell(row=row, column=1, value="Exported On")
    ws.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    safe_name = slugify((data.get("bill_no") or data.get("voucher_number") or "order"))[:30]
    filename = f"Order_{safe_name}_Details.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@jwt_login_required
@require_http_methods(["GET"])
def vouchers_export_all_view(request):
    """Export all vouchers/orders for the user's company as an Excel file."""
    try:
        user = request.user
        queryset = _base_queryset(user)
        vouchers = list(queryset)

        rows = []
        for v in vouchers:
            data = VoucherSerializer(v, context={"request": request}).data
            # prefer human-readable account/item names
            if isinstance(data.get("account"), dict) and data.get("account").get("name"):
                data["account"] = data.get("account").get("name")
            if isinstance(data.get("item_name"), dict) and data.get("item_name").get("name"):
                data["item_name"] = data.get("item_name").get("name")
            rows.append(_flatten_dict(data))

        cols = []
        seen = set()
        for r in rows:
            for k in r.keys():
                if k not in seen:
                    cols.append(k)
                    seen.add(k)

        wb = Workbook()
        ws = wb.active
        ws.title = "Orders"

        for c_idx, col in enumerate(cols, start=1):
            header = col.replace(".", " ").replace("_", " ").title()
            ws.cell(row=1, column=c_idx, value=header)

        for r_idx, r in enumerate(rows, start=2):
            for c_idx, col in enumerate(cols, start=1):
                raw_val = r.get(col, "")
                val = _excel_safe_value(raw_val)
                ws.cell(row=r_idx, column=c_idx, value=val)

        from openpyxl.utils import get_column_letter

        for i, col in enumerate(cols, start=1):
            letter = get_column_letter(i)
            maxlen = len(col)
            for cell in ws[letter]:
                if cell.value is not None:
                    maxlen = max(maxlen, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(100, max(10, maxlen + 2))

        from datetime import datetime

        footer_row = len(rows) + 3
        ws.cell(row=footer_row, column=1, value="Exported On")
        ws.cell(row=footer_row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

        filename = "orders_data.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    except Exception as exc:
        import traceback

        tb = traceback.format_exc()
        return JsonResponse({"detail": str(exc), "trace": tb}, status=500)


@jwt_login_required
@require_http_methods(["GET"])
def voucher_receipt_view(request, pk):
    """Return a compact payload suitable for printing a receipt for the given order.

    Response shape:
    {
      "order": { ...VoucherSerializer fields... },
      "company": { "id": "..", "name": "..", "display_name": "..", "registration_number": "..", "gstin": ".." }
    }
    """
    user = request.user
    queryset = _base_queryset(user)
    try:
        voucher = queryset.get(pk=pk)
    except Voucher.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    company = getattr(user, "company", None) or getattr(voucher, "company", None)

    data = VoucherSerializer(voucher, context={"request": request}).data
    company_data = None
    if company:
        company_data = {
            "id": str(company.id),
            "name": getattr(company, "name", None),
            "display_name": getattr(company, "display_name", None),
            "registration_number": getattr(company, "registration_number", None),
            "gstin": getattr(company, "gstin", None),
        }

    # Prefer a human-readable account object in the order payload so
    # frontend receipt rendering can show the account name instead of a raw UUID.
    try:
        acct = getattr(voucher, "account", None)
        if acct is not None:
            acct_payload = {"id": str(getattr(acct, "id", ""))}
            # Account model uses `account_name`; map it to `name` for frontend compatibility
            acct_name = (
                getattr(acct, "account_name", None)
                or getattr(acct, "name", None)
                or getattr(acct, "display_name", None)
            )
            if acct_name:
                acct_payload["name"] = acct_name
            # also include account_no if available (useful fallback)
            acct_no = getattr(acct, "account_no", None)
            if acct_no:
                acct_payload["account_no"] = acct_no
            data["account"] = acct_payload
    except Exception:
        # best-effort: if anything goes wrong, leave serializer data as-is
        pass

    return JsonResponse({"order": data, "company": company_data}, status=200)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def pur_and_approval_collection_view(request):
    """List or create PurAndApprovalM entries for the current user's company."""
    user = request.user
    queryset = PurAndApprovalM.objects.select_related(
        "company", "account", "item_name", "unit", "shape", "created_by", "updated_by"
    ).order_by("-created_at")
    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    if request.method == "GET":
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(
                Q(order_no__icontains=search)
                | Q(remark__icontains=search)
                | Q(tag_no__icontains=search)
            )

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [PurAndApprovalMSerializer(v).data for v in queryset[start:end]]

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    # Map incoming FK aliases to write-only *_id fields
    if isinstance(payload.get("account"), dict):
        payload["account_id"] = payload["account"].get("id") or payload["account"].get("pk")
    elif payload.get("account"):
        payload["account_id"] = payload.get("account")

    if isinstance(payload.get("item_name"), dict):
        payload["item_name_id"] = payload["item_name"].get("id") or payload["item_name"].get("pk")
    elif payload.get("item_name"):
        payload["item_name_id"] = payload.get("item_name")

    if isinstance(payload.get("unit"), dict):
        payload["unit_id"] = payload["unit"].get("id") or payload["unit"].get("pk")
    elif payload.get("unit"):
        payload["unit_id"] = payload.get("unit")

    if isinstance(payload.get("shape"), dict):
        payload["shape_id"] = payload["shape"].get("id") or payload["shape"].get("pk")
    elif payload.get("shape"):
        payload["shape_id"] = payload.get("shape")

    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    serializer = PurAndApprovalMSerializer(data=payload)
    if serializer.is_valid():
        obj = serializer.save(company=company, created_by=user)
        return JsonResponse(PurAndApprovalMSerializer(obj).data, status=201)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PATCH", "DELETE"])
def pur_and_approval_detail_view(request, pk):
    user = request.user
    try:
        obj = PurAndApprovalM.objects.select_related(
            "company", "account", "item_name", "unit", "shape", "created_by", "updated_by"
        ).get(pk=pk)
    except PurAndApprovalM.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    if getattr(user, "company_id", None) and obj.company_id != user.company_id:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(PurAndApprovalMSerializer(obj).data, status=200)

    if request.method == "DELETE":
        obj.delete()
        return HttpResponse(status=204)

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    if isinstance(payload.get("account"), dict):
        payload["account_id"] = payload["account"].get("id") or payload["account"].get("pk")
    elif payload.get("account"):
        payload["account_id"] = payload.get("account")

    if isinstance(payload.get("item_name"), dict):
        payload["item_name_id"] = payload["item_name"].get("id") or payload["item_name"].get("pk")
    elif payload.get("item_name"):
        payload["item_name_id"] = payload.get("item_name")

    if isinstance(payload.get("unit"), dict):
        payload["unit_id"] = payload["unit"].get("id") or payload["unit"].get("pk")
    elif payload.get("unit"):
        payload["unit_id"] = payload.get("unit")

    if isinstance(payload.get("shape"), dict):
        payload["shape_id"] = payload["shape"].get("id") or payload["shape"].get("pk")
    elif payload.get("shape"):
        payload["shape_id"] = payload.get("shape")

    serializer = PurAndApprovalMSerializer(obj, data=payload, partial=True)
    if serializer.is_valid():
        updated = serializer.save()
        return JsonResponse(PurAndApprovalMSerializer(updated).data, status=200)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def pur_return_collection_view(request):
    """List or create PurReturn entries for the current user's company."""
    user = request.user
    queryset = PurReturn.objects.select_related(
        "company",
        "account",
        "item_name",
        "stamp",
        "unit",
        "shape",
        "clarity",
        "created_by",
        "updated_by",
    ).order_by("-created_at")
    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    if request.method == "GET":
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(
                Q(order_no__icontains=search)
                | Q(remark__icontains=search)
                | Q(tag_no__icontains=search)
            )

        # Date filtering
        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        if date_from and date_to:
            if date_from > date_to:
                return JsonResponse({"detail": "From date cannot be after To date"}, status=400)
            queryset = queryset.filter(date__gte=date_from, date__lte=date_to)
        elif date_from:
            queryset = queryset.filter(date__gte=date_from)
        elif date_to:
            queryset = queryset.filter(date__lte=date_to)

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [PurReturnSerializer(v).data for v in queryset[start:end]]

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    # Map incoming FK aliases to write-only *_id fields
    if isinstance(payload.get("account"), dict):
        payload["account_id"] = payload["account"].get("id") or payload["account"].get("pk")
    elif payload.get("account"):
        payload["account_id"] = payload.get("account")

    if isinstance(payload.get("item_name"), dict):
        payload["item_name_id"] = payload["item_name"].get("id") or payload["item_name"].get("pk")
    elif payload.get("item_name"):
        payload["item_name_id"] = payload.get("item_name")

    if isinstance(payload.get("stamp"), dict):
        payload["stamp_id"] = payload["stamp"].get("id") or payload["stamp"].get("pk")
    elif payload.get("stamp"):
        payload["stamp_id"] = payload.get("stamp")

    if isinstance(payload.get("unit"), dict):
        payload["unit_id"] = payload["unit"].get("id") or payload["unit"].get("pk")
    elif payload.get("unit"):
        payload["unit_id"] = payload.get("unit")

    if isinstance(payload.get("shape"), dict):
        payload["shape_id"] = payload["shape"].get("id") or payload["shape"].get("pk")
    elif payload.get("shape"):
        payload["shape_id"] = payload.get("shape")

    if isinstance(payload.get("clarity"), dict):
        payload["clarity_id"] = payload["clarity"].get("id") or payload["clarity"].get("pk")
    elif payload.get("clarity"):
        payload["clarity_id"] = payload.get("clarity")

    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    serializer = PurReturnSerializer(data=payload)
    if serializer.is_valid():
        obj = serializer.save(company=company, created_by=user)
        return JsonResponse(PurReturnSerializer(obj).data, status=201)

    return JsonResponse({"errors": serializer.errors}, status=400)


@jwt_login_required
@require_http_methods(["GET"])
def pur_return_overview_view(request):
    """Return aggregates and recent entries for Pur.Return overview page."""
    user = request.user
    qs = PurReturn.objects.select_related("company", "account").order_by("-created_at")
    if user.is_authenticated and getattr(user, "company_id", None):
        qs = qs.filter(company_id=user.company_id)
    else:
        qs = qs.none()

    total_count = qs.count()
    # sum piece and estimate total value as sum(rate * net_wt) where available
    total_piece = 0
    total_value = 0
    for rec in qs.values("net_wt", "rate", "piece"):
        nw = rec.get("net_wt") or 0
        rt = rec.get("rate") or 0
        pc = rec.get("piece") or 0
        try:
            # piece is expected to be an integer count
            total_piece += int(pc)
        except Exception:
            try:
                total_piece += int(float(pc))
            except Exception:
                pass
        try:
            total_value += float(rt) * float(nw)
        except Exception:
            pass

    recent_qs = qs[:5]
    recent = [PurReturnSerializer(r).data for r in recent_qs]

    payload = {
        "total_count": total_count,
        "total_piece": total_piece,
        "total_value": total_value,
        "recent": recent,
    }
    return JsonResponse(payload, status=200)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PUT", "PATCH", "DELETE"])
def pur_return_detail_view(request, pk):
    user = request.user
    try:
        obj = PurReturn.objects.select_related(
            "company",
            "account",
            "item_name",
            "stamp",
            "unit",
            "shape",
            "clarity",
            "created_by",
            "updated_by",
        ).get(pk=pk)
    except PurReturn.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    if getattr(user, "company_id", None) and obj.company_id != user.company_id:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(PurReturnSerializer(obj).data, status=200)

    if request.method == "DELETE":
        obj.delete()
        return HttpResponse(status=204)

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    # Normalize FK aliases
    if isinstance(payload.get("account"), dict):
        payload["account_id"] = payload["account"].get("id") or payload["account"].get("pk")
    elif payload.get("account"):
        payload["account_id"] = payload.get("account")

    if isinstance(payload.get("item_name"), dict):
        payload["item_name_id"] = payload["item_name"].get("id") or payload["item_name"].get("pk")
    elif payload.get("item_name"):
        payload["item_name_id"] = payload.get("item_name")

    if isinstance(payload.get("stamp"), dict):
        payload["stamp_id"] = payload["stamp"].get("id") or payload["stamp"].get("pk")
    elif payload.get("stamp"):
        payload["stamp_id"] = payload.get("stamp")

    if isinstance(payload.get("unit"), dict):
        payload["unit_id"] = payload["unit"].get("id") or payload["unit"].get("pk")
    elif payload.get("unit"):
        payload["unit_id"] = payload.get("unit")

    if isinstance(payload.get("shape"), dict):
        payload["shape_id"] = payload["shape"].get("id") or payload["shape"].get("pk")
    elif payload.get("shape"):
        payload["shape_id"] = payload.get("shape")

    if isinstance(payload.get("clarity"), dict):
        payload["clarity_id"] = payload["clarity"].get("id") or payload["clarity"].get("pk")
    elif payload.get("clarity"):
        payload["clarity_id"] = payload.get("clarity")

    serializer = PurReturnSerializer(obj, data=payload, partial=True)
    if serializer.is_valid():
        updated = serializer.save()
        return JsonResponse(PurReturnSerializer(updated).data, status=200)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def receive_collection_view(request):
    """List or create Receive entries for the current user's company."""
    user = request.user
    queryset = Receive.objects.select_related(
        "company", "account", "item_name", "stamp", "unit", "created_by", "updated_by"
    ).order_by("-created_at")
    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    if request.method == "GET":
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(
                Q(tag_no__icontains=search)
                | Q(remark__icontains=search)
                | Q(item_name__icontains=search)
            )

        # support date range filtering
        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        if date_from and date_to:
            if date_from > date_to:
                return JsonResponse({"detail": "From date cannot be after To date"}, status=400)
            queryset = queryset.filter(date__gte=date_from, date__lte=date_to)
        elif date_from:
            queryset = queryset.filter(date__gte=date_from)
        elif date_to:
            queryset = queryset.filter(date__lte=date_to)

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [ReceiveSerializer(v).data for v in queryset[start:end]]

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    # Map incoming FK aliases to write-only *_id fields
    if isinstance(payload.get("account"), dict):
        payload["account_id"] = payload["account"].get("id") or payload["account"].get("pk")
    elif payload.get("account"):
        payload["account_id"] = payload.get("account")

    if isinstance(payload.get("item_name"), dict):
        payload["item_name_id"] = payload["item_name"].get("id") or payload["item_name"].get("pk")
    elif payload.get("item_name"):
        payload["item_name_id"] = payload.get("item_name")

    if isinstance(payload.get("stamp"), dict):
        payload["stamp_id"] = payload["stamp"].get("id") or payload["stamp"].get("pk")
    elif payload.get("stamp"):
        payload["stamp_id"] = payload.get("stamp")

    if isinstance(payload.get("unit"), dict):
        payload["unit_id"] = payload["unit"].get("id") or payload["unit"].get("pk")
    elif payload.get("unit"):
        payload["unit_id"] = payload.get("unit")

    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    serializer = ReceiveSerializer(data=payload)
    if serializer.is_valid():
        obj = serializer.save(company=company, created_by=user)
        return JsonResponse(ReceiveSerializer(obj).data, status=201)

    return JsonResponse({"errors": serializer.errors}, status=400)


@jwt_login_required
@require_http_methods(["GET"])
def receive_overview_view(request):
    """Return aggregates and recent entries for Receive overview page."""
    user = request.user
    qs = Receive.objects.select_related("company", "account").order_by("-created_at")
    if user.is_authenticated and getattr(user, "company_id", None):
        qs = qs.filter(company_id=user.company_id)
    else:
        qs = qs.none()

    total_count = qs.count()
    total_piece = 0
    total_value = 0
    for rec in qs.values("net_wt", "rate", "pc"):
        nw = rec.get("net_wt") or 0
        rt = rec.get("rate") or 0
        pc = rec.get("pc") or 0
        try:
            total_piece += int(pc)
        except Exception:
            try:
                total_piece += int(float(pc))
            except Exception:
                pass
        try:
            total_value += float(rt) * float(nw)
        except Exception:
            pass

    recent_qs = qs[:5]
    recent = [ReceiveSerializer(r).data for r in recent_qs]

    payload = {
        "total_count": total_count,
        "total_piece": total_piece,
        "total_value": total_value,
        "recent": recent,
    }
    return JsonResponse(payload, status=200)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PATCH", "DELETE"])
def receive_detail_view(request, pk):
    user = request.user
    try:
        obj = Receive.objects.select_related(
            "company", "account", "item_name", "stamp", "unit", "created_by", "updated_by"
        ).get(pk=pk)
    except Receive.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    if getattr(user, "company_id", None) and obj.company_id != user.company_id:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(ReceiveSerializer(obj).data, status=200)

    if request.method == "DELETE":
        obj.delete()
        return HttpResponse(status=204)

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    # Normalize FK aliases
    if isinstance(payload.get("account"), dict):
        payload["account_id"] = payload["account"].get("id") or payload["account"].get("pk")
    elif payload.get("account"):
        payload["account_id"] = payload.get("account")

    if isinstance(payload.get("item_name"), dict):
        payload["item_name_id"] = payload["item_name"].get("id") or payload["item_name"].get("pk")
    elif payload.get("item_name"):
        payload["item_name_id"] = payload.get("item_name")

    if isinstance(payload.get("stamp"), dict):
        payload["stamp_id"] = payload["stamp"].get("id") or payload["stamp"].get("pk")
    elif payload.get("stamp"):
        payload["stamp_id"] = payload.get("stamp")

    if isinstance(payload.get("unit"), dict):
        payload["unit_id"] = payload["unit"].get("id") or payload["unit"].get("pk")
    elif payload.get("unit"):
        payload["unit_id"] = payload.get("unit")

    serializer = ReceiveSerializer(obj, data=payload, partial=True)
    if serializer.is_valid():
        updated = serializer.save()
        return JsonResponse(ReceiveSerializer(updated).data, status=200)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def purchase_m_collection_view(request):
    """List or create PurchaseM entries for the current user's company."""
    user = request.user
    try:
        from .models import PurchaseM
    except Exception:
        PurchaseM = None

    queryset = (
        PurchaseM.objects.select_related(
            "company", "account", "item_name", "stamp", "unit", "created_by", "updated_by"
        ).order_by("-created_at")
        if PurchaseM
        else []
    )
    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none() if hasattr(queryset, "none") else []

    if request.method == "GET":
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(
                Q(tag_no__icontains=search)
                | Q(order_no__icontains=search)
                | Q(remark__icontains=search)
            )

        # Ordering
        ordering = request.GET.get("ordering")
        if ordering:
            if ordering == "date_asc":
                queryset = queryset.order_by("date")
            elif ordering == "date_desc":
                queryset = queryset.order_by("-date")
            elif ordering == "party_asc":
                queryset = queryset.order_by("account__account_name")
            elif ordering == "party_desc":
                queryset = queryset.order_by("-account__account_name")

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count() if hasattr(queryset, "count") else 0
        start = (page - 1) * page_size
        end = start + page_size
        results = (
            [PurchaseMSerializer(v).data for v in queryset[start:end]]
            if hasattr(queryset, "__iter__")
            else []
        )

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    # Map incoming FK aliases to write-only *_id fields
    if isinstance(payload.get("account"), dict):
        payload["account_id"] = payload["account"].get("id") or payload["account"].get("pk")
    elif payload.get("account"):
        payload["account_id"] = payload.get("account")

    if isinstance(payload.get("item_name"), dict):
        payload["item_name_id"] = payload["item_name"].get("id") or payload["item_name"].get("pk")
    elif payload.get("item_name"):
        payload["item_name_id"] = payload.get("item_name")

    if isinstance(payload.get("stamp"), dict):
        payload["stamp_id"] = payload["stamp"].get("id") or payload["stamp"].get("pk")
    elif payload.get("stamp"):
        payload["stamp_id"] = payload.get("stamp")

    if isinstance(payload.get("unit"), dict):
        payload["unit_id"] = payload["unit"].get("id") or payload["unit"].get("pk")
    elif payload.get("unit"):
        payload["unit_id"] = payload.get("unit")

    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    serializer = PurchaseMSerializer(data=payload)
    if serializer.is_valid():
        obj = serializer.save(company=company, created_by=user)
        return JsonResponse(PurchaseMSerializer(obj).data, status=201)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PATCH", "DELETE"])
def purchase_m_detail_view(request, pk):
    user = request.user
    try:
        from .models import PurchaseM
    except Exception:
        PurchaseM = None

    try:
        obj = PurchaseM.objects.select_related(
            "company", "account", "item_name", "stamp", "unit", "created_by", "updated_by"
        ).get(pk=pk)
    except Exception:
        return JsonResponse({"detail": "Not found"}, status=404)

    if getattr(user, "company_id", None) and obj.company_id != user.company_id:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(PurchaseMSerializer(obj).data, status=200)

    if request.method == "DELETE":
        obj.delete()
        return HttpResponse(status=204)

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    if isinstance(payload.get("account"), dict):
        payload["account_id"] = payload["account"].get("id") or payload["account"].get("pk")
    elif payload.get("account"):
        payload["account_id"] = payload.get("account")

    if isinstance(payload.get("item_name"), dict):
        payload["item_name_id"] = payload["item_name"].get("id") or payload["item_name"].get("pk")
    elif payload.get("item_name"):
        payload["item_name_id"] = payload.get("item_name")

    if isinstance(payload.get("stamp"), dict):
        payload["stamp_id"] = payload["stamp"].get("id") or payload["stamp"].get("pk")
    elif payload.get("stamp"):
        payload["stamp_id"] = payload.get("stamp")

    if isinstance(payload.get("unit"), dict):
        payload["unit_id"] = payload["unit"].get("id") or payload["unit"].get("pk")
    elif payload.get("unit"):
        payload["unit_id"] = payload.get("unit")

    serializer = PurchaseMSerializer(obj, data=payload, partial=True)
    if serializer.is_valid():
        updated = serializer.save()
        return JsonResponse(PurchaseMSerializer(updated).data, status=200)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def purchase_tagwise_collection_view(request):
    """List or create PurchaseTagwiseM entries for the current user's company."""
    user = request.user
    try:
        from .models import PurchaseTagwiseM
    except Exception:
        PurchaseTagwiseM = None

    queryset = (
        PurchaseTagwiseM.objects.select_related(
            "company", "account", "item_name", "stamp", "unit", "created_by", "updated_by"
        ).order_by("-created_at")
        if PurchaseTagwiseM
        else []
    )
    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none() if hasattr(queryset, "none") else []

    if request.method == "GET":
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(
                Q(order_no__icontains=search)
                | Q(remark__icontains=search)
                | Q(design__icontains=search)
            )

        # Ordering
        ordering = request.GET.get("ordering")
        if ordering:
            if ordering == "date_asc":
                queryset = queryset.order_by("date")
            elif ordering == "date_desc":
                queryset = queryset.order_by("-date")
            elif ordering == "party_asc":
                queryset = queryset.order_by("account__account_name")
            elif ordering == "party_desc":
                queryset = queryset.order_by("-account__account_name")

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count() if hasattr(queryset, "count") else 0
        start = (page - 1) * page_size
        end = start + page_size
        results = (
            [PurchaseTagwiseMSerializer(v).data for v in queryset[start:end]]
            if hasattr(queryset, "__iter__")
            else []
        )

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    if isinstance(payload.get("account"), dict):
        payload["account_id"] = payload["account"].get("id") or payload["account"].get("pk")
    elif payload.get("account"):
        payload["account_id"] = payload.get("account")

    if isinstance(payload.get("item_name"), dict):
        payload["item_name_id"] = payload["item_name"].get("id") or payload["item_name"].get("pk")
    elif payload.get("item_name"):
        payload["item_name_id"] = payload.get("item_name")

    if isinstance(payload.get("stamp"), dict):
        payload["stamp_id"] = payload["stamp"].get("id") or payload["stamp"].get("pk")
    elif payload.get("stamp"):
        payload["stamp_id"] = payload.get("stamp")

    if isinstance(payload.get("unit"), dict):
        payload["unit_id"] = payload["unit"].get("id") or payload["unit"].get("pk")
    elif payload.get("unit"):
        payload["unit_id"] = payload.get("unit")

    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    serializer = PurchaseTagwiseMSerializer(data=payload)
    if serializer.is_valid():
        obj = serializer.save(company=company, created_by=user)
        return JsonResponse(PurchaseTagwiseMSerializer(obj).data, status=201)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PATCH", "DELETE"])
def purchase_tagwise_detail_view(request, pk):
    user = request.user
    try:
        from .models import PurchaseTagwiseM
    except Exception:
        PurchaseTagwiseM = None

    try:
        obj = PurchaseTagwiseM.objects.select_related(
            "company", "account", "item_name", "stamp", "unit", "created_by", "updated_by"
        ).get(pk=pk)
    except Exception:
        return JsonResponse({"detail": "Not found"}, status=404)

    if getattr(user, "company_id", None) and obj.company_id != user.company_id:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(PurchaseTagwiseMSerializer(obj).data, status=200)

    if request.method == "DELETE":
        obj.delete()
        return HttpResponse(status=204)

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    if isinstance(payload.get("account"), dict):
        payload["account_id"] = payload["account"].get("id") or payload["account"].get("pk")
    elif payload.get("account"):
        payload["account_id"] = payload.get("account")

    if isinstance(payload.get("item_name"), dict):
        payload["item_name_id"] = payload["item_name"].get("id") or payload["item_name"].get("pk")
    elif payload.get("item_name"):
        payload["item_name_id"] = payload.get("item_name")

    if isinstance(payload.get("stamp"), dict):
        payload["stamp_id"] = payload["stamp"].get("id") or payload["stamp"].get("pk")
    elif payload.get("stamp"):
        payload["stamp_id"] = payload.get("stamp")

    if isinstance(payload.get("unit"), dict):
        payload["unit_id"] = payload["unit"].get("id") or payload["unit"].get("pk")
    elif payload.get("unit"):
        payload["unit_id"] = payload.get("unit")

    serializer = PurchaseTagwiseMSerializer(obj, data=payload, partial=True)
    if serializer.is_valid():
        updated = serializer.save()
        return JsonResponse(PurchaseTagwiseMSerializer(updated).data, status=200)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def purchase_diamond_collection_view(request):
    """List or create PurchaseDiamond entries for the current user's company."""
    user = request.user
    queryset = PurchaseDiamond.objects.select_related(
        "company",
        "account",
        "item_name",
        "shape",
        "size",
        "colour",
        "clarity",
        "lab",
        "created_by",
        "updated_by",
    ).order_by("-created_at")
    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    if request.method == "GET":
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(
                Q(batch__icontains=search)
                | Q(remark__icontains=search)
                | Q(item_name__icontains=search)
            )

        # Ordering
        ordering = request.GET.get("ordering")
        if ordering:
            if ordering == "date_asc":
                queryset = queryset.order_by("date")
            elif ordering == "date_desc":
                queryset = queryset.order_by("-date")
            elif ordering == "party_asc":
                queryset = queryset.order_by("account__account_name")
            elif ordering == "party_desc":
                queryset = queryset.order_by("-account__account_name")

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [PurchaseDiamondSerializer(v).data for v in queryset[start:end]]

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    # Map incoming FK aliases to write-only *_id fields
    if isinstance(payload.get("account"), dict):
        payload["account_id"] = payload["account"].get("id") or payload["account"].get("pk")
    elif payload.get("account"):
        payload["account_id"] = payload.get("account")

    if isinstance(payload.get("item_name"), dict):
        payload["item_name_id"] = payload["item_name"].get("id") or payload["item_name"].get("pk")
    elif payload.get("item_name"):
        payload["item_name_id"] = payload.get("item_name")

    if isinstance(payload.get("shape"), dict):
        payload["shape_id"] = payload["shape"].get("id") or payload["shape"].get("pk")
    elif payload.get("shape"):
        payload["shape_id"] = payload.get("shape")

    if isinstance(payload.get("size"), dict):
        payload["size_id"] = payload["size"].get("id") or payload["size"].get("pk")
    elif payload.get("size"):
        payload["size_id"] = payload.get("size")

    if isinstance(payload.get("colour"), dict):
        payload["colour_id"] = payload["colour"].get("id") or payload["colour"].get("pk")
    elif payload.get("colour"):
        payload["colour_id"] = payload.get("colour")

    if isinstance(payload.get("clarity"), dict):
        payload["clarity_id"] = payload["clarity"].get("id") or payload["clarity"].get("pk")
    elif payload.get("clarity"):
        payload["clarity_id"] = payload.get("clarity")

    if isinstance(payload.get("lab"), dict):
        payload["lab_id"] = payload["lab"].get("id") or payload["lab"].get("pk")
    elif payload.get("lab"):
        payload["lab_id"] = payload.get("lab")

    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    serializer = PurchaseDiamondSerializer(data=payload)
    if serializer.is_valid():
        obj = serializer.save(company=company, created_by=user)
        return JsonResponse(PurchaseDiamondSerializer(obj).data, status=201)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PATCH", "DELETE"])
def purchase_diamond_detail_view(request, pk):
    user = request.user
    try:
        obj = PurchaseDiamond.objects.select_related(
            "company",
            "account",
            "item_name",
            "shape",
            "size",
            "colour",
            "clarity",
            "lab",
            "created_by",
            "updated_by",
        ).get(pk=pk)
    except PurchaseDiamond.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    if getattr(user, "company_id", None) and obj.company_id != user.company_id:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(PurchaseDiamondSerializer(obj).data, status=200)

    if request.method == "DELETE":
        obj.delete()
        return HttpResponse(status=204)

    # PATCH
    payload = _parse_json(request.body)
    if payload is None:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    serializer = PurchaseDiamondSerializer(obj, data=payload, partial=True)
    if serializer.is_valid():
        updated = serializer.save(updated_by=user)
        return JsonResponse(PurchaseDiamondSerializer(updated).data, status=200)

    return JsonResponse({"errors": serializer.errors}, status=400)


@api_view(["GET", "POST"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def sales_collection_view(request):
    user = request.user
    queryset = Sale.objects.select_related(
        "company", "account", "created_by", "updated_by"
    ).order_by("-created_at")
    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    # Exclude sales that have been converted to orders
    # A sale is "converted" if it has an OrderDraft with status='confirmed' and a final_order
    from vouchers.models import OrderDraft

    converted_sale_ids = OrderDraft.objects.filter(
        status="confirmed", final_order__isnull=False
    ).values_list("source_sale_id", flat=True)
    queryset = queryset.exclude(id__in=converted_sale_ids)

    if request.method == "GET":
        # simple list with optional search
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(
                Q(order_no__icontains=search) | Q(item_name__icontains=search)
            )

        # Date filtering (using created_at since Sale model doesn't have date field)
        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        if date_from and date_to:
            if date_from > date_to:
                return Response({"detail": "From date cannot be after To date"}, status=400)
            queryset = queryset.filter(
                created_at__date__gte=date_from, created_at__date__lte=date_to
            )
        elif date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        elif date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)

        # Ordering
        ordering = request.GET.get("ordering")
        if ordering:
            if ordering == "date_asc":
                queryset = queryset.order_by("created_at")
            elif ordering == "date_desc":
                queryset = queryset.order_by("-created_at")
            elif ordering == "party_asc":
                queryset = queryset.order_by("account__account_name")
            elif ordering == "party_desc":
                queryset = queryset.order_by("-account__account_name")

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [SaleSerializer(s).data for s in queryset[start:end]]

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return Response(payload, status=200)

    payload = _parse_json(request.body)
    if payload is None:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    if isinstance(payload.get("account"), dict):
        payload["account"] = payload["account"].get("id") or payload["account"].get("pk")

    company = getattr(user, "company", None)
    if company is None:
        return Response({"errors": {"company": ["User does not belong to a company"]}}, status=400)

    serializer = SaleSerializer(data=payload)
    if serializer.is_valid():
        sale = serializer.save(company=company, created_by=user)
        return Response(SaleSerializer(sale).data, status=201)

    return Response({"errors": serializer.errors}, status=400)


@api_view(["GET", "PATCH", "DELETE"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def sale_detail_view(request, pk):
    user = request.user
    queryset = Sale.objects.select_related(
        "company", "account", "created_by", "updated_by"
    ).order_by("-created_at")
    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    try:
        sale = queryset.get(pk=pk)
    except Sale.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return Response(SaleSerializer(sale).data, status=200)

    elif request.method == "PATCH":
        payload = _parse_json(request.body)
        if payload is None:
            return Response({"detail": "Invalid JSON"}, status=400)

        if isinstance(payload.get("account"), dict):
            payload["account"] = payload["account"].get("id") or payload["account"].get("pk")

        serializer = SaleSerializer(sale, data=payload, partial=True)
        if serializer.is_valid():
            updated_sale = serializer.save(updated_by=user)
            return Response(SaleSerializer(updated_sale).data, status=200)

        return Response({"errors": serializer.errors}, status=400)

    elif request.method == "DELETE":
        sale.delete()
        return Response({"detail": "Sale deleted successfully"}, status=204)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET"])
def sale_estimates_view(request, pk):
    """Get all estimates linked to a sale."""
    user = request.user
    company = getattr(user, "company", None)

    if not company:
        return JsonResponse({"error": "User does not belong to a company"}, status=400)

    try:
        sale = Sale.objects.get(pk=pk, company=company)
    except Sale.DoesNotExist:
        return JsonResponse({"detail": "Sale not found"}, status=404)

    # Get estimates linked to this sale
    estimates = EstimateVoucher.objects.filter(company=company, sale=sale).order_by("-created_at")

    # Also get estimates matching the item_name (for backward compatibility)
    # Use exact word matching to avoid "Ring" matching "Earring"
    if sale.item_name:
        all_type_estimates = EstimateVoucher.objects.filter(
            company=company,
            sale__isnull=True,
            sales_query__isnull=True,
        ).order_by("-created_at")

        # Filter by exact jewelry type match (case-insensitive)
        item_name_lower = sale.item_name.lower()
        type_matched_list = []
        for estimate in all_type_estimates:
            if estimate.item_name:
                # Check if the jewelry type matches as a whole word
                estimate_name_lower = estimate.item_name.lower()
                # Split by common separators and check if jewelry type is one of the words
                estimate_words = estimate_name_lower.replace("-", " ").replace("_", " ").split()
                if (
                    item_name_lower in estimate_words
                    or estimate_name_lower.startswith(item_name_lower + " ")
                    or estimate_name_lower == item_name_lower
                ):
                    type_matched_list.append(estimate)

        # Convert list back to queryset-like behavior
        type_matched_ids = [est.id for est in type_matched_list]
        type_matched = EstimateVoucher.objects.filter(id__in=type_matched_ids).order_by(
            "-created_at"
        )
    else:
        type_matched = EstimateVoucher.objects.none()

    # Combine and serialize
    all_estimate_ids = set()
    estimates_data = []

    for estimate in estimates:
        if estimate.id not in all_estimate_ids:
            all_estimate_ids.add(estimate.id)
            est_data = EstimateVoucherSerializer(estimate).data
            est_data["is_linked"] = True
            estimates_data.append(est_data)

    for estimate in type_matched:
        if estimate.id not in all_estimate_ids:
            all_estimate_ids.add(estimate.id)
            est_data = EstimateVoucherSerializer(estimate).data
            est_data["is_linked"] = False
            estimates_data.append(est_data)

    return JsonResponse(
        {
            "sale_id": str(sale.id),
            "item_name": sale.item_name,
            "selected_estimate_id": str(sale.selected_estimate_id)
            if sale.selected_estimate_id
            else None,
            "all_estimates": estimates_data,
            "estimates_count": len(estimates_data),
        },
        status=200,
    )


@csrf_exempt
@jwt_login_required
@require_http_methods(["POST"])
def sale_select_estimate_view(request, pk):
    """Select an estimate for a sale (marks it as primary)."""
    user = request.user
    company = getattr(user, "company", None)

    if not company:
        return JsonResponse({"error": "User does not belong to a company"}, status=400)

    try:
        sale = Sale.objects.get(pk=pk, company=company)
    except Sale.DoesNotExist:
        return JsonResponse({"detail": "Sale not found"}, status=404)

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    estimate_id = payload.get("estimate_id")
    if not estimate_id:
        return JsonResponse({"detail": "estimate_id is required"}, status=400)

    try:
        estimate = EstimateVoucher.objects.get(pk=estimate_id, company=company)
    except EstimateVoucher.DoesNotExist:
        return JsonResponse({"detail": "Estimate not found"}, status=404)

    # Verify estimate is linked to this sale
    if estimate.sale_id != sale.id:
        return JsonResponse({"detail": "Estimate is not linked to this sale"}, status=400)

    # Update all estimates for this sale: set selected one to 'selected', others to 'draft'
    EstimateVoucher.objects.filter(sale=sale).exclude(pk=estimate_id).update(status="draft")
    estimate.status = "selected"
    estimate.save()

    # Update sale's selected_estimate
    sale.selected_estimate = estimate
    sale.save()

    return JsonResponse(
        {
            "message": "Estimate selected successfully",
            "sale_id": str(sale.id),
            "selected_estimate_id": str(estimate.id),
        },
        status=200,
    )


@csrf_exempt
@jwt_login_required
@require_http_methods(["POST"])
def sale_deselect_estimate_view(request, pk):
    """Deselect the currently selected estimate for a sale."""
    user = request.user
    company = getattr(user, "company", None)

    if not company:
        return JsonResponse({"error": "User does not belong to a company"}, status=400)

    try:
        sale = Sale.objects.get(pk=pk, company=company)
    except Sale.DoesNotExist:
        return JsonResponse({"detail": "Sale not found"}, status=404)

    if not sale.selected_estimate:
        return JsonResponse({"detail": "No estimate is currently selected"}, status=400)

    # Reset the selected estimate's status to draft
    selected_estimate = sale.selected_estimate
    selected_estimate.status = "draft"
    selected_estimate.save()

    # Clear the selection
    sale.selected_estimate = None
    sale.save()

    return JsonResponse(
        {
            "message": "Estimate deselected successfully",
            "sale_id": str(sale.id),
        },
        status=200,
    )


@csrf_exempt
@jwt_login_required
@require_http_methods(["POST"])
def sale_convert_to_order_view(request, pk):
    """Convert the selected estimate to an Order."""
    user = request.user
    company = getattr(user, "company", None)

    if not company:
        return JsonResponse({"error": "User does not belong to a company"}, status=400)

    try:
        sale = Sale.objects.get(pk=pk, company=company)
    except Sale.DoesNotExist:
        return JsonResponse({"detail": "Sale not found"}, status=404)

    if not sale.selected_estimate:
        return JsonResponse(
            {"detail": "No estimate is selected. Please select an estimate first."}, status=400
        )

    estimate = sale.selected_estimate

    # Create Order from estimate data
    order = Order.objects.create(
        company=company,
        account=sale.account,
        item_name=estimate.item_name,
        date=estimate.date,
        created_by=user,
        order_type="STOCK_JEWELRY",
        # Copy relevant fields from estimate
        gold_rate=0,  # Can be updated later
        number_of_pieces=1,
    )

    # Update estimate status to indicate it's been converted
    estimate.status = "selected"  # Keep as selected since it was converted
    estimate.save()

    return JsonResponse(
        {
            "message": "Order created successfully from estimate",
            "sale_id": str(sale.id),
            "order_id": str(order.id),
            "order_bill_no": order.bill_no,
            "order_job_no": order.job_no,
        },
        status=201,
    )


@jwt_login_required
@require_http_methods(["GET"])
def sale_export_view(request, pk):
    """Export a single sale record as an Excel (.xlsx) file."""
    user = request.user
    queryset = Sale.objects.select_related("company", "account").order_by("-created_at")
    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    try:
        sale = queryset.get(pk=pk)
    except Sale.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    data = SaleSerializer(sale).data

    wb = Workbook()
    ws = wb.active
    ws.title = "Sale Details"
    row = 1

    def w(label, value):
        nonlocal row
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value if value is not None else "")
        row += 1

    w("Invoice/Tag No", data.get("tag_no") or data.get("order_no") or data.get("bill_no"))
    w("Date", data.get("date"))
    w("Item Name", data.get("item_name"))
    w("Order No", data.get("order_no"))
    w("Unit", data.get("unit"))
    w("Net Wt", data.get("net_wt"))
    w("Amount", data.get("amount"))
    w("Remarks", data.get("remark"))

    from datetime import datetime

    row += 1
    ws.cell(row=row, column=1, value="Exported On")
    ws.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    safe_name = slugify((data.get("tag_no") or data.get("order_no") or "sale"))[:30]
    filename = f"Sale_{safe_name}_Details.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# =============================================================================
# Sales PDF Generation Endpoint
# =============================================================================


@api_view(["POST"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def sales_pdf_view(request):
    """
    Generate a sales lead PDF using the template overlay approach.

    POST /api/sales/pdf/

    Request JSON: All sales form fields
    Returns: PDF file as application/pdf with download headers
    """
    from .sales_pdf import generate_sales_pdf

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    try:
        # Generate PDF
        pdf_bytes = generate_sales_pdf(payload)

        # Create response with PDF
        response = HttpResponse(pdf_bytes, content_type="application/pdf")

        # Generate filename
        client_name = payload.get("client_name", "")
        if not client_name and payload.get("account"):
            client_name = (
                payload["account"].get("name", "") if isinstance(payload["account"], dict) else ""
            )
        job_no = payload.get("order_no", payload.get("job_no", payload.get("bill_no", "")))

        safe_name = slugify(client_name or job_no or "sale")[:30] or "sale"
        filename = f"Sales_Query_{safe_name}.pdf"

        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"] = len(pdf_bytes)

        return response

    except FileNotFoundError as e:
        return JsonResponse({"detail": str(e)}, status=500)
    except Exception as e:
        import traceback

        traceback.print_exc()
        return JsonResponse({"detail": f"PDF generation failed: {str(e)}"}, status=500)


@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def sales_pdf_calibration_view(request):
    """
    Generate a calibration PDF with grid overlay for coordinate tuning.

    GET /api/sales/pdf/calibration/

    Use this endpoint during development to fine-tune the coordinate
    constants in sales_pdf.py.
    """
    from .sales_pdf import generate_calibration_pdf

    try:
        pdf_bytes = generate_calibration_pdf()

        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="sales_calibration.pdf"'
        response["Content-Length"] = len(pdf_bytes)

        return response

    except Exception as e:
        import traceback

        traceback.print_exc()
        return JsonResponse({"detail": f"Calibration PDF failed: {str(e)}"}, status=500)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def repair_collection_view(request):
    """List or create Repair entries for the current user's company."""
    user = request.user
    try:
        from .models import Repair
    except Exception:
        Repair = None

    queryset = (
        Repair.objects.select_related(
            "company", "account", "item_name", "stamp", "created_by", "updated_by"
        ).order_by("-created_at")
        if Repair
        else []
    )
    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none() if hasattr(queryset, "none") else []

    if request.method == "GET":
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(
                Q(tag_no__icontains=search)
                | Q(remark__icontains=search)
                | Q(item_name__icontains=search)
            )

        # Date filtering
        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        if date_from and date_to:
            if date_from > date_to:
                return JsonResponse({"detail": "From date cannot be after To date"}, status=400)
            queryset = queryset.filter(date__gte=date_from, date__lte=date_to)
        elif date_from:
            queryset = queryset.filter(date__gte=date_from)
        elif date_to:
            queryset = queryset.filter(date__lte=date_to)

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count() if hasattr(queryset, "count") else 0
        start = (page - 1) * page_size
        end = start + page_size
        results = (
            [RepairSerializer(v).data for v in queryset[start:end]]
            if hasattr(queryset, "__iter__")
            else []
        )

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    # Map incoming FK aliases to write-only *_id fields
    if isinstance(payload.get("account"), dict):
        payload["account_id"] = payload["account"].get("id") or payload["account"].get("pk")
    elif payload.get("account"):
        payload["account_id"] = payload.get("account")

    if isinstance(payload.get("item_name"), dict):
        payload["item_name_id"] = payload["item_name"].get("id") or payload["item_name"].get("pk")
    elif payload.get("item_name"):
        payload["item_name_id"] = payload.get("item_name")

    if isinstance(payload.get("stamp"), dict):
        payload["stamp_id"] = payload["stamp"].get("id") or payload["stamp"].get("pk")
    elif payload.get("stamp"):
        payload["stamp_id"] = payload.get("stamp")

    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    serializer = RepairSerializer(data=payload)
    if serializer.is_valid():
        obj = serializer.save(company=company, created_by=user)
        return JsonResponse(RepairSerializer(obj).data, status=201)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def payment_collection_view(request):
    """List or create PaymentEntry records for the current user's company."""
    user = request.user
    try:
        from .models import PaymentEntry
    except Exception:
        PaymentEntry = None

    queryset = (
        PaymentEntry.objects.select_related(
            "company", "account", "party", "sub_account", "created_by", "updated_by"
        ).order_by("-created_at")
        if PaymentEntry
        else []
    )
    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none() if hasattr(queryset, "none") else []

    if request.method == "GET":
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(Q(narration__icontains=search))

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count() if hasattr(queryset, "count") else 0
        start = (page - 1) * page_size
        end = start + page_size
        results = (
            [PaymentEntrySerializer(v).data for v in queryset[start:end]]
            if hasattr(queryset, "__iter__")
            else []
        )

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    # Handle both JSON body and FormData
    if request.content_type and "application/json" in request.content_type:
        # JSON body (no file upload)
        try:
            payload = json.loads(request.body.decode("utf-8")) if request.body else {}
        except Exception as e:
            return JsonResponse({"detail": f"Invalid JSON: {str(e)}"}, status=400)
    else:
        # FormData (with or without file upload)
        payload = request.POST.dict()
        # Handle file separately
        if "proof_image" in request.FILES:
            payload["proof_image"] = request.FILES["proof_image"]

    # Map incoming FK aliases -> write-only *_id fields
    if isinstance(payload.get("account"), dict):
        payload["account_id"] = payload["account"].get("id") or payload["account"].get("pk")
    elif payload.get("account"):
        payload["account_id"] = payload.get("account")

    if isinstance(payload.get("party"), dict):
        payload["party_id"] = payload["party"].get("id") or payload["party"].get("pk")
    elif payload.get("party"):
        payload["party_id"] = payload.get("party")

    if isinstance(payload.get("sub_account"), dict):
        payload["sub_account_id"] = payload["sub_account"].get("id") or payload["sub_account"].get(
            "pk"
        )
    elif payload.get("sub_account"):
        payload["sub_account_id"] = payload.get("sub_account")

    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    serializer = PaymentEntrySerializer(data=payload)
    if serializer.is_valid():
        obj = serializer.save(company=company, created_by=user)
        return JsonResponse(PaymentEntrySerializer(obj).data, status=201)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PATCH", "DELETE"])
def payment_detail_view(request, pk):
    user = request.user
    try:
        from .models import PaymentEntry
    except Exception:
        PaymentEntry = None

    try:
        obj = PaymentEntry.objects.select_related(
            "company", "account", "party", "sub_account", "created_by", "updated_by"
        ).get(pk=pk)
    except Exception:
        return JsonResponse({"detail": "Not found"}, status=404)

    if getattr(user, "company_id", None) and obj.company_id != user.company_id:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(PaymentEntrySerializer(obj).data, status=200)

    if request.method == "DELETE":
        obj.delete()
        return HttpResponse(status=204)

    # Handle both JSON body and FormData for PATCH
    if request.content_type and "application/json" in request.content_type:
        # JSON body (no file upload)
        try:
            payload = json.loads(request.body.decode("utf-8")) if request.body else {}
        except Exception as e:
            return JsonResponse({"detail": f"Invalid JSON: {str(e)}"}, status=400)
    else:
        # FormData (with or without file upload)
        payload = request.POST.dict()
        # Handle file separately
        if "proof_image" in request.FILES:
            payload["proof_image"] = request.FILES["proof_image"]

    if isinstance(payload.get("account"), dict):
        payload["account_id"] = payload["account"].get("id") or payload["account"].get("pk")
    elif payload.get("account"):
        payload["account_id"] = payload.get("account")

    if isinstance(payload.get("party"), dict):
        payload["party_id"] = payload["party"].get("id") or payload["party"].get("pk")
    elif payload.get("party"):
        payload["party_id"] = payload.get("party")

    if isinstance(payload.get("sub_account"), dict):
        payload["sub_account_id"] = payload["sub_account"].get("id") or payload["sub_account"].get(
            "pk"
        )
    elif payload.get("sub_account"):
        payload["sub_account_id"] = payload.get("sub_account")

    serializer = PaymentEntrySerializer(obj, data=payload, partial=True)
    if serializer.is_valid():
        updated = serializer.save()
        return JsonResponse(PaymentEntrySerializer(updated).data, status=200)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def journal_collection_view(request):
    """List or create JournalEntry records for the current user's company."""
    user = request.user
    try:
        from .models import JournalEntry
    except Exception:
        JournalEntry = None

    queryset = (
        JournalEntry.objects.select_related(
            "company", "account", "party", "sub_account", "created_by", "updated_by"
        ).order_by("-created_at")
        if JournalEntry
        else []
    )
    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none() if hasattr(queryset, "none") else []

    if request.method == "GET":
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(Q(narration__icontains=search))

        # Date filtering
        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        if date_from and date_to:
            if date_from > date_to:
                return JsonResponse({"detail": "From date cannot be after To date"}, status=400)
            queryset = queryset.filter(date__gte=date_from, date__lte=date_to)
        elif date_from:
            queryset = queryset.filter(date__gte=date_from)
        elif date_to:
            queryset = queryset.filter(date__lte=date_to)

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count() if hasattr(queryset, "count") else 0
        start = (page - 1) * page_size
        end = start + page_size
        results = (
            [JournalEntrySerializer(v).data for v in queryset[start:end]]
            if hasattr(queryset, "__iter__")
            else []
        )

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    # Map incoming FK aliases -> write-only *_id fields
    if isinstance(payload.get("account"), dict):
        payload["account_id"] = payload["account"].get("id") or payload["account"].get("pk")
    elif payload.get("account"):
        payload["account_id"] = payload.get("account")

    if isinstance(payload.get("party"), dict):
        payload["party_id"] = payload["party"].get("id") or payload["party"].get("pk")
    elif payload.get("party"):
        payload["party_id"] = payload.get("party")

    if isinstance(payload.get("sub_account"), dict):
        payload["sub_account_id"] = payload["sub_account"].get("id") or payload["sub_account"].get(
            "pk"
        )
    elif payload.get("sub_account"):
        payload["sub_account_id"] = payload.get("sub_account")

    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    serializer = JournalEntrySerializer(data=payload)
    if serializer.is_valid():
        obj = serializer.save(company=company, created_by=user)
        return JsonResponse(JournalEntrySerializer(obj).data, status=201)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PATCH", "DELETE"])
def journal_detail_view(request, pk):
    user = request.user
    try:
        from .models import JournalEntry
    except Exception:
        JournalEntry = None

    try:
        obj = JournalEntry.objects.select_related(
            "company", "account", "party", "sub_account", "created_by", "updated_by"
        ).get(pk=pk)
    except Exception:
        return JsonResponse({"detail": "Not found"}, status=404)

    if getattr(user, "company_id", None) and obj.company_id != user.company_id:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(JournalEntrySerializer(obj).data, status=200)

    if request.method == "DELETE":
        obj.delete()
        return HttpResponse(status=204)

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    if isinstance(payload.get("account"), dict):
        payload["account_id"] = payload["account"].get("id") or payload["account"].get("pk")
    elif payload.get("account"):
        payload["account_id"] = payload.get("account")

    if isinstance(payload.get("party"), dict):
        payload["party_id"] = payload["party"].get("id") or payload["party"].get("pk")
    elif payload.get("party"):
        payload["party_id"] = payload.get("party")

    if isinstance(payload.get("sub_account"), dict):
        payload["sub_account_id"] = payload["sub_account"].get("id") or payload["sub_account"].get(
            "pk"
        )
    elif payload.get("sub_account"):
        payload["sub_account_id"] = payload.get("sub_account")

    serializer = JournalEntrySerializer(obj, data=payload, partial=True)
    if serializer.is_valid():
        updated = serializer.save()
        return JsonResponse(JournalEntrySerializer(updated).data, status=200)

    return JsonResponse({"errors": serializer.errors}, status=400)


# ===========================================================================
# Payment Voucher (new table-based format)
# ===========================================================================


def _voucher_validate_balance(cr_total, dr_total):
    """Return error string if totals don't match, else None."""
    try:
        if round(float(dr_total), 2) != round(float(cr_total), 2):
            return f"Debit total ({dr_total}) must equal credit total ({cr_total})."
    except (TypeError, ValueError):
        pass
    return None


def _normalize_party(payload):
    if isinstance(payload.get("party_name"), dict):
        payload["party_name_id"] = payload["party_name"].get("id") or payload["party_name"].get(
            "pk"
        )
    elif payload.get("party_name") and not payload.get("party_name_id"):
        payload["party_name_id"] = payload.pop("party_name")


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET"])
def payment_voucher_overview_view(request):
    from .models import PaymentVoucher
    from .serializers import PaymentVoucherSerializer

    user = request.user
    company = getattr(user, "company", None)
    if not company:
        return JsonResponse({"detail": "No company"}, status=400)
    qs = PaymentVoucher.objects.filter(company=company).order_by("-created_at")
    total_count = qs.count()
    total_dr = sum(float(v or 0) for v in qs.values_list("dr", flat=True))
    from .models import PaymentCreditEntry

    total_cr = sum(
        float(v or 0)
        for v in PaymentCreditEntry.objects.filter(payment__company=company).values_list(
            "cr", flat=True
        )
    )
    recent = [
        PaymentVoucherSerializer(v).data
        for v in qs.prefetch_related("credit_entries__party").select_related("party_name")[:5]
    ]
    return JsonResponse(
        {
            "total_count": total_count,
            "total_dr": total_dr,
            "total_cr": total_cr,
            "running_balance": total_cr - total_dr,
            "recent": recent,
        },
        status=200,
    )


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET"])
def journal_voucher_overview_view(request):
    from .models import JournalVoucher
    from .serializers import JournalVoucherSerializer

    user = request.user
    company = getattr(user, "company", None)
    if not company:
        return JsonResponse({"detail": "No company"}, status=400)
    qs = JournalVoucher.objects.filter(company=company).order_by("-created_at")
    total_count = qs.count()
    total_dr = sum(float(v or 0) for v in qs.values_list("dr", flat=True))
    from .models import JournalCreditEntry

    total_cr = sum(
        float(v or 0)
        for v in JournalCreditEntry.objects.filter(journal__company=company).values_list(
            "cr", flat=True
        )
    )
    recent = [
        JournalVoucherSerializer(v).data
        for v in qs.prefetch_related("credit_entries__party").select_related("party_name")[:5]
    ]
    return JsonResponse(
        {
            "total_count": total_count,
            "total_dr": total_dr,
            "total_cr": total_cr,
            "running_balance": total_cr - total_dr,
            "recent": recent,
        },
        status=200,
    )


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET"])
def account_ledger_balance_view(request, pk):
    """Calculate running ledger balance for an account across all voucher types."""
    import uuid as _uuid
    from accounts.models import Account, AccountOpeningBalance
    from .models import (
        Receipt,
        ReceiptDebitEntry,
        PaymentVoucher,
        PaymentCreditEntry,
        JournalVoucher,
        JournalCreditEntry,
        ContraVoucher,
        ContraCreditEntry,
    )

    # Guard: pk must be a valid UUID â€” UI sometimes sends account name by mistake
    try:
        _uuid.UUID(str(pk))
    except (ValueError, AttributeError):
        return JsonResponse({"detail": "Invalid account ID."}, status=400)

    user = request.user
    company = getattr(user, "company", None)
    if not company:
        return JsonResponse({"detail": "No company"}, status=400)
    try:
        account = Account.objects.select_related("opening_balance").get(pk=pk, company=company)
    except Account.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    # Opening balance
    ob = getattr(account, "opening_balance", None)
    opening_amount = float(ob.amount) if ob and ob.amount is not None else 0.0
    opening_drcr = ob.amount_drcr if ob else "Dr"
    # Convert to signed: Dr = positive, Cr = negative
    running = opening_amount if opening_drcr == "Dr" else -opening_amount

    # Receipt: account is credit party (cr adds to balance)
    for r in Receipt.objects.filter(company=company, party_name=account):
        running += float(r.cr or 0)
    # Receipt debit entries: account is debit party
    for e in ReceiptDebitEntry.objects.filter(receipt__company=company, party=account):
        running -= float(e.dr or 0)

    # Payment: account is debit party (dr reduces balance)
    for p in PaymentVoucher.objects.filter(company=company, party_name=account):
        running -= float(p.dr or 0)
    # Payment credit entries
    for e in PaymentCreditEntry.objects.filter(payment__company=company, party=account):
        running += float(e.cr or 0)

    # Journal: account is debit party
    for j in JournalVoucher.objects.filter(company=company, party_name=account):
        running -= float(j.dr or 0)
    # Journal credit entries
    for e in JournalCreditEntry.objects.filter(journal__company=company, party=account):
        running += float(e.cr or 0)

    # Contra: no longer linked to account master (internal Cash/Bank only)
    # for c in ContraVoucher.objects.filter(company=company, party_name=account):
    #     running -= float(c.dr or 0)
    # for e in ContraCreditEntry.objects.filter(contra__company=company, party=account):
    #     running += float(e.cr or 0)

    abs_balance = abs(running)
    drcr = "Dr" if running >= 0 else "Cr"
    return JsonResponse(
        {
            "account_id": str(account.id),
            "account_name": account.account_name,
            "balance": abs_balance,
            "balance_drcr": drcr,
            "balance_display": f"{abs_balance:.2f} {drcr}".strip(),
        },
        status=200,
    )


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def payment_voucher_collection_view(request):
    from .models import PaymentVoucher, VoucherSequence
    from .serializers import PaymentVoucherSerializer

    user = request.user
    company = getattr(user, "company", None)
    if not company:
        return JsonResponse({"detail": "No company"}, status=400)

    qs = (
        PaymentVoucher.objects.select_related("company", "party_name", "created_by")
        .prefetch_related("credit_entries__party")
        .filter(company=company)
        .order_by("-created_at")
    )

    if request.method == "GET":
        search = request.GET.get("search")
        if search:
            qs = qs.filter(
                Q(narration__icontains=search)
                | Q(party_name__account_name__icontains=search)
                | Q(voucher_no__icontains=search)
            )

        # Date filtering
        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        if date_from and date_to:
            if date_from > date_to:
                return JsonResponse({"detail": "From date cannot be after To date"}, status=400)
            qs = qs.filter(date__gte=date_from, date__lte=date_to)
        elif date_from:
            qs = qs.filter(date__gte=date_from)
        elif date_to:
            qs = qs.filter(date__lte=date_to)

        # Ordering
        ordering = request.GET.get("ordering")
        if ordering:
            if ordering == "date_asc":
                qs = qs.order_by("date")
            elif ordering == "date_desc":
                qs = qs.order_by("-date")
            elif ordering == "party_asc":
                qs = qs.order_by("party_name__account_name")
            elif ordering == "party_desc":
                qs = qs.order_by("-party_name__account_name")

        page_size = int(request.GET.get("page_size", 50))
        page = max(int(request.GET.get("page", 1) or 1), 1)
        total = qs.count()
        results = [
            PaymentVoucherSerializer(v).data for v in qs[(page - 1) * page_size : page * page_size]
        ]
        return JsonResponse(
            {"count": total, "next": None, "previous": None, "results": results}, status=200
        )

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    _normalize_party(payload)
    credit_entries = payload.get("credit_entries", [])
    dr_amount = payload.get("dr")
    if dr_amount is not None and credit_entries:
        err = _voucher_validate_balance(
            sum(float(e.get("cr") or 0) for e in credit_entries), float(dr_amount)
        )
        if err:
            return JsonResponse({"detail": err}, status=400)

    payload["voucher_no"] = str(VoucherSequence.next(company, "PAYMENT"))

    serializer = PaymentVoucherSerializer(data=payload)
    if serializer.is_valid():
        obj = serializer.save(company=company, created_by=user)
        return JsonResponse(PaymentVoucherSerializer(obj).data, status=201)
    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PUT", "PATCH", "DELETE"])
def payment_voucher_detail_view(request, pk):
    from .models import PaymentVoucher
    from .serializers import PaymentVoucherSerializer

    user = request.user
    try:
        obj = (
            PaymentVoucher.objects.select_related("company", "party_name")
            .prefetch_related("credit_entries__party")
            .get(pk=pk, company__id=user.company_id)
        )
    except PaymentVoucher.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)
    if request.method == "GET":
        return JsonResponse(PaymentVoucherSerializer(obj).data, status=200)
    if request.method == "DELETE":
        obj.delete()
        return HttpResponse(status=204)
    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)
    _normalize_party(payload)
    serializer = PaymentVoucherSerializer(obj, data=payload, partial=True)
    if serializer.is_valid():
        return JsonResponse(
            PaymentVoucherSerializer(serializer.save(updated_by=user)).data, status=200
        )
    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET"])
def payment_voucher_next_no_view(request):
    from .models import VoucherSequence

    company = getattr(request.user, "company", None)
    if not company:
        return JsonResponse({"detail": "No company"}, status=400)
    return JsonResponse({"voucher_no": str(VoucherSequence.peek(company, "PAYMENT"))}, status=200)


# ===========================================================================
# Journal Voucher
# ===========================================================================


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET"])
def journal_voucher_next_no_view(request):
    from .models import VoucherSequence

    company = getattr(request.user, "company", None)
    if not company:
        return JsonResponse({"detail": "No company"}, status=400)
    return JsonResponse({"voucher_no": str(VoucherSequence.peek(company, "JOURNAL"))}, status=200)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET"])
def contra_voucher_next_no_view(request):
    from .models import VoucherSequence

    company = getattr(request.user, "company", None)
    if not company:
        return JsonResponse({"detail": "No company"}, status=400)
    return JsonResponse({"voucher_no": str(VoucherSequence.peek(company, "CONTRA"))}, status=200)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def journal_voucher_collection_view(request):
    from .models import JournalVoucher, VoucherSequence
    from .serializers import JournalVoucherSerializer

    user = request.user
    company = getattr(user, "company", None)
    if not company:
        return JsonResponse({"detail": "No company"}, status=400)

    qs = (
        JournalVoucher.objects.select_related("company", "party_name", "created_by")
        .prefetch_related("credit_entries__party")
        .filter(company=company)
        .order_by("-created_at")
    )

    if request.method == "GET":
        search = request.GET.get("search")
        if search:
            qs = qs.filter(
                Q(narration__icontains=search)
                | Q(party_name__account_name__icontains=search)
                | Q(voucher_no__icontains=search)
            )

        # Date filtering
        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        if date_from and date_to:
            if date_from > date_to:
                return JsonResponse({"detail": "From date cannot be after To date"}, status=400)
            qs = qs.filter(date__gte=date_from, date__lte=date_to)
        elif date_from:
            qs = qs.filter(date__gte=date_from)
        elif date_to:
            qs = qs.filter(date__lte=date_to)

        # Ordering
        ordering = request.GET.get("ordering")
        if ordering:
            if ordering == "date_asc":
                qs = qs.order_by("date")
            elif ordering == "date_desc":
                qs = qs.order_by("-date")
            elif ordering == "party_asc":
                qs = qs.order_by("party_name__account_name")
            elif ordering == "party_desc":
                qs = qs.order_by("-party_name__account_name")

        page_size = int(request.GET.get("page_size", 50))
        page = max(int(request.GET.get("page", 1) or 1), 1)
        total = qs.count()
        results = [
            JournalVoucherSerializer(v).data for v in qs[(page - 1) * page_size : page * page_size]
        ]
        return JsonResponse(
            {"count": total, "next": None, "previous": None, "results": results}, status=200
        )

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    _normalize_party(payload)
    credit_entries = payload.get("credit_entries", [])
    dr_amount = payload.get("dr")
    if dr_amount is not None and credit_entries:
        err = _voucher_validate_balance(
            sum(float(e.get("cr") or 0) for e in credit_entries), float(dr_amount)
        )
        if err:
            return JsonResponse({"detail": err}, status=400)

    payload["voucher_no"] = str(VoucherSequence.next(company, "JOURNAL"))

    serializer = JournalVoucherSerializer(data=payload)
    if serializer.is_valid():
        obj = serializer.save(company=company, created_by=user)
        return JsonResponse(JournalVoucherSerializer(obj).data, status=201)
    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PUT", "PATCH", "DELETE"])
def journal_voucher_detail_view(request, pk):
    from .models import JournalVoucher
    from .serializers import JournalVoucherSerializer

    user = request.user
    try:
        obj = (
            JournalVoucher.objects.select_related("company", "party_name")
            .prefetch_related("credit_entries__party")
            .get(pk=pk, company__id=user.company_id)
        )
    except JournalVoucher.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)
    if request.method == "GET":
        return JsonResponse(JournalVoucherSerializer(obj).data, status=200)
    if request.method == "DELETE":
        obj.delete()
        return HttpResponse(status=204)
    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)
    _normalize_party(payload)
    serializer = JournalVoucherSerializer(obj, data=payload, partial=True)
    if serializer.is_valid():
        return JsonResponse(
            JournalVoucherSerializer(serializer.save(updated_by=user)).data, status=200
        )
    return JsonResponse({"errors": serializer.errors}, status=400)


# ===========================================================================
# Contra Voucher
# ===========================================================================


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def contra_voucher_collection_view(request):
    from .models import ContraVoucher, VoucherSequence
    from .serializers import ContraVoucherSerializer

    user = request.user
    company = getattr(user, "company", None)
    if not company:
        return JsonResponse({"detail": "No company"}, status=400)

    qs = (
        ContraVoucher.objects.select_related("company", "created_by", "updated_by")
        .prefetch_related("credit_entries")
        .filter(company=company)
        .order_by("-created_at")
    )

    if request.method == "GET":
        search = request.GET.get("search")
        if search:
            qs = qs.filter(
                Q(narration__icontains=search)
                | Q(party_name__icontains=search)
                | Q(voucher_no__icontains=search)
            )

        # Date filtering
        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        if date_from and date_to:
            if date_from > date_to:
                return JsonResponse({"detail": "From date cannot be after To date"}, status=400)
            qs = qs.filter(date__gte=date_from, date__lte=date_to)
        elif date_from:
            qs = qs.filter(date__gte=date_from)
        elif date_to:
            qs = qs.filter(date__lte=date_to)

        # Ordering
        ordering = request.GET.get("ordering")
        if ordering:
            if ordering == "date_asc":
                qs = qs.order_by("date")
            elif ordering == "date_desc":
                qs = qs.order_by("-date")
            elif ordering == "party_asc":
                qs = qs.order_by("party_name")
            elif ordering == "party_desc":
                qs = qs.order_by("-party_name")

        page_size = int(request.GET.get("page_size", 50))
        page = max(int(request.GET.get("page", 1) or 1), 1)
        total = qs.count()
        results = [
            ContraVoucherSerializer(v).data for v in qs[(page - 1) * page_size : page * page_size]
        ]
        return JsonResponse(
            {"count": total, "next": None, "previous": None, "results": results}, status=200
        )

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    credit_entries = payload.get("credit_entries", [])
    dr_amount = payload.get("dr")
    if dr_amount is not None and credit_entries:
        err = _voucher_validate_balance(
            sum(float(e.get("cr") or 0) for e in credit_entries), float(dr_amount)
        )
        if err:
            return JsonResponse({"detail": err}, status=400)

    payload["voucher_no"] = str(VoucherSequence.next(company, "CONTRA"))

    serializer = ContraVoucherSerializer(data=payload)
    if serializer.is_valid():
        obj = serializer.save(company=company, created_by=user)
        return JsonResponse(ContraVoucherSerializer(obj).data, status=201)
    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET"])
def contra_voucher_overview_view(request):
    """Return aggregates and recent entries for Contra voucher overview page."""
    from .models import ContraVoucher
    from .serializers import ContraVoucherSerializer

    user = request.user
    company = getattr(user, "company", None)
    if not company:
        return JsonResponse({"detail": "No company"}, status=400)

    qs = ContraVoucher.objects.filter(company=company).order_by("-created_at")
    total_count = qs.count()
    total_dr = sum(float(v or 0) for v in qs.values_list("dr", flat=True))
    total_cr_entries = 0
    for cv in qs.prefetch_related("credit_entries"):
        for e in cv.credit_entries.all():
            total_cr_entries += float(e.cr or 0)

    recent = [ContraVoucherSerializer(v).data for v in qs.prefetch_related("credit_entries")[:5]]
    return JsonResponse(
        {
            "total_count": total_count,
            "total_dr": total_dr,
            "total_cr": total_cr_entries,
            "running_balance": total_cr_entries - total_dr,
            "recent": recent,
        },
        status=200,
    )


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PUT", "PATCH", "DELETE"])
def contra_voucher_detail_view(request, pk):
    from .models import ContraVoucher
    from .serializers import ContraVoucherSerializer

    user = request.user
    try:
        obj = (
            ContraVoucher.objects.select_related("company", "created_by", "updated_by")
            .prefetch_related("credit_entries")
            .get(pk=pk, company__id=user.company_id)
        )
    except ContraVoucher.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)
    if request.method == "GET":
        return JsonResponse(ContraVoucherSerializer(obj).data, status=200)
    if request.method == "DELETE":
        obj.delete()
        return HttpResponse(status=204)
    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)
    serializer = ContraVoucherSerializer(obj, data=payload, partial=True)
    if serializer.is_valid():
        return JsonResponse(
            ContraVoucherSerializer(serializer.save(updated_by=user)).data, status=200
        )
    return JsonResponse({"errors": serializer.errors}, status=400)


# ===========================================================================
# Shared next-voucher-no endpoint for all types
# ===========================================================================


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET"])
def voucher_next_no_view(request):
    """GET /api/vouchers/next-no/?type=PAYMENT|JOURNAL|CONTRA|RECEIPT"""
    from .models import VoucherSequence

    company = getattr(request.user, "company", None)
    if not company:
        return JsonResponse({"detail": "No company"}, status=400)
    vtype = request.GET.get("type", "").upper()
    valid = {"RECEIPT", "PAYMENT", "JOURNAL", "CONTRA"}
    if vtype not in valid:
        return JsonResponse({"detail": f"type must be one of {valid}"}, status=400)
    # Peek at next number without consuming it
    seq, _ = VoucherSequence.objects.get_or_create(
        company=company, voucher_type=vtype, defaults={"last_number": 0}
    )
    return JsonResponse({"voucher_no": str(seq.last_number + 1), "type": vtype}, status=200)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET"])
def receipt_next_voucher_no_view(request):
    """Return the next auto-incremented voucher number for a new receipt."""
    user = request.user
    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse({"detail": "No company"}, status=400)
    return JsonResponse({"voucher_no": str(_peek_receipt_voucher_no(company))}, status=200)


def _next_receipt_voucher_no(company):
    """Consume and return the next receipt voucher number (call only on actual save)."""
    from .models import VoucherSequence

    return VoucherSequence.next(company, "RECEIPT")


def _peek_receipt_voucher_no(company):
    """Preview next receipt voucher number without consuming it (safe for form load)."""
    from .models import VoucherSequence

    return VoucherSequence.peek(company, "RECEIPT")


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def receipt_collection_view(request):
    """List or create Receipt vouchers for the current user's company."""
    user = request.user
    try:
        queryset = (
            Receipt.objects.select_related("company", "party_name", "created_by", "updated_by")
            .prefetch_related("debit_entries__party")
            .order_by("-created_at")
        )
        if user.is_authenticated and getattr(user, "company_id", None):
            queryset = queryset.filter(company_id=user.company_id)
        else:
            queryset = queryset.none()
    except Exception:
        return JsonResponse(
            {
                "detail": "Receipt model unavailable. Please run makemigrations and migrate for the vouchers app."
            },
            status=500,
        )

    if request.method == "GET":
        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")

        if date_from and date_to:
            if date_from > date_to:
                return JsonResponse({"detail": "From date cannot be after To date"}, status=400)
            queryset = queryset.filter(date__gte=date_from, date__lte=date_to)
        elif date_from:
            queryset = queryset.filter(date__gte=date_from)
        elif date_to:
            queryset = queryset.filter(date__lte=date_to)

        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(
                Q(narration__icontains=search)
                | Q(party_name__account_name__icontains=search)
                | Q(voucher_no__icontains=search)
            )

        # Ordering
        ordering = request.GET.get("ordering")
        if ordering:
            if ordering == "date_asc":
                queryset = queryset.order_by("date")
            elif ordering == "date_desc":
                queryset = queryset.order_by("-date")
            elif ordering == "party_asc":
                queryset = queryset.order_by("party_name__account_name")
            elif ordering == "party_desc":
                queryset = queryset.order_by("-party_name__account_name")

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [ReceiptSerializer(v).data for v in queryset[start:end]]
        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    # Normalize party_name FK
    if isinstance(payload.get("party_name"), dict):
        payload["party_name_id"] = payload["party_name"].get("id") or payload["party_name"].get(
            "pk"
        )
    elif payload.get("party_name") and not payload.get("party_name_id"):
        payload["party_name_id"] = payload.get("party_name")

    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    # Validate: debit entries total must equal credit amount
    cr_amount = payload.get("cr")
    debit_entries = payload.get("debit_entries", [])
    if cr_amount is not None and debit_entries:
        try:
            total_dr = sum(float(e.get("dr") or 0) for e in debit_entries)
            if round(total_dr, 2) != round(float(cr_amount), 2):
                return JsonResponse(
                    {"detail": f"Debit total ({total_dr}) must equal credit amount ({cr_amount})."},
                    status=400,
                )
        except (TypeError, ValueError):
            pass

    # Always use sequence counter to guarantee uniqueness
    payload["voucher_no"] = str(_next_receipt_voucher_no(company))

    serializer = ReceiptSerializer(data=payload)
    if serializer.is_valid():
        try:
            obj = serializer.save(company=company, created_by=user)
            return JsonResponse(ReceiptSerializer(obj).data, status=201)
        except OperationalError:
            return JsonResponse(
                {
                    "detail": "Receipt table not found. Please run `python manage.py migrate vouchers`."
                },
                status=500,
            )
        except Exception as exc:
            return JsonResponse({"detail": str(exc)}, status=500)

    return JsonResponse({"errors": serializer.errors}, status=400)


@jwt_login_required
@require_http_methods(["GET"])
def receipt_overview_view(request):
    """Return aggregates and recent entries for Receipt overview page."""
    user = request.user
    try:
        qs = Receipt.objects.select_related("company", "party_name").order_by("-created_at")
        if user.is_authenticated and getattr(user, "company_id", None):
            qs = qs.filter(company_id=user.company_id)
        else:
            qs = qs.none()
    except Exception:
        return JsonResponse(
            {
                "detail": "Receipt model unavailable. Please run makemigrations and migrate for the vouchers app."
            },
            status=500,
        )

    total_count = qs.count()
    total_dr = 0
    total_cr = 0
    for rec in qs.values("dr", "cr"):
        try:
            total_dr += float(rec.get("dr") or 0)
        except Exception:
            pass
        try:
            total_cr += float(rec.get("cr") or 0)
        except Exception:
            pass

    running_balance = total_cr - total_dr

    recent_qs = qs[:5]
    recent = [ReceiptSerializer(r).data for r in recent_qs]

    payload = {
        "total_count": total_count,
        "total_dr": total_dr,
        "total_cr": total_cr,
        "running_balance": running_balance,
        "recent": recent,
    }
    return JsonResponse(payload, status=200)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PATCH", "DELETE"])
def receipt_detail_view(request, pk):
    user = request.user
    try:
        obj = (
            Receipt.objects.select_related("company", "party_name", "created_by", "updated_by")
            .prefetch_related("debit_entries__party")
            .get(pk=pk)
        )
    except Receipt.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)
    except Exception:
        return JsonResponse(
            {
                "detail": "Receipt model unavailable. Please run makemigrations and migrate for the vouchers app."
            },
            status=500,
        )

    if getattr(user, "company_id", None) and obj.company_id != user.company_id:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(ReceiptSerializer(obj).data, status=200)

    if request.method == "DELETE":
        obj.delete()
        return HttpResponse(status=204)

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    if isinstance(payload.get("party_name"), dict):
        payload["party_name_id"] = payload["party_name"].get("id") or payload["party_name"].get(
            "pk"
        )
    elif payload.get("party_name") and not payload.get("party_name_id"):
        payload["party_name_id"] = payload.get("party_name")

    # Validate debit/credit balance on update
    cr_amount = payload.get("cr", obj.cr)
    debit_entries = payload.get("debit_entries")
    if debit_entries is not None and cr_amount is not None:
        try:
            total_dr = sum(float(e.get("dr") or 0) for e in debit_entries)
            if round(total_dr, 2) != round(float(cr_amount), 2):
                return JsonResponse(
                    {"detail": f"Debit total ({total_dr}) must equal credit amount ({cr_amount})."},
                    status=400,
                )
        except (TypeError, ValueError):
            pass

    serializer = ReceiptSerializer(obj, data=payload, partial=True)
    if serializer.is_valid():
        updated = serializer.save(updated_by=request.user)
        return JsonResponse(ReceiptSerializer(updated).data, status=200)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PATCH", "DELETE"])
def repair_detail_view(request, pk):
    user = request.user
    try:
        from .models import Repair
    except Exception:
        Repair = None

    try:
        obj = Repair.objects.select_related(
            "company", "account", "item_name", "stamp", "created_by", "updated_by"
        ).get(pk=pk)
    except Exception:
        return JsonResponse({"detail": "Not found"}, status=404)

    if getattr(user, "company_id", None) and obj.company_id != user.company_id:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(RepairSerializer(obj).data, status=200)

    if request.method == "DELETE":
        obj.delete()
        return HttpResponse(status=204)

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    if isinstance(payload.get("account"), dict):
        payload["account_id"] = payload["account"].get("id") or payload["account"].get("pk")
    elif payload.get("account"):
        payload["account_id"] = payload.get("account")

    if isinstance(payload.get("item_name"), dict):
        payload["item_name_id"] = payload["item_name"].get("id") or payload["item_name"].get("pk")
    elif payload.get("item_name"):
        payload["item_name_id"] = payload.get("item_name")

    if isinstance(payload.get("stamp"), dict):
        payload["stamp_id"] = payload["stamp"].get("id") or payload["stamp"].get("pk")
    elif payload.get("stamp"):
        payload["stamp_id"] = payload.get("stamp")

    serializer = RepairSerializer(obj, data=payload, partial=True)
    if serializer.is_valid():
        updated = serializer.save()
        return JsonResponse(RepairSerializer(updated).data, status=200)

    return JsonResponse({"errors": serializer.errors}, status=400)


@jwt_login_required
@require_http_methods(["GET"])
def repair_export_view(request, pk):
    """Export a single repair record as an Excel (.xlsx) file."""
    user = request.user
    try:
        repair = Repair.objects.select_related("company", "account", "item_name", "stamp").get(
            pk=pk
        )
    except Exception:
        return JsonResponse({"detail": "Not found"}, status=404)

    if getattr(user, "company_id", None) and repair.company_id != user.company_id:
        return JsonResponse({"detail": "Not found"}, status=404)

    data = RepairSerializer(repair).data

    wb = Workbook()
    ws = wb.active
    ws.title = "Repair Details"
    row = 1

    def w(label, value):
        nonlocal row
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value if value is not None else "")
        row += 1

    w("Tag No", data.get("tag_no"))
    w("Date", data.get("date"))
    w(
        "Item Name",
        (data.get("item_name") or {}).get("name")
        if isinstance(data.get("item_name"), dict)
        else data.get("item_name"),
    )
    w(
        "Account",
        (data.get("account") or {}).get("name")
        if isinstance(data.get("account"), dict)
        else data.get("account"),
    )
    for key in ("gr_wt", "net_wt", "rate", "piece", "total", "remark"):
        val = data.get(key)
        if isinstance(val, dict) and "name" in val:
            val = val.get("name")
        w(key.replace("_", " ").title(), val)

    from datetime import datetime

    row += 1
    ws.cell(row=row, column=1, value="Exported On")
    ws.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    safe_name = slugify((data.get("tag_no") or "repair"))[:30]
    filename = f"Repair_{safe_name}_Details.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@jwt_login_required
@require_http_methods(["GET"])
def repairs_export_all_view(request):
    try:
        user = request.user
        queryset = Repair.objects.select_related(
            "company", "account", "item_name", "stamp"
        ).order_by("-created_at")
        if user.is_authenticated and getattr(user, "company_id", None):
            queryset = queryset.filter(company_id=user.company_id)
        else:
            queryset = queryset.none()

        rows = []
        for r in queryset:
            data = RepairSerializer(r).data
            if isinstance(data.get("account"), dict) and data.get("account").get("name"):
                data["account"] = data.get("account").get("name")
            rows.append(_flatten_dict(data))

        cols = []
        seen = set()
        for rr in rows:
            for k in rr.keys():
                if k not in seen:
                    cols.append(k)
                    seen.add(k)

        wb = Workbook()
        ws = wb.active
        ws.title = "Repairs"

        for c_idx, col in enumerate(cols, start=1):
            header = col.replace(".", " ").replace("_", " ").title()
            ws.cell(row=1, column=c_idx, value=header)

        for r_idx, r in enumerate(rows, start=2):
            for c_idx, col in enumerate(cols, start=1):
                raw_val = r.get(col, "")
                val = _excel_safe_value(raw_val)
                ws.cell(row=r_idx, column=c_idx, value=val)

        from openpyxl.utils import get_column_letter

        for i, col in enumerate(cols, start=1):
            letter = get_column_letter(i)
            maxlen = len(col)
            for cell in ws[letter]:
                if cell.value is not None:
                    maxlen = max(maxlen, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(100, max(10, maxlen + 2))

        from datetime import datetime

        footer_row = len(rows) + 3
        ws.cell(row=footer_row, column=1, value="Exported On")
        ws.cell(row=footer_row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

        filename = "repairs_data.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    except Exception as exc:
        import traceback

        tb = traceback.format_exc()
        return JsonResponse({"detail": str(exc), "trace": tb}, status=500)


@jwt_login_required
@require_http_methods(["GET"])
def sales_export_all_view(request):
    """Export all Sale records for the user's company as an Excel file."""
    try:
        user = request.user
        queryset = Sale.objects.select_related("company", "account").order_by("-created_at")
        if user.is_authenticated and getattr(user, "company_id", None):
            queryset = queryset.filter(company_id=user.company_id)
        else:
            queryset = queryset.none()

        rows = []
        for s in queryset:
            data = SaleSerializer(s).data
            # prefer human-readable account/item names
            if isinstance(data.get("account"), dict) and data.get("account").get("name"):
                data["account"] = data.get("account").get("name")
            rows.append(_flatten_dict(data))

        cols = []
        seen = set()
        for r in rows:
            for k in r.keys():
                if k not in seen:
                    cols.append(k)
                    seen.add(k)

        wb = Workbook()
        ws = wb.active
        ws.title = "Sales"

        for c_idx, col in enumerate(cols, start=1):
            header = col.replace(".", " ").replace("_", " ").title()
            ws.cell(row=1, column=c_idx, value=header)

        for r_idx, r in enumerate(rows, start=2):
            for c_idx, col in enumerate(cols, start=1):
                raw_val = r.get(col, "")
                val = _excel_safe_value(raw_val)
                ws.cell(row=r_idx, column=c_idx, value=val)

        from openpyxl.utils import get_column_letter

        for i, col in enumerate(cols, start=1):
            letter = get_column_letter(i)
            maxlen = len(col)
            for cell in ws[letter]:
                if cell.value is not None:
                    maxlen = max(maxlen, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(100, max(10, maxlen + 2))

        from datetime import datetime

        footer_row = len(rows) + 3
        ws.cell(row=footer_row, column=1, value="Exported On")
        ws.cell(row=footer_row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

        filename = "sales_data.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    except Exception as exc:
        import traceback

        tb = traceback.format_exc()
        return JsonResponse({"detail": str(exc), "trace": tb}, status=500)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET"])
def vouchers_aggregates_view(request):
    """Return simple aggregates for the current user's company.

    Response shape:
    {
        "total": int,
        "with_advance": int,
        "recent_7_days": int
    }
    """
    user = request.user
    queryset = _base_queryset(user)

    total = queryset.count()
    # advance_payment_received is now a YES/NO choice field
    with_advance = queryset.filter(advance_payment_received="YES").count()
    week_ago = timezone.now() - timedelta(days=7)
    recent = queryset.filter(created_at__gte=week_ago).count()

    payload = {"total": total, "with_advance": with_advance, "recent_7_days": recent}
    return JsonResponse(payload, status=200)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def approval_loose_collection_view(request):
    """List or create ApprovalLooseM entries for the current user's company."""
    user = request.user
    queryset = ApprovalLooseM.objects.select_related(
        "company", "account", "item_name", "stamp", "unit", "created_by", "updated_by"
    ).order_by("-created_at")
    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    if request.method == "GET":
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(
                Q(order_number__icontains=search) | Q(remark__icontains=search)
            )

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [ApprovalLooseMSerializer(v).data for v in queryset[start:end]]

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    # POST
    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    # Normalize FK aliases -> use write-only *_id fields expected by serializer
    if isinstance(payload.get("account"), dict):
        payload["account_id"] = payload["account"].get("id") or payload["account"].get("pk")
    elif payload.get("account"):
        payload["account_id"] = payload.get("account")

    if isinstance(payload.get("item_name"), dict):
        payload["item_name_id"] = payload["item_name"].get("id") or payload["item_name"].get("pk")
    elif payload.get("item_name"):
        payload["item_name_id"] = payload.get("item_name")

    if isinstance(payload.get("stamp"), dict):
        payload["stamp_id"] = payload["stamp"].get("id") or payload["stamp"].get("pk")
    elif payload.get("stamp"):
        payload["stamp_id"] = payload.get("stamp")

    if isinstance(payload.get("unit"), dict):
        payload["unit_id"] = payload["unit"].get("id") or payload["unit"].get("pk")
    elif payload.get("unit"):
        payload["unit_id"] = payload.get("unit")

    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    serializer = ApprovalLooseMSerializer(data=payload)
    if serializer.is_valid():
        obj = serializer.save(company=company, created_by=user)
        return JsonResponse(ApprovalLooseMSerializer(obj).data, status=201)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def approval_tag_collection_view(request):
    """List or create ApprovalTagM entries for the current user's company."""
    user = request.user
    queryset = ApprovalTagM.objects.select_related(
        "company", "item_name", "stamp", "unit", "created_by", "updated_by"
    ).order_by("-created_at")
    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    if request.method == "GET":
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(
                Q(order_number__icontains=search)
                | Q(remark__icontains=search)
                | Q(design__icontains=search)
            )

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [ApprovalTagMSerializer(v).data for v in queryset[start:end]]

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    # Map incoming FK aliases to write-only *_id fields
    if isinstance(payload.get("account"), dict):
        payload["account_id"] = payload["account"].get("id") or payload["account"].get("pk")
    elif payload.get("account"):
        payload["account_id"] = payload.get("account")

    if isinstance(payload.get("item_name"), dict):
        payload["item_name_id"] = payload["item_name"].get("id") or payload["item_name"].get("pk")
    elif payload.get("item_name"):
        payload["item_name_id"] = payload.get("item_name")

    if isinstance(payload.get("stamp"), dict):
        payload["stamp_id"] = payload["stamp"].get("id") or payload["stamp"].get("pk")
    elif payload.get("stamp"):
        payload["stamp_id"] = payload.get("stamp")

    if isinstance(payload.get("unit"), dict):
        payload["unit_id"] = payload["unit"].get("id") or payload["unit"].get("pk")
    elif payload.get("unit"):
        payload["unit_id"] = payload.get("unit")

    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    serializer = ApprovalTagMSerializer(data=payload)
    if serializer.is_valid():
        obj = serializer.save(company=company, created_by=user)
        return JsonResponse(ApprovalTagMSerializer(obj).data, status=201)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PATCH", "DELETE"])
def approval_tag_detail_view(request, pk):
    user = request.user
    try:
        obj = ApprovalTagM.objects.select_related(
            "company", "item_name", "stamp", "unit", "created_by", "updated_by"
        ).get(pk=pk)
    except ApprovalTagM.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    if getattr(user, "company_id", None) and obj.company_id != user.company_id:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(ApprovalTagMSerializer(obj).data, status=200)

    if request.method == "DELETE":
        obj.delete()
        return HttpResponse(status=204)

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    if isinstance(payload.get("account"), dict):
        payload["account_id"] = payload["account"].get("id") or payload["account"].get("pk")
    elif payload.get("account"):
        payload["account_id"] = payload.get("account")

    if isinstance(payload.get("item_name"), dict):
        payload["item_name_id"] = payload["item_name"].get("id") or payload["item_name"].get("pk")
    elif payload.get("item_name"):
        payload["item_name_id"] = payload.get("item_name")

    if isinstance(payload.get("stamp"), dict):
        payload["stamp_id"] = payload["stamp"].get("id") or payload["stamp"].get("pk")
    elif payload.get("stamp"):
        payload["stamp_id"] = payload.get("stamp")

    if isinstance(payload.get("unit"), dict):
        payload["unit_id"] = payload["unit"].get("id") or payload["unit"].get("pk")
    elif payload.get("unit"):
        payload["unit_id"] = payload.get("unit")

    serializer = ApprovalTagMSerializer(obj, data=payload, partial=True)
    if serializer.is_valid():
        updated = serializer.save()
        return JsonResponse(ApprovalTagMSerializer(updated).data, status=200)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PATCH", "DELETE"])
def approval_loose_detail_view(request, pk):
    user = request.user
    try:
        obj = ApprovalLooseM.objects.select_related(
            "company", "account", "item_name", "stamp", "unit", "created_by", "updated_by"
        ).get(pk=pk)
    except ApprovalLooseM.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    if getattr(user, "company_id", None) and obj.company_id != user.company_id:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(ApprovalLooseMSerializer(obj).data, status=200)

    if request.method == "DELETE":
        obj.delete()
        return HttpResponse(status=204)

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    # Normalize FK aliases -> use write-only *_id fields expected by serializer
    if isinstance(payload.get("account"), dict):
        payload["account_id"] = payload["account"].get("id") or payload["account"].get("pk")
    elif payload.get("account"):
        payload["account_id"] = payload.get("account")

    if isinstance(payload.get("item_name"), dict):
        payload["item_name_id"] = payload["item_name"].get("id") or payload["item_name"].get("pk")
    elif payload.get("item_name"):
        payload["item_name_id"] = payload.get("item_name")

    if isinstance(payload.get("stamp"), dict):
        payload["stamp_id"] = payload["stamp"].get("id") or payload["stamp"].get("pk")
    elif payload.get("stamp"):
        payload["stamp_id"] = payload.get("stamp")

    if isinstance(payload.get("unit"), dict):
        payload["unit_id"] = payload["unit"].get("id") or payload["unit"].get("pk")
    elif payload.get("unit"):
        payload["unit_id"] = payload.get("unit")

    serializer = ApprovalLooseMSerializer(obj, data=payload, partial=True)
    if serializer.is_valid():
        updated = serializer.save()
        return JsonResponse(ApprovalLooseMSerializer(updated).data, status=200)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def vouchers_master_list_view(request):
    """Return master data for vouchers/orders (series, stamps, base metals).

    Response shape:
    {
        "series": [{"id": "...","name": "..."}, ...],
        "stamps": [{"id":"...","name":"...","code":"..."}, ...],
        "base_metals": [{"id":"...","name":"...","code":"..."}, ...]
    }
    """
    # Fix: Properly load the company relationship
    try:
        from users.models import User

        user = User.objects.select_related("company").get(id=request.user.id)
        company = user.company
    except User.DoesNotExist:
        return Response({"error": "User not found"}, status=400)

    if not company:
        return Response({"error": "No company found"}, status=400)

    # Return company-scoped masters plus global masters (company IS NULL) so
    # frontend users see both shared and tenant-specific options.
    if company:
        series_qs = SeriesMaster.objects.filter(
            Q(company=company) | Q(company__isnull=True)
        ).order_by("name")
        stamps_qs = StampMaster.objects.filter(
            Q(company=company) | Q(company__isnull=True)
        ).order_by("name")
        metals_qs = BaseMetalMaster.objects.filter(
            Q(company=company) | Q(company__isnull=True)
        ).order_by("name")
        item_qs = ItemNameMaster.objects.filter(
            Q(company=company) | Q(company__isnull=True)
        ).order_by("name")
        clarity_qs = ClarityMaster.objects.filter(
            Q(company=company) | Q(company__isnull=True)
        ).order_by("name")
        shape_qs = ShapeMaster.objects.filter(
            Q(company=company) | Q(company__isnull=True)
        ).order_by("name")
        unit_qs = UnitMaster.objects.filter(Q(company=company) | Q(company__isnull=True)).order_by(
            "name"
        )
        size_qs = SizeMaster.objects.filter(Q(company=company) | Q(company__isnull=True)).order_by(
            "name"
        )
        colour_qs = ColourMaster.objects.filter(
            Q(company=company) | Q(company__isnull=True)
        ).order_by("name")
        lab_qs = LabMaster.objects.filter(Q(company=company) | Q(company__isnull=True)).order_by(
            "name"
        )
    else:
        series_qs = SeriesMaster.objects.order_by("name")
        stamps_qs = StampMaster.objects.order_by("name")
        metals_qs = BaseMetalMaster.objects.order_by("name")
        item_qs = ItemNameMaster.objects.order_by("name")
        clarity_qs = ClarityMaster.objects.order_by("name")
        shape_qs = ShapeMaster.objects.order_by("name")
        unit_qs = UnitMaster.objects.order_by("name")
        size_qs = SizeMaster.objects.order_by("name")
        colour_qs = ColourMaster.objects.order_by("name")
        lab_qs = LabMaster.objects.order_by("name")

    payload = {
        "series": [{"id": str(s.id), "name": s.name} for s in series_qs],
        "stamps": [{"id": str(s.id), "name": s.name, "code": s.code} for s in stamps_qs],
        "base_metals": [{"id": str(m.id), "name": m.name, "code": m.code} for m in metals_qs],
        "item_names": [{"id": str(i.id), "name": i.name, "code": i.code} for i in item_qs],
        "clarities": [{"id": str(c.id), "name": c.name, "code": c.code} for c in clarity_qs],
        "shapes": [{"id": str(s.id), "name": s.name, "code": s.code} for s in shape_qs],
        "units": [{"id": str(u.id), "name": u.name, "code": u.code} for u in unit_qs],
        "sizes": [{"id": str(s.id), "name": s.name} for s in size_qs],
        "colours": [{"id": str(c.id), "name": c.name} for c in colour_qs],
        "labs": [{"id": str(lab.id), "name": lab.name} for lab in lab_qs],
    }

    return Response(payload, status=200)


@jwt_login_required
@require_http_methods(["GET"])
def payment_master_view(request):
    """Return dropdown data needed by the frontend Payment Add page.

    Response shape:
    {
        "accounts": [{"id": "..", "name": ".."}, ...],
        "sub_accounts": [{"id":"..","name":"..","account_id":".."}, ...]
    }
    """
    user = request.user
    company = getattr(user, "company", None)

    if company:
        accounts_qs = Account.objects.filter(company=company).order_by("account_name")
        sub_qs = SubAccount.objects.filter(created_by__company=company).order_by("sub_account_name")
    else:
        accounts_qs = Account.objects.none()
        sub_qs = SubAccount.objects.none()

    accounts = [
        {"id": str(a.id), "name": getattr(a, "account_name", None) or getattr(a, "name", None)}
        for a in accounts_qs
    ]
    sub_accounts = [
        {
            "id": str(s.id),
            "name": s.sub_account_name,
            "account_id": str(s.account_id) if s.account_id else None,
        }
        for s in sub_qs
    ]

    return JsonResponse({"accounts": accounts, "sub_accounts": sub_accounts}, status=200)


@jwt_login_required
@require_http_methods(["GET"])
def payment_aggregates_view(request):
    """Return simple aggregates for Payments overview page."""
    user = request.user
    try:
        from .models import PaymentEntry
    except Exception:
        PaymentEntry = None

    if not PaymentEntry:
        return JsonResponse({"total": 0, "recent_7_days": 0, "sum_dr": 0, "sum_cr": 0}, status=200)

    if user.is_authenticated and getattr(user, "company_id", None):
        qs = PaymentEntry.objects.filter(company_id=user.company_id)
    else:
        qs = PaymentEntry.objects.none()

    total = qs.count()
    week_ago = timezone.now() - timedelta(days=7)
    recent = qs.filter(created_at__gte=week_ago).count()
    sums = qs.aggregate(sum_dr=Sum("dr"), sum_cr=Sum("cr"))
    sum_dr = float(sums.get("sum_dr") or 0)
    sum_cr = float(sums.get("sum_cr") or 0)

    return JsonResponse(
        {"total": total, "recent_7_days": recent, "sum_dr": sum_dr, "sum_cr": sum_cr}, status=200
    )


@jwt_login_required
@require_http_methods(["GET"])
def journal_master_view(request):
    """Return dropdown data needed by the frontend Journal Add page.

    Reuses the same account/sub-account lists as payments.
    """
    return payment_master_view(request)


@jwt_login_required
@require_http_methods(["GET"])
def journal_aggregates_view(request):
    """Return simple aggregates for Journals overview page."""
    user = request.user
    try:
        from .models import JournalEntry
    except Exception:
        JournalEntry = None

    if not JournalEntry:
        return JsonResponse({"total": 0, "recent_7_days": 0, "sum_dr": 0, "sum_cr": 0}, status=200)

    if user.is_authenticated and getattr(user, "company_id", None):
        qs = JournalEntry.objects.filter(company_id=user.company_id)
    else:
        qs = JournalEntry.objects.none()

    total = qs.count()
    week_ago = timezone.now() - timedelta(days=7)
    recent = qs.filter(created_at__gte=week_ago).count()
    sums = qs.aggregate(sum_dr=Sum("dr"), sum_cr=Sum("cr"))
    sum_dr = float(sums.get("sum_dr") or 0)
    sum_cr = float(sums.get("sum_cr") or 0)

    return JsonResponse(
        {"total": total, "recent_7_days": recent, "sum_dr": sum_dr, "sum_cr": sum_cr}, status=200
    )


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET"])
def archives_list_view(request):
    """Return archived orders for the current user's company.

    This mirrors the vouchers list but reads from the Archives model which
    stores records where advance_payment_received == "NO".
    """
    user = request.user
    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    queryset = Archives.objects.select_related(
        "company", "account", "created_by", "updated_by"
    ).order_by("-created_at")
    if getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    search = request.GET.get("search")
    if search:
        queryset = queryset.filter(Q(bill_no__icontains=search) | Q(item_name__icontains=search))

    page_size = int(request.GET.get("page_size", 50))
    page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
    page = page if page > 0 else 1
    total = queryset.count()
    start = (page - 1) * page_size
    end = start + page_size
    results = [
        ArchivesSerializer(v, context={"request": request}).data for v in queryset[start:end]
    ]

    payload = {
        "count": total,
        "next": None,
        "previous": None,
        "results": results,
    }
    return JsonResponse(payload, status=200)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "DELETE"])
def archives_detail_view(request, pk):
    """Detail / delete view for an Archives record.

    Provides a DELETE handler so the frontend can remove pending query records
    using an archives-specific URL (vouchers/archives/<pk>/).
    """
    user = request.user
    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    try:
        archive = Archives.objects.select_related(
            "company", "account", "created_by", "updated_by"
        ).get(pk=pk)
    except Archives.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    if getattr(user, "company_id", None) and archive.company_id != user.company_id:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(
            ArchivesSerializer(archive, context={"request": request}).data, status=200
        )

    # DELETE
    archive.delete()
    return HttpResponse(status=204)


@jwt_login_required
@require_http_methods(["GET"])
def archives_export_view(request, pk):
    """Export a single archived order as an Excel file (field | value)."""
    user = request.user
    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    try:
        archive = Archives.objects.get(pk=pk)
    except Archives.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    # ensure company scope
    if getattr(user, "company_id", None) and archive.company_id != user.company_id:
        return JsonResponse({"detail": "Not found"}, status=404)

    data = ArchivesSerializer(archive, context={"request": request}).data

    wb = Workbook()
    ws = wb.active
    ws.title = "Archived Order Details"
    row = 1

    def w(label, value):
        nonlocal row
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value if value is not None else "")
        row += 1

    w("Order Number", data.get("bill_no") or data.get("voucher_number"))
    w("Date", data.get("date"))
    # try to surface item_name/account names where available
    item = (
        (data.get("item_name") or {}).get("name")
        if isinstance(data.get("item_name"), dict)
        else data.get("item_name")
    )
    acct = (
        (data.get("account") or {}).get("name")
        if isinstance(data.get("account"), dict)
        else data.get("account")
    )
    w("Item Name", item)
    w("Account", acct)

    for key in (
        "rate",
        "gross_wt",
        "net_wt",
        "unit",
        "tunch",
        "remark",
        "amount",
        "advance_payment_received",
    ):
        val = data.get(key)
        if isinstance(val, dict) and "name" in val:
            val = val.get("name")
        w(key.replace("_", " ").title(), val)

    from datetime import datetime

    row += 1
    ws.cell(row=row, column=1, value="Exported On")
    ws.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    safe_name = slugify((data.get("bill_no") or data.get("voucher_number") or "archive"))[:30]
    filename = f"Archived_Order_{safe_name}_Details.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def item_names_master_view(request):
    """Return item name masters (tenant + global) as a small list.

    This endpoint allows the frontend to fetch item names on-demand so admin
    changes become visible immediately to the UI.
    """
    user = request.user
    company = getattr(user, "company", None)
    if company:
        qs = ItemNameMaster.objects.filter(Q(company=company) | Q(company__isnull=True)).order_by(
            "name"
        )
    else:
        qs = ItemNameMaster.objects.order_by("name")

    payload = [{"id": str(i.id), "name": i.name, "code": i.code} for i in qs]
    return Response({"item_names": payload}, status=200)


@csrf_exempt
@csrf_exempt
@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def clarity_master_view(request):
    """Return clarity masters for raw material purchase."""
    try:
        from users.models import User

        # Fix: Properly load the company relationship
        user = User.objects.select_related("company").get(id=request.user.id)
        company = getattr(user, "company", None)

        if company:
            qs = ClarityMaster.objects.filter(
                Q(company=company) | Q(company__isnull=True)
            ).order_by("name")
        else:
            qs = ClarityMaster.objects.order_by("name")

        payload = [{"id": str(c.id), "name": c.name, "code": getattr(c, "code", None)} for c in qs]
        return Response({"clarities": payload}, status=200)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def shape_master_view(request):
    # Fix: Properly load the company relationship
    try:
        from users.models import User

        user = User.objects.select_related("company").get(id=request.user.id)
        company = user.company
    except User.DoesNotExist:
        return Response({"error": "User not found"}, status=400)

    if company:
        qs = ShapeMaster.objects.filter(Q(company=company) | Q(company__isnull=True)).order_by(
            "name"
        )
    else:
        qs = ShapeMaster.objects.order_by("name")

    payload = [{"id": str(s.id), "name": s.name, "code": s.code} for s in qs]
    return Response({"shapes": payload}, status=200)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET"])
def unit_master_view(request):
    user = request.user
    company = getattr(user, "company", None)
    if company:
        qs = UnitMaster.objects.filter(Q(company=company) | Q(company__isnull=True)).order_by(
            "name"
        )
    else:
        qs = UnitMaster.objects.order_by("name")

    payload = [{"id": str(u.id), "name": u.name, "code": u.code} for u in qs]
    return JsonResponse({"units": payload}, status=200)


@csrf_exempt
@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def colour_master_view(request):
    """Return colour masters for raw material purchase."""
    try:
        from users.models import User

        # Fix: Properly load the company relationship
        user = User.objects.select_related("company").get(id=request.user.id)
        company = getattr(user, "company", None)

        if company:
            qs = ColourMaster.objects.filter(Q(company=company) | Q(company__isnull=True)).order_by(
                "name"
            )
        else:
            qs = ColourMaster.objects.order_by("name")

        payload = [{"id": str(c.id), "name": c.name, "code": getattr(c, "code", None)} for c in qs]
        return Response({"colours": payload}, status=200)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@csrf_exempt
@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def lab_master_view(request):
    """Return lab masters for raw material purchase."""
    try:
        from users.models import User

        # Fix: Properly load the company relationship
        user = User.objects.select_related("company").get(id=request.user.id)
        company = getattr(user, "company", None)

        if company:
            qs = LabMaster.objects.filter(Q(company=company) | Q(company__isnull=True)).order_by(
                "name"
            )
        else:
            qs = LabMaster.objects.order_by("name")

        payload = [
            {"id": str(lab.id), "name": lab.name, "code": getattr(lab, "code", None)} for lab in qs
        ]
        return Response({"labs": payload}, status=200)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


# =============================================================================
# Enhanced Master Data APIs for Raw Material Purchase
# =============================================================================


@csrf_exempt
@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def gemstone_master_view(request):
    """Return gemstone type masters for raw material purchase."""
    try:
        from users.models import User
        from core.models import GemstoneMaster

        user = User.objects.select_related("company").get(id=request.user.id)
        company = user.company
    except User.DoesNotExist:
        return Response({"error": "User not found"}, status=400)

    if company:
        qs = GemstoneMaster.objects.filter(Q(company=company) | Q(company__isnull=True)).order_by(
            "name"
        )
    else:
        qs = GemstoneMaster.objects.order_by("name")

    payload = [{"id": str(g.id), "name": g.name, "code": g.code} for g in qs]
    return Response({"gemstones": payload}, status=200)


@csrf_exempt
@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def gemstone_shape_master_view(request):
    """Return gemstone shape masters for raw material purchase."""
    try:
        from users.models import User
        from core.models import GemstoneShapeMaster

        user = User.objects.select_related("company").get(id=request.user.id)
        company = user.company
    except User.DoesNotExist:
        return Response({"error": "User not found"}, status=400)

    if company:
        qs = GemstoneShapeMaster.objects.filter(
            Q(company=company) | Q(company__isnull=True)
        ).order_by("name")
    else:
        qs = GemstoneShapeMaster.objects.order_by("name")

    payload = [{"id": str(s.id), "name": s.name, "code": s.code} for s in qs]
    return Response({"gemstone_shapes": payload}, status=200)


@csrf_exempt
@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def gemstone_color_master_view(request):
    """Return gemstone color masters for raw material purchase."""
    try:
        from users.models import User
        from core.models import GemstoneColorMaster

        user = User.objects.select_related("company").get(id=request.user.id)
        company = user.company
    except User.DoesNotExist:
        return Response({"error": "User not found"}, status=400)

    if company:
        qs = GemstoneColorMaster.objects.filter(
            Q(company=company) | Q(company__isnull=True)
        ).order_by("name")
    else:
        qs = GemstoneColorMaster.objects.order_by("name")

    payload = [{"id": str(c.id), "name": c.name, "code": c.code} for c in qs]
    return Response({"gemstone_colors": payload}, status=200)


@csrf_exempt
@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def gemstone_clarity_master_view(request):
    """Return gemstone clarity masters for raw material purchase."""
    try:
        from users.models import User
        from core.models import GemstoneClarityMaster

        user = User.objects.select_related("company").get(id=request.user.id)
        company = user.company
    except User.DoesNotExist:
        return Response({"error": "User not found"}, status=400)

    if company:
        qs = GemstoneClarityMaster.objects.filter(
            Q(company=company) | Q(company__isnull=True)
        ).order_by("name")
    else:
        qs = GemstoneClarityMaster.objects.order_by("name")

    payload = [{"id": str(c.id), "name": c.name, "code": c.code} for c in qs]
    return Response({"gemstone_clarities": payload}, status=200)


@csrf_exempt
@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def gemstone_treatment_master_view(request):
    """Return gemstone treatment masters for raw material purchase."""
    try:
        from users.models import User
        from core.models import GemstoneTreatmentMaster

        user = User.objects.select_related("company").get(id=request.user.id)
        company = user.company
    except User.DoesNotExist:
        return Response({"error": "User not found"}, status=400)

    if company:
        qs = GemstoneTreatmentMaster.objects.filter(
            Q(company=company) | Q(company__isnull=True)
        ).order_by("name")
    else:
        qs = GemstoneTreatmentMaster.objects.order_by("name")

    payload = [{"id": str(t.id), "name": t.name, "code": t.code} for t in qs]
    return Response({"gemstone_treatments": payload}, status=200)


@csrf_exempt
@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def origin_master_view(request):
    """Return origin masters for raw material purchase."""
    try:
        from users.models import User
        from core.models import OriginMaster

        user = User.objects.select_related("company").get(id=request.user.id)
        company = user.company
    except User.DoesNotExist:
        return Response({"error": "User not found"}, status=400)

    # Support filtering by material type
    material_type = request.GET.get("material_type", "").lower()

    if company:
        qs = OriginMaster.objects.filter(Q(company=company) | Q(company__isnull=True))
    else:
        qs = OriginMaster.objects.all()

    # Filter by material type if provided
    if material_type in ["diamond", "gemstone"]:
        qs = qs.filter(Q(material_type=material_type) | Q(material_type="all"))

    qs = qs.order_by("name")

    payload = [{"id": str(o.id), "name": o.name, "material_type": o.material_type} for o in qs]
    return Response({"origins": payload}, status=200)


@csrf_exempt
@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def metal_type_master_view(request):
    """Return metal type masters for raw material purchase."""
    try:
        from users.models import User
        from core.models import MetalTypeMaster

        user = User.objects.select_related("company").get(id=request.user.id)
        company = user.company
    except User.DoesNotExist:
        return Response({"error": "User not found"}, status=400)

    if company:
        qs = MetalTypeMaster.objects.filter(Q(company=company) | Q(company__isnull=True)).order_by(
            "name"
        )
    else:
        qs = MetalTypeMaster.objects.order_by("name")

    payload = [{"id": str(m.id), "name": m.name} for m in qs]
    return Response({"metal_types": payload}, status=200)


from .models import EstimateVoucher
from .serializers import EstimateVoucherSerializer


@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def estimate_gold_qualities_view(request):
    """Get hardcoded gold quality options for estimate dropdown (supports custom input)"""
    GOLD_QUALITIES = [
        {"id": "24kt", "name": "24KT", "code": "24"},
        {"id": "22kt", "name": "22KT", "code": "22"},
        {"id": "18kt", "name": "18KT", "code": "18"},
        {"id": "14kt", "name": "14KT", "code": "14"},
        {"id": "10kt", "name": "10KT", "code": "10"},
        {"id": "9kt", "name": "9KT", "code": "9"},
    ]

    return Response(
        {
            "gold_qualities": GOLD_QUALITIES,
            "count": len(GOLD_QUALITIES),
            "message": "Standard gold quality options. Custom values are also accepted.",
        }
    )


@api_view(["GET", "POST"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def estimate_collection_view(request):
    """List or create EstimateVoucher entries for the current user's company."""
    import logging

    logger = logging.getLogger(__name__)

    user = request.user

    # Check if user has a company
    if not user.is_authenticated:
        logger.error("User not authenticated")
        return Response({"detail": "Authentication required"}, status=401)

    if not getattr(user, "company_id", None):
        logger.error(f"User {user.email} has no company assigned")
        return Response({"detail": "User has no company assigned"}, status=400)

    queryset = (
        EstimateVoucher.objects.select_related("company", "account", "created_by", "updated_by")
        .prefetch_related("line_items")
        .order_by("-created_at")
    )

    queryset = queryset.filter(company_id=user.company_id)

    if request.method == "GET":
        try:
            # Support filtering by account
            account_id = request.GET.get("account_id")
            if account_id:
                queryset = queryset.filter(account_id=account_id)

            # Support filtering by sales lead
            sales_query_id = request.GET.get("sales_query_id")
            if sales_query_id:
                queryset = queryset.filter(sales_query_id=sales_query_id)

            # Support filtering by jewelry type
            jewellery_type = request.GET.get("jewellery_type")
            if jewellery_type:
                queryset = queryset.filter(jewellery_type__iexact=jewellery_type)

            # Support search
            search = request.GET.get("search")
            if search:
                queryset = queryset.filter(
                    Q(item_name__icontains=search) | Q(account__account_name__icontains=search)
                )

            # Support date range filtering
            start_date = request.GET.get("start_date")
            end_date = request.GET.get("end_date")
            if start_date:
                queryset = queryset.filter(date__gte=start_date)
            if end_date:
                queryset = queryset.filter(date__lte=end_date)

            page_size = int(request.GET.get("page_size", 50))
            page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
            page = page if page > 0 else 1

            total = queryset.count()
            start = (page - 1) * page_size
            end = start + page_size

            # Serialize results with error handling
            results = []
            for estimate in queryset[start:end]:
                try:
                    results.append(EstimateVoucherSerializer(estimate).data)
                except Exception as e:
                    logger.error(f"Error serializing estimate {estimate.id}: {str(e)}")
                    # Skip problematic estimates rather than failing the entire request
                    continue

            payload = {"count": total, "next": None, "previous": None, "results": results}
            return Response(payload, status=200)

        except Exception as e:
            logger.error(f"Error in estimate collection GET: {str(e)}")
            return Response({"detail": f"Error retrieving estimates: {str(e)}"}, status=500)

    # POST - Create new estimate
    # Accept either application/json or multipart/form-data for image uploads
    if request.content_type and "multipart" in request.content_type:
        # Handle multipart form data (for file uploads)
        try:
            raw = {
                k: v[0] if isinstance(v, (list, tuple)) and len(v) == 1 else v
                for k, v in request.POST.lists()
            }
        except Exception:
            raw = dict(request.POST)

        # Try to parse JSON-encoded fields (common when frontend stringifies objects)
        for k, v in list(raw.items()):
            if isinstance(v, str) and (v.startswith("{") or v.startswith("[")):
                try:
                    raw[k] = json.loads(v)
                except Exception:
                    pass

        payload = raw
        # Include uploaded file so serializer can accept it
        if request.FILES.get("product_image"):
            payload["product_image"] = request.FILES.get("product_image")
    else:
        # Handle JSON data
        try:
            payload = json.loads(request.body.decode("utf-8")) if request.body else {}
        except Exception:
            return Response({"detail": "Invalid JSON"}, status=400)

    # Normalize account FK alias
    if isinstance(payload.get("account"), dict):
        payload["account_id"] = payload["account"].get("id") or payload["account"].get("pk")
    elif payload.get("account"):
        payload["account_id"] = payload.get("account")

    company = getattr(user, "company", None)
    if company is None:
        return Response({"errors": {"company": ["User does not belong to a company"]}}, status=400)

    # Log the incoming payload for debugging
    import logging

    logger = logging.getLogger(__name__)
    logger.info(f"Estimate create payload: {payload}")

    # Validate sales_query_id exists and belongs to user's company
    sales_query_id = payload.get("sales_query_id")
    if sales_query_id:
        logger.info(f"Creating estimate for sales_query_id: {sales_query_id}")
        from sales_queries.models import SalesQuery

        try:
            sq_exists = SalesQuery.objects.filter(
                id=sales_query_id, account__company=company
            ).exists()
            logger.info(f"Sales Lead {sales_query_id} exists for company {company.id}: {sq_exists}")
            if not sq_exists:
                logger.error(f"Sales Lead {sales_query_id} not found for company {company.id}")
                return Response(
                    {"detail": f"Sales Lead {sales_query_id} not found or not accessible"},
                    status=400,
                )
        except Exception as e:
            logger.error(f"Error checking sales lead {sales_query_id}: {str(e)}")
            return Response({"detail": f"Error validating sales lead: {str(e)}"}, status=400)
    else:
        logger.info("Creating standalone estimate (no sales_query_id)")

    serializer = EstimateVoucherSerializer(data=payload)
    if serializer.is_valid():
        # Save estimate with company and created_by
        obj = serializer.save(company=company, created_by=user)
        return Response(EstimateVoucherSerializer(obj).data, status=201)

    # Log detailed validation errors for debugging
    logger.error(f"Estimate validation failed: {serializer.errors}")

    # Return errors in a format the frontend can understand
    error_detail = "Validation failed: " + str(serializer.errors)
    return Response({"detail": error_detail, "errors": serializer.errors}, status=400)


@api_view(["GET", "PUT", "PATCH", "DELETE"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def estimate_detail_view(request, pk):
    """Retrieve, update, or delete an EstimateVoucher."""
    user = request.user
    try:
        obj = (
            EstimateVoucher.objects.select_related("company", "account", "created_by", "updated_by")
            .prefetch_related("line_items")
            .get(pk=pk)
        )
    except EstimateVoucher.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)

    # Ensure user can only access their company's estimates
    if getattr(user, "company_id", None) and obj.company_id != user.company_id:
        return Response({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return Response(EstimateVoucherSerializer(obj).data, status=200)

    if request.method == "DELETE":
        obj.delete()
        return HttpResponse(status=204)

    # PUT or PATCH - Update estimate
    # Accept either application/json or multipart/form-data for image uploads
    if request.content_type and "multipart" in request.content_type:
        # Handle multipart form data (for file uploads)
        try:
            raw = {
                k: v[0] if isinstance(v, (list, tuple)) and len(v) == 1 else v
                for k, v in request.POST.lists()
            }
        except Exception:
            raw = dict(request.POST)

        # Try to parse JSON-encoded fields (common when frontend stringifies objects)
        for k, v in list(raw.items()):
            if isinstance(v, str) and (v.startswith("{") or v.startswith("[")):
                try:
                    raw[k] = json.loads(v)
                except Exception:
                    pass

        payload = raw
        # Include uploaded file so serializer can accept it
        if request.FILES.get("product_image"):
            payload["product_image"] = request.FILES.get("product_image")
    else:
        # Handle JSON data
        try:
            payload = json.loads(request.body.decode("utf-8")) if request.body else {}
        except Exception:
            return Response({"detail": "Invalid JSON"}, status=400)

    # Normalize account FK alias
    if isinstance(payload.get("account"), dict):
        payload["account_id"] = payload["account"].get("id") or payload["account"].get("pk")
    elif payload.get("account"):
        payload["account_id"] = payload.get("account")

    partial = request.method == "PATCH"
    serializer = EstimateVoucherSerializer(obj, data=payload, partial=partial)
    if serializer.is_valid():
        updated = serializer.save(updated_by=user)
        return Response(EstimateVoucherSerializer(updated).data, status=200)

    return Response({"errors": serializer.errors}, status=400)


# =============================================================================
# Estimate PDF Generation Endpoint
# =============================================================================


@api_view(["POST"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def estimate_pdf_view(request):
    """
    Generate a pixel-perfect estimate PDF using the template overlay approach.

    POST /api/estimates/pdf/

    Request JSON:
    {
        "item_name": "BRACELET 18K",
        "line_items": [
            {
                "particulars": "DIAMOND",
                "shape": "RD",
                "colour": "F-G",
                "clarity": "VVS",
                "pc": 234,
                "weight": 14.52,
                "unit": "CT",
                "rate": 48000.00,
                "amount": 696960.00
            },
            ...
        ],
        "totals": {
            "taxable_value": 1139484.00,
            "gst": 34184.52,
            "grand_total": 1173668.52
        },
        "image_base64": "..." (optional)
    }

    Returns:
        PDF file as application/pdf with download headers
    """
    from .estimate_pdf import generate_estimate_pdf
    import base64

    try:
        payload = json.loads(request.body.decode("utf-8")) if request.body else {}
    except Exception:
        return JsonResponse({"detail": "Invalid JSON"}, status=400)

    # Validate required fields
    item_name = payload.get("item_name", "")
    if not item_name:
        return JsonResponse({"detail": "item_name is required"}, status=400)

    line_items = payload.get("line_items", [])
    if not isinstance(line_items, list):
        return JsonResponse({"detail": "line_items must be a list"}, status=400)

    totals = payload.get("totals", {})
    if not isinstance(totals, dict):
        return JsonResponse({"detail": "totals must be an object"}, status=400)

    # Ensure totals have required fields with defaults
    totals_normalized = {
        "taxable_value": float(totals.get("taxable_value", 0) or 0),
        "gst": float(totals.get("gst", 0) or 0),
        "grand_total": float(totals.get("grand_total", 0) or 0),
    }

    # Normalize line items
    normalized_items = []
    for item in line_items:
        normalized_items.append(
            {
                "particulars": str(item.get("particulars", "")),
                "shape": str(item.get("shape", "")),
                "colour": str(item.get("colour", "")),
                "clarity": str(item.get("clarity", "")),
                "pc": item.get("pc"),
                "weight": item.get("weight"),
                "unit": str(item.get("unit", "")),
                "rate": item.get("rate"),
                "amount": float(item.get("amount", 0) or 0),
            }
        )

    # Handle optional image
    image_data = None
    image_base64 = payload.get("image_base64")
    if image_base64:
        try:
            # Remove data URL prefix if present
            if "," in image_base64:
                image_base64 = image_base64.split(",")[1]
            image_data = base64.b64decode(image_base64)
        except Exception as e:
            # Log but don't fail - image is optional
            print(f"Warning: Could not decode image: {e}")

    try:
        # Generate PDF
        pdf_bytes = generate_estimate_pdf(
            item_name=item_name,
            line_items=normalized_items,
            totals=totals_normalized,
            image_data=image_data,
            size_details=payload.get("size_details"),
        )

        # Create response with PDF
        response = HttpResponse(pdf_bytes, content_type="application/pdf")

        # Generate filename
        safe_name = slugify(item_name)[:30] or "estimate"
        filename = f"Estimate_{safe_name}.pdf"

        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"] = len(pdf_bytes)

        return response

    except FileNotFoundError as e:
        return JsonResponse({"detail": str(e)}, status=500)
    except Exception as e:
        import traceback

        traceback.print_exc()
        return JsonResponse({"detail": f"PDF generation failed: {str(e)}"}, status=500)


@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def estimate_pdf_calibration_view(request):
    """
    Generate a calibration PDF with grid overlay for coordinate tuning.

    GET /api/estimates/pdf/calibration/

    Use this endpoint during development to fine-tune the coordinate
    constants in estimate_pdf.py.
    """
    from .estimate_pdf import generate_calibration_pdf

    try:
        pdf_bytes = generate_calibration_pdf()

        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="estimate_calibration.pdf"'
        response["Content-Length"] = len(pdf_bytes)

        return response

    except Exception as e:
        import traceback

        traceback.print_exc()
        return JsonResponse({"detail": f"Calibration PDF generation failed: {str(e)}"}, status=500)


@csrf_exempt
@api_view(["POST"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def estimate_landscape_pdf_view(request):
    """Generate landscape estimate PDF."""
    try:
        from .estimate_landscape_pdf import generate_estimate_landscape_pdf
        import base64

        # Handle both DRF request.data and raw JSON
        try:
            if hasattr(request, "data") and request.data:
                payload = request.data
            else:
                import json

                payload = json.loads(request.body.decode("utf-8")) if request.body else {}
        except Exception as e:
            return Response({"detail": f"Invalid request data: {str(e)}"}, status=400)

        # Validate required fields
        item_name = payload.get("item_name", "")
        if not item_name:
            return Response({"detail": "item_name is required"}, status=400)

        line_items = payload.get("line_items", [])
        if not isinstance(line_items, list):
            return Response({"detail": "line_items must be a list"}, status=400)

        totals = payload.get("totals", {})
        if not isinstance(totals, dict):
            return Response({"detail": "totals must be an object"}, status=400)

        # Normalize totals
        totals_normalized = {
            "taxable_value": float(totals.get("taxable_value", 0) or 0),
            "gst": float(totals.get("gst", 0) or 0),
            "discount": float(totals.get("discount", 0) or 0),
            "discount_percent": float(totals.get("discount_percent", 0) or 0),
            "grand_total": float(totals.get("grand_total", 0) or 0),
        }

        # Normalize line items
        normalized_items = []
        for i, item in enumerate(line_items):
            try:
                normalized_items.append(
                    {
                        "particulars": str(item.get("particulars", "")),
                        "shape": str(item.get("shape", "")),
                        "colour": str(item.get("colour", "")),
                        "clarity": str(item.get("clarity", "")),
                        "pc": item.get("pc"),
                        "weight": item.get("weight"),
                        "unit": str(item.get("unit", "")),
                        "rate": item.get("rate"),
                        "amount": float(item.get("amount", 0) or 0),
                    }
                )
            except Exception as e:
                return Response({"detail": f"Invalid line item at index {i}: {str(e)}"}, status=400)

        # Handle optional fields
        image_data = None
        image_base64 = payload.get("image_base64")
        if image_base64:
            try:
                if "," in image_base64:
                    image_base64 = image_base64.split(",")[1]
                image_data = base64.b64decode(image_base64)
            except Exception:
                pass  # Image is optional

        customer_details = payload.get("customer_details")
        jewellery_details = payload.get("jewellery_details")
        estimate_details = payload.get("estimate_details")

        # If estimate_id is provided, auto-populate customer, jewellery, and estimate details
        estimate_id = payload.get("estimate_id")
        print(f"[PDF DEBUG] estimate_id received: {estimate_id}")
        if estimate_id:
            try:
                from .models import EstimateVoucher

                estimate = EstimateVoucher.objects.select_related(
                    "account", "account__contact", "sub_account_record", "sales_query"
                ).get(id=estimate_id)

                # Always fetch phone from account contact regardless of what UI sends
                phone = ""
                if (
                    estimate.account
                    and hasattr(estimate.account, "contact")
                    and estimate.account.contact
                    and estimate.account.contact.phone
                ):
                    phone = estimate.account.contact.phone.strip()
                    print(f"[PDF DEBUG] Phone from account.contact: {phone}")
                elif estimate.phone_number:
                    phone = estimate.phone_number
                    print(f"[PDF DEBUG] Phone from estimate.phone_number: {phone}")
                else:
                    print(f"[PDF DEBUG] No phone found!")

                sub_account_name = ""
                if estimate.sub_account_record:
                    sub_account_name = estimate.sub_account_record.sub_account_name
                elif estimate.sub_account:
                    sub_account_name = estimate.sub_account

                customer_details = {
                    "main_account": estimate.account.account_name
                    if estimate.account
                    else (customer_details or {}).get("main_account", ""),
                    "sub_account": sub_account_name
                    or (customer_details or {}).get("sub_account", ""),
                    "phone": phone,
                    "sales_person_name": estimate.sales_person_name
                    or (estimate.sales_query.sales_person if estimate.sales_query else "")
                    or (customer_details or {}).get("sales_person_name", ""),
                    "nsj_representative": estimate.nsj_representative
                    or (customer_details or {}).get("nsj_representative", ""),
                }
                print(f"[PDF DEBUG] Final customer_details: {customer_details}")
            except Exception as e:
                print(f"Warning: Could not auto-populate customer details: {e}")

        # If estimate_id is provided, auto-populate jewellery details
        if estimate_id and not jewellery_details:
            try:
                from .models import EstimateVoucher

                estimate = EstimateVoucher.objects.get(id=estimate_id)

                jewellery_details = {
                    "jewellery_type": estimate.jewellery_type or "",
                    "size_details": estimate.size_details or "",
                    "gold_quality": estimate.gold_quality or "",
                }
            except Exception as e:
                print(f"Warning: Could not auto-populate jewellery details: {e}")

        # If estimate_id is provided, auto-populate estimate details (date, expiry_date)
        if estimate_id:
            try:
                from .models import EstimateVoucher

                estimate = EstimateVoucher.objects.get(id=estimate_id)

                # Format dates for PDF
                date_str = ""
                if estimate.date:
                    date_str = estimate.date.strftime("%d-%m-%Y")

                expiry_date_str = ""
                if estimate.expiry_date:
                    expiry_date_str = estimate.expiry_date.strftime("%d-%m-%Y")

                estimate_details = {
                    "date": date_str or (estimate_details or {}).get("date", ""),
                    "expiry_date": expiry_date_str or (estimate_details or {}).get("expiry_date", ""),
                }
                print(f"[PDF DEBUG] estimate_details from DB: {estimate_details}")
            except Exception as e:
                print(f"Warning: Could not auto-populate estimate details: {e}")

        # Generate PDF
        pdf_bytes = generate_estimate_landscape_pdf(
            item_name=item_name,
            line_items=normalized_items,
            totals=totals_normalized,
            image_data=image_data,
            customer_details=customer_details,
            jewellery_details=jewellery_details,
            estimate_details=estimate_details,
        )

        # Return PDF response
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        safe_name = slugify(item_name)[:30] or "estimate"
        filename = f"Estimate_Landscape_{safe_name}.pdf"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"] = len(pdf_bytes)
        return response

    except Exception as e:
        import traceback

        error_msg = f"Landscape PDF generation failed: {str(e)}"
        print(f"ERROR: {error_msg}")
        print(f"TRACEBACK: {traceback.format_exc()}")
        return Response({"detail": error_msg}, status=500)


@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def estimate_landscape_pdf_calibration_view(request):
    """
    Generate a calibration PDF with grid overlay for coordinate tuning (landscape version).

    GET /api/estimates/landscape-pdf/calibration/

    Use this endpoint during development to fine-tune the coordinate
    constants in estimate_landscape_pdf.py.
    """
    from .estimate_landscape_pdf import generate_calibration_landscape_pdf

    try:
        pdf_bytes = generate_calibration_landscape_pdf()

        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = (
            'attachment; filename="estimate_landscape_calibration.pdf"'
        )
        response["Content-Length"] = len(pdf_bytes)

        return response

    except Exception as e:
        import traceback

        traceback.print_exc()
        return Response(
            {"detail": f"Landscape calibration PDF generation failed: {str(e)}"}, status=500
        )


# =============================================================================
# Dashboard Statistics Endpoint
# =============================================================================


@csrf_exempt
@require_http_methods(["GET"])
def dashboard_stats_view(request):
    """
    Get dashboard statistics for the main dashboard.
    Returns active orders, completed orders, total sales, and average order value.

    Supports department filtering via ?department=<department_id> query parameter.
    This endpoint is public (no login required) for dashboard display.
    """
    from django.utils import timezone
    from datetime import timedelta

    try:
        # Get all orders (or filter by company if user is authenticated)
        orders = Voucher.objects.all()

        if request.user.is_authenticated and getattr(request.user, "company_id", None):
            orders = orders.filter(company_id=request.user.company_id)

        # Filter by department if provided
        department_id = request.GET.get("department")
        department_name = None
        if department_id:
            try:
                from core.models import Department

                department = Department.objects.get(id=department_id)
                department_name = department.name

                # Filter orders based on department
                # This is a placeholder - adjust based on your actual data model
                # For now, we'll just add the department info to the response
                # You may need to add a department field to Voucher model or filter differently
            except Department.DoesNotExist:
                pass

        # Calculate statistics
        total_orders = orders.count()

        # Active orders = orders without advance payment (pending)
        # Use safe field access
        try:
            active_orders = orders.filter(advance_payment_received="NO").count()
        except:
            active_orders = 0

        # Completed orders = orders with advance payment received
        try:
            completed_orders = orders.filter(advance_payment_received="YES").count()
        except:
            completed_orders = 0

        # Calculate total sales value - simplified to avoid field errors
        total_sales = 0
        avg_order_value = 0

        # Calculate trends (compare with last month)
        today = timezone.now().date()
        last_month_start = (today.replace(day=1) - timedelta(days=1)).replace(day=1)
        last_month_end = today.replace(day=1) - timedelta(days=1)

        try:
            this_month_orders = orders.filter(created_at__date__gte=today.replace(day=1)).count()
            last_month_orders = orders.filter(
                created_at__date__gte=last_month_start, created_at__date__lte=last_month_end
            ).count()

            # Calculate trend percentage
            if last_month_orders > 0:
                orders_trend = round(
                    ((this_month_orders - last_month_orders) / last_month_orders) * 100, 1
                )
            else:
                orders_trend = 0 if this_month_orders == 0 else 100
        except:
            orders_trend = 0

        # Monthly data for chart (last 12 months)
        monthly_data = []
        try:
            for i in range(11, -1, -1):
                month_date = today - timedelta(days=i * 30)
                month_start = month_date.replace(day=1)
                if month_date.month == 12:
                    month_end = month_date.replace(
                        year=month_date.year + 1, month=1, day=1
                    ) - timedelta(days=1)
                else:
                    month_end = month_date.replace(month=month_date.month + 1, day=1) - timedelta(
                        days=1
                    )

                month_orders = orders.filter(
                    created_at__date__gte=month_start, created_at__date__lte=month_end
                ).count()

                monthly_data.append({"month": month_start.strftime("%b"), "orders": month_orders})
        except:
            # Fallback to empty monthly data
            for i in range(12):
                month_date = today - timedelta(days=i * 30)
                monthly_data.append({"month": month_date.strftime("%b"), "orders": 0})

        response_data = {
            "active_orders": active_orders,
            "completed_orders": completed_orders,
            "total_orders": total_orders,
            "total_sales": round(total_sales, 2),
            "avg_order_value": round(avg_order_value, 2),
            "orders_trend": orders_trend,
            "monthly_data": monthly_data,
        }

        # Add department info if filtered
        if department_name:
            response_data["department"] = department_name

        return JsonResponse(response_data)

    except Exception as e:
        # Return safe defaults on any error
        import sys

        print(f"Dashboard stats error: {e}", file=sys.stderr)
        return JsonResponse(
            {
                "active_orders": 0,
                "completed_orders": 0,
                "total_orders": 0,
                "total_sales": 0,
                "avg_order_value": 0,
                "orders_trend": 0,
                "monthly_data": [],
                "error": str(e),
            }
        )


# =============================================================================
# Live Rates API Endpoints (Gold, Silver, Platinum, Currency Exchange)
# =============================================================================


@csrf_exempt
@require_http_methods(["GET"])
def gold_rate_view(request):
    """
    Get current gold rate (24K per gram in INR).
    Fetches from goldapi.io and caches the result.
    """
    import requests
    from django.core.cache import cache

    # Try to get cached rate first (cache for 5 minutes)
    cached_rate = cache.get("gold_rate_24k")
    if cached_rate:
        cached_rate["cached"] = True
        return JsonResponse(cached_rate)

    try:
        import os

        gold_api_key = os.getenv("GOLD_API_KEY")

        # Fetch from goldapi.io - returns price_gram_24k directly in INR
        response = requests.get(
            "https://www.goldapi.io/api/XAU/INR",
            headers={"x-access-token": gold_api_key},
            timeout=10,
        )

        if response.status_code == 200:
            data = response.json()

            # Extract 24K price per gram (already in INR)
            price_24k = data.get("price_gram_24k", 0)
            change = data.get("ch", 0)  # Daily change
            change_percent = data.get("chp", 0)  # Daily change percent

            result = {
                "price_per_gram_24k": round(price_24k, 2),
                "change": round(change, 2),
                "change_percent": round(change_percent, 2),
                "source": "goldapi.io",
                "cached": False,
            }

            # Cache for 5 minutes
            cache.set("gold_rate_24k", result, 300)

            return JsonResponse(result)
    except Exception:
        pass

    # Fallback to approximate rate if API fails
    fallback_result = {
        "price_per_gram_24k": 7900,
        "change": 0,
        "change_percent": 0,
        "source": "fallback",
        "cached": False,
    }

    return JsonResponse(fallback_result)


@csrf_exempt
@require_http_methods(["GET"])
def live_rates_view(request):
    """
    Get comprehensive live rates for all materials and currencies.
    Includes Gold, Silver, Platinum, and USD-INR exchange rate.
    """
    import requests
    from django.core.cache import cache

    # Try to get cached rates first (cache for 5 minutes)
    cached_rates = cache.get("live_rates_comprehensive")
    if cached_rates:
        cached_rates["cached"] = True
        return JsonResponse(cached_rates)

    rates_data = {
        "timestamp": None,
        "cached": False,
        "gold": {},
        "silver": {},
        "platinum": {},
        "exchange_rate": {},
        "sources": [],
    }

    # 1. Get USD-INR Exchange Rate (Free API)
    try:
        # Using exchangerate-api.com (free tier: 1500 requests/month)
        response = requests.get("https://api.exchangerate-api.com/v4/latest/USD", timeout=10)
        if response.status_code == 200:
            data = response.json()
            inr_rate = data.get("rates", {}).get("INR", 83.0)
            rates_data["exchange_rate"] = {
                "usd_to_inr": round(inr_rate, 2),
                "source": "exchangerate-api.com",
                "last_updated": data.get("date", ""),
            }
            rates_data["sources"].append("exchangerate-api.com")
    except Exception:
        # Fallback exchange rate
        rates_data["exchange_rate"] = {"usd_to_inr": 83.0, "source": "fallback", "last_updated": ""}

    # 2. Get Gold Rate (using goldapi.io with environment variable)
    try:
        import os

        gold_api_key = os.getenv("GOLD_API_KEY", "goldapi-h03osmjmslj39-io")  # Fallback to old key

        # First get USD price, then convert to INR
        response = requests.get(
            "https://www.goldapi.io/api/XAU/USD",
            headers={"x-access-token": gold_api_key},
            timeout=10,
        )
        if response.status_code == 200:
            data = response.json()
            usd_to_inr = rates_data["exchange_rate"]["usd_to_inr"]

            # Convert USD prices to INR (per gram)
            price_24k_usd = data.get("price_gram_24k", 0)
            price_22k_usd = data.get("price_gram_22k", 0)
            price_18k_usd = data.get("price_gram_18k", 0)

            rates_data["gold"] = {
                "price_per_gram_24k": round(price_24k_usd * usd_to_inr, 2)
                if price_24k_usd
                else 7900,
                "price_per_gram_22k": round(price_22k_usd * usd_to_inr, 2)
                if price_22k_usd
                else 7200,
                "price_per_gram_18k": round(price_18k_usd * usd_to_inr, 2)
                if price_18k_usd
                else 5900,
                "change": round(data.get("ch", 0), 2),
                "change_percent": round(data.get("chp", 0), 2),
                "source": "goldapi.io",
                "usd_prices": {
                    "price_per_gram_24k_usd": round(price_24k_usd, 2),
                    "price_per_gram_22k_usd": round(price_22k_usd, 2),
                    "price_per_gram_18k_usd": round(price_18k_usd, 2),
                },
            }
            rates_data["sources"].append("goldapi.io")
    except Exception:
        rates_data["gold"] = {
            "price_per_gram_24k": 7900,
            "price_per_gram_22k": 7200,
            "price_per_gram_18k": 5900,
            "change": 0,
            "change_percent": 0,
            "source": "fallback",
        }

    # 3. Get Silver Rate (using goldapi.io with same API key)
    try:
        import os

        gold_api_key = os.getenv("GOLD_API_KEY", "goldapi-h03osmjmslj39-io")

        response = requests.get(
            "https://www.goldapi.io/api/XAG/USD",
            headers={"x-access-token": gold_api_key},
            timeout=10,
        )
        if response.status_code == 200:
            data = response.json()
            usd_to_inr = rates_data["exchange_rate"]["usd_to_inr"]

            # Get silver price per gram in USD (use price_gram_24k for silver)
            silver_usd_per_gram = data.get("price_gram_24k", 0)
            silver_usd_per_oz = data.get("price", 0)

            # Convert to INR per gram
            silver_inr_per_gram = silver_usd_per_gram * usd_to_inr

            rates_data["silver"] = {
                "price_per_gram": round(silver_inr_per_gram, 2),
                "price_per_oz_usd": round(silver_usd_per_oz, 2),
                "change": round(data.get("ch", 0), 2),
                "change_percent": round(data.get("chp", 0), 2),
                "source": "goldapi.io",
                "usd_prices": {
                    "price_per_gram_usd": round(silver_usd_per_gram, 4),
                },
            }
            rates_data["sources"].append("goldapi.io")
        else:
            raise Exception("Silver API failed")

    except Exception:
        # Fallback silver rate (approximate)
        rates_data["silver"] = {
            "price_per_gram": 95.0,
            "price_per_oz_usd": 25.0,
            "source": "fallback",
        }

    # 4. Get Platinum Rate (using goldapi.io with same API key)
    try:
        import os

        gold_api_key = os.getenv("GOLD_API_KEY", "goldapi-h03osmjmslj39-io")

        response = requests.get(
            "https://www.goldapi.io/api/XPT/USD",
            headers={"x-access-token": gold_api_key},
            timeout=10,
        )
        if response.status_code == 200:
            data = response.json()
            usd_to_inr = rates_data["exchange_rate"]["usd_to_inr"]

            # Get platinum price per gram in USD (use price_gram_24k for platinum)
            platinum_usd_per_gram = data.get("price_gram_24k", 0)
            platinum_usd_per_oz = data.get("price", 0)

            # Convert to INR per gram
            platinum_inr_per_gram = platinum_usd_per_gram * usd_to_inr

            rates_data["platinum"] = {
                "price_per_gram": round(platinum_inr_per_gram, 2),
                "price_per_oz_usd": round(platinum_usd_per_oz, 2),
                "change": round(data.get("ch", 0), 2),
                "change_percent": round(data.get("chp", 0), 2),
                "source": "goldapi.io",
                "usd_prices": {
                    "price_per_gram_usd": round(platinum_usd_per_gram, 4),
                },
            }
            rates_data["sources"].append("goldapi.io")
        else:
            raise Exception("Platinum API failed")

    except Exception:
        # Fallback platinum rate (approximate)
        rates_data["platinum"] = {
            "price_per_gram": 2700.0,
            "price_per_oz_usd": 1000.0,
            "source": "fallback",
        }

    # Add timestamp
    from datetime import datetime

    rates_data["timestamp"] = datetime.now().isoformat()

    # Cache for 5 minutes
    cache.set("live_rates_comprehensive", rates_data, 300)

    return JsonResponse(rates_data)


@csrf_exempt
@require_http_methods(["GET"])
def currency_exchange_view(request):
    """
    Get current USD to INR exchange rate.
    """
    import requests
    from django.core.cache import cache

    # Try to get cached rate first (cache for 10 minutes)
    cached_rate = cache.get("usd_inr_exchange")
    if cached_rate:
        cached_rate["cached"] = True
        return JsonResponse(cached_rate)

    try:
        # Using exchangerate-api.com (free tier)
        response = requests.get("https://api.exchangerate-api.com/v4/latest/USD", timeout=10)
        if response.status_code == 200:
            data = response.json()
            inr_rate = data.get("rates", {}).get("INR", 83.0)

            result = {
                "usd_to_inr": round(inr_rate, 2),
                "source": "exchangerate-api.com",
                "last_updated": data.get("date", ""),
                "cached": False,
            }

            # Cache for 10 minutes
            cache.set("usd_inr_exchange", result, 600)

            return JsonResponse(result)
    except Exception:
        pass

    # Fallback exchange rate
    fallback_result = {
        "usd_to_inr": 83.0,
        "source": "fallback",
        "last_updated": "",
        "cached": False,
    }

    return JsonResponse(fallback_result)


@api_view(["POST"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def calculate_material_price_view(request):
    """
    Calculate material prices based on current live rates.
    Accepts material type, quantity, and other parameters.
    """
    try:
        data = request.data
        material_type = data.get("material_type", "").lower()

        # Get exchange rate (use provided or default)
        exchange_rate = float(data.get("exchange_rate", 83.0))

        result = {
            "material_type": material_type,
            "exchange_rate": exchange_rate,
            "calculation_date": timezone.now().isoformat(),
        }

        if material_type == "diamond":
            # Diamond calculations
            carat = float(data.get("quantity", 0))
            price_per_ct_usd = float(data.get("price_per_ct_usd", 0))

            price_per_ct_inr = price_per_ct_usd * exchange_rate
            total_inr = carat * price_per_ct_inr

            result.update(
                {
                    "quantity": carat,
                    "unit": "carat",
                    "price_per_ct_usd": price_per_ct_usd,
                    "price_per_ct_inr": round(price_per_ct_inr, 2),
                    "total_inr": round(total_inr, 2),
                    "calculation_formula": "Total = Carat Ã— Price/CT (INR)",
                }
            )

        elif material_type == "gemstone":
            # Gemstone calculations (similar to diamond)
            carat_weight = float(data.get("quantity", 0))
            price_per_ct_usd = float(data.get("price_per_ct_usd", 0))

            price_per_ct_inr = price_per_ct_usd * exchange_rate
            total_inr = carat_weight * price_per_ct_inr

            result.update(
                {
                    "quantity": carat_weight,
                    "unit": "carat",
                    "price_per_ct_usd": price_per_ct_usd,
                    "price_per_ct_inr": round(price_per_ct_inr, 2),
                    "total_inr": round(total_inr, 2),
                    "calculation_formula": "Total = Carat Weight Ã— Price/CT (INR)",
                }
            )

        elif material_type in ["gold", "silver", "platinum"]:
            # Metal calculations
            weight = float(data.get("quantity", 0))
            price_per_gram_inr = float(data.get("price_per_gram_inr", 0))

            total_inr = weight * price_per_gram_inr

            result.update(
                {
                    "quantity": weight,
                    "unit": "grams",
                    "price_per_gram_inr": price_per_gram_inr,
                    "total_inr": round(total_inr, 2),
                    "calculation_formula": "Total = Weight Ã— Price/Gram (INR)",
                }
            )

        else:
            return JsonResponse({"error": "Unsupported material type"}, status=400)

        return JsonResponse(result)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# ============================================================================
# Raw Material Purchase Views
# ============================================================================
from .models import RawMaterialPurchase
from .serializers import RawMaterialPurchaseSerializer


@api_view(["GET", "POST"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def raw_material_purchase_list_create(request):
    """List all raw material purchases or create a new one."""
    # Fix: Properly load the company relationship
    try:
        from users.models import User

        user = User.objects.select_related("company").get(id=request.user.id)
        company = user.company
    except User.DoesNotExist:
        return Response({"error": "User not found"}, status=400)

    if not company:
        return Response({"error": "No company found"}, status=400)

    if request.method == "GET":
        purchases = RawMaterialPurchase.objects.filter(company=company)

        # Pagination
        page = int(request.GET.get("page", 1))
        page_size = int(request.GET.get("page_size", 20))
        start = (page - 1) * page_size
        end = start + page_size

        total = purchases.count()
        purchases = purchases[start:end]

        serializer = RawMaterialPurchaseSerializer(
            purchases, many=True, context={"request": request}
        )
        return JsonResponse(
            {
                "results": serializer.data,
                "count": total,
                "page": page,
                "page_size": page_size,
            }
        )

    elif request.method == "POST":
        try:
            # Check if books are closed for the date
            purchase_date = request.data.get("date")
            if purchase_date:
                book_close = DailyBookClose.objects.filter(
                    company=company, date=purchase_date, is_closed=True
                ).first()
                if book_close:
                    return JsonResponse(
                        {
                            "error": f"Books are closed for {purchase_date}. Cannot create new purchases."
                        },
                        status=400,
                    )

            # Convert string boolean values to actual booleans
            data = request.data.copy() if hasattr(request.data, "copy") else dict(request.data)

            # Handle boolean fields that might come as strings
            for bool_field in ["mode_cash", "mode_bill"]:
                if bool_field in data:
                    val = data[bool_field]
                    if isinstance(val, str):
                        data[bool_field] = val.lower() in ("true", "1", "yes")

            serializer = RawMaterialPurchaseSerializer(data=data, context={"request": request})
            if serializer.is_valid():
                with transaction.atomic():
                    purchase = serializer.save(
                        company=company,
                        created_by=request.user,
                    )
                    # Handle file upload
                    if request.FILES.get("proof_image"):
                        purchase.proof_image = request.FILES["proof_image"]
                        purchase.save()

                    # Auto-generate DIA. ID if not provided
                    if not purchase.dia_id:
                        purchase.dia_id = f"DIA-{purchase.id.hex[:8].upper()}"
                        purchase.save()

                    # Create inventory record automatically
                    # Determine the appropriate weight/carat field based on material type
                    available_quantity = 0
                    if purchase.material_type:
                        material_name = purchase.material_type.name.lower()
                        if material_name == "diamond":
                            available_quantity = purchase.carat or 0
                        elif material_name == "gemstone":
                            available_quantity = (
                                purchase.gemstone_carat_weight or purchase.carat_weight or 0
                            )
                        elif material_name in ["gold", "silver", "platinum"]:
                            # Use new unified metal_weight_grams field, fallback to gold_weight
                            available_quantity = (
                                purchase.metal_weight_grams or purchase.gold_weight or 0
                            )
                        else:
                            # Fallback to carat field
                            available_quantity = purchase.carat or purchase.carat_weight or 0
                    else:
                        # No material type specified, try common fields
                        available_quantity = (
                            purchase.carat
                            or purchase.carat_weight
                            or purchase.metal_weight_grams
                            or purchase.gold_weight
                            or 0
                        )

                    RawMaterialInventory.objects.create(
                        company=company,
                        purchase=purchase,
                        available_carat=available_quantity,
                        original_carat=available_quantity,
                        issued_carat=0,
                        status="AVAILABLE",
                    )

                # Refresh the purchase object to get calculated values
                purchase.refresh_from_db()

                return JsonResponse(
                    RawMaterialPurchaseSerializer(purchase, context={"request": request}).data,
                    status=201,
                )

            # Log validation errors for debugging
            print(f"Validation errors: {serializer.errors}")
            return JsonResponse({"errors": serializer.errors}, status=400)

        except Exception as e:
            import traceback

            print(f"Raw Material Purchase Creation Error: {str(e)}")
            print(f"Traceback: {traceback.format_exc()}")
            return JsonResponse({"error": f"Internal server error: {str(e)}"}, status=500)


@api_view(["GET", "PUT", "PATCH", "DELETE"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def raw_material_purchase_detail(request, pk):
    """Retrieve, update, or delete a raw material purchase."""
    # Fix: Properly load the company relationship
    try:
        from users.models import User

        user = User.objects.select_related("company").get(id=request.user.id)
        company = user.company
    except User.DoesNotExist:
        return Response({"error": "User not found"}, status=400)

    if not company:
        return Response({"error": "No company found"}, status=400)

    try:
        purchase = RawMaterialPurchase.objects.get(pk=pk, company=company)
    except RawMaterialPurchase.DoesNotExist:
        return Response({"error": "Not found"}, status=404)

    if request.method == "GET":
        serializer = RawMaterialPurchaseSerializer(purchase, context={"request": request})
        return Response(serializer.data)

    elif request.method in ["PUT", "PATCH"]:
        # For DRF, use request.data instead of request.POST
        serializer = RawMaterialPurchaseSerializer(
            purchase,
            data=request.data,
            partial=(request.method == "PATCH"),
            context={"request": request},
        )
        if serializer.is_valid():
            purchase = serializer.save(updated_by=request.user)
            # Handle file upload
            if request.FILES.get("proof_image"):
                purchase.proof_image = request.FILES["proof_image"]
                purchase.save()
            return Response(
                RawMaterialPurchaseSerializer(purchase, context={"request": request}).data
            )
        return Response(serializer.errors, status=400)

    elif request.method == "DELETE":
        purchase.delete()
        return Response({"message": "Deleted"}, status=204)


@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def raw_material_purchase_aggregates(request):
    """Get aggregate statistics for raw material purchases."""
    # Fix: Properly load the company relationship
    try:
        from users.models import User

        user = User.objects.select_related("company").get(id=request.user.id)
        company = user.company
    except User.DoesNotExist:
        return Response({"error": "User not found"}, status=400)

    if not company:
        return Response({"error": "No company found"}, status=400)

    purchases = RawMaterialPurchase.objects.filter(company=company)
    total = purchases.count()

    # Recent 7 days
    seven_days_ago = timezone.now() - timedelta(days=7)
    recent_7_days = purchases.filter(created_at__gte=seven_days_ago).count()

    # Total value
    total_value = purchases.aggregate(total=Sum("total"))["total"] or 0

    return Response(
        {
            "total": total,
            "recent_7_days": recent_7_days,
            "total_value": float(total_value),
        }
    )


@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def raw_material_suppliers_view(request):
    """Get list of suppliers for raw material purchases."""
    # Fix: Properly load the company relationship
    try:
        from users.models import User

        user = User.objects.select_related("company").get(id=request.user.id)
        company = user.company
    except User.DoesNotExist:
        return Response({"error": "User not found"}, status=400)

    if not company:
        return Response({"error": "No company found"}, status=400)

    # Get all suppliers that have been used in raw material purchases
    from accounts.models import Account

    suppliers = (
        Account.objects.filter(company=company, group_code="SUPPLIER")
        .values("id", "account_name", "account_no")
        .order_by("account_name")
    )

    return Response({"suppliers": list(suppliers)})


@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def raw_material_field_schema(request):
    """Get field schema based on material type selection."""
    material_type = request.GET.get("material_type", "").lower()

    # Define field schemas for each material type
    field_schemas = {
        "diamond": {
            "fields": [
                {"name": "shape", "type": "select", "label": "Shape", "required": False},
                {"name": "carat", "type": "decimal", "label": "Carat", "required": False},
                {"name": "colour", "type": "select", "label": "Colour", "required": False},
                {"name": "clarity", "type": "select", "label": "Clarity", "required": False},
                {"name": "cut", "type": "text", "label": "Cut", "required": False},
                {"name": "pol", "type": "text", "label": "Polish", "required": False},
                {"name": "sym", "type": "text", "label": "Symmetry", "required": False},
                {"name": "flouro", "type": "text", "label": "Fluorescence", "required": False},
                {"name": "lab", "type": "select", "label": "Lab", "required": False},
                {"name": "cert_no", "type": "text", "label": "Certificate No", "required": False},
                {"name": "rap", "type": "decimal", "label": "RAP ($)", "required": False},
                {"name": "discount", "type": "decimal", "label": "Discount (%)", "required": False},
                {
                    "name": "price_per_ct_usd",
                    "type": "decimal",
                    "label": "Price/CT ($)",
                    "required": False,
                },
                {
                    "name": "exchange_rate",
                    "type": "decimal",
                    "label": "Exchange Rate",
                    "required": False,
                },
            ]
        },
        "gemstone": {
            "fields": [
                {
                    "name": "gemstone_type",
                    "type": "text",
                    "label": "Gemstone Type",
                    "required": False,
                },
                {"name": "gemstone_shape", "type": "text", "label": "Shape", "required": False},
                {
                    "name": "gemstone_carat_weight",
                    "type": "decimal",
                    "label": "Carat Weight",
                    "required": False,
                },
                {
                    "name": "gemstone_number_of_pieces",
                    "type": "number",
                    "label": "Number of Pieces",
                    "required": False,
                },
                {"name": "gemstone_color", "type": "text", "label": "Color", "required": False},
                {"name": "gemstone_clarity", "type": "text", "label": "Clarity", "required": False},
                {
                    "name": "gemstone_treatment",
                    "type": "text",
                    "label": "Treatment",
                    "required": False,
                },
                {"name": "gemstone_lab", "type": "text", "label": "Lab", "required": False},
                {
                    "name": "gemstone_certificate_number",
                    "type": "text",
                    "label": "Certificate Number",
                    "required": False,
                },
                {"name": "gemstone_cut", "type": "text", "label": "Cut", "required": False},
                {"name": "gemstone_size", "type": "text", "label": "Size", "required": False},
                {
                    "name": "gemstone_purchase_budget_total",
                    "type": "decimal",
                    "label": "Purchase Budget Total",
                    "required": False,
                },
                {
                    "name": "gemstone_purchase_budget_per_carat",
                    "type": "decimal",
                    "label": "Purchase Budget Per Carat",
                    "required": False,
                },
                {
                    "name": "gemstone_suggested_supplier",
                    "type": "text",
                    "label": "Suggested Supplier",
                    "required": False,
                },
                {
                    "name": "gemstone_additional_details",
                    "type": "textarea",
                    "label": "Additional Details",
                    "required": False,
                },
            ]
        },
        "gold": {
            "fields": [
                {
                    "name": "gold_purity",
                    "type": "select",
                    "label": "Gold Purity",
                    "required": False,
                    "options": ["24K", "22K", "18K", "14K", "10K"],
                },
                {
                    "name": "gold_weight",
                    "type": "decimal",
                    "label": "Weight (grams)",
                    "required": False,
                },
                {
                    "name": "gold_mode",
                    "type": "select",
                    "label": "Mode",
                    "required": False,
                    "options": ["Cast", "Mill", "Wire", "Sheet", "Tube"],
                },
                {
                    "name": "gold_suggested_supplier",
                    "type": "text",
                    "label": "Suggested Supplier",
                    "required": False,
                },
            ]
        },
        "silver": {
            "fields": [
                {
                    "name": "gold_purity",
                    "type": "select",
                    "label": "Silver Purity",
                    "required": False,
                    "options": ["999", "925", "900", "800"],
                },
                {
                    "name": "gold_weight",
                    "type": "decimal",
                    "label": "Weight (grams)",
                    "required": False,
                },
                {
                    "name": "gold_mode",
                    "type": "select",
                    "label": "Mode",
                    "required": False,
                    "options": ["Cast", "Mill", "Wire", "Sheet", "Tube"],
                },
                {
                    "name": "gold_suggested_supplier",
                    "type": "text",
                    "label": "Suggested Supplier",
                    "required": False,
                },
            ]
        },
        "platinum": {
            "fields": [
                {
                    "name": "gold_purity",
                    "type": "select",
                    "label": "Platinum Purity",
                    "required": False,
                    "options": ["950", "900", "850"],
                },
                {
                    "name": "gold_weight",
                    "type": "decimal",
                    "label": "Weight (grams)",
                    "required": False,
                },
                {
                    "name": "gold_mode",
                    "type": "select",
                    "label": "Mode",
                    "required": False,
                    "options": ["Cast", "Mill", "Wire", "Sheet"],
                },
                {
                    "name": "gold_suggested_supplier",
                    "type": "text",
                    "label": "Suggested Supplier",
                    "required": False,
                },
            ]
        },
    }

    # Common fields that appear for all material types
    common_fields = [
        {"name": "date", "type": "date", "label": "Date", "required": True},
        {"name": "supplier_id", "type": "select", "label": "Supplier", "required": True},
        {"name": "order_id", "type": "text", "label": "Order ID", "required": False},
        {"name": "master_size", "type": "text", "label": "Master Size", "required": False},
        {"name": "origin", "type": "text", "label": "Origin", "required": False},
        {"name": "carat_weight", "type": "decimal", "label": "Carat Weight", "required": False},
        {
            "name": "number_of_pieces",
            "type": "number",
            "label": "Number of Pieces",
            "required": False,
        },
        {
            "name": "certificate_number",
            "type": "text",
            "label": "Certificate Number",
            "required": False,
        },
        {"name": "size", "type": "text", "label": "Size", "required": False},
        {"name": "fluorescence", "type": "text", "label": "Fluorescence", "required": False},
        {
            "name": "purchase_budget_total",
            "type": "decimal",
            "label": "Purchase Budget Total",
            "required": False,
        },
        {
            "name": "purchase_budget_per_carat",
            "type": "decimal",
            "label": "Purchase Budget Per Carat",
            "required": False,
        },
        {
            "name": "suggested_supplier",
            "type": "text",
            "label": "Suggested Supplier",
            "required": False,
        },
        {
            "name": "additional_details",
            "type": "textarea",
            "label": "Additional Details",
            "required": False,
        },
        {"name": "proof_image", "type": "file", "label": "Proof Image", "required": False},
    ]

    # Get material-specific fields
    material_fields = field_schemas.get(material_type, {"fields": []})["fields"]

    # Combine common and material-specific fields
    all_fields = common_fields + material_fields

    return Response(
        {
            "material_type": material_type,
            "fields": all_fields,
            "available_material_types": list(field_schemas.keys()),
        }
    )


# ============================================================================
# Raw Material Inventory Views
# ============================================================================
from .models import RawMaterialInventory, RawMaterialIssuance, DailyBookClose, DailyReport
from .serializers import (
    RawMaterialInventorySerializer,
    RawMaterialIssuanceSerializer,
    DailyBookCloseSerializer,
    DailyReportSerializer,
)
from django.db import transaction


@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def raw_material_inventory_list(request):
    """List all inventory items with available stock."""
    # Fix: Properly load the company relationship
    try:
        from users.models import User

        user = User.objects.select_related("company").get(id=request.user.id)
        company = user.company
    except User.DoesNotExist:
        return Response({"error": "User not found"}, status=400)

    if not company:
        return Response({"error": "No company found"}, status=400)

    inventory = RawMaterialInventory.objects.filter(company=company).select_related(
        "purchase", "purchase__supplier", "purchase__shape", "purchase__colour", "purchase__clarity"
    )

    # Filter by status
    status = request.GET.get("status")
    if status:
        inventory = inventory.filter(status=status)

    # Only show available items by default
    show_all = request.GET.get("show_all", "false").lower() == "true"
    if not show_all:
        inventory = inventory.exclude(status="FULLY_ISSUED")

    # Pagination
    page = int(request.GET.get("page", 1))
    page_size = int(request.GET.get("page_size", 20))
    start = (page - 1) * page_size
    end = start + page_size

    total = inventory.count()
    inventory = inventory[start:end]

    serializer = RawMaterialInventorySerializer(inventory, many=True)
    return Response(
        {
            "results": serializer.data,
            "count": total,
            "page": page,
            "page_size": page_size,
        }
    )


@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def raw_material_inventory_summary(request):
    """Get inventory summary statistics."""
    # Fix: Properly load the company relationship
    try:
        from users.models import User

        user = User.objects.select_related("company").get(id=request.user.id)
        company = user.company
    except User.DoesNotExist:
        return Response({"error": "User not found"}, status=400)

    if not company:
        return Response({"error": "No company found"}, status=400)

    inventory = RawMaterialInventory.objects.filter(company=company)

    total_items = inventory.count()
    available_items = inventory.filter(status="AVAILABLE").count()
    partially_issued = inventory.filter(status="PARTIALLY_ISSUED").count()
    fully_issued = inventory.filter(status="FULLY_ISSUED").count()

    total_available_carat = inventory.aggregate(total=Sum("available_carat"))["total"] or 0
    total_original_carat = inventory.aggregate(total=Sum("original_carat"))["total"] or 0
    total_issued_carat = inventory.aggregate(total=Sum("issued_carat"))["total"] or 0

    return Response(
        {
            "total_items": total_items,
            "available_items": available_items,
            "partially_issued": partially_issued,
            "fully_issued": fully_issued,
            "total_available_carat": float(total_available_carat),
            "total_original_carat": float(total_original_carat),
            "total_issued_carat": float(total_issued_carat),
        }
    )


# ============================================================================
# Raw Material Issuance Views
# ============================================================================
@csrf_exempt
@require_http_methods(["GET", "POST"])
def raw_material_issuance_list_create(request):
    """List all issuances or create a new one."""
    company = getattr(request.user, "company", None)
    if not company:
        return JsonResponse({"error": "No company found"}, status=400)

    if request.method == "GET":
        issuances = RawMaterialIssuance.objects.filter(company=company).select_related(
            "inventory", "inventory__purchase", "job", "issued_by"
        )

        # Pagination
        page = int(request.GET.get("page", 1))
        page_size = int(request.GET.get("page_size", 20))
        start = (page - 1) * page_size
        end = start + page_size

        total = issuances.count()
        issuances = issuances[start:end]

        serializer = RawMaterialIssuanceSerializer(issuances, many=True)
        return JsonResponse(
            {
                "results": serializer.data,
                "count": total,
                "page": page,
                "page_size": page_size,
            }
        )

    elif request.method == "POST":
        # Check if books are closed for the date
        try:
            data = (
                json.loads(request.body)
                if request.content_type == "application/json"
                else request.POST.dict()
            )
        except json.JSONDecodeError:
            data = request.POST.dict()

        issue_date = data.get("date")
        if issue_date:
            book_close = DailyBookClose.objects.filter(
                company=company, date=issue_date, is_closed=True
            ).first()
            if book_close:
                return JsonResponse(
                    {"error": f"Books are closed for {issue_date}. Cannot create new issuances."},
                    status=400,
                )

        serializer = RawMaterialIssuanceSerializer(data=data)
        if serializer.is_valid():
            with transaction.atomic():
                inventory_id = data.get("inventory_id")
                issued_carat = decimal.Decimal(str(data.get("issued_carat", 0)))

                # Get and lock inventory record
                inventory = RawMaterialInventory.objects.select_for_update().get(pk=inventory_id)

                # Validate stock availability
                if issued_carat > inventory.available_carat:
                    return JsonResponse(
                        {
                            "error": f"Insufficient stock. Only {inventory.available_carat} ct available."
                        },
                        status=400,
                    )

                # Create issuance
                issuance = serializer.save(
                    company=company,
                    issued_by=request.user,
                    inventory=inventory,
                )

                # Update inventory
                inventory.available_carat -= issued_carat
                inventory.issued_carat += issued_carat
                inventory.update_status()

                return JsonResponse(
                    RawMaterialIssuanceSerializer(issuance).data,
                    status=201,
                )
        return JsonResponse(serializer.errors, status=400)


@csrf_exempt
@require_http_methods(["GET", "DELETE"])
def raw_material_issuance_detail(request, pk):
    """Get or delete an issuance."""
    company = getattr(request.user, "company", None)
    if not company:
        return JsonResponse({"error": "No company found"}, status=400)

    try:
        issuance = RawMaterialIssuance.objects.get(pk=pk, company=company)
    except RawMaterialIssuance.DoesNotExist:
        return JsonResponse({"error": "Not found"}, status=404)

    if request.method == "GET":
        serializer = RawMaterialIssuanceSerializer(issuance)
        return JsonResponse(serializer.data)

    elif request.method == "DELETE":
        # Reverse the issuance - add back to inventory
        with transaction.atomic():
            inventory = issuance.inventory
            inventory.available_carat += issuance.issued_carat
            inventory.issued_carat -= issuance.issued_carat
            inventory.update_status()
            issuance.delete()
        return JsonResponse({"message": "Deleted and inventory restored"}, status=200)


# ============================================================================
# Daily Book Close Views
# ============================================================================
@csrf_exempt
@require_http_methods(["GET", "POST"])
def daily_book_close_list_create(request):
    """List book close records or close books for a date."""
    company = getattr(request.user, "company", None)
    if not company:
        return JsonResponse({"error": "No company found"}, status=400)

    if request.method == "GET":
        closes = DailyBookClose.objects.filter(company=company)

        # Pagination
        page = int(request.GET.get("page", 1))
        page_size = int(request.GET.get("page_size", 20))
        start = (page - 1) * page_size
        end = start + page_size

        total = closes.count()
        closes = closes[start:end]

        serializer = DailyBookCloseSerializer(closes, many=True)
        return JsonResponse(
            {
                "results": serializer.data,
                "count": total,
                "page": page,
                "page_size": page_size,
            }
        )

    elif request.method == "POST":
        try:
            data = (
                json.loads(request.body)
                if request.content_type == "application/json"
                else request.POST.dict()
            )
        except json.JSONDecodeError:
            data = request.POST.dict()

        close_date = data.get("date")
        if not close_date:
            return JsonResponse({"error": "Date is required"}, status=400)

        # Check if already closed
        existing = DailyBookClose.objects.filter(company=company, date=close_date).first()
        if existing and existing.is_closed:
            return JsonResponse({"error": f"Books already closed for {close_date}"}, status=400)

        # Calculate summary data
        from tasks.models import Task

        purchases_today = RawMaterialPurchase.objects.filter(
            company=company, date=close_date
        ).count()
        issuances_today = RawMaterialIssuance.objects.filter(
            company=company, date=close_date
        ).count()
        purchase_value = (
            RawMaterialPurchase.objects.filter(company=company, date=close_date).aggregate(
                total=Sum("total")
            )["total"]
            or 0
        )
        tasks_completed = Task.objects.filter(
            company=company, updated_at__date=close_date, status="COMPLETED"
        ).count()
        orders_today = Order.objects.filter(company=company, date=close_date).count()

        # Create or update book close record
        book_close, created = DailyBookClose.objects.update_or_create(
            company=company,
            date=close_date,
            defaults={
                "is_closed": True,
                "closed_at": timezone.now(),
                "closed_by": request.user,
                "total_purchases": purchases_today,
                "total_issuances": issuances_today,
                "total_purchase_value": purchase_value,
                "total_tasks_completed": tasks_completed,
                "total_orders": orders_today,
                "notes": data.get("notes", ""),
            },
        )

        return JsonResponse(DailyBookCloseSerializer(book_close).data, status=201)


@csrf_exempt
@require_http_methods(["GET"])
def daily_book_close_status(request):
    """Check if books are closed for a specific date."""
    company = getattr(request.user, "company", None)
    if not company:
        return JsonResponse({"error": "No company found"}, status=400)

    check_date = request.GET.get("date", str(timezone.now().date()))

    book_close = DailyBookClose.objects.filter(company=company, date=check_date).first()

    return JsonResponse(
        {
            "date": check_date,
            "is_closed": book_close.is_closed if book_close else False,
            "closed_at": str(book_close.closed_at) if book_close and book_close.closed_at else None,
            "closed_by": book_close.closed_by.name if book_close and book_close.closed_by else None,
        }
    )


# ============================================================================
# Daily Report Views
# ============================================================================
@csrf_exempt
@require_http_methods(["GET", "POST"])
def daily_report_list_create(request):
    """List daily reports or create/update one."""
    company = getattr(request.user, "company", None)
    if not company:
        return JsonResponse({"error": "No company found"}, status=400)

    if request.method == "GET":
        reports = DailyReport.objects.filter(company=company).select_related("user")

        # Filter by date
        report_date = request.GET.get("date")
        if report_date:
            reports = reports.filter(date=report_date)

        # Filter by user
        user_id = request.GET.get("user_id")
        if user_id:
            reports = reports.filter(user_id=user_id)

        # Pagination
        page = int(request.GET.get("page", 1))
        page_size = int(request.GET.get("page_size", 20))
        start = (page - 1) * page_size
        end = start + page_size

        total = reports.count()
        reports = reports[start:end]

        serializer = DailyReportSerializer(reports, many=True)
        return JsonResponse(
            {
                "results": serializer.data,
                "count": total,
                "page": page,
                "page_size": page_size,
            }
        )

    elif request.method == "POST":
        try:
            data = (
                json.loads(request.body)
                if request.content_type == "application/json"
                else request.POST.dict()
            )
        except json.JSONDecodeError:
            data = request.POST.dict()

        report_date = data.get("date", str(timezone.now().date()))
        is_submitted = data.get("is_submitted", False)

        # Create or update report for current user
        report, created = DailyReport.objects.update_or_create(
            company=company,
            date=report_date,
            user=request.user,
            defaults={
                "tasks_completed": data.get("tasks_completed", 0),
                "tasks_pending": data.get("tasks_pending", 0),
                "orders_processed": data.get("orders_processed", 0),
                "materials_issued": data.get("materials_issued", 0),
                "summary": data.get("summary", ""),
                "challenges": data.get("challenges", ""),
                "next_day_plan": data.get("next_day_plan", ""),
                "is_submitted": is_submitted,
                "submitted_at": timezone.now() if is_submitted else None,
            },
        )

        # Send email if report is submitted
        if is_submitted:
            print("[EMAIL DEBUG] Report submitted, attempting to send email...")
            print(f"[EMAIL DEBUG] is_submitted = {is_submitted}")

            try:
                from .email_utils import send_daily_report_email

                # Get recipient emails from environment or use default
                recipient_emails_str = os.getenv("DAILY_REPORT_RECIPIENTS", "")
                print(f"[EMAIL DEBUG] DAILY_REPORT_RECIPIENTS from env: '{recipient_emails_str}'")

                recipient_emails = [
                    email.strip() for email in recipient_emails_str.split(",") if email.strip()
                ]
                print(f"[EMAIL DEBUG] Parsed recipient emails: {recipient_emails}")

                # If no recipients configured, try to send to company admins or superusers
                if not recipient_emails:
                    print("[EMAIL DEBUG] No recipients in env, checking for admin users...")
                    from users.models import User

                    admin_users = User.objects.filter(company=company, is_staff=True).exclude(
                        email=""
                    )
                    recipient_emails = [user.email for user in admin_users if user.email]
                    print(f"[EMAIL DEBUG] Admin user emails: {recipient_emails}")

                # Send email
                if recipient_emails:
                    print(f"[EMAIL DEBUG] Attempting to send email to: {recipient_emails}")
                    user_name = getattr(request.user, "name", None) or getattr(
                        request.user, "username", "Unknown"
                    )
                    print(f"[EMAIL DEBUG] User: {user_name}")

                    email_sent = send_daily_report_email(report, request.user, recipient_emails)
                    if email_sent:
                        print(
                            f"[EMAIL SUCCESS] Daily report email sent to {', '.join(recipient_emails)}"
                        )
                    else:
                        print("[EMAIL ERROR] Failed to send daily report email (returned False)")
                else:
                    print("[EMAIL ERROR] No recipient emails found!")
            except Exception as e:
                # Don't let email errors break report submission
                print(f"[EMAIL ERROR] Exception while sending email: {e}")
                import traceback

                traceback.print_exc()
        else:
            print(
                f"[EMAIL DEBUG] Report not submitted (is_submitted={is_submitted}), skipping email"
            )

        return JsonResponse(DailyReportSerializer(report).data, status=201)


@csrf_exempt
@require_http_methods(["GET"])
def daily_report_dashboard(request):
    """Get aggregated dashboard data for daily reports."""
    company = getattr(request.user, "company", None)
    if not company:
        return JsonResponse({"error": "No company found"}, status=400)

    report_date = request.GET.get("date", str(timezone.now().date()))

    reports = DailyReport.objects.filter(company=company, date=report_date)

    total_users = reports.count()
    submitted_reports = reports.filter(is_submitted=True).count()

    # Aggregate data
    total_tasks_completed = reports.aggregate(total=Sum("tasks_completed"))["total"] or 0
    total_tasks_pending = reports.aggregate(total=Sum("tasks_pending"))["total"] or 0
    total_orders_processed = reports.aggregate(total=Sum("orders_processed"))["total"] or 0
    total_materials_issued = reports.aggregate(total=Sum("materials_issued"))["total"] or 0

    # Get book close status
    book_close = DailyBookClose.objects.filter(company=company, date=report_date).first()

    return JsonResponse(
        {
            "date": report_date,
            "total_users_with_reports": total_users,
            "submitted_reports": submitted_reports,
            "pending_reports": total_users - submitted_reports,
            "total_tasks_completed": total_tasks_completed,
            "total_tasks_pending": total_tasks_pending,
            "total_orders_processed": total_orders_processed,
            "total_materials_issued": total_materials_issued,
            "books_closed": book_close.is_closed if book_close else False,
        }
    )


# ============================================================================
# Orders Dropdown for Raw Material Purchase
# ============================================================================
@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def orders_dropdown(request):
    """Get list of orders for dropdown selection with account details."""
    # Fix: Properly load the company relationship
    try:
        from users.models import User

        user = User.objects.select_related("company").get(id=request.user.id)
        company = user.company
    except User.DoesNotExist:
        return Response({"error": "User not found"}, status=400)

    if not company:
        return Response({"error": "No company found"}, status=400)

    orders = (
        Order.objects.filter(company=company)
        .select_related("account")
        .order_by("-date", "-created_at")[:100]
    )

    orders_data = []
    for order in orders:
        order_dict = {
            "id": str(order.id),
            "bill_no": order.bill_no,
            "job_no": order.job_no,
            "date": order.date.isoformat() if order.date else None,
            "item_name": order.item_name,
        }
        # Include account details if available
        if order.account:
            order_dict["account"] = {
                "id": str(order.account.id),
                "account_name": order.account.account_name,
                "email": getattr(order.account, "email", None),
                "phone": getattr(order.account, "phone", None),
                "address": getattr(order.account, "address", None),
                "city": getattr(order.account, "city", None),
            }
        else:
            order_dict["account"] = None
        orders_data.append(order_dict)

    return Response({"orders": orders_data})


@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def search_order_by_id(request):
    """
    Search for an order by bill_no or job_no and return order details with query orders.

    GET /api/orders/search/?order_id=<bill_no or job_no>

    Returns:
    - Order basic info (id, bill_no, job_no, date, item_name, account)
    - Query orders (previous, current, upcoming)
    - Progress statistics
    """
    try:
        user = request.user
        company = getattr(user, "company", None)

        if not company:
            return Response({"error": "User must be assigned to a company"}, status=400)

        # Get order_id from query params
        order_id = request.GET.get("order_id", "").strip()

        if not order_id:
            return Response({"error": "order_id parameter is required"}, status=400)

        # Search by bill_no or job_no
        order = (
            Order.objects.filter(company=company)
            .filter(Q(bill_no__iexact=order_id) | Q(job_no__iexact=order_id))
            .select_related("account")
            .first()
        )

        if not order:
            return Response(
                {"found": False, "message": f"No order found with ID: {order_id}"}, status=404
            )

        # Get query orders
        steps = order.process_steps.all().order_by("position")

        # Check lock level
        lock_level = "UNLOCKED"
        courier_dispatched = False
        try:
            from .models import OrderProcessLock

            process_lock = OrderProcessLock.objects.get(order=order)
            lock_level = process_lock.lock_level
            courier_dispatched = process_lock.courier_dispatched
        except:
            pass

        # Categorize steps
        previous_steps = []
        current_step = None
        upcoming_steps = []

        for step in steps:
            step_data = {
                "id": str(step.id),
                "step_name": step.step_name,
                "description": step.description,
                "department": step.department,
                "position": step.position,
                "status": step.status,
                "completed_at": step.completed_at.isoformat() if step.completed_at else None,
                "notes": step.notes,
                "reference_id": str(step.reference_id) if step.reference_id else None,
            }

            if step.status == "COMPLETED":
                previous_steps.append(step_data)
            elif step.status == "IN_PROGRESS":
                current_step = step_data
            else:  # PENDING
                upcoming_steps.append(step_data)

        # Calculate progress
        total_steps = steps.count()
        completed_steps = steps.filter(status="COMPLETED").count()
        progress_percentage = (completed_steps / total_steps * 100) if total_steps > 0 else 0

        # Build response
        response_data = {
            "found": True,
            "order": {
                "id": str(order.id),
                "bill_no": order.bill_no,
                "job_no": order.job_no,
                "date": order.date.isoformat() if order.date else None,
                "item_name": order.item_name,
                "account": {
                    "id": str(order.account.id),
                    "account_name": order.account.account_name,
                    "email": getattr(order.account, "email", None),
                    "phone": getattr(order.account, "phone", None),
                }
                if order.account
                else None,
            },
            "process_steps": {
                "previous": previous_steps,
                "current": current_step,
                "upcoming": upcoming_steps,
            },
            "progress": {
                "total_steps": total_steps,
                "completed_steps": completed_steps,
                "in_progress_steps": 1 if current_step else 0,
                "pending_steps": len(upcoming_steps),
                "progress_percentage": round(progress_percentage, 2),
            },
            "lock_info": {
                "lock_level": lock_level,
                "courier_dispatched": courier_dispatched,
                "can_update_status": lock_level != "FULLY_LOCKED",
            },
        }

        return Response(response_data)

    except Exception as e:
        import logging

        logger = logging.getLogger(__name__)
        logger.error(f"Error searching order: {str(e)}", exc_info=True)
        return Response({"error": f"An error occurred: {str(e)}"}, status=500)


# 3D Design views
from .models import ThreeDDesign, ThreeDDesignImage
from .serializers import ThreeDDesignSerializer


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET"])
def three_d_design_overview_view(request):
    """
    Get overview/statistics for 3D Designs

    GET /api/3d-designs/overview/

    Returns:
    {
        "total_designs": 150,
        "designs_with_images": 120,
        "designs_with_approved_images": 80,
        "recent_designs_count": 10,
        "latest_designs": [...],
        "completion_rate": 53.33
    }
    """
    user = request.user
    queryset = ThreeDDesign.objects.all()

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    # Calculate statistics
    total_designs = queryset.count()
    designs_with_images = (
        queryset.filter(design_image__isnull=False).exclude(design_image="").count()
    )
    designs_with_approved = (
        queryset.filter(approved_design_image__isnull=False)
        .exclude(approved_design_image="")
        .count()
    )

    # Recent designs (last 7 days)
    from datetime import timedelta

    seven_days_ago = timezone.now() - timedelta(days=7)
    recent_designs = queryset.filter(created_at__gte=seven_days_ago).count()

    # Get latest 5 designs
    latest_designs = queryset.order_by("-created_at")[:5]
    latest_designs_data = [
        {
            "id": str(design.id),
            "account_order_id": design.account_order_id,
            "has_design_image": bool(design.design_image),
            "has_approved_image": bool(design.approved_design_image),
            "created_at": design.created_at.isoformat() if design.created_at else None,
            "created_by": design.created_by.email if design.created_by else None,
        }
        for design in latest_designs
    ]

    return JsonResponse(
        {
            "total_designs": total_designs,
            "designs_with_images": designs_with_images,
            "designs_with_approved_images": designs_with_approved,
            "recent_designs_count": recent_designs,
            "latest_designs": latest_designs_data,
            "completion_rate": round(
                (designs_with_approved / total_designs * 100) if total_designs > 0 else 0, 2
            ),
        },
        status=200,
    )


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def three_d_design_collection_view(request):
    """List and create 3D Design records."""
    user = request.user
    queryset = ThreeDDesign.objects.select_related("company", "created_by", "updated_by").order_by(
        "-created_at"
    )

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    if request.method == "GET":
        # Simple list with optional search
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(Q(account_order_id__icontains=search))

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [
            ThreeDDesignSerializer(d, context={"request": request}).data
            for d in queryset[start:end]
        ]

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    # POST - create new 3D Design
    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    # Determine if this is a draft save (no validation) or a final save (validate required fields)
    is_draft = request.POST.get("is_draft", "false").lower() in ("true", "1", "yes")

    # Extract image-related parameters before passing to serializer
    final_design_image_id = request.POST.get("final_design_image_id")
    final_approved_design_image_id = request.POST.get("final_approved_design_image_id")

    # Create directly to avoid serializer issues with extra POST keys
    from .models import ThreeDDesign as ThreeDDesignModel

    design = ThreeDDesignModel.objects.create(
        company=company,
        created_by=user,
        is_draft=is_draft,
        account_order_id=request.POST.get("account_order_id") or None,
    )
    # Get final file indices
    final_design_index = None
    final_approved_index = None
    try:
        if "final_design_file_index" in request.POST:
            final_design_index = int(request.POST.get("final_design_file_index"))
        if "final_approved_design_file_index" in request.POST:
            final_approved_index = int(request.POST.get("final_approved_design_file_index"))
    except (ValueError, TypeError):
        pass

    # Generate one log_group UUID for this save batch
    batch_log_group = _new_log_group() if request.FILES else None

    design_files = request.FILES.getlist("design_images[]")
    for idx, file in enumerate(design_files):
        is_final = final_design_index is not None and idx == final_design_index
        ThreeDDesignImage.objects.create(
            three_d_design=design,
            image=file,
            field_type="design",
            is_final_design=is_final,
            log_group=batch_log_group,
            company=company,
            uploaded_by=user,
        )

    approved_files = request.FILES.getlist("approved_design_images[]")
    for idx, file in enumerate(approved_files):
        is_final = final_approved_index is not None and idx == final_approved_index
        ThreeDDesignImage.objects.create(
            three_d_design=design,
            image=file,
            field_type="approved",
            is_final_approved=is_final,
            log_group=batch_log_group,
            company=company,
            uploaded_by=user,
        )

    # Legacy single-file backwards compatibility
    if "design_image" in request.FILES:
        design.design_image = request.FILES["design_image"]
        design.save(update_fields=["design_image"])
    if "approved_design_image" in request.FILES:
        design.approved_design_image = request.FILES["approved_design_image"]
        design.save(update_fields=["approved_design_image"])

    # If final save (not draft), validate required fields
    if not is_draft:
        is_valid, missing = _check_form_record_fields(
            design, "ThreeDDesign", uploaded_files=request.FILES
        )
        if not is_valid:
            design.delete()
            return JsonResponse(
                {
                    "errors": {
                        "validation": missing,
                        "message": (
                            f"Cannot save â€” the following required fields are missing: "
                            f"{', '.join(missing)}. Use 'Save as Draft' to save incomplete forms."
                        ),
                        "missing_fields": missing,
                    }
                },
                status=400,
            )

    data = ThreeDDesignSerializer(design, context={"request": request}).data
    data["is_draft"] = design.is_draft
    return JsonResponse(data, status=201)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PATCH", "PUT", "DELETE"])
def three_d_design_detail_view(request, pk):
    """Retrieve, update, or delete a 3D Design record."""
    user = request.user
    queryset = ThreeDDesign.objects.select_related("company", "created_by", "updated_by").order_by(
        "-created_at"
    )

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    try:
        design = queryset.get(pk=pk)
    except ThreeDDesign.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(
            ThreeDDesignSerializer(design, context={"request": request}).data, status=200
        )

    if request.method in ["PATCH", "PUT"]:
        # Parse PUT/PATCH multipart data
        parse_put_request(request)

        is_draft = request.POST.get("is_draft", "false").lower() in ("true", "1", "yes")

        final_design_image_id = request.POST.get("final_design_image_id")
        final_approved_design_image_id = request.POST.get("final_approved_design_image_id")

        # Update the record directly â€” avoids any serializer validation issues
        # with unrecognised POST keys (order_id, select_log_group, etc.)
        update_fields = ["is_draft", "updated_by", "updated_at"]
        design.is_draft = is_draft
        design.updated_by = user
        if request.POST.get("account_order_id") is not None:
            design.account_order_id = request.POST.get("account_order_id") or None
            update_fields.append("account_order_id")
        design.save(update_fields=update_fields)
        updated_design = design

        # Handle keep lists â€” delete images not in keep list
        if "keep_design_image_ids[]" in request.POST:
            keep_design_ids = set(request.POST.getlist("keep_design_image_ids[]"))
            updated_design.images.filter(field_type="design").exclude(
                id__in=keep_design_ids
            ).delete()
        if "keep_approved_design_image_ids[]" in request.POST:
            keep_approved_ids = set(request.POST.getlist("keep_approved_design_image_ids[]"))
            updated_design.images.filter(field_type="approved").exclude(
                id__in=keep_approved_ids
            ).delete()

        # Get final file indices
        final_design_index = None
        final_approved_index = None
        try:
            if "final_design_file_index" in request.POST:
                final_design_index = int(request.POST.get("final_design_file_index"))
            if "final_approved_design_file_index" in request.POST:
                final_approved_index = int(request.POST.get("final_approved_design_file_index"))
        except (ValueError, TypeError):
            pass

        # Generate one log_group UUID for this PATCH batch
        patch_log_group = _new_log_group() if request.FILES else None

        # Handle multi-image uploads for design images
        design_files = request.FILES.getlist("design_images[]")
        for idx, file in enumerate(design_files):
            is_final = final_design_index is not None and idx == final_design_index
            ThreeDDesignImage.objects.create(
                three_d_design=updated_design,
                image=file,
                field_type="design",
                is_final_design=is_final,
                log_group=patch_log_group,
                company=updated_design.company,
                uploaded_by=user,
            )

        # Handle multi-image uploads for approved design images
        approved_files = request.FILES.getlist("approved_design_images[]")
        for idx, file in enumerate(approved_files):
            is_final = final_approved_index is not None and idx == final_approved_index
            ThreeDDesignImage.objects.create(
                three_d_design=updated_design,
                image=file,
                field_type="approved",
                is_final_approved=is_final,
                log_group=patch_log_group,
                company=updated_design.company,
                uploaded_by=user,
            )

        # Handle independent log group selections for design + approved image types
        save_fields = []
        select_log = request.POST.get("select_log_group")
        if select_log is not None:
            updated_design.selected_log_group = select_log or None
            save_fields.append("selected_log_group")

        select_secondary = request.POST.get("select_secondary_log_group")
        if select_secondary is not None:
            updated_design.selected_secondary_log_group = select_secondary or None
            save_fields.append("selected_secondary_log_group")

        if save_fields:
            updated_design.save(update_fields=save_fields)

        # Update final image selection for existing images if provided
        if final_design_image_id:
            updated_design.images.filter(field_type="design").update(is_final_design=False)
            try:
                final_img = ThreeDDesignImage.objects.get(
                    id=final_design_image_id, field_type="design"
                )
                final_img.is_final_design = True
                final_img.save()
            except ThreeDDesignImage.DoesNotExist:
                pass

        if final_approved_design_image_id:
            updated_design.images.filter(field_type="approved").update(is_final_approved=False)
            try:
                final_img = ThreeDDesignImage.objects.get(
                    id=final_approved_design_image_id, field_type="approved"
                )
                final_img.is_final_approved = True
                final_img.save()
            except ThreeDDesignImage.DoesNotExist:
                pass

        # Keep old single-file uploads for backwards compatibility
        files_updated = False
        if "design_image" in request.FILES:
            updated_design.design_image = request.FILES["design_image"]
            files_updated = True
        if "approved_design_image" in request.FILES:
            updated_design.approved_design_image = request.FILES["approved_design_image"]
            files_updated = True

        if files_updated:
            updated_design.save()
            updated_design.refresh_from_db()

        # If final save (not draft), validate required fields
        if not is_draft:
            is_valid, missing = _check_form_record_fields(
                updated_design, "ThreeDDesign", uploaded_files=request.FILES
            )
            if not is_valid:
                return JsonResponse(
                    {
                        "errors": {
                            "validation": missing,
                            "message": (
                                f"Cannot save â€” the following required fields are missing: "
                                f"{', '.join(missing)}. Use 'Save as Draft' to save incomplete forms."
                            ),
                            "missing_fields": missing,
                        }
                    },
                    status=400,
                )

        # Always return the updated design data
        updated_design.refresh_from_db()
        data = ThreeDDesignSerializer(updated_design, context={"request": request}).data
        data["is_draft"] = updated_design.is_draft
        return JsonResponse(data, status=200)

    if request.method == "DELETE":
        design.delete()
        return JsonResponse({"detail": "Deleted successfully"}, status=200)


# 3D Printing/CAM Piece views
from .models import ThreeDPrintingCAM, ThreeDPrintingCAMImage
from .serializers import ThreeDPrintingCAMSerializer


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET"])
def three_d_printing_cam_overview_view(request):
    """Get overview/statistics for 3D Printing/CAM Pieces"""
    user = request.user
    queryset = ThreeDPrintingCAM.objects.all()

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    from datetime import timedelta

    seven_days_ago = timezone.now() - timedelta(days=7)

    total_records = queryset.count()
    recent_records = queryset.filter(created_at__gte=seven_days_ago).count()

    latest_records = queryset.order_by("-created_at")[:5]
    latest_data = [
        {
            "id": str(record.id),
            "order_id": str(record.order_id) if record.order_id else None,
            "created_at": record.created_at.isoformat() if record.created_at else None,
        }
        for record in latest_records
    ]

    return JsonResponse(
        {
            "total_records": total_records,
            "recent_records_count": recent_records,
            "latest_records": latest_data,
        },
        status=200,
    )


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def three_d_printing_cam_collection_view(request):
    """List and create 3D Printing/CAM Piece records."""
    user = request.user
    queryset = ThreeDPrintingCAM.objects.select_related(
        "company", "order", "created_by", "updated_by"
    ).order_by("-created_at")

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    if request.method == "GET":
        # Simple list with optional search
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(Q(account_order_id__icontains=search))

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [
            ThreeDPrintingCAMSerializer(d, context={"request": request}).data
            for d in queryset[start:end]
        ]

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    # POST - create new 3D Printing/CAM Piece
    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    is_draft = request.POST.get("is_draft", "false").lower() in ("true", "1", "yes")

    # Extract image-related parameters before passing to serializer
    final_cam_piece_id = request.POST.get("final_cam_piece_image_id")
    final_approved_cam_id = request.POST.get("final_approved_cam_image_id")

    # Get final file indices (no defaults - only final if explicitly specified)
    final_cam_piece_index = None
    final_approved_cam_index = None
    try:
        if "final_cam_piece_file_index" in request.POST:
            final_cam_piece_index = int(request.POST.get("final_cam_piece_file_index"))
        if "final_approved_cam_file_index" in request.POST:
            final_approved_cam_index = int(request.POST.get("final_approved_cam_file_index"))
    except (ValueError, TypeError):
        pass

    serializer = ThreeDPrintingCAMSerializer(data=request.POST, context={"request": request})
    if serializer.is_valid():
        record = serializer.save(company=company, created_by=user, is_draft=is_draft)

        batch_log_group = _new_log_group() if request.FILES else None

        # Create image records for CAM piece images
        cam_piece_files = request.FILES.getlist("cam_piece_images[]")
        for idx, file in enumerate(cam_piece_files):
            is_final = final_cam_piece_index is not None and idx == final_cam_piece_index
            ThreeDPrintingCAMImage.objects.create(
                three_d_printing_cam=record,
                image=file,
                field_type="cam_piece",
                is_final=is_final,
                log_group=batch_log_group,
                company=company,
                uploaded_by=user,
            )

        # Create image records for approved CAM images
        approved_cam_files = request.FILES.getlist("approved_cam_images[]")
        for idx, file in enumerate(approved_cam_files):
            is_final = final_approved_cam_index is not None and idx == final_approved_cam_index
            ThreeDPrintingCAMImage.objects.create(
                three_d_printing_cam=record,
                image=file,
                field_type="approved_cam",
                is_final=is_final,
                log_group=batch_log_group,
                company=company,
                uploaded_by=user,
            )

        # Handle old single-file uploads for backwards compatibility
        if "cam_piece_image" in request.FILES:
            record.cam_piece_image = request.FILES["cam_piece_image"]
        if "approved_cam_piece" in request.FILES:
            record.approved_cam_piece = request.FILES["approved_cam_piece"]
        if "carry_forward_image" in request.FILES:
            record.carry_forward_image = request.FILES["carry_forward_image"]
        record.save()

        if not is_draft:
            is_valid, missing = _check_form_record_fields(
                record, "ThreeDPrintingCAM", uploaded_files=request.FILES
            )
            if not is_valid:
                record.delete()
                return JsonResponse(
                    {
                        "errors": {
                            "validation": missing,
                            "message": (
                                f"Cannot save â€” the following required fields are missing: "
                                f"{', '.join(missing)}. Use 'Save as Draft' to save incomplete forms."
                            ),
                            "missing_fields": missing,
                        }
                    },
                    status=400,
                )

        data = ThreeDPrintingCAMSerializer(record, context={"request": request}).data
        data["is_draft"] = record.is_draft
        return JsonResponse(data, status=201)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PATCH", "PUT", "DELETE"])
def three_d_printing_cam_detail_view(request, pk):
    """Retrieve, update, or delete a 3D Printing/CAM Piece record."""
    user = request.user
    queryset = ThreeDPrintingCAM.objects.select_related(
        "company", "order", "created_by", "updated_by"
    ).order_by("-created_at")

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    try:
        record = queryset.get(pk=pk)
    except ThreeDPrintingCAM.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(
            ThreeDPrintingCAMSerializer(record, context={"request": request}).data, status=200
        )

    if request.method in ["PATCH", "PUT"]:
        # Parse PUT/PATCH multipart data
        parse_put_request(request)

        # Extract image-related parameters
        final_cam_piece_id = request.POST.get("final_cam_piece_image_id")
        final_approved_cam_id = request.POST.get("final_approved_cam_image_id")

        # Get final file indices (no defaults)
        final_cam_piece_index = None
        final_approved_cam_index = None
        try:
            if "final_cam_piece_file_index" in request.POST:
                final_cam_piece_index = int(request.POST.get("final_cam_piece_file_index"))
            if "final_approved_cam_file_index" in request.POST:
                final_approved_cam_index = int(request.POST.get("final_approved_cam_file_index"))
        except (ValueError, TypeError):
            pass

        # Keep lists - delete images not in keep list
        keep_cam_ids = set(request.POST.getlist("keep_cam_piece_image_ids[]"))
        keep_approved_ids = set(request.POST.getlist("keep_approved_cam_image_ids[]"))

        serializer = ThreeDPrintingCAMSerializer(
            record, data=request.POST, partial=True, context={"request": request}
        )
        if serializer.is_valid():
            updated_record = serializer.save(updated_by=user)

            # Delete images not in keep list
            if "keep_cam_piece_image_ids[]" in request.POST:
                keep_cam_ids = set(request.POST.getlist("keep_cam_piece_image_ids[]"))
                ThreeDPrintingCAMImage.objects.filter(
                    three_d_printing_cam=updated_record, field_type="cam_piece"
                ).exclude(id__in=keep_cam_ids).delete()
            if "keep_approved_cam_image_ids[]" in request.POST:
                keep_approved_ids = set(request.POST.getlist("keep_approved_cam_image_ids[]"))
                ThreeDPrintingCAMImage.objects.filter(
                    three_d_printing_cam=updated_record, field_type="approved_cam"
                ).exclude(id__in=keep_approved_ids).delete()

            # Update is_final on existing images
            if final_cam_piece_id:
                ThreeDPrintingCAMImage.objects.filter(
                    three_d_printing_cam=updated_record, field_type="cam_piece"
                ).update(is_final=False)
                ThreeDPrintingCAMImage.objects.filter(
                    id=final_cam_piece_id, three_d_printing_cam=updated_record
                ).update(is_final=True)
            if final_approved_cam_id:
                ThreeDPrintingCAMImage.objects.filter(
                    three_d_printing_cam=updated_record, field_type="approved_cam"
                ).update(is_final=False)
                ThreeDPrintingCAMImage.objects.filter(
                    id=final_approved_cam_id, three_d_printing_cam=updated_record
                ).update(is_final=True)

            patch_log_group = _new_log_group() if request.FILES else None

            # Create new image records for CAM piece images
            cam_piece_files = request.FILES.getlist("cam_piece_images[]")
            for idx, file in enumerate(cam_piece_files):
                is_final = final_cam_piece_index is not None and idx == final_cam_piece_index
                ThreeDPrintingCAMImage.objects.create(
                    three_d_printing_cam=updated_record,
                    image=file,
                    field_type="cam_piece",
                    is_final=is_final,
                    log_group=patch_log_group,
                    company=updated_record.company,
                    uploaded_by=user,
                )

            # Create new image records for approved CAM images
            approved_cam_files = request.FILES.getlist("approved_cam_images[]")
            for idx, file in enumerate(approved_cam_files):
                is_final = final_approved_cam_index is not None and idx == final_approved_cam_index
                ThreeDPrintingCAMImage.objects.create(
                    three_d_printing_cam=updated_record,
                    image=file,
                    field_type="approved_cam",
                    is_final=is_final,
                    log_group=patch_log_group,
                    company=updated_record.company,
                    uploaded_by=user,
                )

            # Handle independent log selections for cam_piece + approved_cam
            save_lg = []
            select_log = request.POST.get("select_log_group")
            if select_log is not None:
                updated_record.selected_log_group = select_log or None
                save_lg.append("selected_log_group")

            select_secondary = request.POST.get("select_secondary_log_group")
            if select_secondary is not None:
                updated_record.selected_secondary_log_group = select_secondary or None
                save_lg.append("selected_secondary_log_group")

            # Handle legacy single-file uploads
            if "cam_piece_image" in request.FILES:
                updated_record.cam_piece_image = request.FILES["cam_piece_image"]
                save_lg.append("cam_piece_image")
            if "approved_cam_piece" in request.FILES:
                updated_record.approved_cam_piece = request.FILES["approved_cam_piece"]
                save_lg.append("approved_cam_piece")

            if save_lg:
                updated_record.save(update_fields=save_lg)
            updated_record.refresh_from_db()

            return JsonResponse(
                ThreeDPrintingCAMSerializer(updated_record, context={"request": request}).data,
                status=200,
            )
        return JsonResponse({"errors": serializer.errors}, status=400)

    if request.method == "DELETE":
        record.delete()
        return JsonResponse({"detail": "Deleted successfully"}, status=200)


# Ghat Approval views
from .models import GhatApproval
from .serializers import GhatApprovalSerializer


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET"])
def ghat_approval_overview_view(request):
    """Get overview/statistics for Ghat Approvals"""
    user = request.user
    queryset = GhatApproval.objects.all()

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    from datetime import timedelta

    seven_days_ago = timezone.now() - timedelta(days=7)

    total_records = queryset.count()
    recent_records = queryset.filter(created_at__gte=seven_days_ago).count()

    latest_records = queryset.order_by("-created_at")[:5]
    latest_data = [
        {
            "id": str(record.id),
            "order_id": str(record.order_id) if record.order_id else None,
            "created_at": record.created_at.isoformat() if record.created_at else None,
        }
        for record in latest_records
    ]

    return JsonResponse(
        {
            "total_records": total_records,
            "recent_records_count": recent_records,
            "latest_records": latest_data,
        },
        status=200,
    )


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def ghat_approval_collection_view(request):
    """List and create Ghat Approval records."""
    user = request.user
    queryset = GhatApproval.objects.select_related(
        "company", "order", "created_by", "updated_by"
    ).order_by("-created_at")

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    if request.method == "GET":
        # Simple list with optional search
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(Q(account_order_id__icontains=search))

        # Date filtering (using created_at since no date field)
        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        if date_from and date_to:
            if date_from > date_to:
                return JsonResponse({"detail": "From date cannot be after To date"}, status=400)
            queryset = queryset.filter(
                created_at__date__gte=date_from, created_at__date__lte=date_to
            )
        elif date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        elif date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [
            GhatApprovalSerializer(d, context={"request": request}).data
            for d in queryset[start:end]
        ]

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    # POST - create new Ghat Approval
    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    is_draft = request.POST.get("is_draft", "false").lower() in ("true", "1", "yes")

    serializer = GhatApprovalSerializer(data=request.POST, context={"request": request})
    if serializer.is_valid():
        record = serializer.save(company=company, created_by=user, is_draft=is_draft)

        # Handle multi-image carry forward uploads
        from .models import GhatApprovalImage

        batch_log_group = _new_log_group() if request.FILES else None
        carry_forward_files = request.FILES.getlist("carry_forward_images[]")
        final_carry_index = None
        try:
            if "final_carry_forward_file_index" in request.POST:
                final_carry_index = int(request.POST.get("final_carry_forward_file_index"))
        except (ValueError, TypeError):
            pass
        for idx, file in enumerate(carry_forward_files):
            is_final = final_carry_index is not None and idx == final_carry_index
            GhatApprovalImage.objects.create(
                ghat_approval=record,
                image=file,
                is_final=is_final,
                log_group=batch_log_group,
                company=company,
                uploaded_by=user,
            )

        if not is_draft:
            is_valid, missing = _check_form_record_fields(
                record, "GhatApproval", uploaded_files=request.FILES
            )
            if not is_valid:
                record.delete()
                return JsonResponse(
                    {
                        "errors": {
                            "validation": missing,
                            "message": (
                                f"Cannot save â€” the following required fields are missing: "
                                f"{', '.join(missing)}. Use 'Save as Draft' to save incomplete forms."
                            ),
                            "missing_fields": missing,
                        }
                    },
                    status=400,
                )

        data = GhatApprovalSerializer(record, context={"request": request}).data
        data["is_draft"] = record.is_draft
        return JsonResponse(data, status=201)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PATCH", "PUT", "DELETE"])
def ghat_approval_detail_view(request, pk):
    """Retrieve, update, or delete a Ghat Approval record."""
    user = request.user
    queryset = GhatApproval.objects.select_related(
        "company", "order", "created_by", "updated_by"
    ).order_by("-created_at")

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    try:
        record = queryset.get(pk=pk)
    except GhatApproval.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(
            GhatApprovalSerializer(record, context={"request": request}).data, status=200
        )

    if request.method in ["PATCH", "PUT"]:
        # Parse PUT/PATCH multipart data
        parse_put_request(request)

        serializer = GhatApprovalSerializer(
            record, data=request.POST, partial=True, context={"request": request}
        )
        if serializer.is_valid():
            updated_record = serializer.save(updated_by=user)

            # Handle keep list and new multi-image uploads
            from .models import GhatApprovalImage

            if "keep_carry_forward_image_ids[]" in request.POST:
                keep_carry_ids = set(request.POST.getlist("keep_carry_forward_image_ids[]"))
                GhatApprovalImage.objects.filter(ghat_approval=updated_record).exclude(
                    id__in=keep_carry_ids
                ).delete()

            final_carry_id = request.POST.get("final_carry_forward_image_id")
            if final_carry_id:
                GhatApprovalImage.objects.filter(ghat_approval=updated_record).update(
                    is_final=False
                )
                GhatApprovalImage.objects.filter(
                    id=final_carry_id, ghat_approval=updated_record
                ).update(is_final=True)

            final_carry_index = None
            try:
                if "final_carry_forward_file_index" in request.POST:
                    final_carry_index = int(request.POST.get("final_carry_forward_file_index"))
            except (ValueError, TypeError):
                pass

            patch_log_group = _new_log_group() if request.FILES else None
            carry_forward_files = request.FILES.getlist("carry_forward_images[]")
            for idx, file in enumerate(carry_forward_files):
                is_final = final_carry_index is not None and idx == final_carry_index
                GhatApprovalImage.objects.create(
                    ghat_approval=updated_record,
                    image=file,
                    is_final=is_final,
                    log_group=patch_log_group,
                    company=updated_record.company,
                    uploaded_by=user,
                )

            select_log = request.POST.get("select_log_group")
            if select_log is not None:
                updated_record.selected_log_group = select_log or None
                updated_record.save(update_fields=["selected_log_group"])

            updated_record.refresh_from_db()
            return JsonResponse(
                GhatApprovalSerializer(updated_record, context={"request": request}).data,
                status=200,
            )
        return JsonResponse({"errors": serializer.errors}, status=400)

    if request.method == "DELETE":
        record.delete()
        return JsonResponse({"detail": "Deleted successfully"}, status=200)


# Ghat Quality Check views
from .models import GhatQualityCheck
from .serializers import GhatQualityCheckSerializer


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET"])
def ghat_quality_check_overview_view(request):
    """Get overview/statistics for Ghat Quality Checks"""
    user = request.user
    queryset = GhatQualityCheck.objects.all()

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    from datetime import timedelta

    seven_days_ago = timezone.now() - timedelta(days=7)

    total_records = queryset.count()
    recent_records = queryset.filter(created_at__gte=seven_days_ago).count()

    latest_records = queryset.order_by("-created_at")[:5]
    latest_data = [
        {
            "id": str(record.id),
            "order_id": str(record.order_id) if record.order_id else None,
            "created_at": record.created_at.isoformat() if record.created_at else None,
        }
        for record in latest_records
    ]

    return JsonResponse(
        {
            "total_records": total_records,
            "recent_records_count": recent_records,
            "latest_records": latest_data,
        },
        status=200,
    )


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def ghat_quality_check_collection_view(request):
    """List and create Ghat Quality Check records."""
    user = request.user
    queryset = GhatQualityCheck.objects.select_related(
        "company", "order", "created_by", "updated_by"
    ).order_by("-created_at")

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    if request.method == "GET":
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(Q(account_order_id__icontains=search))

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [
            GhatQualityCheckSerializer(d, context={"request": request}).data
            for d in queryset[start:end]
        ]

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    is_draft = request.POST.get("is_draft", "false").lower() in ("true", "1", "yes")

    serializer = GhatQualityCheckSerializer(data=request.POST, context={"request": request})
    if serializer.is_valid():
        record = serializer.save(company=company, created_by=user, is_draft=is_draft)

        # Handle multi-image carry forward uploads
        from .models import GhatQualityCheckImage

        batch_log_group = _new_log_group() if request.FILES else None
        carry_forward_files = request.FILES.getlist("carry_forward_images[]")
        final_carry_index = None
        try:
            if "final_carry_forward_file_index" in request.POST:
                final_carry_index = int(request.POST.get("final_carry_forward_file_index"))
        except (ValueError, TypeError):
            pass
        for idx, file in enumerate(carry_forward_files):
            is_final = final_carry_index is not None and idx == final_carry_index
            GhatQualityCheckImage.objects.create(
                ghat_quality_check=record,
                image=file,
                is_final=is_final,
                log_group=batch_log_group,
                company=company,
                uploaded_by=user,
            )

        if not is_draft:
            is_valid, missing = _check_form_record_fields(
                record, "GhatQualityCheck", uploaded_files=request.FILES
            )
            if not is_valid:
                record.delete()
                return JsonResponse(
                    {
                        "errors": {
                            "validation": missing,
                            "message": (
                                f"Cannot save â€” the following required fields are missing: "
                                f"{', '.join(missing)}. Use 'Save as Draft' to save incomplete forms."
                            ),
                            "missing_fields": missing,
                        }
                    },
                    status=400,
                )

        data = GhatQualityCheckSerializer(record, context={"request": request}).data
        data["is_draft"] = record.is_draft
        return JsonResponse(data, status=201)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PATCH", "PUT", "DELETE"])
def ghat_quality_check_detail_view(request, pk):
    """Retrieve, update, or delete a Ghat Quality Check record."""
    user = request.user
    queryset = GhatQualityCheck.objects.select_related(
        "company", "order", "created_by", "updated_by"
    ).order_by("-created_at")

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    try:
        record = queryset.get(pk=pk)
    except GhatQualityCheck.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(
            GhatQualityCheckSerializer(record, context={"request": request}).data, status=200
        )

    if request.method in ["PATCH", "PUT"]:
        # Parse PUT/PATCH multipart data
        parse_put_request(request)

        is_draft = request.POST.get("is_draft", "false").lower() in ("true", "1", "yes")

        # Only pass known GhatQualityCheck model fields to serializer
        serializer_data = {}
        if request.POST.get("account_order_id") is not None:
            serializer_data["account_order_id"] = request.POST.get("account_order_id")

        serializer = GhatQualityCheckSerializer(
            record, data=serializer_data, partial=True, context={"request": request}
        )
        if serializer.is_valid():
            updated_record = serializer.save(updated_by=user, is_draft=is_draft)

            # Handle keep list and new multi-image carry forward uploads
            from .models import GhatQualityCheckImage

            if "keep_carry_forward_image_ids[]" in request.POST:
                keep_carry_ids = set(request.POST.getlist("keep_carry_forward_image_ids[]"))
                GhatQualityCheckImage.objects.filter(ghat_quality_check=updated_record).exclude(
                    id__in=keep_carry_ids
                ).delete()

            final_carry_id = request.POST.get("final_carry_forward_image_id")
            if final_carry_id:
                GhatQualityCheckImage.objects.filter(ghat_quality_check=updated_record).update(
                    is_final=False
                )
                GhatQualityCheckImage.objects.filter(
                    id=final_carry_id, ghat_quality_check=updated_record
                ).update(is_final=True)

            final_carry_index = None
            try:
                if "final_carry_forward_file_index" in request.POST:
                    final_carry_index = int(request.POST.get("final_carry_forward_file_index"))
            except (ValueError, TypeError):
                pass

            patch_log_group = _new_log_group() if request.FILES else None
            carry_forward_files = request.FILES.getlist("carry_forward_images[]")
            for idx, file in enumerate(carry_forward_files):
                is_final = final_carry_index is not None and idx == final_carry_index
                GhatQualityCheckImage.objects.create(
                    ghat_quality_check=updated_record,
                    image=file,
                    is_final=is_final,
                    log_group=patch_log_group,
                    company=updated_record.company,
                    uploaded_by=user,
                )

            select_log = request.POST.get("select_log_group")
            if select_log is not None:
                updated_record.selected_log_group = select_log or None
                updated_record.save(update_fields=["selected_log_group"])

            updated_record.refresh_from_db()

            return JsonResponse(
                GhatQualityCheckSerializer(updated_record, context={"request": request}).data,
                status=200,
            )
        return JsonResponse({"errors": serializer.errors}, status=400)

    if request.method == "DELETE":
        record.delete()
        return JsonResponse({"detail": "Deleted successfully"}, status=200)


# Stone Demand to Bagging views
from .models import StoneDemandToBagging, StoneDemandToBaggingImage
from .serializers import StoneDemandToBaggingSerializer


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET"])
def stone_demand_to_bagging_overview_view(request):
    """Get overview/statistics for Stone Demand to Bagging"""
    user = request.user
    queryset = StoneDemandToBagging.objects.all()

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    from datetime import timedelta

    seven_days_ago = timezone.now() - timedelta(days=7)

    total_records = queryset.count()
    recent_records = queryset.filter(created_at__gte=seven_days_ago).count()

    latest_records = queryset.order_by("-created_at")[:5]
    latest_data = [
        {
            "id": str(record.id),
            "order_id": str(record.order_id) if record.order_id else None,
            "created_at": record.created_at.isoformat() if record.created_at else None,
        }
        for record in latest_records
    ]

    return JsonResponse(
        {
            "total_records": total_records,
            "recent_records_count": recent_records,
            "latest_records": latest_data,
        },
        status=200,
    )


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def stone_demand_to_bagging_collection_view(request):
    """List and create Stone Demand to Bagging records. Supports multiple items in one submission."""
    user = request.user
    queryset = StoneDemandToBagging.objects.select_related(
        "company", "order", "created_by", "updated_by"
    ).order_by("-created_at")

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    if request.method == "GET":
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(Q(account_order_id__icontains=search))

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [
            StoneDemandToBaggingSerializer(d, context={"request": request}).data
            for d in queryset[start:end]
        ]

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    # Check if stone_items is provided (multiple items mode)
    stone_items_json = request.POST.get("stone_items")

    if stone_items_json:
        # MULTIPLE ITEMS MODE
        try:
            stone_items = json.loads(stone_items_json)
            if not isinstance(stone_items, list):
                return JsonResponse(
                    {"errors": {"stone_items": ["Must be a list of items"]}}, status=400
                )

            if len(stone_items) == 0:
                return JsonResponse(
                    {"errors": {"stone_items": ["At least one item is required"]}}, status=400
                )

            # Get common fields
            account_order_id = request.POST.get("account_order_id")
            if not account_order_id:
                return JsonResponse(
                    {"errors": {"account_order_id": ["This field is required"]}}, status=400
                )

            is_draft = request.POST.get("is_draft", "false").lower() in ("true", "1", "yes")

            # Get multi-image files and final indices (shared across all items)
            approved_bagging_files = request.FILES.getlist("approved_bagging_images[]")
            carry_forward_files = request.FILES.getlist("carry_forward_images[]")

            final_approved_index = None
            final_carry_index = None
            try:
                if "final_approved_bagging_file_index" in request.POST:
                    final_approved_index = int(
                        request.POST.get("final_approved_bagging_file_index")
                    )
                if "final_carry_forward_file_index" in request.POST:
                    final_carry_index = int(request.POST.get("final_carry_forward_file_index"))
            except (ValueError, TypeError):
                pass

            # Create one record per stone item; images are attached to the FIRST record only
            created_records = []
            for item_index, item in enumerate(stone_items):
                record_data = {
                    "account_order_id": account_order_id,
                    "diamond_color_stone": item.get("diamond_color_stone"),
                    "batch_id": item.get("batch_id"),
                    "master_size": item.get("master_size"),
                    "shape": item.get("shape"),
                    "mm_size": item.get("mm_size"),
                    "no_of_pieces": item.get("no_of_pieces"),
                    "estimated_total_carat_weight": item.get("estimated_total_carat_weight"),
                }

                serializer = StoneDemandToBaggingSerializer(
                    data=record_data, context={"request": request}
                )
                if serializer.is_valid():
                    record = serializer.save(company=company, created_by=user, is_draft=is_draft)

                    # Attach images to the first record only
                    if item_index == 0:
                        batch_log_group = (
                            _new_log_group()
                            if (approved_bagging_files or carry_forward_files)
                            else None
                        )
                        for idx, file in enumerate(approved_bagging_files):
                            is_final = (
                                final_approved_index is not None and idx == final_approved_index
                            )
                            StoneDemandToBaggingImage.objects.create(
                                stone_demand=record,
                                image=file,
                                field_type="approved_bagging",
                                is_final=is_final,
                                log_group=batch_log_group,
                                company=company,
                                uploaded_by=user,
                            )
                        for idx, file in enumerate(carry_forward_files):
                            is_final = final_carry_index is not None and idx == final_carry_index
                            StoneDemandToBaggingImage.objects.create(
                                stone_demand=record,
                                image=file,
                                field_type="carry_forward",
                                is_final=is_final,
                                log_group=batch_log_group,
                                company=company,
                                uploaded_by=user,
                            )
                        if batch_log_group:
                            select_log = request.POST.get("select_log_group")
                            select_secondary = request.POST.get("select_secondary_log_group")
                            save_lg = []
                            if select_log or not record.selected_log_group:
                                record.selected_log_group = select_log or batch_log_group
                                save_lg.append("selected_log_group")
                            if select_secondary or not record.selected_secondary_log_group:
                                record.selected_secondary_log_group = (
                                    select_secondary or batch_log_group
                                )
                                save_lg.append("selected_secondary_log_group")
                            if save_lg:
                                record.save(update_fields=save_lg)

                    record.refresh_from_db()
                    created_records.append(
                        StoneDemandToBaggingSerializer(record, context={"request": request}).data
                    )
                else:
                    return JsonResponse({"errors": serializer.errors}, status=400)

            return JsonResponse(
                {
                    "message": f"Successfully created {len(created_records)} stone demand records",
                    "count": len(created_records),
                    "records": created_records,
                },
                status=201,
            )

        except json.JSONDecodeError:
            return JsonResponse({"errors": {"stone_items": ["Invalid JSON format"]}}, status=400)

    else:
        # SINGLE ITEM MODE (backward compatibility)
        data = request.POST.copy()
        if "measurement_details" in data:
            try:
                data["measurement_details"] = json.loads(data["measurement_details"])
            except (json.JSONDecodeError, TypeError):
                pass
        if "sent" in data:
            try:
                data["sent"] = json.loads(data["sent"])
            except (json.JSONDecodeError, TypeError):
                pass
        if "total" in data:
            try:
                data["total"] = json.loads(data["total"])
            except (json.JSONDecodeError, TypeError):
                pass

        serializer = StoneDemandToBaggingSerializer(data=data, context={"request": request})
        if serializer.is_valid():
            is_draft = request.POST.get("is_draft", "false").lower() in ("true", "1", "yes")
            record = serializer.save(company=company, created_by=user, is_draft=is_draft)

            # Create image records for approved bagging + carry-forward images
            approved_bagging_files = request.FILES.getlist("approved_bagging_images[]")
            carry_forward_files = request.FILES.getlist("carry_forward_images[]")
            batch_log_group = (
                _new_log_group() if (approved_bagging_files or carry_forward_files) else None
            )

            final_approved_index = None
            final_carry_index = None
            try:
                if "final_approved_bagging_file_index" in request.POST:
                    final_approved_index = int(
                        request.POST.get("final_approved_bagging_file_index")
                    )
                if "final_carry_forward_file_index" in request.POST:
                    final_carry_index = int(request.POST.get("final_carry_forward_file_index"))
            except (ValueError, TypeError):
                pass

            for idx, file in enumerate(approved_bagging_files):
                is_final = final_approved_index is not None and idx == final_approved_index
                StoneDemandToBaggingImage.objects.create(
                    stone_demand=record,
                    image=file,
                    field_type="approved_bagging",
                    is_final=is_final,
                    log_group=batch_log_group,
                    company=company,
                    uploaded_by=user,
                )

            for idx, file in enumerate(carry_forward_files):
                is_final = final_carry_index is not None and idx == final_carry_index
                StoneDemandToBaggingImage.objects.create(
                    stone_demand=record,
                    image=file,
                    field_type="carry_forward",
                    is_final=is_final,
                    log_group=batch_log_group,
                    company=company,
                    uploaded_by=user,
                )

            # Auto-select log groups
            if batch_log_group:
                select_log = request.POST.get("select_log_group")
                select_secondary = request.POST.get("select_secondary_log_group")
                record.selected_log_group = select_log or batch_log_group
                record.selected_secondary_log_group = select_secondary or batch_log_group
                record.save(update_fields=["selected_log_group", "selected_secondary_log_group"])

            # Keep old fields for backwards compatibility
            if "approved_bagging_list" in request.FILES:
                record.approved_bagging_list = request.FILES["approved_bagging_list"]
                record.save(update_fields=["approved_bagging_list"])
            if "carry_forward_image" in request.FILES:
                record.carry_forward_image = request.FILES["carry_forward_image"]
                record.save(update_fields=["carry_forward_image"])

            if not is_draft:
                is_valid, missing = _check_form_record_fields(
                    record, "StoneDemandToBagging", uploaded_files=request.FILES
                )
                if not is_valid:
                    record.delete()
                    return JsonResponse(
                        {
                            "errors": {
                                "validation": missing,
                                "message": (
                                    f"Cannot save â€” the following required fields are missing: "
                                    f"{', '.join(missing)}. Use 'Save as Draft' to save incomplete forms."
                                ),
                                "missing_fields": missing,
                            }
                        },
                        status=400,
                    )

            data_resp = StoneDemandToBaggingSerializer(record, context={"request": request}).data
            data_resp["is_draft"] = record.is_draft
            return JsonResponse(data_resp, status=201)

        return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PATCH", "PUT", "DELETE"])
def stone_demand_to_bagging_detail_view(request, pk):
    """Retrieve, update, or delete a Stone Demand to Bagging record."""
    user = request.user
    queryset = StoneDemandToBagging.objects.select_related(
        "company", "order", "created_by", "updated_by"
    ).order_by("-created_at")

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    try:
        record = queryset.get(pk=pk)
    except StoneDemandToBagging.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(
            StoneDemandToBaggingSerializer(record, context={"request": request}).data, status=200
        )

    if request.method in ["PATCH", "PUT"]:
        # Parse PUT/PATCH multipart data
        parse_put_request(request)

        # Check if this is a multiple items update
        stone_items_json = request.POST.get("stone_items")

        if stone_items_json:
            # MULTIPLE ITEMS UPDATE MODE
            try:
                stone_items = json.loads(stone_items_json)
                if not isinstance(stone_items, list):
                    return JsonResponse(
                        {"errors": {"stone_items": ["Must be a list of items"]}}, status=400
                    )

                if len(stone_items) == 0:
                    return JsonResponse(
                        {"errors": {"stone_items": ["At least one item is required"]}}, status=400
                    )

                account_order_id = record.account_order_id
                company = record.company
                is_draft = request.POST.get("is_draft", "false").lower() == "true"

                # Get final file indices from request
                final_approved_index = None
                final_carry_index = None
                try:
                    if "final_approved_bagging_file_index" in request.POST:
                        final_approved_index = int(
                            request.POST.get("final_approved_bagging_file_index")
                        )
                    if "final_carry_forward_file_index" in request.POST:
                        final_carry_index = int(request.POST.get("final_carry_forward_file_index"))
                except (ValueError, TypeError):
                    pass

                # Get keep lists - image IDs the user wants to retain
                keep_approved_ids = set(request.POST.getlist("keep_approved_bagging_image_ids[]"))
                keep_carry_ids = set(request.POST.getlist("keep_carry_forward_image_ids[]"))

                # Get final image IDs (existing images marked as final)
                final_approved_id = request.POST.get("final_approved_bagging_image_id")
                final_carry_id = request.POST.get("final_carry_forward_image_id")

                # Update the original record with first item data (PRESERVE the ID)
                first_item = stone_items[0]
                record.diamond_color_stone = (
                    first_item.get("diamond_color_stone") or record.diamond_color_stone
                )
                record.batch_id = first_item.get("batch_id") or record.batch_id
                record.master_size = first_item.get("master_size") or record.master_size
                record.shape = first_item.get("shape") or record.shape
                record.mm_size = first_item.get("mm_size") or record.mm_size
                record.no_of_pieces = first_item.get("no_of_pieces") or record.no_of_pieces
                record.estimated_total_carat_weight = (
                    first_item.get("estimated_total_carat_weight")
                    or record.estimated_total_carat_weight
                )
                record.is_draft = is_draft
                record.updated_by = user
                record.save()

                # Handle image deletions â€” delete images not in keep lists
                if "keep_approved_bagging_image_ids[]" in request.POST:
                    keep_approved_ids = set(
                        request.POST.getlist("keep_approved_bagging_image_ids[]")
                    )
                    StoneDemandToBaggingImage.objects.filter(
                        stone_demand=record, field_type="approved_bagging"
                    ).exclude(id__in=keep_approved_ids).delete()
                if "keep_carry_forward_image_ids[]" in request.POST:
                    keep_carry_ids = set(request.POST.getlist("keep_carry_forward_image_ids[]"))
                    StoneDemandToBaggingImage.objects.filter(
                        stone_demand=record, field_type="carry_forward"
                    ).exclude(id__in=keep_carry_ids).delete()

                # Update is_final flags on existing images
                if final_approved_id:
                    StoneDemandToBaggingImage.objects.filter(
                        stone_demand=record, field_type="approved_bagging"
                    ).update(is_final=False)
                    StoneDemandToBaggingImage.objects.filter(
                        id=final_approved_id, stone_demand=record
                    ).update(is_final=True)

                if final_carry_id:
                    StoneDemandToBaggingImage.objects.filter(
                        stone_demand=record, field_type="carry_forward"
                    ).update(is_final=False)
                    StoneDemandToBaggingImage.objects.filter(
                        id=final_carry_id, stone_demand=record
                    ).update(is_final=True)

                # Add new approved bagging + carry-forward images with log_group
                approved_bagging_files = request.FILES.getlist("approved_bagging_images[]")
                carry_forward_files = request.FILES.getlist("carry_forward_images[]")
                patch_log_group = (
                    _new_log_group() if (approved_bagging_files or carry_forward_files) else None
                )

                for idx, file in enumerate(approved_bagging_files):
                    is_final = final_approved_index is not None and idx == final_approved_index
                    StoneDemandToBaggingImage.objects.create(
                        stone_demand=record,
                        image=file,
                        field_type="approved_bagging",
                        is_final=is_final,
                        log_group=patch_log_group,
                        company=company,
                        uploaded_by=user,
                    )

                for idx, file in enumerate(carry_forward_files):
                    is_final = final_carry_index is not None and idx == final_carry_index
                    StoneDemandToBaggingImage.objects.create(
                        stone_demand=record,
                        image=file,
                        field_type="carry_forward",
                        is_final=is_final,
                        log_group=patch_log_group,
                        company=company,
                        uploaded_by=user,
                    )

                # Handle independent log selections
                save_lg = []
                select_log = request.POST.get("select_log_group")
                if select_log is not None:
                    record.selected_log_group = select_log or None
                    save_lg.append("selected_log_group")
                select_secondary = request.POST.get("select_secondary_log_group")
                if select_secondary is not None:
                    record.selected_secondary_log_group = select_secondary or None
                    save_lg.append("selected_secondary_log_group")
                if save_lg:
                    record.save(update_fields=save_lg)

                record.refresh_from_db()
                return JsonResponse(
                    {
                        "message": "Successfully updated 1 stone demand records",
                        "count": 1,
                        "records": [
                            StoneDemandToBaggingSerializer(
                                record, context={"request": request}
                            ).data
                        ],
                    },
                    status=200,
                )

            except json.JSONDecodeError:
                return JsonResponse(
                    {"errors": {"stone_items": ["Invalid JSON format"]}}, status=400
                )

        else:
            # SINGLE ITEM UPDATE MODE (backward compatible)
            data = request.POST.copy()
            if "measurement_details" in data:
                try:
                    data["measurement_details"] = json.loads(data["measurement_details"])
                except (json.JSONDecodeError, TypeError):
                    pass
            if "sent" in data:
                try:
                    data["sent"] = json.loads(data["sent"])
                except (json.JSONDecodeError, TypeError):
                    pass
            if "total" in data:
                try:
                    data["total"] = json.loads(data["total"])
                except (json.JSONDecodeError, TypeError):
                    pass

            serializer = StoneDemandToBaggingSerializer(
                record, data=data, partial=True, context={"request": request}
            )
            if serializer.is_valid():
                updated_record = serializer.save(updated_by=user)

                # Get final indices and keep lists
                final_approved_index = None
                final_carry_index = None
                try:
                    if "final_approved_bagging_file_index" in request.POST:
                        final_approved_index = int(
                            request.POST.get("final_approved_bagging_file_index")
                        )
                    if "final_carry_forward_file_index" in request.POST:
                        final_carry_index = int(request.POST.get("final_carry_forward_file_index"))
                except (ValueError, TypeError):
                    pass

                keep_approved_ids = set(request.POST.getlist("keep_approved_bagging_image_ids[]"))
                keep_carry_ids = set(request.POST.getlist("keep_carry_forward_image_ids[]"))
                final_approved_id = request.POST.get("final_approved_bagging_image_id")
                final_carry_id = request.POST.get("final_carry_forward_image_id")

                # Delete images not in keep list
                if "keep_approved_bagging_image_ids[]" in request.POST:
                    keep_approved_ids = set(
                        request.POST.getlist("keep_approved_bagging_image_ids[]")
                    )
                    StoneDemandToBaggingImage.objects.filter(
                        stone_demand=updated_record, field_type="approved_bagging"
                    ).exclude(id__in=keep_approved_ids).delete()
                if "keep_carry_forward_image_ids[]" in request.POST:
                    keep_carry_ids = set(request.POST.getlist("keep_carry_forward_image_ids[]"))
                    StoneDemandToBaggingImage.objects.filter(
                        stone_demand=updated_record, field_type="carry_forward"
                    ).exclude(id__in=keep_carry_ids).delete()

                # Update is_final on existing images
                if final_approved_id:
                    StoneDemandToBaggingImage.objects.filter(
                        stone_demand=updated_record, field_type="approved_bagging"
                    ).update(is_final=False)
                    StoneDemandToBaggingImage.objects.filter(
                        id=final_approved_id, stone_demand=updated_record
                    ).update(is_final=True)

                if final_carry_id:
                    StoneDemandToBaggingImage.objects.filter(
                        stone_demand=updated_record, field_type="carry_forward"
                    ).update(is_final=False)
                    StoneDemandToBaggingImage.objects.filter(
                        id=final_carry_id, stone_demand=updated_record
                    ).update(is_final=True)

                # Create image records for new approved bagging + carry-forward images
                approved_bagging_files = request.FILES.getlist("approved_bagging_images[]")
                carry_forward_files_check = request.FILES.getlist("carry_forward_images[]")
                patch_log_group_single = (
                    _new_log_group()
                    if (approved_bagging_files or carry_forward_files_check)
                    else None
                )

                for idx, file in enumerate(approved_bagging_files):
                    is_final = final_approved_index is not None and idx == final_approved_index
                    StoneDemandToBaggingImage.objects.create(
                        stone_demand=updated_record,
                        image=file,
                        field_type="approved_bagging",
                        is_final=is_final,
                        log_group=patch_log_group_single,
                        company=updated_record.company,
                        uploaded_by=user,
                    )

                patch_log_group_carry = patch_log_group_single
                carry_forward_files = request.FILES.getlist("carry_forward_images[]")
                for idx, file in enumerate(carry_forward_files):
                    is_final = final_carry_index is not None and idx == final_carry_index
                    StoneDemandToBaggingImage.objects.create(
                        stone_demand=updated_record,
                        image=file,
                        field_type="carry_forward",
                        is_final=is_final,
                        log_group=patch_log_group_carry,
                        company=updated_record.company,
                        uploaded_by=user,
                    )

                # Handle independent log selections for approved_bagging + carry_forward
                save_lg = []
                select_log = request.POST.get("select_log_group")
                if select_log is not None:
                    updated_record.selected_log_group = select_log or None
                    save_lg.append("selected_log_group")
                elif patch_log_group_single and not updated_record.selected_log_group:
                    updated_record.selected_log_group = patch_log_group_single
                    save_lg.append("selected_log_group")
                select_secondary = request.POST.get("select_secondary_log_group")
                if select_secondary is not None:
                    updated_record.selected_secondary_log_group = select_secondary or None
                    save_lg.append("selected_secondary_log_group")
                elif patch_log_group_single and not updated_record.selected_secondary_log_group:
                    updated_record.selected_secondary_log_group = patch_log_group_single
                    save_lg.append("selected_secondary_log_group")
                if save_lg:
                    updated_record.save(update_fields=save_lg)

                updated_record.refresh_from_db()
                return JsonResponse(
                    StoneDemandToBaggingSerializer(
                        updated_record, context={"request": request}
                    ).data,
                    status=200,
                )
            return JsonResponse({"errors": serializer.errors}, status=400)

    if request.method == "DELETE":
        record.delete()
        return JsonResponse({"detail": "Deleted successfully"}, status=200)


# ============================================================================
# Pre-Rhodium Quality Check Views
# ============================================================================

from .models import PreRhodiumQualityCheck
from .serializers import PreRhodiumQualityCheckSerializer


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET"])
def pre_rhodium_quality_check_overview_view(request):
    """Get overview/statistics for Pre-Rhodium Quality Checks"""
    user = request.user
    queryset = PreRhodiumQualityCheck.objects.all()

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    # Calculate statistics
    total_records = queryset.count()

    # Recent records (last 7 days)
    from datetime import timedelta

    seven_days_ago = timezone.now() - timedelta(days=7)
    recent_records_count = queryset.filter(created_at__gte=seven_days_ago).count()

    # Get latest 5 records
    latest_records = queryset.order_by("-created_at")[:5]
    latest_records_data = [
        {
            "id": str(record.id),
            "account_order_id": record.account_order_id,
            "quality_check": record.quality_check,
            "has_image": bool(record.quality_check_image),
            "created_at": record.created_at.isoformat() if record.created_at else None,
            "created_by": record.created_by.email if record.created_by else None,
        }
        for record in latest_records
    ]

    return JsonResponse(
        {
            "total_records": total_records,
            "recent_records_count": recent_records_count,
            "latest_records": latest_records_data,
        },
        status=200,
    )


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def pre_rhodium_quality_check_collection_view(request):
    """List and create Pre-Rhodium Quality Check records."""
    user = request.user
    queryset = PreRhodiumQualityCheck.objects.select_related(
        "company", "order", "created_by", "updated_by"
    ).order_by("-created_at")

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    if request.method == "GET":
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(Q(account_order_id__icontains=search))

        # Date filtering (using created_at since no date field)
        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        if date_from and date_to:
            if date_from > date_to:
                return JsonResponse({"detail": "From date cannot be after To date"}, status=400)
            queryset = queryset.filter(
                created_at__date__gte=date_from, created_at__date__lte=date_to
            )
        elif date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        elif date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [
            PreRhodiumQualityCheckSerializer(d, context={"request": request}).data
            for d in queryset[start:end]
        ]

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    data = request.POST.copy()
    # Convert quality_check string to boolean
    if "quality_check" in data:
        data["quality_check"] = data["quality_check"].lower() in ("true", "1", "yes")

    is_draft = request.POST.get("is_draft", "false").lower() in ("true", "1", "yes")

    serializer = PreRhodiumQualityCheckSerializer(data=data, context={"request": request})
    if serializer.is_valid():
        record = serializer.save(company=company, created_by=user, is_draft=is_draft)

        from .models import PreRhodiumQualityCheckImage

        batch_log_group = _new_log_group() if request.FILES else None
        qc_files = request.FILES.getlist("quality_check_images[]")
        final_index = None
        try:
            if "final_quality_check_file_index" in request.POST:
                final_index = int(request.POST.get("final_quality_check_file_index"))
        except (ValueError, TypeError):
            pass
        for idx, file in enumerate(qc_files):
            PreRhodiumQualityCheckImage.objects.create(
                pre_rhodium_qc=record,
                image=file,
                is_final=(final_index is not None and idx == final_index),
                log_group=batch_log_group,
                company=company,
                uploaded_by=user,
            )

        if not is_draft:
            is_valid, missing = _check_form_record_fields(
                record, "PreRhodiumQualityCheck", uploaded_files=request.FILES
            )
            if not is_valid:
                record.delete()
                return JsonResponse(
                    {
                        "errors": {
                            "validation": missing,
                            "message": (
                                f"Cannot save â€” the following required fields are missing: "
                                f"{', '.join(missing)}. Use 'Save as Draft' to save incomplete forms."
                            ),
                            "missing_fields": missing,
                        }
                    },
                    status=400,
                )

        data_resp = PreRhodiumQualityCheckSerializer(record, context={"request": request}).data
        data_resp["is_draft"] = record.is_draft
        return JsonResponse(data_resp, status=201)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PATCH", "PUT", "DELETE"])
def pre_rhodium_quality_check_detail_view(request, pk):
    """Retrieve, update, or delete a Pre-Rhodium Quality Check record."""
    user = request.user
    queryset = PreRhodiumQualityCheck.objects.select_related(
        "company", "order", "created_by", "updated_by"
    ).order_by("-created_at")

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    try:
        record = queryset.get(pk=pk)
    except PreRhodiumQualityCheck.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(
            PreRhodiumQualityCheckSerializer(record, context={"request": request}).data, status=200
        )

    if request.method in ["PATCH", "PUT"]:
        # Parse PUT/PATCH multipart data
        parse_put_request(request)

        is_draft = request.POST.get("is_draft", "false").lower() in ("true", "1", "yes")

        # Only pass known fields to serializer
        serializer_data = {}
        if request.POST.get("account_order_id") is not None:
            serializer_data["account_order_id"] = request.POST.get("account_order_id")
        if "quality_check" in request.POST:
            serializer_data["quality_check"] = request.POST["quality_check"].lower() in (
                "true",
                "1",
                "yes",
            )

        serializer = PreRhodiumQualityCheckSerializer(
            record, data=serializer_data, partial=True, context={"request": request}
        )
        if serializer.is_valid():
            updated_record = serializer.save(updated_by=user, is_draft=is_draft)

            from .models import PreRhodiumQualityCheckImage

            if "keep_quality_check_image_ids[]" in request.POST:
                keep_ids = set(request.POST.getlist("keep_quality_check_image_ids[]"))
            else:
                keep_ids = set()
            if keep_ids:
                PreRhodiumQualityCheckImage.objects.filter(pre_rhodium_qc=updated_record).exclude(
                    id__in=keep_ids
                ).delete()
            final_id = request.POST.get("final_quality_check_image_id")
            if final_id:
                PreRhodiumQualityCheckImage.objects.filter(pre_rhodium_qc=updated_record).update(
                    is_final=False
                )
                PreRhodiumQualityCheckImage.objects.filter(
                    id=final_id, pre_rhodium_qc=updated_record
                ).update(is_final=True)
            final_index = None
            try:
                if "final_quality_check_file_index" in request.POST:
                    final_index = int(request.POST.get("final_quality_check_file_index"))
            except (ValueError, TypeError):
                pass
            patch_log_group = _new_log_group() if request.FILES else None
            for idx, file in enumerate(request.FILES.getlist("quality_check_images[]")):
                PreRhodiumQualityCheckImage.objects.create(
                    pre_rhodium_qc=updated_record,
                    image=file,
                    is_final=(final_index is not None and idx == final_index),
                    log_group=patch_log_group,
                    company=updated_record.company,
                    uploaded_by=user,
                )

            select_log = request.POST.get("select_log_group")
            if select_log is not None:
                updated_record.selected_log_group = select_log or None
                updated_record.save(update_fields=["selected_log_group"])

            updated_record.refresh_from_db()
            return JsonResponse(
                PreRhodiumQualityCheckSerializer(updated_record, context={"request": request}).data,
                status=200,
            )
        return JsonResponse({"errors": serializer.errors}, status=400)

    if request.method == "DELETE":
        record.delete()
        return JsonResponse({"detail": "Deleted successfully"}, status=200)


# ============================================================================
# Final Quality Check Views
# ============================================================================

from .models import FinalQualityCheck
from .serializers import FinalQualityCheckSerializer


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def final_quality_check_collection_view(request):
    """List and create Final Quality Check records."""
    user = request.user
    queryset = FinalQualityCheck.objects.select_related(
        "company", "order", "created_by", "updated_by"
    ).order_by("-created_at")

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    if request.method == "GET":
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(Q(account_order_id__icontains=search))

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [
            FinalQualityCheckSerializer(d, context={"request": request}).data
            for d in queryset[start:end]
        ]

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    data = request.POST.copy()
    # Convert final_quality_check string to boolean
    if "final_quality_check" in data:
        data["final_quality_check"] = data["final_quality_check"].lower() in ("true", "1", "yes")

    is_draft = request.POST.get("is_draft", "false").lower() in ("true", "1", "yes")

    serializer = FinalQualityCheckSerializer(data=data, context={"request": request})
    if serializer.is_valid():
        record = serializer.save(company=company, created_by=user, is_draft=is_draft)

        if "final_quality_check_image" in request.FILES:
            record.final_quality_check_image = request.FILES["final_quality_check_image"]
            record.save()

        if not is_draft:
            is_valid, missing = _check_form_record_fields(
                record, "FinalQualityCheck", uploaded_files=request.FILES
            )
            if not is_valid:
                record.delete()
                return JsonResponse(
                    {
                        "errors": {
                            "validation": missing,
                            "message": (
                                f"Cannot save â€” the following required fields are missing: "
                                f"{', '.join(missing)}. Use 'Save as Draft' to save incomplete forms."
                            ),
                            "missing_fields": missing,
                        }
                    },
                    status=400,
                )

        data_resp = FinalQualityCheckSerializer(record, context={"request": request}).data
        data_resp["is_draft"] = record.is_draft
        return JsonResponse(data_resp, status=201)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PATCH", "PUT", "DELETE"])
def final_quality_check_detail_view(request, pk):
    """Retrieve, update, or delete a Final Quality Check record."""
    user = request.user
    queryset = FinalQualityCheck.objects.select_related(
        "company", "order", "created_by", "updated_by"
    ).order_by("-created_at")

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    try:
        record = queryset.get(pk=pk)
    except FinalQualityCheck.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(
            FinalQualityCheckSerializer(record, context={"request": request}).data, status=200
        )

    if request.method in ["PATCH", "PUT"]:
        # Parse PUT/PATCH multipart data
        parse_put_request(request)

        data = request.POST.copy()
        # Convert final_quality_check string to boolean
        if "final_quality_check" in data:
            data["final_quality_check"] = data["final_quality_check"].lower() in (
                "true",
                "1",
                "yes",
            )

        serializer = FinalQualityCheckSerializer(
            record, data=data, partial=True, context={"request": request}
        )
        if serializer.is_valid():
            updated_record = serializer.save(updated_by=user)

            if "final_quality_check_image" in request.FILES:
                updated_record.final_quality_check_image = request.FILES[
                    "final_quality_check_image"
                ]
                updated_record.save()
                updated_record.refresh_from_db()

            return JsonResponse(
                FinalQualityCheckSerializer(updated_record, context={"request": request}).data,
                status=200,
            )
        return JsonResponse({"errors": serializer.errors}, status=400)

    if request.method == "DELETE":
        record.delete()
        return JsonResponse({"detail": "Deleted successfully"}, status=200)


# ============================================================================
# Item Final Packing List Views
# ============================================================================

from .models import ItemFinalPackingList
from .serializers import ItemFinalPackingListSerializer


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET"])
def item_final_packing_list_overview_view(request):
    """Get overview/statistics for Item Final Packing Lists"""
    user = request.user
    queryset = ItemFinalPackingList.objects.all()

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    from datetime import timedelta

    seven_days_ago = timezone.now() - timedelta(days=7)

    total_records = queryset.count()
    recent_records = queryset.filter(created_at__gte=seven_days_ago).count()

    latest_records = queryset.order_by("-created_at")[:5]
    latest_data = [
        {
            "id": str(record.id),
            "order_id": str(record.order_id) if record.order_id else None,
            "created_at": record.created_at.isoformat() if record.created_at else None,
        }
        for record in latest_records
    ]

    return JsonResponse(
        {
            "total_records": total_records,
            "recent_records_count": recent_records,
            "latest_records": latest_data,
        },
        status=200,
    )


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def item_final_packing_list_collection_view(request):
    """List and create Item Final Packing List records."""
    user = request.user
    queryset = ItemFinalPackingList.objects.select_related(
        "company", "order", "created_by", "updated_by"
    ).order_by("-created_at")

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    if request.method == "GET":
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(Q(account_order_id__icontains=search))

        # Date filtering (using created_at since no date field)
        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        if date_from and date_to:
            if date_from > date_to:
                return JsonResponse({"detail": "From date cannot be after To date"}, status=400)
            queryset = queryset.filter(
                created_at__date__gte=date_from, created_at__date__lte=date_to
            )
        elif date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        elif date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [
            ItemFinalPackingListSerializer(d, context={"request": request}).data
            for d in queryset[start:end]
        ]

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    data = request.POST.copy()

    is_draft = request.POST.get("is_draft", "false").lower() in ("true", "1", "yes")

    serializer = ItemFinalPackingListSerializer(data=data, context={"request": request})
    if serializer.is_valid():
        record = serializer.save(company=company, created_by=user, is_draft=is_draft)

        from .models import ItemFinalPackingListImage

        batch_log_group = _new_log_group() if request.FILES else None
        jp_files = request.FILES.getlist("jewellery_piece_images[]")
        final_index = None
        try:
            if "final_jewellery_piece_file_index" in request.POST:
                final_index = int(request.POST.get("final_jewellery_piece_file_index"))
        except (ValueError, TypeError):
            pass
        for idx, file in enumerate(jp_files):
            ItemFinalPackingListImage.objects.create(
                packing_list=record,
                image=file,
                is_final=(final_index is not None and idx == final_index),
                log_group=batch_log_group,
                company=company,
                uploaded_by=user,
            )

        if not is_draft:
            is_valid, missing = _check_form_record_fields(
                record, "ItemFinalPackingList", uploaded_files=request.FILES
            )
            if not is_valid:
                record.delete()
                return JsonResponse(
                    {
                        "errors": {
                            "validation": missing,
                            "message": (
                                f"Cannot save â€” the following required fields are missing: "
                                f"{', '.join(missing)}. Use 'Save as Draft' to save incomplete forms."
                            ),
                            "missing_fields": missing,
                        }
                    },
                    status=400,
                )

        data_resp = ItemFinalPackingListSerializer(record, context={"request": request}).data
        data_resp["is_draft"] = record.is_draft
        return JsonResponse(data_resp, status=201)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PATCH", "PUT", "DELETE"])
def item_final_packing_list_detail_view(request, pk):
    """Retrieve, update, or delete an Item Final Packing List record."""
    user = request.user
    queryset = ItemFinalPackingList.objects.select_related(
        "company", "order", "created_by", "updated_by"
    ).order_by("-created_at")

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    try:
        record = queryset.get(pk=pk)
    except ItemFinalPackingList.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(
            ItemFinalPackingListSerializer(record, context={"request": request}).data, status=200
        )

    if request.method in ["PATCH", "PUT"]:
        # Parse PUT/PATCH multipart data
        parse_put_request(request)

        is_draft = request.POST.get("is_draft", "false").lower() in ("true", "1", "yes")

        # Only pass known fields to serializer
        serializer_data = {}
        if request.POST.get("account_order_id") is not None:
            serializer_data["account_order_id"] = request.POST.get("account_order_id")

        serializer = ItemFinalPackingListSerializer(
            record, data=serializer_data, partial=True, context={"request": request}
        )
        if serializer.is_valid():
            updated_record = serializer.save(updated_by=user, is_draft=is_draft)

            from .models import ItemFinalPackingListImage

            if "keep_jewellery_piece_image_ids[]" in request.POST:
                keep_ids = set(request.POST.getlist("keep_jewellery_piece_image_ids[]"))
            else:
                keep_ids = set()
            if keep_ids:
                ItemFinalPackingListImage.objects.filter(packing_list=updated_record).exclude(
                    id__in=keep_ids
                ).delete()
            final_id = request.POST.get("final_jewellery_piece_image_id")
            if final_id:
                ItemFinalPackingListImage.objects.filter(packing_list=updated_record).update(
                    is_final=False
                )
                ItemFinalPackingListImage.objects.filter(
                    id=final_id, packing_list=updated_record
                ).update(is_final=True)
            final_index = None
            try:
                if "final_jewellery_piece_file_index" in request.POST:
                    final_index = int(request.POST.get("final_jewellery_piece_file_index"))
            except (ValueError, TypeError):
                pass
            patch_log_group = _new_log_group() if request.FILES else None
            for idx, file in enumerate(request.FILES.getlist("jewellery_piece_images[]")):
                ItemFinalPackingListImage.objects.create(
                    packing_list=updated_record,
                    image=file,
                    is_final=(final_index is not None and idx == final_index),
                    log_group=patch_log_group,
                    company=updated_record.company,
                    uploaded_by=user,
                )

            select_log = request.POST.get("select_log_group")
            if select_log is not None:
                updated_record.selected_log_group = select_log or None
                updated_record.save(update_fields=["selected_log_group"])

            updated_record.refresh_from_db()
            return JsonResponse(
                ItemFinalPackingListSerializer(updated_record, context={"request": request}).data,
                status=200,
            )
        return JsonResponse({"errors": serializer.errors}, status=400)

    if request.method == "DELETE":
        record.delete()
        return JsonResponse({"detail": "Deleted successfully"}, status=200)


# ============================================================================
# Raw Material Tally Views
# ============================================================================

from .models import RawMaterialTally
from .serializers import RawMaterialTallySerializer


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET"])
def raw_material_tally_overview_view(request):
    """Get overview/statistics for Raw Material Tally"""
    user = request.user
    queryset = RawMaterialTally.objects.all()

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    # Calculate statistics
    total_records = queryset.count()

    # Recent records (last 7 days)
    from datetime import timedelta

    seven_days_ago = timezone.now() - timedelta(days=7)
    recent_records_count = queryset.filter(created_at__gte=seven_days_ago).count()

    # Get latest 5 records
    latest_records = queryset.order_by("-created_at")[:5]
    latest_records_data = [
        {
            "id": str(record.id),
            "account_order_id": record.account_order_id,
            "created_at": record.created_at.isoformat() if record.created_at else None,
            "created_by": record.created_by.email if record.created_by else None,
        }
        for record in latest_records
    ]

    return JsonResponse(
        {
            "total_records": total_records,
            "recent_records_count": recent_records_count,
            "latest_records": latest_records_data,
        },
        status=200,
    )


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def raw_material_tally_collection_view(request):
    """List and create Raw Material Tally records. Supports multiple entries in one submission."""
    user = request.user
    queryset = RawMaterialTally.objects.select_related(
        "company", "order", "created_by", "updated_by"
    ).order_by("-created_at")

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    if request.method == "GET":
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(Q(account_order_id__icontains=search))

        # Date filtering (using created_at since no date field)
        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        if date_from and date_to:
            if date_from > date_to:
                return JsonResponse({"detail": "From date cannot be after To date"}, status=400)
            queryset = queryset.filter(
                created_at__date__gte=date_from, created_at__date__lte=date_to
            )
        elif date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        elif date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [
            RawMaterialTallySerializer(d, context={"request": request}).data
            for d in queryset[start:end]
        ]

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    # Check if material_entries is provided (multiple entries mode)
    material_entries_json = request.POST.get("material_entries")

    if material_entries_json:
        # MULTIPLE ENTRIES MODE
        try:
            material_entries = json.loads(material_entries_json)
            if not isinstance(material_entries, list):
                return JsonResponse(
                    {"errors": {"material_entries": ["Must be a list of entries"]}}, status=400
                )

            if len(material_entries) == 0:
                return JsonResponse(
                    {"errors": {"material_entries": ["At least one entry is required"]}}, status=400
                )

            # Get common fields
            account_order_id = request.POST.get("account_order_id")
            if not account_order_id:
                return JsonResponse(
                    {"errors": {"account_order_id": ["This field is required"]}}, status=400
                )

            # Get carry forward image (shared across all entries)
            carry_forward_image = request.FILES.get("carry_forward_image")

            # Create multiple records
            created_records = []
            for entry in material_entries:
                # Each entry becomes a separate record with its own raw_material_movement
                record_data = {
                    "account_order_id": account_order_id,
                    "raw_material_movement": entry,  # Store the entire entry as JSON
                }

                serializer = RawMaterialTallySerializer(
                    data=record_data, context={"request": request}
                )
                if serializer.is_valid():
                    record = serializer.save(company=company, created_by=user)

                    # Attach image to all records
                    if carry_forward_image:
                        record.carry_forward_image = carry_forward_image
                    record.save()

                    created_records.append(
                        RawMaterialTallySerializer(record, context={"request": request}).data
                    )
                else:
                    return JsonResponse({"errors": serializer.errors}, status=400)

            return JsonResponse(
                {
                    "message": f"Successfully created {len(created_records)} raw material tally records",
                    "count": len(created_records),
                    "records": created_records,
                },
                status=201,
            )

        except json.JSONDecodeError:
            return JsonResponse(
                {"errors": {"material_entries": ["Invalid JSON format"]}}, status=400
            )

    else:
        # SINGLE ENTRY MODE (backward compatibility)
        data = request.POST.copy()
        # Note: raw_material_movement parsing is now handled by serializer validation

        is_draft = request.POST.get("is_draft", "false").lower() in ("true", "1", "yes")

        serializer = RawMaterialTallySerializer(data=data, context={"request": request})
        if serializer.is_valid():
            record = serializer.save(company=company, created_by=user, is_draft=is_draft)

            from .models import RawMaterialTallyImage

            batch_log_group = _new_log_group() if request.FILES else None
            cf_files = request.FILES.getlist("carry_forward_images[]")
            final_index = None
            try:
                if "final_carry_forward_file_index" in request.POST:
                    final_index = int(request.POST.get("final_carry_forward_file_index"))
            except (ValueError, TypeError):
                pass
            for idx, file in enumerate(cf_files):
                RawMaterialTallyImage.objects.create(
                    raw_material_tally=record,
                    image=file,
                    is_final=(final_index is not None and idx == final_index),
                    log_group=batch_log_group,
                    company=company,
                    uploaded_by=user,
                )

            if not is_draft:
                is_valid, missing = _check_form_record_fields(
                    record, "RawMaterialTally", uploaded_files=request.FILES
                )
                if not is_valid:
                    record.delete()
                    return JsonResponse(
                        {
                            "errors": {
                                "validation": missing,
                                "message": (
                                    f"Cannot save â€” the following required fields are missing: "
                                    f"{', '.join(missing)}. Use 'Save as Draft' to save incomplete forms."
                                ),
                                "missing_fields": missing,
                            }
                        },
                        status=400,
                    )

            data_resp = RawMaterialTallySerializer(record, context={"request": request}).data
            data_resp["is_draft"] = record.is_draft
            return JsonResponse(data_resp, status=201)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PATCH", "PUT", "DELETE"])
def raw_material_tally_detail_view(request, pk):
    """Retrieve, update, or delete a Raw Material Tally record."""
    user = request.user
    queryset = RawMaterialTally.objects.select_related(
        "company", "order", "created_by", "updated_by"
    ).order_by("-created_at")

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    try:
        record = queryset.get(pk=pk)
    except RawMaterialTally.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(
            RawMaterialTallySerializer(record, context={"request": request}).data, status=200
        )

    if request.method in ["PATCH", "PUT"]:
        # Parse PUT/PATCH multipart data
        parse_put_request(request)

        # Build serializer_data with only known model fields to avoid validation errors
        is_draft = request.POST.get("is_draft", "false").lower() in ("true", "1", "yes")
        serializer_data = {}
        if request.POST.get("account_order_id") is not None:
            serializer_data["account_order_id"] = request.POST.get("account_order_id")
        if request.POST.get("raw_material_movement") is not None:
            serializer_data["raw_material_movement"] = request.POST.get("raw_material_movement")

        serializer = RawMaterialTallySerializer(
            record, data=serializer_data, partial=True, context={"request": request}
        )
        if serializer.is_valid():
            updated_record = serializer.save(updated_by=user, is_draft=is_draft)

            from .models import RawMaterialTallyImage

            # Delete images not in keep list
            if "keep_carry_forward_image_ids[]" in request.POST:
                keep_ids = set(request.POST.getlist("keep_carry_forward_image_ids[]"))
                RawMaterialTallyImage.objects.filter(raw_material_tally=updated_record).exclude(
                    id__in=keep_ids
                ).delete()

            # Create new images with log_group
            cf_files = request.FILES.getlist("carry_forward_images[]")
            patch_log_group = _new_log_group() if cf_files else None
            final_index = None
            try:
                if "final_carry_forward_file_index" in request.POST:
                    final_index = int(request.POST.get("final_carry_forward_file_index"))
            except (ValueError, TypeError):
                pass
            for idx, file in enumerate(cf_files):
                RawMaterialTallyImage.objects.create(
                    raw_material_tally=updated_record,
                    image=file,
                    is_final=(final_index is not None and idx == final_index),
                    log_group=patch_log_group,
                    company=updated_record.company,
                    uploaded_by=user,
                )

            # Handle log group selection
            save_lg = []
            select_log = request.POST.get("select_log_group")
            if select_log is not None:
                updated_record.selected_log_group = select_log or None
                save_lg.append("selected_log_group")
            if save_lg:
                updated_record.save(update_fields=save_lg)

            updated_record.refresh_from_db()
            return JsonResponse(
                RawMaterialTallySerializer(updated_record, context={"request": request}).data,
                status=200,
            )
        return JsonResponse({"errors": serializer.errors}, status=400)

    if request.method == "DELETE":
        record.delete()
        return JsonResponse({"detail": "Deleted successfully"}, status=200)


# ============================================================================
# Metal Issue Views
# ============================================================================

from .models import MetalIssue
from .serializers import MetalIssueSerializer


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def metal_issue_collection_view(request):
    """List and create Metal Issue records."""
    user = request.user
    queryset = MetalIssue.objects.select_related(
        "company", "order", "created_by", "updated_by"
    ).order_by("-created_at")

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    if request.method == "GET":
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(Q(account_order_id__icontains=search))

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [
            MetalIssueSerializer(d, context={"request": request}).data for d in queryset[start:end]
        ]

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    data = request.POST.copy()

    is_draft = request.POST.get("is_draft", "false").lower() in ("true", "1", "yes")

    serializer = MetalIssueSerializer(data=data, context={"request": request})
    if serializer.is_valid():
        record = serializer.save(company=company, created_by=user, is_draft=is_draft)

        if "carry_forward_image" in request.FILES:
            record.carry_forward_image = request.FILES["carry_forward_image"]
            record.save()

        if not is_draft:
            is_valid, missing = _check_form_record_fields(
                record, "MetalIssue", uploaded_files=request.FILES
            )
            if not is_valid:
                record.delete()
                return JsonResponse(
                    {
                        "errors": {
                            "validation": missing,
                            "message": (
                                f"Cannot save â€” the following required fields are missing: "
                                f"{', '.join(missing)}. Use 'Save as Draft' to save incomplete forms."
                            ),
                            "missing_fields": missing,
                        }
                    },
                    status=400,
                )

        data_resp = MetalIssueSerializer(record, context={"request": request}).data
        data_resp["is_draft"] = record.is_draft
        return JsonResponse(data_resp, status=201)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PATCH", "PUT", "DELETE"])
def metal_issue_detail_view(request, pk):
    """Retrieve, update, or delete a Metal Issue record."""
    user = request.user
    queryset = MetalIssue.objects.select_related(
        "company", "order", "created_by", "updated_by"
    ).order_by("-created_at")

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    try:
        record = queryset.get(pk=pk)
    except MetalIssue.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(
            MetalIssueSerializer(record, context={"request": request}).data, status=200
        )

    if request.method in ["PATCH", "PUT"]:
        # Parse PUT/PATCH multipart data
        parse_put_request(request)

        data = request.POST.copy()

        serializer = MetalIssueSerializer(
            record, data=data, partial=True, context={"request": request}
        )
        if serializer.is_valid():
            updated_record = serializer.save(updated_by=user)

            if "carry_forward_image" in request.FILES:
                updated_record.carry_forward_image = request.FILES["carry_forward_image"]
                updated_record.save()
                updated_record.refresh_from_db()

            return JsonResponse(
                MetalIssueSerializer(updated_record, context={"request": request}).data, status=200
            )
        return JsonResponse({"errors": serializer.errors}, status=400)

    if request.method == "DELETE":
        record.delete()
        return JsonResponse({"detail": "Deleted successfully"}, status=200)


# ============================================================================
# Bagging Ready Views
# ============================================================================

from .models import BaggingReady
from .serializers import BaggingReadySerializer


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def bagging_ready_collection_view(request):
    """List and create Bagging Ready records."""
    user = request.user
    queryset = BaggingReady.objects.select_related(
        "company", "order", "created_by", "updated_by"
    ).order_by("-created_at")

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    if request.method == "GET":
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(Q(account_order_id__icontains=search))

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [
            BaggingReadySerializer(d, context={"request": request}).data
            for d in queryset[start:end]
        ]

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    data = request.POST.copy()

    is_draft = request.POST.get("is_draft", "false").lower() in ("true", "1", "yes")

    serializer = BaggingReadySerializer(data=data, context={"request": request})
    if serializer.is_valid():
        record = serializer.save(company=company, created_by=user, is_draft=is_draft)

        if "carry_forward_image" in request.FILES:
            record.carry_forward_image = request.FILES["carry_forward_image"]
            record.save()

        if not is_draft:
            is_valid, missing = _check_form_record_fields(
                record, "BaggingReady", uploaded_files=request.FILES
            )
            if not is_valid:
                record.delete()
                return JsonResponse(
                    {
                        "errors": {
                            "validation": missing,
                            "message": (
                                f"Cannot save â€” the following required fields are missing: "
                                f"{', '.join(missing)}. Use 'Save as Draft' to save incomplete forms."
                            ),
                            "missing_fields": missing,
                        }
                    },
                    status=400,
                )

        data_resp = BaggingReadySerializer(record, context={"request": request}).data
        data_resp["is_draft"] = record.is_draft
        return JsonResponse(data_resp, status=201)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PATCH", "PUT", "DELETE"])
def bagging_ready_detail_view(request, pk):
    """Retrieve, update, or delete a Bagging Ready record."""
    user = request.user
    queryset = BaggingReady.objects.select_related(
        "company", "order", "created_by", "updated_by"
    ).order_by("-created_at")

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    try:
        record = queryset.get(pk=pk)
    except BaggingReady.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(
            BaggingReadySerializer(record, context={"request": request}).data, status=200
        )

    if request.method in ["PATCH", "PUT"]:
        # Parse PUT/PATCH multipart data
        parse_put_request(request)

        data = request.POST.copy()

        serializer = BaggingReadySerializer(
            record, data=data, partial=True, context={"request": request}
        )
        if serializer.is_valid():
            updated_record = serializer.save(updated_by=user)

            if "carry_forward_image" in request.FILES:
                updated_record.carry_forward_image = request.FILES["carry_forward_image"]
                updated_record.save()
                updated_record.refresh_from_db()

            return JsonResponse(
                BaggingReadySerializer(updated_record, context={"request": request}).data,
                status=200,
            )
        return JsonResponse({"errors": serializer.errors}, status=400)

    if request.method == "DELETE":
        record.delete()
        return JsonResponse({"detail": "Deleted successfully"}, status=200)


# ============================================================================
# Diamond Purchase/Issue Views
# ============================================================================

from .models import DiamondPurchaseIssue
from .serializers import DiamondPurchaseIssueSerializer


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def diamond_purchase_issue_collection_view(request):
    """List and create Diamond Purchase/Issue records."""
    user = request.user
    queryset = DiamondPurchaseIssue.objects.select_related(
        "company", "order", "created_by", "updated_by"
    ).order_by("-created_at")

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    if request.method == "GET":
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(Q(account_order_id__icontains=search))

        # Ordering
        ordering = request.GET.get("ordering")
        if ordering:
            if ordering == "date_asc":
                queryset = queryset.order_by("created_at")
            elif ordering == "date_desc":
                queryset = queryset.order_by("-created_at")

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [
            DiamondPurchaseIssueSerializer(d, context={"request": request}).data
            for d in queryset[start:end]
        ]

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    data = request.POST.copy()

    is_draft = request.POST.get("is_draft", "false").lower() in ("true", "1", "yes")

    serializer = DiamondPurchaseIssueSerializer(data=data, context={"request": request})
    if serializer.is_valid():
        record = serializer.save(company=company, created_by=user, is_draft=is_draft)

        if "carry_forward_image" in request.FILES:
            record.carry_forward_image = request.FILES["carry_forward_image"]
            record.save()

        if not is_draft:
            is_valid, missing = _check_form_record_fields(
                record, "DiamondPurchaseIssue", uploaded_files=request.FILES
            )
            if not is_valid:
                record.delete()
                return JsonResponse(
                    {
                        "errors": {
                            "validation": missing,
                            "message": (
                                f"Cannot save â€” the following required fields are missing: "
                                f"{', '.join(missing)}. Use 'Save as Draft' to save incomplete forms."
                            ),
                            "missing_fields": missing,
                        }
                    },
                    status=400,
                )

        data_resp = DiamondPurchaseIssueSerializer(record, context={"request": request}).data
        data_resp["is_draft"] = record.is_draft
        return JsonResponse(data_resp, status=201)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PATCH", "PUT", "DELETE"])
def diamond_purchase_issue_detail_view(request, pk):
    """Retrieve, update, or delete a Diamond Purchase/Issue record."""
    user = request.user
    queryset = DiamondPurchaseIssue.objects.select_related(
        "company", "order", "created_by", "updated_by"
    ).order_by("-created_at")

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    try:
        record = queryset.get(pk=pk)
    except DiamondPurchaseIssue.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(
            DiamondPurchaseIssueSerializer(record, context={"request": request}).data, status=200
        )

    if request.method in ["PATCH", "PUT"]:
        # Parse PUT/PATCH multipart data
        parse_put_request(request)

        data = request.POST.copy()

        serializer = DiamondPurchaseIssueSerializer(
            record, data=data, partial=True, context={"request": request}
        )
        if serializer.is_valid():
            updated_record = serializer.save(updated_by=user)

            if "carry_forward_image" in request.FILES:
                updated_record.carry_forward_image = request.FILES["carry_forward_image"]
                updated_record.save()
                updated_record.refresh_from_db()

            return JsonResponse(
                DiamondPurchaseIssueSerializer(updated_record, context={"request": request}).data,
                status=200,
            )
        return JsonResponse({"errors": serializer.errors}, status=400)

    if request.method == "DELETE":
        record.delete()
        return JsonResponse({"detail": "Deleted successfully"}, status=200)


# ============================================================================
# Gemstone Purchase/Issue Views
# ============================================================================

from .models import GemstonePurchaseIssue
from .serializers import GemstonePurchaseIssueSerializer


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "POST"])
def gemstone_purchase_issue_collection_view(request):
    """List and create Gemstone Purchase/Issue records."""
    user = request.user
    queryset = GemstonePurchaseIssue.objects.select_related(
        "company", "order", "created_by", "updated_by"
    ).order_by("-created_at")

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    if request.method == "GET":
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(Q(account_order_id__icontains=search))

        # Ordering
        ordering = request.GET.get("ordering")
        if ordering:
            if ordering == "date_asc":
                queryset = queryset.order_by("created_at")
            elif ordering == "date_desc":
                queryset = queryset.order_by("-created_at")

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [
            GemstonePurchaseIssueSerializer(d, context={"request": request}).data
            for d in queryset[start:end]
        ]

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return JsonResponse(payload, status=200)

    company = getattr(user, "company", None)
    if company is None:
        return JsonResponse(
            {"errors": {"company": ["User does not belong to a company"]}}, status=400
        )

    data = request.POST.copy()

    is_draft = request.POST.get("is_draft", "false").lower() in ("true", "1", "yes")

    serializer = GemstonePurchaseIssueSerializer(data=data, context={"request": request})
    if serializer.is_valid():
        record = serializer.save(company=company, created_by=user, is_draft=is_draft)

        if "carry_forward_image" in request.FILES:
            record.carry_forward_image = request.FILES["carry_forward_image"]
            record.save()

        if not is_draft:
            is_valid, missing = _check_form_record_fields(
                record, "GemstonePurchaseIssue", uploaded_files=request.FILES
            )
            if not is_valid:
                record.delete()
                return JsonResponse(
                    {
                        "errors": {
                            "validation": missing,
                            "message": (
                                f"Cannot save â€” the following required fields are missing: "
                                f"{', '.join(missing)}. Use 'Save as Draft' to save incomplete forms."
                            ),
                            "missing_fields": missing,
                        }
                    },
                    status=400,
                )

        data_resp = GemstonePurchaseIssueSerializer(record, context={"request": request}).data
        data_resp["is_draft"] = record.is_draft
        return JsonResponse(data_resp, status=201)

    return JsonResponse({"errors": serializer.errors}, status=400)


@csrf_exempt
@jwt_login_required
@require_http_methods(["GET", "PATCH", "PUT", "DELETE"])
def gemstone_purchase_issue_detail_view(request, pk):
    """Retrieve, update, or delete a Gemstone Purchase/Issue record."""
    user = request.user
    queryset = GemstonePurchaseIssue.objects.select_related(
        "company", "order", "created_by", "updated_by"
    ).order_by("-created_at")

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()

    try:
        record = queryset.get(pk=pk)
    except GemstonePurchaseIssue.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(
            GemstonePurchaseIssueSerializer(record, context={"request": request}).data, status=200
        )

    if request.method in ["PATCH", "PUT"]:
        # Parse PUT/PATCH multipart data
        parse_put_request(request)

        data = request.POST.copy()

        serializer = GemstonePurchaseIssueSerializer(
            record, data=data, partial=True, context={"request": request}
        )
        if serializer.is_valid():
            updated_record = serializer.save(updated_by=user)

            if "carry_forward_image" in request.FILES:
                updated_record.carry_forward_image = request.FILES["carry_forward_image"]
                updated_record.save()
                updated_record.refresh_from_db()

            return JsonResponse(
                GemstonePurchaseIssueSerializer(updated_record, context={"request": request}).data,
                status=200,
            )
        return JsonResponse({"errors": serializer.errors}, status=400)

    if request.method == "DELETE":
        record.delete()
        return JsonResponse({"detail": "Deleted successfully"}, status=200)


@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def order_id_preview_view(request):
    """
    Preview the next Order ID for a given order type.
    Useful for UI display before order creation.

    Query Parameters:
    - order_type: One of STOCK_JEWELRY, BESPOKE_NATURAL, BESPOKE_CVD, LOOSE_DIAMONDS

    Returns:
    - next_order_id: The next Order ID that would be generated
    - prefix: The prefix letter for this order type
    - current_sequence: Current sequence number
    """
    from .order_id_generator import OrderIDGenerator
    from .models import OrderSequence

    order_type = request.GET.get("order_type")

    if not order_type:
        return Response(
            {
                "error": "order_type parameter is required",
                "valid_types": [choice[0] for choice in OrderSequence.ORDER_TYPE_CHOICES],
            },
            status=400,
        )

    # Validate order type
    valid_types = [choice[0] for choice in OrderSequence.ORDER_TYPE_CHOICES]
    if order_type not in valid_types:
        return Response(
            {"error": f"Invalid order_type: {order_type}", "valid_types": valid_types}, status=400
        )

    try:
        # Get preview without incrementing sequence
        next_order_id = OrderIDGenerator.get_next_order_id_preview(order_type)
        prefix = OrderIDGenerator.get_prefix_for_order_type(order_type)

        # Get current sequence
        try:
            sequence_obj = OrderSequence.objects.get(order_type=order_type)
            current_sequence = sequence_obj.current_sequence
        except OrderSequence.DoesNotExist:
            current_sequence = 0

        return Response(
            {
                "next_order_id": next_order_id,
                "prefix": prefix,
                "current_sequence": current_sequence,
                "order_type": order_type,
                "order_type_display": dict(OrderSequence.ORDER_TYPE_CHOICES).get(order_type),
            }
        )

    except Exception as e:
        return Response({"error": f"Failed to preview Order ID: {str(e)}"}, status=500)


@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def order_sequence_status_view(request):
    """
    Get the current status of all Order ID sequences.
    Useful for admin monitoring and debugging.

    Returns:
    - sequences: Status of all order type sequences
    """
    from .order_id_generator import OrderIDGenerator

    try:
        status = OrderIDGenerator.get_sequence_status()

        return Response(
            {
                "sequences": status,
                "total_types": len(status),
                "timestamp": timezone.now().isoformat(),
            }
        )

    except Exception as e:
        return Response({"error": f"Failed to get sequence status: {str(e)}"}, status=500)


# =============================================================================
# INVOICE VIEWS
# =============================================================================

from .models import Invoice, InvoiceLineItem, InvoicePayment
from .serializers import (
    InvoiceSerializer,
    InvoiceLineItemSerializer,
    InvoicePaymentSerializer,
    CreateInvoiceFromSaleSerializer,
)


@api_view(["GET", "POST"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def invoice_collection_view(request):
    """List or create invoices for the current user's company."""
    user = request.user
    company = getattr(user, "company", None)

    if not company:
        return Response({"errors": {"company": ["User does not belong to a company"]}}, status=400)

    if request.method == "GET":
        # List invoices
        queryset = (
            Invoice.objects.filter(company=company)
            .select_related("sale", "estimate", "sales_query", "company", "created_by")
            .prefetch_related("line_items", "payments")
            .order_by("-invoice_date", "-created_at")
        )

        # Filter by status
        status = request.GET.get("status")
        if status:
            queryset = queryset.filter(status=status)

        # Search
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(
                Q(invoice_number__icontains=search)
                | Q(customer_name__icontains=search)
                | Q(customer_email__icontains=search)
            )

        # Pagination
        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size

        results = [
            InvoiceSerializer(inv, context={"request": request}).data for inv in queryset[start:end]
        ]

        return Response(
            {
                "count": total,
                "next": None,
                "previous": None,
                "results": results,
            }
        )

    # POST - Create invoice
    payload = _parse_json(request.body)
    if payload is None:
        return Response({"detail": "Invalid JSON"}, status=400)

    serializer = InvoiceSerializer(data=payload, context={"request": request})
    if serializer.is_valid():
        invoice = serializer.save(company=company, created_by=user)
        return Response(InvoiceSerializer(invoice, context={"request": request}).data, status=201)

    return Response({"errors": serializer.errors}, status=400)


@api_view(["GET", "PATCH", "DELETE"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def invoice_detail_view(request, pk):
    """Retrieve, update, or delete an invoice."""
    user = request.user
    company = getattr(user, "company", None)

    try:
        invoice = (
            Invoice.objects.filter(company=company)
            .select_related("sale", "estimate", "sales_query", "company", "created_by")
            .prefetch_related("line_items", "payments")
            .get(pk=pk)
        )
    except Invoice.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return Response(InvoiceSerializer(invoice, context={"request": request}).data)

    if request.method == "PATCH":
        payload = _parse_json(request.body)
        if payload is None:
            return Response({"detail": "Invalid JSON"}, status=400)

        serializer = InvoiceSerializer(
            invoice, data=payload, partial=True, context={"request": request}
        )
        if serializer.is_valid():
            updated = serializer.save(updated_by=user)
            return Response(InvoiceSerializer(updated, context={"request": request}).data)

        return Response({"errors": serializer.errors}, status=400)

    if request.method == "DELETE":
        invoice.delete()
        return Response(status=204)


@api_view(["POST"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def invoice_from_sale_view(request):
    """Create an invoice from a completed sale with estimate."""
    user = request.user
    company = getattr(user, "company", None)

    if not company:
        return Response({"errors": {"company": ["User does not belong to a company"]}}, status=400)

    payload = _parse_json(request.body)
    if payload is None:
        return Response({"detail": "Invalid JSON"}, status=400)

    serializer = CreateInvoiceFromSaleSerializer(
        data=payload, context={"user": user, "company": company}
    )

    if serializer.is_valid():
        invoice = serializer.save()
        return Response(InvoiceSerializer(invoice, context={"request": request}).data, status=201)

    return Response({"errors": serializer.errors}, status=400)


@api_view(["POST"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def invoice_from_estimate_view(request):
    """Create an invoice directly from an estimate (without sale)."""
    user = request.user
    company = getattr(user, "company", None)

    if not company:
        return Response({"errors": {"company": ["User does not belong to a company"]}}, status=400)

    payload = _parse_json(request.body)
    if payload is None:
        return Response({"detail": "Invalid JSON"}, status=400)

    from .serializers import CreateInvoiceFromEstimateSerializer

    serializer = CreateInvoiceFromEstimateSerializer(
        data=payload, context={"user": user, "company": company}
    )

    if serializer.is_valid():
        invoice = serializer.save()
        return Response(InvoiceSerializer(invoice, context={"request": request}).data, status=201)

    return Response({"errors": serializer.errors}, status=400)


@api_view(["POST"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def invoice_record_payment_view(request, pk):
    """Record a payment for an invoice."""
    user = request.user
    company = getattr(user, "company", None)

    try:
        invoice = Invoice.objects.filter(company=company).get(pk=pk)
    except Invoice.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)

    payload = _parse_json(request.body)
    if payload is None:
        return Response({"detail": "Invalid JSON"}, status=400)

    # Add invoice to payload
    payload["invoice"] = str(invoice.id)

    serializer = InvoicePaymentSerializer(data=payload)
    if serializer.is_valid():
        payment = serializer.save(invoice=invoice, recorded_by=user)

        # Return updated invoice
        invoice.refresh_from_db()
        return Response(InvoiceSerializer(invoice, context={"request": request}).data, status=201)

    return Response({"errors": serializer.errors}, status=400)


@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def invoice_pdf_view(request, pk):
    """Generate and download invoice PDF using template overlay."""
    user = request.user
    company = getattr(user, "company", None)

    try:
        invoice = (
            Invoice.objects.filter(company=company)
            .select_related("sale", "estimate", "sales_lead")
            .prefetch_related("line_items")
            .get(pk=pk)
        )
    except Invoice.DoesNotExist:
        return Response({"detail": "Invoice not found"}, status=404)

    try:
        # Check if template exists
        import os

        template_path = os.path.join(
            os.path.dirname(__file__), "..", "Satyam Jewellery-Invoice-2026.pdf"
        )

        if not os.path.exists(template_path):
            # Fallback to ReportLab if template not found
            from .invoice_pdf import generate_invoice_pdf

            line_items = []
            for line in invoice.line_items.all():
                line_items.append(
                    {
                        "particulars": line.particulars,
                        "description": getattr(line, "description", "") or "",
                        "shape": getattr(line, "shape", "") or "",
                        "colour": getattr(line, "colour", "") or "",
                        "clarity": getattr(line, "clarity", "") or "",
                        "quantity": line.quantity,
                        "weight": float(line.weight) if line.weight else 0,
                        "unit": line.unit or "GMS",
                        "rate": float(line.rate),
                        "amount": float(line.amount),
                    }
                )

            pdf_bytes = generate_invoice_pdf(
                invoice_number=invoice.invoice_number,
                invoice_date=invoice.invoice_date.strftime("%d/%m/%Y"),
                due_date=invoice.due_date.strftime("%d/%m/%Y"),
                customer_name=invoice.customer_name,
                customer_phone=invoice.phone_number or "",
                customer_email=invoice.email or "",
                customer_address=invoice.address or "",
                customer_gstin=invoice.pan_gstin or "",
                item_name=invoice.item_name,
                line_items=line_items,
                subtotal=float(invoice.subtotal),
                gst_rate=float(invoice.gst_rate),
                gst_amount=float(invoice.gst_amount),
                total_amount=float(invoice.total_amount),
                amount_paid=float(invoice.amount_paid),
                balance_due=float(invoice.balance_due),
                notes=invoice.notes or "",
                terms_and_conditions=invoice.terms_and_conditions or "",
                company_name=company.name if company else "Satyam Jewellery",
                company_address=getattr(company, "address", "") if company else "",
                company_phone=getattr(company, "phone", "") if company else "",
                company_email=getattr(company, "email", "") if company else "",
                company_gstin=getattr(company, "gstin", "") if company else "",
                image_data=None,
            )
        else:
            # Use template overlay with text-only redactions
            from .invoice_template_pdf import generate_invoice_template_pdf

            # Prepare line items â€” pull directly from ESTIMATE for accurate data
            line_items = []
            if invoice.estimate:
                # Pull from estimate line items (source of truth)
                estimate_items = invoice.estimate.line_items.all()
                print(
                    f"ðŸ“‹ PDF PAGE 1: Using ESTIMATE line items (estimate={invoice.estimate.id})"
                )
                print(f"ðŸ“‹ Found {estimate_items.count()} estimate line items")
                for line in estimate_items:
                    weight_val = float(line.weight) if line.weight else 0
                    rate_val = float(line.rate) if line.rate else 0
                    # Dynamically compute amount = rate Ã— weight
                    computed_amount = rate_val * weight_val
                    print(
                        f"  â†’ {line.particulars}: rate={rate_val}, weight={weight_val}, amount={computed_amount}"
                    )
                    line_items.append(
                        {
                            "particulars": line.particulars,
                            "hsn_code": "",
                            "quantity": line.pc or 1,
                            "weight": weight_val,
                            "rate": rate_val,
                            "unit": line.unit if line.unit else "GMS",
                            "amount": computed_amount,
                        }
                    )
            else:
                print(f"âš ï¸ PDF PAGE 1: No estimate linked, using INVOICE line items (fallback)")
                # Fallback: use invoice line items if no estimate linked
                for line in invoice.line_items.all():
                    print(
                        f"  â†’ {line.particulars}: rate={float(line.rate)}, weight={float(line.weight) if line.weight else 0}"
                    )
                    line_items.append(
                        {
                            "particulars": line.particulars,
                            "hsn_code": getattr(line, "hsn_code", "") or "",
                            "quantity": line.quantity,
                            "weight": float(line.weight) if line.weight else 0,
                            "rate": float(line.rate),
                            "unit": line.unit if line.unit else "GMS",
                            "amount": float(line.amount),
                        }
                    )
            print(f"ðŸ“‹ Total line items prepared for PDF page 1: {len(line_items)}")

            # Get company details
            company_name = company.name if company else "Niti Shah Jewels"
            company_address = (
                getattr(company, "address", "2/63 JVPD, NS Road No. 1, Vile Parle West, Mumbai")
                if company
                else "2/63 JVPD, NS Road No. 1, Vile Parle West, Mumbai"
            )
            company_phone = (
                getattr(company, "phone", "+919987520906") if company else "+919987520906"
            )
            company_email = (
                getattr(company, "email", "hello@nitishahjewels.com")
                if company
                else "hello@nitishahjewels.com"
            )
            company_gstin = (
                getattr(company, "gstin", "27FSFPS4058J1Z5") if company else "27FSFPS4058J1Z5"
            )
            company_state = getattr(company, "state", "Maharashtra") if company else "Maharashtra"
            company_state_code = getattr(company, "state_code", "27") if company else "27"
            company_pan = getattr(company, "pan", "AEFPG4541P") if company else "AEFPG4541P"

            # Get customer state info
            customer_state = ""
            customer_state_code = ""
            if invoice.address:
                address_lower = invoice.address.lower()
                if "punjab" in address_lower:
                    customer_state = "Punjab"
                    customer_state_code = "03"
                elif "maharashtra" in address_lower or "mumbai" in address_lower:
                    customer_state = "Maharashtra"
                    customer_state_code = "27"

            # Get product image from estimate (if available)
            product_image_path = None
            estimate_number = ""
            item_name = ""
            estimate_line_items = []

            if invoice.estimate:
                # Product image
                if invoice.estimate.product_image:
                    try:
                        product_image_path = invoice.estimate.product_image.path
                        if not os.path.exists(product_image_path):
                            product_image_path = None
                    except Exception:
                        product_image_path = None

                # Packing list data from estimate
                estimate_number = str(invoice.estimate.id)[:8].upper()  # First 8 chars of UUID
                item_name = invoice.estimate.item_name or ""
                estimate_date = (
                    invoice.estimate.date.strftime("%d/%m/%Y")
                    if invoice.estimate.date
                    else invoice.invoice_date.strftime("%d/%m/%Y")
                )

                # Prepare estimate line items for packing list
                for line in invoice.estimate.line_items.all():
                    # Check unit to decide where to put the weight (Carats or Grams)
                    is_carat = line.unit == "CT"
                    estimate_line_items.append(
                        {
                            "particulars": line.particulars,
                            "shape": line.shape or "",
                            "colour": line.colour or "",
                            "clarity": line.clarity or "",
                            "quantity": line.pc or 0,
                            "carats": float(line.weight) if is_carat and line.weight else "",
                            "weight": float(line.weight) if not is_carat and line.weight else 0,
                            "rate": float(line.rate) if line.rate else 0,
                            "amount": float(line.amount),
                        }
                    )
            else:
                estimate_date = invoice.invoice_date.strftime("%d/%m/%Y")

            # Generate PDF using template with text-only redactions
            pdf_bytes = generate_invoice_template_pdf(
                invoice_number=invoice.invoice_number,
                invoice_date=invoice.invoice_date.strftime("%d/%m/%Y"),
                customer_name=invoice.customer_name,
                customer_address=invoice.address or "",
                customer_gstin=invoice.pan_gstin or "",
                customer_state=customer_state,
                customer_state_code=customer_state_code,
                line_items=line_items,
                subtotal=float(invoice.estimate.total_taxable_value)
                if invoice.estimate and invoice.estimate.total_taxable_value
                else float(invoice.subtotal),
                gst_rate=float(invoice.gst_rate),
                gst_amount=float(invoice.estimate.gst_amount)
                if invoice.estimate and invoice.estimate.gst_amount
                else float(invoice.gst_amount),
                total_amount=float(invoice.estimate.grand_total)
                if invoice.estimate and invoice.estimate.grand_total
                else float(invoice.total_amount),
                payment_terms=invoice.get_payment_terms_display()
                if hasattr(invoice, "get_payment_terms_display")
                else "",
                place_of_supply=customer_state_code,
                company_name=company_name,
                company_address=company_address,
                company_phone=company_phone,
                company_email=company_email,
                company_gstin=company_gstin,
                company_pan=company_pan,
                company_state=company_state,
                company_state_code=company_state_code,
                product_image_path=product_image_path,
                # Packing list parameters (page 2)
                estimate_number=estimate_number,
                item_name=item_name,
                estimate_line_items=estimate_line_items,
                estimate_date=estimate_date,  # Add estimate date
            )

        # Return PDF response
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="Invoice_{invoice.invoice_number}.pdf"'
        )
        return response

    except Exception as e:
        import traceback

        error_trace = traceback.format_exc()
        print(f"PDF Generation Error: {error_trace}")  # Log to console
        return Response(
            {"detail": f"Failed to generate PDF: {str(e)}", "trace": error_trace}, status=500
        )


# ==================== ORDER PROCESS MANAGEMENT VIEWS ====================
# Moved from order_process_views.py

"""
Order Process Management Views - Query â†’ Order Conversion Workflow

BUSINESS FLOW:
Query â†’ Advance Received â†’ Conversion Form â†’ Query Order Confirmation â†’ Order Creation

IMPLEMENTED RULES:
1. Query orders finalized BEFORE order creation
2. Steps editable ONLY during confirmation stage
3. After order creation: LOCKED_FOR_EDIT
4. After courier dispatch: FULLY_LOCKED
5. Auto status: PENDING â†’ IN_PROGRESS â†’ COMPLETED
6. Color coding: PENDING (gray), IN_PROGRESS (blue), COMPLETED (green)
"""

from rest_framework import status
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.utils import timezone
import logging

from .models import (
    Order,
    OrderDraft,
    OrderProcessStep,
    OrderProcessLock,
    DEFAULT_ORDER_PROCESS_STEPS,
    Sale,
    Receipt,
    ThreeDDesign,
    ThreeDPrintingCAM,
    TwoDDesign,
    GhatApproval,
    GhatQualityCheck,
    StoneDemandToBagging,
    PreRhodiumQualityCheck,
    FinalQualityCheck,
    ItemFinalPackingList,
    RawMaterialTally,
    MetalIssue,
    BaggingReady,
    DiamondPurchaseIssue,
    GemstonePurchaseIssue,
)
from users.models import User


logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Step-form required fields registry
#
# Maps each step name to the model that holds its form data and the list of
# field names that MUST be non-empty before the step can be marked as Done.
#
# "non-empty" means:
#   - CharField / TextField: not None and not ""
#   - ImageField: not None (i.e. a file has been uploaded)
#   - BooleanField: True  (checkbox must be checked)
#   - IntegerField / DecimalField: not None
#   - JSONField: not None and not empty list/dict
# ---------------------------------------------------------------------------
STEP_FORM_REQUIRED_FIELDS = {
    "3D Design": {
        "model": ThreeDDesign,
        "fk_field": None,  # linked via reference_id (UUID pk)
        "required": [
            ("account_order_id", "Account / Order ID"),
            ("images", "3D Design Images"),
        ],
    },
    "2D Design Approval": {
        "model": TwoDDesign,
        "fk_field": None,  # linked via reference_id (UUID pk)
        "required": [
            ("account_order_id", "Account / Order ID"),
            ("images", "2D Design Images"),
        ],
    },
    "3D Printing/CAM Piece": {
        "model": ThreeDPrintingCAM,
        "fk_field": None,
        "required": [
            ("account_order_id", "Account / Order ID"),
            ("images", "CAM Piece Image"),
        ],
    },
    "Ghat Trial Approval": {
        "model": GhatApproval,
        "fk_field": None,
        "required": [
            ("account_order_id", "Account / Order ID"),
            ("images", "Carry-Forward Image"),
        ],
    },
    "Ghat QC": {
        "model": GhatQualityCheck,
        "fk_field": None,
        "required": [
            ("account_order_id", "Account / Order ID"),
            ("images", "Carry-Forward Image"),
        ],
    },
    "Stone Demand to Bagging": {
        "model": StoneDemandToBagging,
        "fk_field": None,
        "required": [
            ("account_order_id", "Account / Order ID"),
            ("diamond_color_stone", "Diamond / Color Stone"),
            ("no_of_pieces", "Number of Pieces"),
            ("images", "Images"),
        ],
    },
    "Pre Rhodium QC": {
        "model": PreRhodiumQualityCheck,
        "fk_field": None,
        "required": [
            ("account_order_id", "Account / Order ID"),
            ("images", "Quality Check Image"),
        ],
    },
    "Item with Final Packing List In": {
        "model": ItemFinalPackingList,
        "fk_field": None,
        "required": [
            ("account_order_id", "Account / Order ID"),
            ("images", "Jewellery Piece Image"),
        ],
    },
    "Raw Material Tally": {
        "model": RawMaterialTally,
        "fk_field": None,
        "required": [
            ("account_order_id", "Account / Order ID"),
            ("raw_material_movement", "Raw Material Movement"),
        ],
    },
    "Metal Issue": {
        "model": MetalIssue,
        "fk_field": None,
        "required": [
            ("account_order_id", "Account / Order ID"),
            ("carry_forward_image", "Carry-Forward Image"),
        ],
    },
    "Bagging Ready": {
        "model": BaggingReady,
        "fk_field": None,
        "required": [
            ("account_order_id", "Account / Order ID"),
            ("carry_forward_image", "Carry-Forward Image"),
        ],
    },
    "Diamond Purchase/Issue": {
        "model": DiamondPurchaseIssue,
        "fk_field": None,
        "required": [
            ("account_order_id", "Account / Order ID"),
            ("carry_forward_image", "Carry-Forward Image"),
        ],
    },
    "Gemstone Purchase/Issue": {
        "model": GemstonePurchaseIssue,
        "fk_field": None,
        "required": [
            ("account_order_id", "Account / Order ID"),
            ("carry_forward_image", "Carry-Forward Image"),
        ],
    },
}


# ---------------------------------------------------------------------------
# Department mapper â€” converts query order department strings to the
# Task model's DEPARTMENT_CHOICES keys.
# ---------------------------------------------------------------------------
_DEPT_MAP = {
    "accounts": "ACCOUNTS",
    "admin": "ADMINISTRATION",
    "design": "PRODUCT_DESIGN",
    "sales": "SALES",
    "production": "PRODUCTION",
    "inventory": "RAW_MATERIAL_INVENTORY",
    "purchase": "SOURCING",
    "qc": "PRODUCTION",
    "packing": "PRODUCTION",
    "marketing": "SALES",
    "compliance": "ADMINISTRATION",
    "logistics": "LOGISTICS",
    "vendor": "PRODUCTION",
    "raw material": "RAW_MATERIAL_INVENTORY",
}


def _map_department(dept_str: str) -> str:
    """Map a query order department string to a Task DEPARTMENT_CHOICES key."""
    if not dept_str:
        return "ADMINISTRATION"
    key = dept_str.lower().strip()
    for fragment, choice in _DEPT_MAP.items():
        if fragment in key:
            return choice
    return "ADMINISTRATION"


def _validate_step_form_completion(step):
    """
    Validate that all required fields of a step's linked form are filled.

    Returns a tuple: (is_valid: bool, missing_fields: list[str])

    - If the step has no form registry entry, it is considered valid (no form
      to validate).
    - If the step has a registry entry but no reference_id, it means the form
      was never saved â†’ invalid.
    - Otherwise, load the form record and check each required field.
    """
    config = STEP_FORM_REQUIRED_FIELDS.get(step.step_name)

    # Steps without a form (e.g. "Generate Order ID", "Advance Received") are
    # always valid â€” they just need to have been saved (reference_id or
    # IN_PROGRESS status).
    if config is None:
        return True, []

    # Form exists but was never linked
    if not step.reference_id:
        return False, ["Form has not been saved yet â€” please fill and save the form first"]

    # Load the form record
    model = config["model"]
    try:
        record = model.objects.get(pk=step.reference_id)
    except model.DoesNotExist:
        return False, ["Linked form record not found — please re-save the form"]

    missing = []
    for field_name, label in config["required"]:
        # Special handling for "images" field (used by models with child image records)
        if field_name == "images":
            if hasattr(record, "images"):
                images = record.images.all()
                if not images.exists():
                    missing.append(label)
                elif hasattr(record, "selected_log_group") and not record.selected_log_group:
                    missing.append(f"{label} - please select a final log before marking as done")
            else:
                missing.append(label)
            continue

        value = getattr(record, field_name, None)

        # ImageField: check if a file is stored
        field_obj = record._meta.get_field(field_name)
        from django.db.models import ImageField as DjangoImageField, FileField as DjangoFileField

        if isinstance(field_obj, (DjangoImageField, DjangoFileField)):
            if not value or not value.name:
                missing.append(label)
            continue

        # BooleanField: must be True
        from django.db.models import BooleanField as DjangoBooleanField

        if isinstance(field_obj, DjangoBooleanField):
            if not value:
                missing.append(label)
            continue

        # JSONField: must not be None / empty
        from django.db.models import JSONField as DjangoJSONField

        if isinstance(field_obj, DjangoJSONField):
            if value is None or value == [] or value == {}:
                missing.append(label)
            continue

        # Everything else (CharField, TextField, IntegerField, DecimalFieldâ€¦)
        if value is None or value == "":
            missing.append(label)

    return len(missing) == 0, missing


# ---------------------------------------------------------------------------
# Required-field definitions for each step-form model
# Used by the form collection/detail views to enforce validation on final save
# (is_draft=False). Draft saves bypass this check entirely.
# ---------------------------------------------------------------------------
FORM_MODEL_REQUIRED_FIELDS = {
    "TwoDDesign": [
        ("account_order_id", "Account / Order ID"),
        ("images", "2D Design Image"),
    ],
    "ThreeDDesign": [
        ("account_order_id", "Account / Order ID"),
        ("images", "3D Design Image"),
    ],
    "ThreeDPrintingCAM": [
        ("account_order_id", "Account / Order ID"),
        ("images", "CAM Piece Image"),
    ],
    "GhatApproval": [
        ("account_order_id", "Account / Order ID"),
        ("images", "Carry-Forward Image"),
    ],
    "GhatQualityCheck": [
        ("account_order_id", "Account / Order ID"),
        ("images", "Carry-Forward Image"),
    ],
    "StoneDemandToBagging": [
        ("account_order_id", "Account / Order ID"),
        ("diamond_color_stone", "Diamond / Color Stone"),
        ("no_of_pieces", "Number of Pieces"),
        ("images", "Image"),
    ],
    "PreRhodiumQualityCheck": [
        ("account_order_id", "Account / Order ID"),
        ("images", "Quality Check Image"),
    ],
    "FinalQualityCheck": [
        ("account_order_id", "Account / Order ID"),
        ("images", "Final Quality Check Image"),
    ],
    "ItemFinalPackingList": [
        ("account_order_id", "Account / Order ID"),
        ("images", "Jewellery Piece Image"),
    ],
    "RawMaterialTally": [
        ("account_order_id", "Account / Order ID"),
        ("raw_material_movement", "Raw Material Movement"),
        ("images", "Carry-Forward Image"),
    ],
    "MetalIssue": [
        ("account_order_id", "Account / Order ID"),
        ("images", "Carry-Forward Image"),
    ],
    "BaggingReady": [
        ("account_order_id", "Account / Order ID"),
        ("images", "Carry-Forward Image"),
    ],
    "DiamondPurchaseIssue": [
        ("account_order_id", "Account / Order ID"),
        ("images", "Carry-Forward Image"),
    ],
    "GemstonePurchaseIssue": [
        ("account_order_id", "Account / Order ID"),
        ("images", "Carry-Forward Image"),
    ],
}


def _check_form_record_fields(record, model_class_name, uploaded_files=None):
    """
    Validate required fields on a saved form record.

    Called by collection/detail views when is_draft=False.
    Returns (is_valid: bool, missing_fields: list[str]).

    `uploaded_files` is request.FILES â€” used to check image fields that were
    just uploaded in this request (not yet persisted to the record).
    """
    from django.db.models import ImageField as DjangoImageField, FileField as DjangoFileField
    from django.db.models import BooleanField as DjangoBooleanField
    from django.db.models import JSONField as DjangoJSONField

    required = FORM_MODEL_REQUIRED_FIELDS.get(model_class_name, [])
    if not required:
        return True, []

    uploaded_files = uploaded_files or {}
    missing = []

    for field_name, label in required:
        # If the file was just uploaded in this request, it counts as filled
        if field_name in uploaded_files:
            continue

        value = getattr(record, field_name, None)

        # Handle related managers (like "images" from multi-image support)
        try:
            field_obj = record._meta.get_field(field_name)
        except Exception:
            # If it's not a model field, check if it's a related manager
            if field_name == "images" and hasattr(record, "images"):
                # Require at least one image exists
                has_any = record.images.exists()
                if not has_any:
                    missing.append(label)
                    continue
                # Require a log group is selected as final
                if hasattr(record, "selected_log_group") and not record.selected_log_group:
                    missing.append(f"{label} — please select a final log")
                continue
            # If we can't find the field, mark as missing
            missing.append(label)
            continue

        if isinstance(field_obj, (DjangoImageField, DjangoFileField)):
            if not value or not value.name:
                missing.append(label)
            continue

        if isinstance(field_obj, DjangoBooleanField):
            # BooleanField is intentionally NOT required here â€” a False value
            # is a valid user choice (e.g. "QC failed"). Only None would be
            # missing, but BooleanField can't be None unless null=True.
            continue

        if isinstance(field_obj, DjangoJSONField):
            if value is None or value == [] or value == {}:
                missing.append(label)
            continue

        if value is None or value == "":
            missing.append(label)

    return len(missing) == 0, missing


# Helper function to get authentication classes
def get_auth_classes():
    """Get authentication classes, with JWT if available"""
    from nsj_backend.authentication import CsrfExemptSessionAuthentication

    auth_classes = [CsrfExemptSessionAuthentication]
    try:
        from rest_framework_simplejwt.authentication import JWTAuthentication

        auth_classes.insert(0, JWTAuthentication)
    except ImportError:
        pass
    return auth_classes


def _get_status_color(status_value):
    """Get color coding for status"""
    return {"PENDING": "gray", "IN_PROGRESS": "blue", "COMPLETED": "green"}.get(
        status_value, "gray"
    )


def _recalculate_step_statuses(order):
    """
    AUTO STATUS LOGIC:
    - First incomplete step â†’ IN_PROGRESS
    - Completed steps â†’ COMPLETED
    - Remaining steps â†’ PENDING
    """
    steps = order.process_steps.all().order_by("position")

    found_in_progress = False
    for step in steps:
        if step.status == "COMPLETED":
            continue  # Keep completed
        elif not found_in_progress:
            # First incomplete step becomes IN_PROGRESS
            if step.status != "IN_PROGRESS":
                step.status = "IN_PROGRESS"
                step.save()
            found_in_progress = True
        else:
            # Rest remain PENDING
            if step.status != "PENDING":
                step.status = "PENDING"
                step.save()


# ==================== STAGE 1: Initiate Conversion (Query â†’ Draft) ====================


@api_view(["POST"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def initiate_order_conversion(request, sale_id):
    """
    STAGE 1: Initiate order conversion from Query

    FLOW: Query â†’ Advance Received â†’ Conversion Form â†’ [THIS ENDPOINT]

    Creates OrderDraft with:
    - Receipt voucher (advance payment)
    - Order data from conversion form
    - Default 29 query orders

    POST /api/sales/{sale_id}/initiate-order-conversion/
    Body: {
        "receipt_voucher_id": "uuid",
        "advance_amount": 25000.00,
        "advance_notes": "Advance payment received",
        "order_data": {
            "item_name": "Diamond Ring",
            "design": "Custom design"
        }
    }
    """
    try:
        user = request.user
        company = getattr(user, "company", None)

        if not company:
            return Response(
                {"error": "User does not belong to a company"}, status=status.HTTP_400_BAD_REQUEST
            )

        # VALIDATION: Get the sale (Query)
        try:
            sale = Sale.objects.get(id=sale_id, company=company)
        except Sale.DoesNotExist:
            return Response(
                {"error": "Sale not found or does not belong to your company"},
                status=status.HTTP_404_NOT_FOUND,
            )

        # VALIDATION: Check if sale is still in QUERY stage
        # (Assuming Sale model has a status field or similar)
        # Add your specific validation here based on your Sale model

        # IDEMPOTENCY CHECK: Prevent duplicate draft creation
        existing_draft = OrderDraft.objects.filter(
            source_sale=sale, status="pending_confirmation"
        ).first()

        if existing_draft:
            logger.info(f"Draft already exists for sale {sale_id}")
            steps = existing_draft.process_steps.all().order_by("position")
            steps_data = [
                {
                    "id": str(step.id),
                    "step_name": step.step_name,
                    "description": step.description,
                    "department": step.department,
                    "position": step.position,
                    "status": step.status,
                    "color": _get_status_color(step.status),
                }
                for step in steps
            ]

            return Response(
                {
                    "success": True,
                    "message": "Order draft already exists. Continue editing.",
                    "draft_id": str(existing_draft.id),
                    "status": existing_draft.status,
                    "receipt_voucher_id": str(existing_draft.receipt_voucher.id)
                    if existing_draft.receipt_voucher
                    else None,
                    "advance_amount": str(existing_draft.advance_amount)
                    if existing_draft.advance_amount
                    else None,
                    "process_steps": steps_data,
                    "total_steps": len(steps_data),
                    "is_existing": True,
                },
                status=status.HTTP_200_OK,
            )

        # Get data from request
        receipt_voucher_id = request.data.get("receipt_voucher_id")
        advance_amount = request.data.get("advance_amount")
        advance_notes = request.data.get("advance_notes", "")
        order_data = request.data.get("order_data", {})

        # VALIDATION: Receipt voucher required
        receipt_voucher = None
        if receipt_voucher_id:
            try:
                receipt_voucher = Receipt.objects.get(id=receipt_voucher_id, company=company)
            except Receipt.DoesNotExist:
                return Response(
                    {"error": "Receipt voucher not found"}, status=status.HTTP_400_BAD_REQUEST
                )

        # VALIDATION: If receipt voucher is provided, advance amount should be provided
        # But both are optional - order can be created without advance payment
        if receipt_voucher_id and not advance_amount:
            return Response(
                {"error": "Advance amount is required when receipt voucher is provided"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        with transaction.atomic():
            # Create OrderDraft
            order_draft = OrderDraft.objects.create(
                source_sale=sale,
                company=company,
                created_by=user,
                status="pending_confirmation",
                receipt_voucher=receipt_voucher,
                advance_amount=advance_amount,
                advance_notes=advance_notes,
                order_data=order_data,
            )

            # Generate default 29 query orders (bulk insert for performance)
            steps_to_create = []
            for step_data in DEFAULT_ORDER_PROCESS_STEPS:
                step_status = "PENDING"
                if step_data["position"] == 1 and advance_amount:
                    step_status = "COMPLETED"

                steps_to_create.append(
                    OrderProcessStep(
                        order_draft=order_draft,
                        step_name=step_data["name"],
                        description=step_data["description"],
                        department=step_data.get("department", ""),
                        position=step_data["position"],
                        status=step_status,
                        completed_at=timezone.now() if step_status == "COMPLETED" else None,
                    )
                )
            OrderProcessStep.objects.bulk_create(steps_to_create)

            # Get all created steps
            steps = order_draft.process_steps.all().order_by("position")
            steps_data = [
                {
                    "id": str(step.id),
                    "step_name": step.step_name,
                    "description": step.description,
                    "department": step.department,
                    "position": step.position,
                    "status": step.status,
                    "color": _get_status_color(step.status),
                }
                for step in steps
            ]

        logger.info(f"Order draft created for sale {sale_id} by user {user.email}")

        return Response(
            {
                "success": True,
                "message": "Order draft created. Please review and confirm query orders.",
                "draft_id": str(order_draft.id),
                "status": "pending_confirmation",
                "receipt_voucher_id": str(receipt_voucher.id) if receipt_voucher else None,
                "advance_amount": str(order_draft.advance_amount),
                "process_steps": steps_data,
                "total_steps": len(steps_data),
            },
            status=status.HTTP_201_CREATED,
        )

    except Exception as e:
        logger.error(f"Error initiating order conversion: {str(e)}", exc_info=True)
        return Response(
            {"error": f"An error occurred: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# ==================== STAGE 2: Query Order Management (Editable) ====================


@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def invoice_stats_view(request):
    """Get invoice statistics for the current user's company."""
    user = request.user
    company = getattr(user, "company", None)

    if not company:
        return Response({"detail": "User does not belong to a company"}, status=400)

    from django.db.models import Sum, Count, Q

    queryset = Invoice.objects.filter(company=company)

    # Overall stats
    total_count = queryset.count()
    total_amount = queryset.aggregate(total=Sum("total_amount"))["total"] or 0
    total_paid = queryset.aggregate(total=Sum("amount_paid"))["total"] or 0
    total_due = queryset.aggregate(total=Sum("balance_due"))["total"] or 0

    # Status breakdown
    status_counts = {
        "pending": queryset.filter(status="pending").count(),
        "partially_paid": queryset.filter(status="partially_paid").count(),
        "paid": queryset.filter(status="paid").count(),
        "overdue": queryset.filter(status="overdue").count(),
        "cancelled": queryset.filter(status="cancelled").count(),
    }

    # Recent invoices
    recent = queryset.select_related("sale", "estimate").order_by("-invoice_date", "-created_at")[
        :5
    ]
    recent_data = [InvoiceSerializer(inv, context={"request": request}).data for inv in recent]

    return Response(
        {
            "total_count": total_count,
            "total_amount": float(total_amount),
            "total_paid": float(total_paid),
            "total_due": float(total_due),
            "status_counts": status_counts,
            "recent_invoices": recent_data,
        }
    )


@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def get_order_draft(request, draft_id):
    """
    STAGE 2: Get order draft with query orders

    RULE: Steps are editable during draft stage

    GET /api/order-drafts/{draft_id}/
    """
    try:
        user = request.user
        company = getattr(user, "company", None)

        order_draft = get_object_or_404(OrderDraft, id=draft_id, company=company)

        steps = order_draft.process_steps.all().order_by("position")
        steps_data = [
            {
                "id": str(step.id),
                "step_name": step.step_name,
                "description": step.description,
                "department": step.department,
                "position": step.position,
                "status": step.status,
                "notes": step.notes,
                "completed_at": step.completed_at.isoformat() if step.completed_at else None,
                "color": _get_status_color(step.status),
            }
            for step in steps
        ]

        return Response(
            {
                "draft_id": str(order_draft.id),
                "status": order_draft.status,
                "receipt_voucher_id": str(order_draft.receipt_voucher.id)
                if order_draft.receipt_voucher
                else None,
                "advance_amount": str(order_draft.advance_amount)
                if order_draft.advance_amount
                else None,
                "advance_notes": order_draft.advance_notes,
                "order_data": order_draft.order_data,
                "process_steps": steps_data,
                "total_steps": len(steps_data),
                "created_at": order_draft.created_at.isoformat(),
                "can_modify": order_draft.status == "pending_confirmation",
                "sales_query_id": str(order_draft.sales_query_id)
                if order_draft.sales_query_id
                else None,
            }
        )

    except Exception as e:
        logger.error(f"Error getting order draft: {str(e)}", exc_info=True)
        return Response(
            {"error": f"An error occurred: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(["PUT"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def update_process_steps(request, draft_id):
    """
    STAGE 2: Update query orders (reorder, add, remove)

    RULE: Only allowed when status is 'pending_confirmation'
    ALLOWED: Drag & reorder, Add new steps, Remove steps, Edit steps

    PUT /api/order-drafts/{draft_id}/process-steps/
    Body: {
        "steps": [
            {"step_name": "...", "description": "...", "department": "...", "position": 1},
            ...
        ]
    }
    """
    try:
        user = request.user
        company = getattr(user, "company", None)

        order_draft = get_object_or_404(OrderDraft, id=draft_id, company=company)

        # VALIDATION: Check if can modify
        if order_draft.status != "pending_confirmation":
            return Response(
                {
                    "error": f"Cannot modify steps. Draft status is: {order_draft.status}",
                    "message": "Query orders can only be modified during pending confirmation stage",
                },
                status=status.HTTP_403_FORBIDDEN,
            )

        steps_data = request.data.get("steps", [])

        if not steps_data:
            return Response({"error": "Steps data is required"}, status=status.HTTP_400_BAD_REQUEST)

        # VALIDATION: At least one step required
        if len(steps_data) < 1:
            return Response(
                {"error": "At least one query order is required"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Capture existing step reference_ids by step_name before deleting
        existing_references = {
            step.step_name: step.reference_id
            for step in order_draft.process_steps.all()
            if step.reference_id
        }

        with transaction.atomic():
            # Delete existing steps
            order_draft.process_steps.all().delete()

            # Create new steps (preserve status from UI + reference_id)
            for step_data in steps_data:
                step_name = step_data.get("step_name", "")
                OrderProcessStep.objects.create(
                    order_draft=order_draft,
                    step_name=step_name,
                    description=step_data.get("description", ""),
                    department=step_data.get("department", ""),
                    position=step_data.get("position", 0),
                    status=step_data.get("status", "PENDING"),
                    notes=step_data.get("notes", ""),
                    completed_at=step_data.get("completed_at"),
                    reference_id=step_data.get("reference_id")
                    or existing_references.get(step_name),
                )

            # Get updated steps
            steps = order_draft.process_steps.all().order_by("position")
            steps_response = [
                {
                    "id": str(step.id),
                    "step_name": step.step_name,
                    "description": step.description,
                    "department": step.department,
                    "position": step.position,
                    "status": step.status,
                    "notes": step.notes,
                    "completed_at": step.completed_at.isoformat() if step.completed_at else None,
                    "color": _get_status_color(step.status),
                }
                for step in steps
            ]

        logger.info(f"Query orders updated for draft {draft_id} by user {user.email}")

        return Response(
            {
                "success": True,
                "message": "Query orders updated successfully",
                "process_steps": steps_response,
                "total_steps": len(steps_response),
            }
        )

    except Exception as e:
        logger.error(f"Error updating query orders: {str(e)}", exc_info=True)
        return Response(
            {"error": f"An error occurred: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# ==================== STAGE 3: Confirmation & Order Creation ====================


@api_view(["POST"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def confirm_and_create_order(request, draft_id):
    """
    STAGE 3: Confirm query orders and create final Order

    CRITICAL FLOW:
    1. Validate: steps exist, order not yet created, still in QUERY stage
    2. Show confirmation: "Are you sure? Steps will be frozen after order creation"
    3. Create order
    4. Transfer query orders to order
    5. Set first step to IN_PROGRESS (AUTO STATUS)
    6. Create LOCKED_FOR_EDIT lock

    POST /api/order-drafts/{draft_id}/confirm/
    """
    try:
        user = request.user
        company = getattr(user, "company", None)

        order_draft = get_object_or_404(OrderDraft, id=draft_id, company=company)

        # VALIDATION: Check status
        if order_draft.status != "pending_confirmation":
            return Response(
                {
                    "error": f"Cannot confirm. Draft status is: {order_draft.status}",
                    "message": "Only pending drafts can be confirmed",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # VALIDATION: Check if steps exist
        draft_steps = order_draft.process_steps.all().order_by("position")
        if not draft_steps.exists():
            return Response(
                {"error": "No query orders found. Cannot create order."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # IDEMPOTENCY CHECK: Prevent duplicate order creation
        if order_draft.final_order:
            logger.warning(f"Draft {draft_id} already has an order: {order_draft.final_order.id}")
            return Response(
                {
                    "error": "Order already created for this draft",
                    "order_id": str(order_draft.final_order.id),
                    "order_bill_no": order_draft.final_order.bill_no,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        with transaction.atomic():
            # Create final Order
            order_data = order_draft.order_data

            # Get account from source_sale or from order_data
            account = None
            if order_draft.source_sale:
                account = order_draft.source_sale.account
            else:
                # For direct orders, get account from order_data
                account_id = order_data.get("account")
                if account_id:
                    from accounts.models import Account

                    try:
                        account = Account.objects.get(id=account_id, company=company)
                    except Account.DoesNotExist:
                        pass

            order = Order.objects.create(
                company=company,
                account=account,
                item_name=order_data.get("item_name", ""),
                design=order_data.get("design", ""),
                stamp=order_data.get("stamp", "") or "",
                size=order_data.get("size", "") or "",
                gold_rate=order_data.get("gold_rate", 0) or 0,
                order_type=order_data.get("order_type", "STOCK_JEWELRY") or "STOCK_JEWELRY",
                number_of_pieces=order_data.get("number_of_pieces", 1) or 1,
                date=timezone.now().date(),
                created_by=user,
                advance_payment_received="YES" if order_draft.advance_amount else "NO",
            )

            # Transfer query orders from draft to order (bulk insert for performance)
            created_steps = OrderProcessStep.objects.bulk_create(
                [
                    OrderProcessStep(
                        order=order,
                        step_name=draft_step.step_name,
                        description=draft_step.description,
                        department=draft_step.department,
                        position=draft_step.position,
                        status=draft_step.status,
                        completed_at=draft_step.completed_at,
                        reference_id=draft_step.reference_id,
                        notes=draft_step.notes,
                        saved_at=draft_step.saved_at,
                        saved_by=draft_step.saved_by,
                        marked_done_at=draft_step.marked_done_at,
                        marked_done_by=draft_step.marked_done_by,
                    )
                    for draft_step in draft_steps
                ]
            )

            steps_to_update = []
            first_incomplete_step = None

            for step in created_steps:
                if step.status == "PENDING" and first_incomplete_step is None:
                    first_incomplete_step = step
                    step.status = "IN_PROGRESS"
                    steps_to_update.append(step)
                elif step.step_name == "Generate Order ID":
                    step.status = "COMPLETED"
                    step.completed_at = timezone.now()
                    step.notes = f"Order ID generated: {order.bill_no}"
                    steps_to_update.append(step)

            if steps_to_update:
                OrderProcessStep.objects.bulk_update(
                    steps_to_update, fields=["status", "completed_at", "notes"], batch_size=100
                )

            # CREATE LOCK: LOCKED_FOR_EDIT (no add/remove/reorder)
            OrderProcessLock.objects.create(
                order=order, locked_by=user, lock_level="LOCKED_FOR_EDIT"
            )

            # Update draft status
            order_draft.status = "confirmed"
            order_draft.confirmed_at = timezone.now()
            order_draft.final_order = order
            order_draft.save()

            # Update SalesLead workflow_status and final_order_id
            if order_draft.sales_query_id:
                try:
                    from sales_queries.models import SalesQuery

                    sales_lead = SalesQuery.objects.filter(id=order_draft.sales_query_id).first()

                    if sales_lead:
                        sales_lead.workflow_status = "converted_to_order"
                        sales_lead.final_order_id = order.id  # Store the final order ID
                        sales_lead.save()
                        logger.info(
                            f"Updated SalesLead {sales_lead.id} status to converted_to_order and linked to Order {order.id}"
                        )
                    else:
                        logger.warning(f"SalesLead {order_draft.sales_query_id} not found")
                except Exception as e:
                    logger.warning(f"Failed to update SalesLead: {str(e)}")

        logger.info(f"Order {order.id} created from draft {draft_id} by user {user.email}")

        # Check if Estimate Approval step has a linked estimate
        estimate_approval_linked = any(
            s.step_name == "Estimate Approval" and s.reference_id for s in created_steps
        )

        return Response(
            {
                "success": True,
                "message": "Order created successfully. Query orders are now locked.",
                "order_id": str(order.id),
                "order_bill_no": order.bill_no,
                "order_job_no": order.job_no,
                "process_locked": True,
                "lock_level": "LOCKED_FOR_EDIT",
                "total_steps": len(created_steps),
                "first_step_status": "IN_PROGRESS",
                "has_linked_estimate": estimate_approval_linked,
            },
            status=status.HTTP_201_CREATED,
        )

    except Exception as e:
        error_trace = traceback.format_exc()
        logger.error(f"Error confirming order: {str(e)}\n{error_trace}")
        return Response(
            {"error": f"An error occurred: {str(e)}", "details": error_trace},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


# ==================== STAGE 4: Process Tracking (Locked, Status Updates Only) ====================


@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def get_order_process_steps(request, order_id):
    """
    STAGE 4: Get query orders for an order

    RULE: Steps are locked, only status updates allowed

    GET /api/orders/{order_id}/process-steps/
    """
    try:
        user = request.user
        company = getattr(user, "company", None)

        # Check if order exists
        try:
            order = Order.objects.get(id=order_id, company=company)
        except Order.DoesNotExist:
            return Response(
                {"error": "Order not found or does not belong to your company"},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Check lock level
        lock_level = "UNLOCKED"
        courier_dispatched = False
        try:
            process_lock = OrderProcessLock.objects.get(order=order)
            lock_level = process_lock.lock_level
            courier_dispatched = process_lock.courier_dispatched
        except OrderProcessLock.DoesNotExist:
            pass

        # Get steps
        steps = order.process_steps.all().order_by("position")

        # HANDLE LEGACY ORDERS
        if not steps.exists():
            return Response(
                {
                    "order_id": str(order.id),
                    "lock_level": "UNLOCKED",
                    "courier_dispatched": False,
                    "can_modify_steps": False,
                    "can_update_status": False,
                    "process_steps": [],
                    "total_steps": 0,
                    "completed_steps": 0,
                    "in_progress_steps": 0,
                    "pending_steps": 0,
                    "progress_percentage": 0,
                    "is_legacy_order": True,
                    "message": "This order was created before the process management feature.",
                }
            )

        steps_data = [
            {
                "id": str(step.id),
                "step_name": step.step_name,
                "description": step.description,
                "department": step.department,
                "position": step.position,
                "status": step.status,
                "color": _get_status_color(step.status),
                "completed_at": step.completed_at.isoformat() if step.completed_at else None,
                "notes": step.notes,
                "reference_id": str(step.reference_id) if step.reference_id else None,
                # Lock/save tracking fields for UI
                "is_locked": step.status == "COMPLETED",
                "has_been_saved": step.saved_at is not None or step.reference_id is not None,
                "saved_at": step.saved_at.isoformat() if step.saved_at else None,
                "marked_done_at": step.marked_done_at.isoformat() if step.marked_done_at else None,
            }
            for step in steps
        ]

        # Calculate progress
        total_steps = steps.count()
        completed_steps = steps.filter(status="COMPLETED").count()
        progress_percentage = (completed_steps / total_steps * 100) if total_steps > 0 else 0

        return Response(
            {
                "order_id": str(order.id),
                "lock_level": lock_level,
                "courier_dispatched": courier_dispatched,
                "can_modify_steps": False,  # Always false after order creation
                "can_update_status": lock_level != "FULLY_LOCKED",  # Can update unless fully locked
                "process_steps": steps_data,
                "total_steps": total_steps,
                "completed_steps": completed_steps,
                "in_progress_steps": steps.filter(status="IN_PROGRESS").count(),
                "pending_steps": steps.filter(status="PENDING").count(),
                "progress_percentage": round(progress_percentage, 2),
                "is_legacy_order": False,
            }
        )

    except Exception as e:
        logger.error(f"Error getting order orders: {str(e)}", exc_info=True)
        return Response(
            {"error": f"An error occurred: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(["PATCH"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def update_step_status(request, order_id, step_id):
    """
    STAGE 4: Update query order status

    RULES:
    - Only status updates allowed (no structural changes)
    - Cannot update if FULLY_LOCKED (after courier dispatch)
    - AUTO RECALCULATION: After update, recalculate all step statuses

    PATCH /api/orders/{order_id}/process-steps/{step_id}/status/
    Body: {
        "status": "COMPLETED",
        "notes": "Step completed successfully"
    }
    """
    try:
        user = request.user
        company = getattr(user, "company", None)

        # Check if order exists
        try:
            order = Order.objects.get(id=order_id, company=company)
        except Order.DoesNotExist:
            return Response(
                {"error": "Order not found or does not belong to your company"},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Check if step exists
        try:
            step = OrderProcessStep.objects.get(id=step_id, order=order)
        except OrderProcessStep.DoesNotExist:
            return Response({"error": "Query order not found"}, status=status.HTTP_404_NOT_FOUND)

        # Check lock level
        try:
            process_lock = OrderProcessLock.objects.get(order=order)
            if process_lock.lock_level == "FULLY_LOCKED":
                return Response(
                    {
                        "error": "Process is fully locked after courier dispatch",
                        "message": "No changes allowed after courier dispatch",
                    },
                    status=status.HTTP_403_FORBIDDEN,
                )
        except OrderProcessLock.DoesNotExist:
            pass

        new_status = request.data.get("status")
        notes = request.data.get("notes", "")

        if not new_status:
            return Response({"error": "Status is required"}, status=status.HTTP_400_BAD_REQUEST)

        # VALIDATION: Valid status values
        if new_status not in ["PENDING", "IN_PROGRESS", "COMPLETED"]:
            return Response(
                {"error": "Invalid status. Must be PENDING, IN_PROGRESS, or COMPLETED"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        with transaction.atomic():
            old_status = step.status
            step.status = new_status
            step.notes = notes

            # AUTO TIMESTAMP: Set completed_at when status becomes COMPLETED
            if new_status == "COMPLETED" and old_status != "COMPLETED":
                # Validate required fields before allowing completion
                is_valid, missing_fields = _validate_step_form_completion(step)
                if not is_valid:
                    return Response(
                        {
                            "success": False,
                            "error": "Cannot mark step as done â€” required fields are missing",
                            "missing_fields": missing_fields,
                            "step_name": step.step_name,
                            "message": (
                                f"The following required fields must be filled before marking "
                                f"'{step.step_name}' as done: {', '.join(missing_fields)}"
                            ),
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                step.completed_at = timezone.now()
                step.marked_done_at = timezone.now()
                step.marked_done_by = user
            elif new_status != "COMPLETED":
                step.completed_at = None
                step.marked_done_at = None
                step.marked_done_by = None

            step.save()

            # AUTO STATUS RECALCULATION
            _recalculate_step_statuses(order)

        logger.info(f"Step {step_id} status updated from {old_status} to {new_status}")

        # Get updated step
        step.refresh_from_db()

        return Response(
            {
                "success": True,
                "message": f"Step status updated to {step.status}",
                "step": {
                    "id": str(step.id),
                    "step_name": step.step_name,
                    "status": step.status,
                    "color": _get_status_color(step.status),
                    "completed_at": step.completed_at.isoformat() if step.completed_at else None,
                    "notes": step.notes,
                },
            }
        )

    except Exception as e:
        logger.error(f"Error updating step status: {str(e)}", exc_info=True)
        return Response(
            {"error": f"An error occurred: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# ==================== COURIER DISPATCH LOCK ====================


@api_view(["POST"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def mark_courier_dispatched(request, order_id):
    """
    Mark order as courier dispatched

    RULE: After courier dispatch â†’ FULLY_LOCKED (no changes at all)

    POST /api/orders/{order_id}/courier-dispatched/
    """
    try:
        user = request.user
        company = getattr(user, "company", None)

        # Check if order exists
        try:
            order = Order.objects.get(id=order_id, company=company)
        except Order.DoesNotExist:
            return Response(
                {"error": "Order not found or does not belong to your company"},
                status=status.HTTP_404_NOT_FOUND,
            )

        try:
            process_lock = OrderProcessLock.objects.get(order=order)

            if process_lock.lock_level == "FULLY_LOCKED":
                return Response(
                    {
                        "message": "Order already marked as courier dispatched",
                        "lock_level": "FULLY_LOCKED",
                    }
                )

            # Upgrade to FULLY_LOCKED
            process_lock.upgrade_to_fully_locked()

            logger.info(f"Order {order_id} marked as courier dispatched by {user.email}")

            return Response(
                {
                    "success": True,
                    "message": "Order marked as courier dispatched. Process is now fully locked.",
                    "lock_level": "FULLY_LOCKED",
                    "courier_dispatched": True,
                    "courier_dispatched_at": process_lock.courier_dispatched_at.isoformat(),
                }
            )

        except OrderProcessLock.DoesNotExist:
            return Response(
                {"error": "Process lock not found for this order"}, status=status.HTTP_404_NOT_FOUND
            )

    except Exception as e:
        logger.error(f"Error marking courier dispatched: {str(e)}", exc_info=True)
        return Response(
            {"error": f"An error occurred: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# ==================== VALIDATION: Prevent Locked Modification ====================


@api_view(["POST", "PUT", "DELETE"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def prevent_locked_step_modification(request, order_id):
    """
    VALIDATION ENDPOINT: Prevent any structural modification to locked query orders

    RULE: After order confirmation, structural changes are forbidden

    POST/PUT/DELETE /api/orders/{order_id}/process-steps/modify/
    """
    return Response(
        {
            "error": "Query orders are locked and cannot be modified after order creation.",
            "message": "Steps were locked when the order was confirmed.",
            "allowed_action": "You can only update step status for tracking purposes.",
            "rule": "AFTER ORDER CONFIRMATION: Query orders become LOCKED_FOR_EDIT",
        },
        status=status.HTTP_403_FORBIDDEN,
    )


# ==================== Auto-Complete Query Orders ====================


@api_view(["POST", "PATCH"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def save_process_step_data(request, order_id, step_name):
    """
    Save query order data without completing the step

    POST /api/orders/{order_id}/process-steps/{step_name}/save/

    Body: {
        "reference_id": "uuid",  // ID of the related record (3D Design, Metal Issue, etc.)
        "notes": "Work in progress notes",
        "data": {...}  // Any additional step-specific data
    }

    Response: {
        "success": true,
        "step_id": "uuid",
        "status": "IN_PROGRESS",  // Status remains unchanged or set to IN_PROGRESS
        "saved_at": "2026-04-16T10:30:00Z",
        "has_been_saved": true
    }

    This endpoint allows users to save their work without marking the step as completed.
    The step status will be set to IN_PROGRESS if currently PENDING.
    """
    try:
        user = request.user
        company = getattr(user, "company", None)

        # Verify order belongs to company
        try:
            order = Order.objects.get(id=order_id, company=company)
        except Order.DoesNotExist:
            return Response(
                {"error": "Order not found or does not belong to your company"},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Find the step by name
        step = OrderProcessStep.objects.filter(order=order, step_name=step_name).first()

        if not step:
            return Response(
                {"error": f"Step '{step_name}' not found for this order"},
                status=status.HTTP_404_NOT_FOUND,
            )

        reference_id = request.data.get("reference_id")
        notes = request.data.get("notes", "")

        # If step is COMPLETED (locked), only allow updating the reference_id silently.
        # This handles the case where the form was saved after the step was already marked done.
        if step.status == "COMPLETED":
            if reference_id and not step.reference_id:
                # Link the record even though step is locked â€” no status change
                step.reference_id = reference_id
                step.save(update_fields=["reference_id", "updated_at"])
                logger.info(
                    f"Linked reference_id to already-completed step '{step_name}' for order {order_id}"
                )
            return Response(
                {
                    "success": True,
                    "step_id": str(step.id),
                    "status": step.status,
                    "is_locked": True,
                    "saved_at": step.saved_at.isoformat() if step.saved_at else None,
                    "has_been_saved": True,
                    "message": "Step is already completed and locked. Reference linked if provided.",
                },
                status=status.HTTP_200_OK,
            )

        # Update step data without completing
        if reference_id:
            step.reference_id = reference_id

        if notes:
            # Append notes if they exist, otherwise set them
            if step.notes:
                step.notes = f"{step.notes}\n{notes}"
            else:
                step.notes = notes

        # Set status to IN_PROGRESS if currently PENDING
        if step.status == "PENDING":
            step.status = "IN_PROGRESS"

        # Track who saved and when
        step.saved_at = timezone.now()
        step.saved_by = user

        step.save()

        logger.info(f"Saved query order '{step_name}' for order {order_id} without completing")

        return Response(
            {
                "success": True,
                "step_id": str(step.id),
                "status": step.status,
                "is_locked": False,
                "saved_at": step.updated_at.isoformat(),
                "has_been_saved": True,
            },
            status=status.HTTP_200_OK,
        )

    except Exception as e:
        logger.error(f"Error saving query order: {str(e)}", exc_info=True)
        return Response(
            {"error": f"An error occurred: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def get_step_status_by_name(request, order_id, step_name):
    """
    Get the current status of a query order by name.

    Used by individual step forms (3D Design, Metal Issue, etc.) to:
    - Check if the step is locked (COMPLETED â†’ edit disabled)
    - Get the reference_id to load existing saved data
    - Know whether to show edit or read-only mode

    GET /api/orders/{order_id}/process-steps/{step_name}/step-status/

    Response: {
        "step_id": "uuid",
        "step_name": "3D Design",
        "status": "IN_PROGRESS",
        "is_locked": false,
        "reference_id": "uuid or null",
        "has_been_saved": true,
        "saved_at": "2026-04-24T10:00:00Z",
        "marked_done_at": null,
        "notes": "..."
    }
    """
    try:
        user = request.user
        company = getattr(user, "company", None)

        try:
            order = Order.objects.get(id=order_id, company=company)
        except Order.DoesNotExist:
            return Response(
                {"error": "Order not found or does not belong to your company"},
                status=status.HTTP_404_NOT_FOUND,
            )

        step = OrderProcessStep.objects.filter(order=order, step_name=step_name).first()

        if not step:
            return Response(
                {"error": f"Step '{step_name}' not found for this order"},
                status=status.HTTP_404_NOT_FOUND,
            )

        if step.reference_id:
            config = STEP_FORM_REQUIRED_FIELDS.get(step.step_name)
            if config:
                model = config["model"]
                if not model.objects.filter(pk=step.reference_id).exists():
                    logger.warning(
                        f"Stale reference_id {step.reference_id} on step "
                        f"'{step.step_name}' for order {order_id} â€” clearing"
                    )
                    step.reference_id = None
                    if step.status == "IN_PROGRESS":
                        step.status = "PENDING"
                    step.saved_at = None
                    step.save(update_fields=["reference_id", "status", "saved_at", "updated_at"])

        return Response(
            {
                "step_id": str(step.id),
                "step_name": step.step_name,
                "status": step.status,
                "is_locked": step.status == "COMPLETED",
                "reference_id": str(step.reference_id) if step.reference_id else None,
                "has_been_saved": step.saved_at is not None or step.reference_id is not None,
                "saved_at": step.saved_at.isoformat() if step.saved_at else None,
                "marked_done_at": step.marked_done_at.isoformat() if step.marked_done_at else None,
                "notes": step.notes,
            }
        )

    except Exception as e:
        logger.error(f"Error getting step status: {str(e)}", exc_info=True)
        return Response(
            {"error": f"An error occurred: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["POST"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def mark_step_done_from_process_tab(request, order_id, step_id):
    """
    Mark a query order as done (COMPLETED) from the Query Orders tab.

    Validates required fields, then requires a deadline + assignee for the
    NEXT step before completing. On success:
    - Marks this step COMPLETED and locks it
    - Creates a Task for the next step assigned to the chosen user
    - Stores deadline + assignee on this step for audit

    POST /api/orders/{order_id}/process-steps/{step_id}/mark-done/

    Body: {
        "notes": "Optional notes",
        "next_step_deadline": "2026-05-10",   // REQUIRED (date YYYY-MM-DD)
        "next_step_assignee_id": "uuid"        // REQUIRED (user UUID)
    }

    Response: {
        "success": true,
        "step_id": "uuid",
        "step_name": "3D Design",
        "status": "COMPLETED",
        "is_locked": true,
        "marked_done_at": "2026-04-24T10:00:00Z",
        "next_step": {
            "step_name": "3D Design Approval",
            "task_id": "uuid",
            "assignee": "user name",
            "deadline": "2026-05-10"
        }
    }
    """
    try:
        from tasks.models import Task

        user = request.user
        company = getattr(user, "company", None)

        try:
            order = Order.objects.get(id=order_id, company=company)
        except Order.DoesNotExist:
            return Response(
                {"error": "Order not found or does not belong to your company"},
                status=status.HTTP_404_NOT_FOUND,
            )

        try:
            step = OrderProcessStep.objects.get(id=step_id, order=order)
        except OrderProcessStep.DoesNotExist:
            return Response({"error": "Query order not found"}, status=status.HTTP_404_NOT_FOUND)

        # Check order-level lock
        try:
            process_lock = OrderProcessLock.objects.get(order=order)
            if process_lock.lock_level == "FULLY_LOCKED":
                return Response(
                    {
                        "error": "Process is fully locked after courier dispatch",
                        "message": "No changes allowed after courier dispatch",
                    },
                    status=status.HTTP_403_FORBIDDEN,
                )
        except OrderProcessLock.DoesNotExist:
            pass

        # Already done â€” idempotent
        if step.status == "COMPLETED":
            return Response(
                {
                    "success": True,
                    "step_id": str(step.id),
                    "step_name": step.step_name,
                    "status": step.status,
                    "is_locked": True,
                    "marked_done_at": step.marked_done_at.isoformat()
                    if step.marked_done_at
                    else None,
                    "message": "Step was already marked as done",
                }
            )

        # ----------------------------------------------------------------
        # STEP 1 â€” Form validation: required fields must be filled
        # ----------------------------------------------------------------
        is_valid, missing_fields = _validate_step_form_completion(step)
        if not is_valid:
            return Response(
                {
                    "success": False,
                    "error": "You are still left with completing your form. You cannot mark as done right now.",
                    "missing_fields": missing_fields,
                    "step_name": step.step_name,
                    "message": (
                        f"You are still left with completing your form. "
                        f"You cannot mark as done right now. "
                        f"Missing: {', '.join(missing_fields)}"
                    ),
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ----------------------------------------------------------------
        # STEP 2 â€” Next-step confirmation: deadline + assignee required
        # ----------------------------------------------------------------
        next_step_deadline = request.data.get("next_step_deadline")
        next_step_assignee_id = request.data.get("next_step_assignee_id")

        # Find the next step (by position)
        next_step = (
            OrderProcessStep.objects.filter(order=order, position__gt=step.position)
            .order_by("position")
            .first()
        )

        # Only require deadline + assignee if there IS a next step
        if next_step:
            if not next_step_deadline:
                return Response(
                    {
                        "success": False,
                        "error": "next_step_deadline is required",
                        "message": "Please select a deadline for the next step before marking as done.",
                        "requires_confirmation": True,
                        "next_step_name": next_step.step_name,
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if not next_step_assignee_id:
                return Response(
                    {
                        "success": False,
                        "error": "next_step_assignee_id is required",
                        "message": "Please assign someone to the next step before marking as done.",
                        "requires_confirmation": True,
                        "next_step_name": next_step.step_name,
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Validate assignee exists and belongs to the same company
            from users.models import User as UserModel

            try:
                assignee = UserModel.objects.get(id=next_step_assignee_id, company=company)
            except UserModel.DoesNotExist:
                return Response(
                    {"error": "Assignee not found or does not belong to your company"},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            assignee = None

        notes = request.data.get("notes", "")

        with transaction.atomic():
            # Mark this step as COMPLETED
            step.status = "COMPLETED"
            step.completed_at = timezone.now()
            step.marked_done_at = timezone.now()
            step.marked_done_by = user
            if notes:
                step.notes = notes

            created_task = None

            if next_step and assignee:
                # Store deadline + assignee on this step for audit
                step.next_step_deadline = next_step_deadline
                step.next_step_assignee = assignee

                # Create a Task for the next step
                created_task = Task.objects.create(
                    company=company,
                    title=f"{next_step.step_name} â€” Order {order.bill_no or str(order.id)[:8]}",
                    description=(
                        f"Query order '{next_step.step_name}' for order "
                        f"{order.bill_no or str(order.id)}. "
                        f"Triggered by completion of '{step.step_name}'."
                    ),
                    deadline=next_step_deadline,
                    assigned_to=assignee,
                    department=_map_department(next_step.department),
                    urgency="MEDIUM",
                    status="PENDING",
                    created_by=user,
                )
                # Also add to M2M assignees for dashboard visibility
                created_task.assignees.add(assignee)

                # Link the task back to this step
                step.next_step_task = created_task

            step.save()

            # Recalculate statuses for remaining steps
            _recalculate_step_statuses(order)

        step.refresh_from_db()
        logger.info(
            f"Step '{step.step_name}' marked as done for order {order_id} by {user.email}. "
            f"Task created: {created_task.id if created_task else 'none'}"
        )

        response_data = {
            "success": True,
            "step_id": str(step.id),
            "step_name": step.step_name,
            "status": step.status,
            "is_locked": True,
            "marked_done_at": step.marked_done_at.isoformat() if step.marked_done_at else None,
            "message": f"Step '{step.step_name}' marked as done and locked",
        }

        if created_task and next_step:
            response_data["next_step"] = {
                "step_name": next_step.step_name,
                "task_id": str(created_task.id),
                "assignee_id": str(assignee.id),
                "assignee_name": getattr(assignee, "name", assignee.email),
                "deadline": str(next_step_deadline),
            }

        return Response(response_data)

    except Exception as e:
        logger.error(f"Error marking step as done: {str(e)}", exc_info=True)
        return Response(
            {"error": f"An error occurred: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["POST"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def auto_complete_process_step(request, order_id):
    """
    Auto-complete a query order when task is completed in another module

    POST /api/orders/{order_id}/auto-complete-step/

    Body: {
        "step_name": "Advance Received",
        "reference_id": "uuid",  // Optional - ID of the related record
        "notes": "Auto-completed from Receipt module"  // Optional
    }

    Example Usage:
    - When Receipt is created â†’ Auto-complete "Advance Received"
    - When 3D Design is created â†’ Auto-complete "3D Design"
    - When Metal Issue is created â†’ Auto-complete "Metal Issue"

    NOTE: Step must be saved at least once before it can be marked as done.
    """
    try:
        from vouchers.auto_complete_steps import auto_complete_step_by_name

        user = request.user
        company = getattr(user, "company", None)

        # Verify order belongs to company
        try:
            order = Order.objects.get(id=order_id, company=company)
        except Order.DoesNotExist:
            return Response(
                {"error": "Order not found or does not belong to your company"},
                status=status.HTTP_404_NOT_FOUND,
            )

        step_name = request.data.get("step_name")
        reference_id = request.data.get("reference_id")
        notes = request.data.get("notes", "Auto-completed")

        if not step_name:
            return Response({"error": "step_name is required"}, status=status.HTTP_400_BAD_REQUEST)

        # Check if step exists and has been saved at least once
        step = OrderProcessStep.objects.filter(order=order, step_name=step_name).first()

        if not step:
            return Response(
                {"error": f"Step '{step_name}' not found for this order"},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Validate that step has been saved at least once (has reference_id or is IN_PROGRESS)
        # Allow completion if step is already IN_PROGRESS or has a reference_id
        if step.status == "PENDING" and not step.reference_id:
            return Response(
                {"error": "Step must be saved at least once before marking as done"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Auto-complete the step with reference_id
        result = auto_complete_step_by_name(order_id, step_name, notes, reference_id=reference_id)

        if result["success"]:
            return Response(result, status=status.HTTP_200_OK)
        else:
            return Response(result, status=status.HTTP_400_BAD_REQUEST)

    except Exception as e:
        logger.error(f"Error auto-completing step: {str(e)}", exc_info=True)
        return Response(
            {"error": f"An error occurred: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# ==================== Estimate Approval Step ====================


@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def get_or_create_estimate_for_approval(request, order_id):
    """
    Get or create the estimate linked to the "Estimate Approval" query order.

    Handles two order creation flows:

    FLOW 1 â€” Sale â†’ Order conversion:
      The order was created from a Sale that had a selected_estimate.
      We find that estimate via: order â†’ source_draft â†’ source_sale â†’ selected_estimate
      and return it (pre-linked). The estimate already exists; we just surface it.

    FLOW 2 â€” Direct order creation:
      No sale, no pre-existing estimate. We create a new EstimateVoucher
      pre-filled with the order's details (account, item_name, date) so the
      user can fill in the line items and save.

    In both flows:
    - If the step already has a reference_id, we return that estimate directly.
    - After the estimate is returned/created, the frontend should call
      POST /api/orders/{order_id}/process-steps/Estimate Approval/save/
      with { "reference_id": "<estimate_id>" } to link it to the step.

    GET /api/orders/{order_id}/estimate-approval/

    Response:
    {
        "estimate": { ...EstimateVoucher fields... },
        "is_new": false,          // true if we just created it (Flow 2)
        "is_locked": false,       // true if step is COMPLETED
        "step_status": "IN_PROGRESS",
        "flow": "sale_conversion" | "direct_order"
    }
    """
    try:
        from .models import EstimateVoucher
        from .serializers import EstimateVoucherSerializer

        user = request.user
        company = getattr(user, "company", None)

        try:
            order = Order.objects.get(id=order_id, company=company)
        except Order.DoesNotExist:
            return Response(
                {"error": "Order not found or does not belong to your company"},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Find the Estimate Approval step
        step = OrderProcessStep.objects.filter(order=order, step_name="Estimate Approval").first()

        if not step:
            return Response(
                {"error": "Estimate Approval step not found for this order"},
                status=status.HTTP_404_NOT_FOUND,
            )

        source_draft = getattr(order, "source_draft", None)

        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        # STEP 1: Resolve an "auto-linked" estimate from any source.
        #         If found â†’ the step is LOCKED automatically.
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        auto_estimate = None
        auto_flow = None

        # (a) Order Query (issues.Query) with linked_estimate_id
        if not auto_estimate and source_draft and source_draft.sales_query_id:
            try:
                from issues.models import Query as CustomerQuery

                cq = CustomerQuery.objects.filter(id=source_draft.sales_query_id).first()
                if cq and cq.linked_estimate_id:
                    auto_estimate = EstimateVoucher.objects.filter(
                        pk=cq.linked_estimate_id, company=company
                    ).first()
                    if auto_estimate:
                        auto_flow = "query_linked"
            except Exception as e:
                logger.warning(f"[EstimateApproval] CustomerQuery lookup failed: {e}")

        # (b) Sales Query (sales_queries.SalesQuery) with selected_estimate_id
        if not auto_estimate and source_draft and source_draft.sales_query_id:
            try:
                from sales_queries.models import SalesQuery

                sq = SalesQuery.objects.filter(id=source_draft.sales_query_id).first()
                if sq and sq.selected_estimate_id:
                    auto_estimate = EstimateVoucher.objects.filter(
                        pk=sq.selected_estimate_id, company=company
                    ).first()
                    if auto_estimate:
                        auto_flow = "sale_conversion"
            except Exception as e:
                logger.warning(f"[EstimateApproval] SalesQuery lookup failed: {e}")

        # (c) Sale with selected_estimate (order came from a Sale object)
        if not auto_estimate and source_draft and source_draft.source_sale:
            sale = source_draft.source_sale
            if sale.selected_estimate:
                auto_estimate = sale.selected_estimate
                auto_flow = "sale_conversion"

        # If an auto-linked estimate was found â†’ lock the step and return
        if auto_estimate and auto_flow:
            if str(step.reference_id) != str(auto_estimate.id) or step.status != "COMPLETED":
                step.reference_id = auto_estimate.id
                step.status = "COMPLETED"
                step.saved_at = timezone.now()
                step.saved_by = user
                step.marked_done_at = timezone.now()
                step.marked_done_by = user
                step.completed_at = timezone.now()
                step.save(
                    update_fields=[
                        "reference_id",
                        "status",
                        "saved_at",
                        "saved_by",
                        "marked_done_at",
                        "marked_done_by",
                        "completed_at",
                    ]
                )
            return Response(
                {
                    "estimate": EstimateVoucherSerializer(
                        auto_estimate, context={"request": request}
                    ).data,
                    "is_new": False,
                    "is_locked": True,
                    "step_status": "COMPLETED",
                    "flow": auto_flow,
                }
            )

        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        # STEP 2: Direct order (no auto-linked estimate).
        #         If step already has a reference_id from a previous visit,
        #         return it respecting its current lock status.
        #         Patch missing pre-fill fields if estimate was created blank.
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        if step.reference_id:
            try:
                estimate = EstimateVoucher.objects.get(pk=step.reference_id, company=company)

                # Patch blank pre-fill fields from order data so the form shows them
                draft_order_data = source_draft.order_data if source_draft else {}
                patch_fields = []
                if not estimate.gold_quality:
                    gq = order.stamp or (draft_order_data or {}).get("stamp", "") or ""
                    if gq:
                        estimate.gold_quality = gq
                        patch_fields.append("gold_quality")
                if not estimate.size_details:
                    sd = order.size or (draft_order_data or {}).get("size", "") or ""
                    if sd:
                        estimate.size_details = sd
                        patch_fields.append("size_details")
                if not estimate.item_name:
                    iname = order.item_name or (draft_order_data or {}).get("item_name", "") or ""
                    if iname:
                        estimate.item_name = iname
                        estimate.jewellery_type = iname
                        patch_fields += ["item_name", "jewellery_type"]
                if patch_fields:
                    estimate.save(update_fields=patch_fields)

                return Response(
                    {
                        "estimate": EstimateVoucherSerializer(
                            estimate, context={"request": request}
                        ).data,
                        "is_new": False,
                        "is_locked": step.status == "COMPLETED",
                        "step_status": step.status,
                        "flow": "direct_order",
                    }
                )
            except EstimateVoucher.DoesNotExist:
                step.reference_id = None
                step.save(update_fields=["reference_id"])

        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        # STEP 3: First visit for a direct order â€” create a blank estimate
        #         pre-filled from order data. Step stays IN_PROGRESS (editable).
        # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
        # Pull extra fields from order_data (stored on draft) as fallback for
        # older orders where Order model fields weren't populated at creation.
        draft_order_data = {}
        if source_draft:
            draft_order_data = source_draft.order_data or {}

        gold_quality = order.stamp or draft_order_data.get("stamp", "") or ""
        size_details = order.size or draft_order_data.get("size", "") or ""
        item_name = order.item_name or draft_order_data.get("item_name", "") or ""

        with transaction.atomic():
            estimate = EstimateVoucher.objects.create(
                company=company,
                account=order.account,
                item_name=item_name,
                date=order.date or timezone.now().date(),
                total_taxable_value=0,
                gst_amount=0,
                grand_total=0,
                status="draft",
                created_by=user,
                jewellery_type=item_name,
                gold_quality=gold_quality,
                size_details=size_details,
            )
            step.reference_id = estimate.id
            step.status = "IN_PROGRESS"
            step.saved_at = timezone.now()
            step.saved_by = user
            step.save()

        return Response(
            {
                "estimate": EstimateVoucherSerializer(estimate, context={"request": request}).data,
                "is_new": True,
                "is_locked": False,
                "step_status": "IN_PROGRESS",
                "flow": "direct_order",
            },
            status=status.HTTP_201_CREATED,
        )

    except Exception as e:
        logger.error(f"Error getting/creating estimate for approval: {str(e)}", exc_info=True)
        return Response(
            {"error": f"An error occurred: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


# ==================== Direct Order Creation (from Order Form) ====================


@api_view(["POST"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def create_order_with_process(request):
    """
    Create Order Draft directly from Order Form

    POST /api/orders/create-with-process/

    Body: {
        "order_data": {
            "item_name": "...",
            "account": "customer-id",
            "date": "2026-02-09",
            "gold_rate": 5000,
            ...
        },
        "receipt_voucher_id": "uuid",  // Optional
        "advance_amount": 25000.00,  // Optional
        "advance_notes": "..."  // Optional
    }

    Response: {
        "draft_id": "uuid",
        "process_steps": [...],
        "status": "pending_confirmation",
        "can_modify": true
    }
    """
    try:
        user = request.user
        company = getattr(user, "company", None)

        if not company:
            return Response(
                {"error": "User does not belong to a company"}, status=status.HTTP_400_BAD_REQUEST
            )

        # Get request data
        order_data = request.data.get("order_data", {})
        receipt_voucher_id = request.data.get("receipt_voucher_id")
        advance_amount = request.data.get("advance_amount", 0)
        advance_notes = request.data.get("advance_notes", "")
        query_id = request.data.get("query_id")
        linked_estimate_id = request.data.get("linked_estimate_id")

        # Validate order_data
        if not order_data:
            return Response({"error": "order_data is required"}, status=status.HTTP_400_BAD_REQUEST)

        # Validate receipt voucher if provided
        receipt = None
        if receipt_voucher_id:
            try:
                receipt = Receipt.objects.get(id=receipt_voucher_id, company=company)
            except Receipt.DoesNotExist:
                return Response(
                    {"error": "Receipt voucher not found"}, status=status.HTTP_404_NOT_FOUND
                )

        with transaction.atomic():
            # Create OrderDraft (without source_sale since this is direct)
            draft = OrderDraft.objects.create(
                source_sale=None,  # No sale for direct orders
                sales_query_id=query_id,
                company=company,
                receipt_voucher=receipt,
                advance_amount=advance_amount,
                advance_notes=advance_notes,
                order_data=order_data,
                status="pending_confirmation",
                created_by=user,
            )

            # Create default query orders (bulk insert for performance)
            steps_to_create = []
            for step_data in DEFAULT_ORDER_PROCESS_STEPS:
                step_status = "PENDING"
                completed_at = None
                if step_data["position"] == 1 and advance_amount and advance_amount > 0:
                    step_status = "COMPLETED"
                    completed_at = timezone.now()

                steps_to_create.append(
                    OrderProcessStep(
                        order_draft=draft,
                        step_name=step_data["name"],
                        description=step_data.get("description", ""),
                        department=step_data.get("department", ""),
                        position=step_data["position"],
                        status=step_status,
                        completed_at=completed_at,
                    )
                )
            OrderProcessStep.objects.bulk_create(steps_to_create)

            # Auto-link a pre-selected estimate to the Estimate Approval step and mark COMPLETED
            if linked_estimate_id:
                try:
                    from .models import EstimateVoucher

                    estimate_obj = EstimateVoucher.objects.get(
                        pk=linked_estimate_id, company=company
                    )
                    approval_step = draft.process_steps.filter(
                        step_name="Estimate Approval"
                    ).first()
                    if approval_step:
                        approval_step.reference_id = estimate_obj.id
                        approval_step.status = "COMPLETED"
                        approval_step.saved_at = timezone.now()
                        approval_step.saved_by = user
                        approval_step.marked_done_at = timezone.now()
                        approval_step.marked_done_by = user
                        approval_step.completed_at = timezone.now()
                        approval_step.save(
                            update_fields=[
                                "reference_id",
                                "status",
                                "saved_at",
                                "saved_by",
                                "marked_done_at",
                                "marked_done_by",
                                "completed_at",
                            ]
                        )
                except Exception as est_err:
                    logger.warning(
                        f"Could not link estimate {linked_estimate_id} to draft {draft.id}: {est_err}"
                    )

            # Prepare response
            steps_data = [
                {
                    "id": str(step.id),
                    "step_name": step.step_name,
                    "description": step.description,
                    "department": step.department,
                    "position": step.position,
                    "status": step.status,
                    "notes": step.notes or "",
                    "completed_at": step.completed_at.isoformat() if step.completed_at else None,
                }
                for step in draft.process_steps.all().order_by("position")
            ]

            logger.info(
                f"Created OrderDraft {draft.id} directly from Order Form by user {user.email}"
            )

            return Response(
                {
                    "draft_id": str(draft.id),
                    "status": draft.status,
                    "process_steps": steps_data,
                    "total_steps": len(steps_data),
                    "message": "Order draft created successfully. Please review and confirm query orders.",
                    "can_modify": True,
                    "advance_amount": str(advance_amount) if advance_amount else None,
                    "advance_notes": advance_notes,
                },
                status=status.HTTP_201_CREATED,
            )

    except Exception as e:
        logger.error(f"Error creating order draft: {str(e)}", exc_info=True)
        return Response(
            {"error": f"An error occurred: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# ==================== RECEIPT API VIEWS ====================
# Moved from receipt_api_views.py

"""
Receipt API Views for Order Process Management

Provides API endpoints to fetch receipt vouchers for the order conversion workflow.
"""

from rest_framework import status
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.db.models import Q
import logging

from .models import Receipt

logger = logging.getLogger(__name__)


# Helper function to get authentication classes
def get_auth_classes():
    """Get authentication classes, with JWT if available"""
    from nsj_backend.authentication import CsrfExemptSessionAuthentication

    auth_classes = [CsrfExemptSessionAuthentication]
    try:
        from rest_framework_simplejwt.authentication import JWTAuthentication

        auth_classes.insert(0, JWTAuthentication)
    except ImportError:
        pass
    return auth_classes


@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def get_receipts_for_dropdown(request):
    """
    Get receipt vouchers for dropdown selection

    Used in Query â†’ Order conversion workflow to select advance payment receipt.

    GET /api/receipts/dropdown/

    Query Parameters:
    - search: Filter by party name or narration
    - type: Filter by receipt type (Cr or Dr) - optional
    - limit: Number of results (default: 50)

    Response:
    [
        {
            "id": "uuid",
            "party_name": "Customer Name",
            "party_account_name": "Account Name",
            "amount": "25000.00",
            "date": "2026-02-06",
            "narration": "Advance payment...",
            "type": "Dr",
            "balance": "25000.00",
            "label": "AALOK SHAH - â‚¹25000.00 (2026-02-06)"
        },
        ...
    ]
    """
    try:
        user = request.user
        company = getattr(user, "company", None)

        if not company:
            return Response(
                {"error": "User does not belong to a company"}, status=status.HTTP_400_BAD_REQUEST
            )

        # Get receipts for this company
        queryset = Receipt.objects.filter(company=company).select_related("party_name")

        # Filter by type if specified (Cr or Dr)
        receipt_type = request.GET.get("type", "")
        if receipt_type and receipt_type in ["Cr", "Dr"]:
            queryset = queryset.filter(type=receipt_type)

        # Search filter
        search = request.GET.get("search", "")
        if search:
            queryset = queryset.filter(
                Q(narration__icontains=search) | Q(party_name__account_name__icontains=search)
            )

        # Limit results
        limit = int(request.GET.get("limit", 50))
        queryset = queryset.order_by("-date")[:limit]

        # Serialize data
        receipts_data = []
        for receipt in queryset:
            party_account_name = ""

            if receipt.party_name:
                # Account model has 'account_name' field, not 'name'
                party_account_name = receipt.party_name.account_name or ""

            receipts_data.append(
                {
                    "id": str(receipt.id),
                    "party_name": party_account_name,  # Using account_name for party_name
                    "party_account_name": party_account_name,
                    "amount": str(receipt.dr) if receipt.dr else "0.00",
                    "date": receipt.date.isoformat() if receipt.date else None,
                    "narration": receipt.narration or "",
                    "type": receipt.type,
                    "balance": str(receipt.balance) if receipt.balance else "0.00",
                    # Formatted label for dropdown
                    "label": f"{party_account_name} - â‚¹{receipt.dr or 0} ({receipt.date})"
                    if receipt.date
                    else f"{party_account_name} - â‚¹{receipt.dr or 0}",
                }
            )

        return Response(receipts_data)

    except Exception as e:
        logger.error(f"Error fetching receipts for dropdown: {str(e)}", exc_info=True)
        return Response(
            {"error": f"An error occurred: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# ==================== SAFE CONVERSION VIEWS ====================
# Moved from safe_conversion_views.py

"""
Safe Order Conversion Views - Handles missing Sale records gracefully
"""

from rest_framework import status
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
import logging

from .models import Sale, EstimateVoucher
from sales_queries.models import SalesQuery

logger = logging.getLogger(__name__)


def get_auth_classes():
    """Get authentication classes, with JWT if available"""
    from nsj_backend.authentication import CsrfExemptSessionAuthentication

    auth_classes = [CsrfExemptSessionAuthentication]
    try:
        from rest_framework_simplejwt.authentication import JWTAuthentication

        auth_classes.insert(0, JWTAuthentication)
    except ImportError:
        pass
    return auth_classes


@api_view(["GET"])
@authentication_classes(get_auth_classes())
@permission_classes([IsAuthenticated])
def get_or_create_sale_from_query(request, query_id):
    """
    SAFE: Get existing Sale or create one from Sales Lead

    This endpoint handles the case where a Sales Lead is marked as
    "converted_to_sale" but no Sale record actually exists.

    GET /api/sales-queries/{query_id}/get-or-create-sale/

    Returns:
    {
        "sale_id": "uuid",
        "created": false,  // true if we had to create it
        "item_name": "Test Ring",
        "message": "Found existing sale" or "Created new sale from query"
    }
    """
    try:
        user = request.user
        company = getattr(user, "company", None)

        if not company:
            return Response(
                {"error": "User does not belong to a company"}, status=status.HTTP_400_BAD_REQUEST
            )

        # Get the sales lead
        try:
            sales_query = SalesQuery.objects.get(id=query_id, company=company)
        except SalesQuery.DoesNotExist:
            return Response({"error": "Sales lead not found"}, status=status.HTTP_404_NOT_FOUND)

        # Try to find existing sale by estimate ID
        sale = None
        if sales_query.selected_estimate_id:
            sale = Sale.objects.filter(
                company=company, selected_estimate_id=sales_query.selected_estimate_id
            ).first()

        # If not found, try by item name and recent date
        if not sale:
            from datetime import timedelta
            from django.utils import timezone

            recent_date = timezone.now() - timedelta(days=30)
            sale = Sale.objects.filter(
                company=company, item_name=sales_query.jewellery_type, created_at__gte=recent_date
            ).first()

        # If still not found, create a new sale
        if not sale:
            logger.info(f"Creating new Sale from SalesLead {query_id}")

            # Get the selected estimate if it exists
            estimate = None
            if sales_query.selected_estimate_id:
                try:
                    estimate = EstimateVoucher.objects.get(
                        id=sales_query.selected_estimate_id, company=company
                    )
                except EstimateVoucher.DoesNotExist:
                    pass

            # Create the sale
            sale = Sale.objects.create(
                company=company,
                item_name=sales_query.jewellery_type or "Unnamed Item",
                account=sales_query.account,
                selected_estimate=estimate,
                created_by=user,
                # Add any other fields from the lead
                remarks=f"Auto-created from Sales Lead {query_id}",
            )

            # Update the sales lead workflow status
            if sales_query.workflow_status != "converted_to_sale":
                sales_query.workflow_status = "converted_to_sale"
                sales_query.save()

            return Response(
                {
                    "sale_id": str(sale.id),
                    "created": True,
                    "item_name": sale.item_name,
                    "message": "Created new sale from query",
                }
            )
        else:
            logger.info(f"Found existing Sale {sale.id} for SalesQuery {query_id}")
            return Response(
                {
                    "sale_id": str(sale.id),
                    "created": False,
                    "item_name": sale.item_name,
                    "message": "Found existing sale",
                }
            )

    except Exception as e:
        logger.error(f"Error in get_or_create_sale_from_query: {str(e)}", exc_info=True)
        return Response(
            {"error": f"An error occurred: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# ============================================================================
# 2D DESIGN ENDPOINTS - WITH MULTI-IMAGE SUPPORT
# ============================================================================

from .models import TwoDDesign, TwoDDesignImage
from .serializers import TwoDDesignSerializer


def _new_log_group():
    """Generate a fresh log-group UUID for grouping images uploaded in one save batch."""
    return uuid.uuid4()


@jwt_login_required
@csrf_exempt
@require_http_methods(["GET"])
def two_d_design_overview_view(request):
    """2D Design overview statistics"""
    company = request.user.company
    designs = TwoDDesign.objects.filter(company=company)

    return JsonResponse(
        {
            "total": designs.count(),
            "with_images": designs.filter(images__isnull=False).distinct().count(),
            "finalized": designs.filter(is_draft=False).count(),
        }
    )


@jwt_login_required
@csrf_exempt
@require_http_methods(["GET", "POST"])
def two_d_design_collection_view(request):
    """List 2D Designs or create new one"""
    company = request.user.company
    user = request.user

    if request.method == "GET":
        designs = TwoDDesign.objects.filter(company=company)

        # Search
        search = request.GET.get("search")
        if search:
            designs = designs.filter(account_order_id__icontains=search)

        # Pagination
        paginator = Paginator(designs.order_by("-created_at"), 10)
        page = request.GET.get("page", 1)

        try:
            page_obj = paginator.page(page)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)

        serializer = TwoDDesignSerializer(page_obj, many=True, context={"request": request})

        return JsonResponse(
            {
                "count": paginator.count,
                "results": serializer.data,
                "total_pages": paginator.num_pages,
                "current_page": page_obj.number,
            }
        )

    elif request.method == "POST":
        # Extract image-related parameters before passing to serializer
        final_design_id = request.POST.get("final_design_image_id")
        final_approved_id = request.POST.get("final_approved_design_image_id")
        is_draft_val = request.POST.get("is_draft") == "true"

        # Prepare data for serializer (only include valid model fields)
        serializer_data = {}
        if request.POST.get("account_order_id"):
            serializer_data["account_order_id"] = request.POST.get("account_order_id")
        if request.POST.get("order_id"):
            serializer_data["order_id"] = request.POST.get("order_id")
        if request.POST.get("is_draft"):
            serializer_data["is_draft"] = request.POST.get("is_draft")

        serializer = TwoDDesignSerializer(data=serializer_data, context={"request": request})

        if serializer.is_valid():
            # Save base record
            record = serializer.save(company=company, created_by=user, is_draft=is_draft_val)

            # Get final file indices
            final_design_index = None
            final_approved_index = None
            try:
                if "final_design_file_index" in request.POST:
                    final_design_index = int(request.POST.get("final_design_file_index"))
                if "final_approved_design_file_index" in request.POST:
                    final_approved_index = int(request.POST.get("final_approved_design_file_index"))
            except (ValueError, TypeError):
                pass

            # Generate one log_group UUID for this entire save batch
            batch_log_group = (
                _new_log_group()
                if (
                    request.FILES.getlist("design_images[]")
                    or request.FILES.getlist("approved_design_images[]")
                )
                else None
            )

            # Create image records for design images
            design_files = request.FILES.getlist("design_images[]")
            for idx, file in enumerate(design_files):
                is_final = final_design_index is not None and idx == final_design_index
                TwoDDesignImage.objects.create(
                    two_d_design=record,
                    image=file,
                    field_type="design",
                    is_final_design=is_final,
                    log_group=batch_log_group,
                    company=company,
                    uploaded_by=user,
                )

            # Create image records for approved design images
            approved_files = request.FILES.getlist("approved_design_images[]")
            for idx, file in enumerate(approved_files):
                is_final = final_approved_index is not None and idx == final_approved_index
                TwoDDesignImage.objects.create(
                    two_d_design=record,
                    image=file,
                    field_type="approved",
                    is_final_approved=is_final,
                    log_group=batch_log_group,
                    company=company,
                    uploaded_by=user,
                )

            # If this is the first upload and no log selected yet, auto-select this log

            # Validate required fields if not draft
            if record.is_draft is False:
                is_valid, missing = _check_form_record_fields(
                    record, "TwoDDesign", uploaded_files=request.FILES
                )
                if not is_valid:
                    record.delete()
                    return JsonResponse(
                        {
                            "errors": {
                                "missing_fields": missing,
                                "message": f"Missing: {', '.join(missing)}",
                            }
                        },
                        status=400,
                    )

            serializer = TwoDDesignSerializer(record, context={"request": request})
            return JsonResponse(serializer.data, status=201)

        return JsonResponse({"errors": serializer.errors}, status=400)


@jwt_login_required
@csrf_exempt
@require_http_methods(["GET", "PATCH", "PUT", "DELETE"])
def two_d_design_detail_view(request, pk):
    """Retrieve, update, or delete a 2D Design"""
    company = request.user.company
    user = request.user

    try:
        record = TwoDDesign.objects.get(pk=pk, company=company)
    except TwoDDesign.DoesNotExist:
        return JsonResponse({"error": "Not found"}, status=404)

    if request.method == "GET":
        serializer = TwoDDesignSerializer(record, context={"request": request})
        return JsonResponse(serializer.data)

    elif request.method in ["PATCH", "PUT"]:
        # Parse multipart data if needed (modifies request in-place, returns None)
        if request.content_type and "multipart/form-data" in request.content_type:
            parse_put_request(request)

        # Extract image-related parameters before passing to serializer
        final_design_id = request.POST.get("final_design_image_id")
        final_approved_id = request.POST.get("final_approved_design_image_id")

        # Prepare data for serializer (only include valid model fields)
        serializer_data = {}
        if request.POST.get("account_order_id"):
            serializer_data["account_order_id"] = request.POST.get("account_order_id")
        if request.POST.get("order_id"):
            serializer_data["order_id"] = request.POST.get("order_id")
        if request.POST.get("is_draft"):
            serializer_data["is_draft"] = request.POST.get("is_draft")

        serializer = TwoDDesignSerializer(
            record,
            data=serializer_data,
            partial=True,
            context={"request": request},
        )

        if serializer.is_valid():
            record = serializer.save(updated_by=user)

            # Handle keep lists â€” delete images not in keep list.
            # Use the key presence check so an empty keep list still deletes all images.
            if "keep_design_image_ids[]" in request.POST:
                keep_design_ids = set(request.POST.getlist("keep_design_image_ids[]"))
                record.images.filter(field_type="design").exclude(id__in=keep_design_ids).delete()
            if "keep_approved_design_image_ids[]" in request.POST:
                keep_approved_ids = set(request.POST.getlist("keep_approved_design_image_ids[]"))
                record.images.filter(field_type="approved").exclude(
                    id__in=keep_approved_ids
                ).delete()

            # Get final file indices
            final_design_index = None
            final_approved_index = None
            try:
                if "final_design_file_index" in request.POST:
                    final_design_index = int(request.POST.get("final_design_file_index"))
                if "final_approved_design_file_index" in request.POST:
                    final_approved_index = int(request.POST.get("final_approved_design_file_index"))
            except (ValueError, TypeError):
                pass

            # Generate one log_group UUID for new images in this PATCH batch
            patch_log_group = (
                _new_log_group()
                if (
                    request.FILES.getlist("design_images[]")
                    or request.FILES.getlist("approved_design_images[]")
                )
                else None
            )

            # Handle new design images
            design_files = request.FILES.getlist("design_images[]")
            for idx, file in enumerate(design_files):
                is_final = final_design_index is not None and idx == final_design_index
                TwoDDesignImage.objects.create(
                    two_d_design=record,
                    image=file,
                    field_type="design",
                    is_final_design=is_final,
                    log_group=patch_log_group,
                    company=company,
                    uploaded_by=user,
                )

            # Handle new approved images
            approved_files = request.FILES.getlist("approved_design_images[]")
            for idx, file in enumerate(approved_files):
                is_final = final_approved_index is not None and idx == final_approved_index
                TwoDDesignImage.objects.create(
                    two_d_design=record,
                    image=file,
                    field_type="approved",
                    is_final_approved=is_final,
                    log_group=patch_log_group,
                    company=company,
                    uploaded_by=user,
                )

            # Handle log group selection from frontend (independent per image type)
            update_fields = []
            select_log = request.POST.get("select_log_group")
            if select_log is not None:
                record.selected_log_group = select_log or None
                update_fields.append("selected_log_group")
            select_secondary_log = request.POST.get("select_secondary_log_group")
            if select_secondary_log is not None:
                record.selected_secondary_log_group = select_secondary_log or None
                update_fields.append("selected_secondary_log_group")
            if update_fields:
                record.save(update_fields=update_fields)

            # Update final selection for existing images if provided
            if final_design_id:
                record.images.filter(field_type="design").update(is_final_design=False)
                try:
                    final_img = TwoDDesignImage.objects.get(id=final_design_id, field_type="design")
                    final_img.is_final_design = True
                    final_img.save()
                except TwoDDesignImage.DoesNotExist:
                    pass

            if final_approved_id:
                record.images.filter(field_type="approved").update(is_final_approved=False)
                try:
                    final_img = TwoDDesignImage.objects.get(
                        id=final_approved_id, field_type="approved"
                    )
                    final_img.is_final_approved = True
                    final_img.save()
                except TwoDDesignImage.DoesNotExist:
                    pass

            # Validate required fields if not draft
            if record.is_draft is False:
                is_valid, missing = _check_form_record_fields(
                    record, "TwoDDesign", uploaded_files=request.FILES
                )
                if not is_valid:
                    return JsonResponse(
                        {
                            "errors": {
                                "missing_fields": missing,
                                "message": f"Missing: {', '.join(missing)}",
                            }
                        },
                        status=400,
                    )

            serializer = TwoDDesignSerializer(record, context={"request": request})
            return JsonResponse(serializer.data)

        return JsonResponse({"errors": serializer.errors}, status=400)

    elif request.method == "DELETE":
        record.delete()
        return JsonResponse({"deleted": True})
