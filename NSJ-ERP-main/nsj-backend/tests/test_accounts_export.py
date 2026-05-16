"""
Comprehensive tests for accounts/views.py export functions
Tests Excel export functionality for accounts and subaccounts.
"""

import pytest
from django.test import RequestFactory
from accounts.models import Account, SubAccount
from accounts.views import (
    account_export_view,
    accounts_export_all_view,
    subaccounts_export_all_view,
    subaccount_export_view,
)
from users.models import User
from core.models import Company, Branch, LocationMaster
from openpyxl import load_workbook
from io import BytesIO


@pytest.fixture
def company(db):
    """Create a test company"""
    return Company.objects.create(name="Test Company", display_name="Test Company")


@pytest.fixture
def branch(db, company):
    """Create a test branch"""
    return Branch.objects.create(name="Main Branch", company=company)


@pytest.fixture
def location(db, company):
    """Create a test location"""
    return LocationMaster.objects.create(name="Test Location", company=company)


@pytest.fixture
def user(db, company):
    """Create a test user"""
    return User.objects.create(
        name="Test User",
        email="test@example.com",
        username="testuser",
        company=company,
        role="EMPLOYEE",
    )


@pytest.fixture
def account(db, company, branch, location, user):
    """Create a test account"""
    return Account.objects.create(
        account_name="Test Account",
        account_no="1001",
        group_code="SUNDRY_DEBTORS",
        company=company,
        branch=branch,
        location=location,
        created_by=user,
        status="ACTIVE",
        remarks="Test remarks",
    )


@pytest.fixture
def subaccount(db, account, user):
    """Create a test subaccount"""
    return SubAccount.objects.create(
        account=account,
        sub_account_name="Test SubAccount",
        address="123 Test St",
        phone_number="1234567890",
        email="subaccount@test.com",
        gender="MALE",
        created_by=user,
    )


@pytest.mark.django_db
class TestAccountExportView:
    """Test single account export to Excel"""

    def test_export_account_success(self, account, user):
        """Test successful export of single account"""
        factory = RequestFactory()
        request = factory.get(f"/api/accounts/{account.id}/export/")
        request.user = user

        response = account_export_view(request, account.id)

        assert response.status_code == 200
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "Account_" in response["Content-Disposition"]
        assert ".xlsx" in response["Content-Disposition"]

        # Verify Excel content
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        assert ws.title == "Account Details"

        # Check key fields are present
        values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
        assert "Account Name" in values
        assert "Test Account" in values
        assert "Account Number" in values
        assert "1001" in values

    def test_export_account_not_found(self, user):
        """Test export with non-existent account ID"""
        import uuid

        factory = RequestFactory()
        fake_uuid = str(uuid.uuid4())
        request = factory.get(f"/api/accounts/{fake_uuid}/export/")
        request.user = user

        response = account_export_view(request, fake_uuid)

        assert response.status_code == 404
        assert response["Content-Type"] == "application/json"

    def test_export_account_with_remarks(self, account, user):
        """Test export includes remarks field"""
        account.remarks = "Important client - handle with care"
        account.save()

        factory = RequestFactory()
        request = factory.get(f"/api/accounts/{account.id}/export/")
        request.user = user

        response = account_export_view(request, account.id)

        assert response.status_code == 200

        # Verify remarks in Excel
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]

        assert "Remarks" in values
        assert "Important client - handle with care" in values


@pytest.mark.django_db
class TestAccountsExportAllView:
    """Test export all accounts to Excel"""

    def test_export_all_accounts_success(self, account, user):
        """Test successful export of all accounts"""
        factory = RequestFactory()
        request = factory.get("/api/accounts/export/")
        request.user = user

        response = accounts_export_all_view(request)

        assert response.status_code == 200
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "accounts_data.xlsx" in response["Content-Disposition"]

        # Verify Excel content
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        assert ws.title == "Accounts"

        # Check header row exists
        header_row = [cell.value for cell in ws[1]]
        assert any("Account Name" in str(val) for val in header_row if val)

    def test_export_all_accounts_multiple(self, company, branch, location, user):
        """Test export with multiple accounts"""
        # Create multiple accounts
        for i in range(3):
            Account.objects.create(
                account_name=f"Account {i}",
                account_no=f"100{i}",
                group_code="SUNDRY_DEBTORS",
                company=company,
                branch=branch,
                location=location,
                created_by=user,
            )

        factory = RequestFactory()
        request = factory.get("/api/accounts/export/")
        request.user = user

        response = accounts_export_all_view(request)

        assert response.status_code == 200

        # Verify multiple accounts in Excel
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active

        # Should have header + 3 data rows + footer
        assert ws.max_row >= 4

    def test_export_all_accounts_empty(self, user):
        """Test export with no accounts"""
        factory = RequestFactory()
        request = factory.get("/api/accounts/export/")
        request.user = user

        response = accounts_export_all_view(request)

        assert response.status_code == 200

        # Verify Excel has header but no data rows
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        assert ws.max_row >= 1  # At least header row


@pytest.mark.django_db
class TestSubAccountExportView:
    """Test single subaccount export to Excel"""

    def test_export_subaccount_success(self, subaccount, user):
        """Test successful export of single subaccount"""
        factory = RequestFactory()
        request = factory.get(f"/api/subaccounts/{subaccount.id}/export/")
        request.user = user

        response = subaccount_export_view(request, subaccount.id)

        assert response.status_code == 200
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "SubAccount_" in response["Content-Disposition"]

        # Verify Excel content
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        assert ws.title == "SubAccount Details"

        # Check key fields
        values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
        assert "Sub Account Name" in values
        assert "Test SubAccount" in values
        assert "Address" in values
        assert "123 Test St" in values

    def test_export_subaccount_not_found(self, user):
        """Test export with non-existent subaccount ID"""
        import uuid

        factory = RequestFactory()
        fake_uuid = str(uuid.uuid4())
        request = factory.get(f"/api/subaccounts/{fake_uuid}/export/")
        request.user = user

        response = subaccount_export_view(request, fake_uuid)

        assert response.status_code == 404

    def test_export_subaccount_with_all_fields(self, subaccount, user):
        """Test export includes all subaccount fields"""
        factory = RequestFactory()
        request = factory.get(f"/api/subaccounts/{subaccount.id}/export/")
        request.user = user

        response = subaccount_export_view(request, subaccount.id)

        assert response.status_code == 200

        # Verify all fields in Excel
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]

        assert "Sub Account Name" in values
        assert "Test SubAccount" in values
        assert "Address" in values
        assert "123 Test St" in values
        assert "Phone Number" in values
        assert "1234567890" in values
        assert "Gender" in values
        assert "MALE" in values


@pytest.mark.django_db
class TestSubAccountsExportAllView:
    """Test export all subaccounts to Excel"""

    def test_export_all_subaccounts_success(self, subaccount, user):
        """Test successful export of all subaccounts"""
        factory = RequestFactory()
        request = factory.get("/api/subaccounts/export/")
        request.user = user

        response = subaccounts_export_all_view(request)

        assert response.status_code == 200
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "subaccounts_data.xlsx" in response["Content-Disposition"]

        # Verify Excel content
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        assert ws.title == "SubAccounts"

    def test_export_all_subaccounts_multiple(self, account, user):
        """Test export with multiple subaccounts"""
        # Create multiple subaccounts
        for i in range(3):
            SubAccount.objects.create(
                account=account,
                sub_account_name=f"SubAccount {i}",
                created_by=user,
            )

        factory = RequestFactory()
        request = factory.get("/api/subaccounts/export/")
        request.user = user

        response = subaccounts_export_all_view(request)

        assert response.status_code == 200

        # Verify multiple subaccounts in Excel
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        assert ws.max_row >= 4  # Header + 3 data rows

    def test_export_all_subaccounts_empty(self, user):
        """Test export with no subaccounts"""
        factory = RequestFactory()
        request = factory.get("/api/subaccounts/export/")
        request.user = user

        response = subaccounts_export_all_view(request)

        assert response.status_code == 200

        # Verify Excel has header but no data rows
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        assert ws.max_row >= 1


@pytest.mark.django_db
class TestExcelHelperFunctions:
    """Test helper functions for Excel export"""

    def test_flatten_dict_simple(self):
        """Test _flatten_dict with simple nested dict"""
        from accounts.views import _flatten_dict

        data = {"name": "Test", "contact": {"phone": "123", "email": "test@test.com"}}
        result = _flatten_dict(data)

        assert result["name"] == "Test"
        assert result["contact.phone"] == "123"
        assert result["contact.email"] == "test@test.com"

    def test_flatten_dict_with_fk_object(self):
        """Test _flatten_dict collapses FK objects to name"""
        from accounts.views import _flatten_dict

        data = {"account": {"id": "123", "name": "Test Account"}}
        result = _flatten_dict(data)

        assert result["account"] == "Test Account"

    def test_excel_safe_value_uuid(self):
        """Test _excel_safe_value converts UUID to string"""
        from accounts.views import _excel_safe_value
        import uuid

        test_uuid = uuid.uuid4()
        result = _excel_safe_value(test_uuid)

        assert isinstance(result, str)
        assert str(test_uuid) == result

    def test_excel_safe_value_decimal(self):
        """Test _excel_safe_value converts Decimal to float"""
        from accounts.views import _excel_safe_value
        from decimal import Decimal

        result = _excel_safe_value(Decimal("123.45"))

        assert isinstance(result, float)
        assert result == 123.45

    def test_excel_safe_value_dict_with_name(self):
        """Test _excel_safe_value extracts name from dict"""
        from accounts.views import _excel_safe_value

        result = _excel_safe_value({"id": "123", "name": "Test"})

        assert result == "Test"

    def test_excel_safe_value_list(self):
        """Test _excel_safe_value converts list to JSON"""
        from accounts.views import _excel_safe_value

        result = _excel_safe_value([1, 2, 3])

        assert isinstance(result, str)
        assert "[1, 2, 3]" in result

    def test_excel_safe_value_none(self):
        """Test _excel_safe_value handles None"""
        from accounts.views import _excel_safe_value

        result = _excel_safe_value(None)

        assert result == ""

    def test_excel_safe_value_primitives(self):
        """Test _excel_safe_value passes through primitives"""
        from accounts.views import _excel_safe_value

        assert _excel_safe_value("test") == "test"
        assert _excel_safe_value(123) == 123
        assert _excel_safe_value(45.67) == 45.67
        assert _excel_safe_value(True) is True
