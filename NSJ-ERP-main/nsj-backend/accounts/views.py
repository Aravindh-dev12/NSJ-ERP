"""Standard Django views that power the accounts JSON API."""

import json
from functools import wraps
from typing import Iterable, List, Optional
import uuid
import decimal

from django.db.models import Q
from django.db import IntegrityError
from django.http import HttpResponse, JsonResponse
from django.utils.text import slugify
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook

from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework_simplejwt.exceptions import InvalidToken, TokenError
from nsj_backend.authentication import CsrfExemptSessionAuthentication

from core.models import Branch, CityMaster, CountryMaster, LocationMaster, StateMaster
from .models import Account, ACGroup, ACGroupMaster, SubAccount
from .serializers import AccountSerializer, ACGroupSerializer, ACGroupMasterSerializer, SubAccountSerializer


def jwt_login_required(view_func):
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            auth_header = request.META.get("HTTP_AUTHORIZATION", "")
            if auth_header.startswith("Bearer "):
                try:
                    jwt_auth = JWTAuthentication()
                    validated_token = jwt_auth.get_validated_token(auth_header[7:])
                    request.user = jwt_auth.get_user(validated_token)
                except (InvalidToken, TokenError):
                    return JsonResponse({"detail": "Authentication required"}, status=401)
            else:
                return JsonResponse({"detail": "Authentication required"}, status=401)
        return view_func(request, *args, **kwargs)

    return _wrapped


SELECT_RELATED_FIELDS: List[str] = [
    "company",
    "branch",
    "location",
    "created_by",
    "contact",
    "bank",
    "tax",
    "opening_balance",
]


def _base_queryset(user) -> Iterable[Account]:
    # Use select_related for required fields, but make created_by optional
    # to handle cases where created_by user no longer exists
    queryset = Account.objects.select_related(
        "company",
        "branch",
        "location",
        "contact",
        "bank",
        "tax",
        "opening_balance",
        "spark_account_group",
    )

    # Add created_by as optional (won't filter out accounts if user doesn't exist)
    queryset = queryset.prefetch_related("created_by")

    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(company_id=user.company_id)
    else:
        queryset = queryset.none()
    # Sort by account_no as integer (descending order - latest first)
    # Using Cast to convert text account_no to integer for proper numeric sorting
    from django.db.models.functions import Cast
    from django.db.models import IntegerField

    return queryset.annotate(account_no_int=Cast("account_no", IntegerField())).order_by(
        "-account_no_int"
    )


def _serialize(account: Account) -> dict:
    data = AccountSerializer(account).data

    branch = account.branch
    if branch:
        data["branch"] = {
            "id": str(branch.id),
            "name": branch.name,
            "code": getattr(branch, "code", None),
        }
    else:
        data["branch"] = None

    location = account.location
    if location:
        data["location"] = {
            "id": str(location.id),
            "name": location.name,
        }
    else:
        data["location"] = None

    contact = getattr(account, "contact", None)
    if contact and data.get("contact"):
        contact_data = data["contact"]
        # All address fields are now text fields (city, state, country)
        # Just return them as strings
        if getattr(contact, "city", None):
            contact_data["city"] = contact.city

        if getattr(contact, "state", None):
            contact_data["state"] = contact.state

        if getattr(contact, "country", None):
            contact_data["country"] = contact.country

    return data


@jwt_login_required
@require_http_methods(["GET"])
def account_export_view(request, pk):
    """Return an Excel (.xlsx) file containing full account details for the requested account."""
    user = request.user
    queryset = _base_queryset(user)
    try:
        account = queryset.get(pk=pk)
    except Account.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    data = _serialize(account)

    # Build an Excel workbook with two columns: Field | Value
    wb = Workbook()
    ws = wb.active
    ws.title = "Account Details"

    row = 1

    def write_row(label, value):
        nonlocal row
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value if value is not None else "")
        row += 1

    # Core fields in a logical order
    write_row("Account Name", data.get("account_name"))
    write_row("Account Number", data.get("account_no"))
    write_row("Group", data.get("group_code"))
    write_row("Status", data.get("status"))
    branch = data.get("branch")
    write_row("Branch", branch.get("name") if isinstance(branch, dict) else branch)
    location = data.get("location")
    write_row("Location", location.get("name") if isinstance(location, dict) else location)
    write_row("Remarks", data.get("remarks"))

    # Contact details
    contact = data.get("contact") or {}
    write_row(
        "Contact - Address Line",
        contact.get("address_line") if isinstance(contact, dict) else contact,
    )
    write_row(
        "Contact - City",
        (
            contact.get("city")["name"]
            if isinstance(contact.get("city"), dict)
            else contact.get("city")
        )
        if isinstance(contact, dict)
        else "",
    )
    write_row(
        "Contact - State",
        (
            contact.get("state")["name"]
            if isinstance(contact.get("state"), dict)
            else contact.get("state")
        )
        if isinstance(contact, dict)
        else "",
    )
    write_row(
        "Contact - Country",
        (
            contact.get("country")["name"]
            if isinstance(contact.get("country"), dict)
            else contact.get("country")
        )
        if isinstance(contact, dict)
        else "",
    )
    write_row("Contact - Phone", contact.get("phone") if isinstance(contact, dict) else "")
    write_row("Contact - Email", contact.get("email") if isinstance(contact, dict) else "")

    # Bank and tax
    bank = data.get("bank") or {}
    write_row("Bank - Name", bank.get("bank_name") if isinstance(bank, dict) else "")
    write_row("Bank - Branch", bank.get("branch") if isinstance(bank, dict) else "")
    write_row("Bank - IFSC", bank.get("ifsc") if isinstance(bank, dict) else "")
    write_row("Bank - Account Number", bank.get("account_number") if isinstance(bank, dict) else "")

    tax = data.get("tax") or {}
    write_row("Tax - GSTIN", tax.get("gstin") if isinstance(tax, dict) else "")
    write_row("Tax - PAN", tax.get("pan") if isinstance(tax, dict) else "")

    # Opening balance
    opening = data.get("opening_balance") or {}
    write_row("Opening - Amount", opening.get("amount") if isinstance(opening, dict) else "")

    # Footer timestamp
    row += 1
    from datetime import datetime

    ws.cell(row=row, column=1, value="Exported On")
    ws.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    # Prepare response
    safe_name = slugify(data.get("account_name") or "account")[:30]
    filename = f"Account_{safe_name}_Details.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _parse_json(body: bytes) -> dict:
    try:
        return json.loads(body.decode("utf-8")) if body else {}
    except (json.JSONDecodeError, UnicodeDecodeError):
        return None


def _normalise_foreign_key_value(value: Optional[object]) -> Optional[object]:
    if isinstance(value, dict):
        return value.get("id") or value.get("pk")
    if isinstance(value, str) and not value.strip():
        return None
    return value


def _apply_fk_alias(payload: dict, canonical: str, aliases: Iterable[str]) -> None:
    for alias in aliases:
        if alias in payload and canonical not in payload:
            payload[canonical] = payload.pop(alias)


def _normalise_account_payload(payload: dict) -> dict:
    if payload is None:
        return None

    payload.pop("company", None)
    payload.pop("created_by", None)

    # top-level foreign keys
    _apply_fk_alias(payload, "branch", ["branch_id", "branchId"])
    _apply_fk_alias(payload, "location", ["location_id", "locationId"])

    if "branch" in payload:
        payload["branch"] = _normalise_foreign_key_value(payload["branch"])
    if "location" in payload:
        payload["location"] = _normalise_foreign_key_value(payload["location"])

    # nested contact
    contact = payload.get("contact")
    if isinstance(contact, dict):
        for key in ("city", "state", "country"):
            _apply_fk_alias(contact, key, [f"{key}_id", f"{key}Id"])
            if key in contact:
                contact[key] = _normalise_foreign_key_value(contact[key])

    return payload


def _serialize_branch(branch: Branch) -> dict:
    return {
        "id": str(branch.id),
        "name": branch.name,
        "code": getattr(branch, "code", None),
    }


def _serialize_location(location: LocationMaster) -> dict:
    return {
        "id": str(location.id),
        "name": location.name,
    }


def _serialize_state(state: StateMaster) -> dict:
    return {
        "id": str(state.id),
        "name": state.name,
        "code": getattr(state, "code", None),
        "country_id": str(state.country_id) if state.country_id else None,
    }


def _serialize_city(city: CityMaster) -> dict:
    return {
        "id": str(city.id),
        "name": city.name,
        "code": getattr(city, "code", None),
        "state_id": str(city.state_id) if city.state_id else None,
        "country_id": str(city.state.country_id) if city.state_id else None,
    }


def _serialize_country(country: CountryMaster) -> dict:
    return {
        "id": str(country.id),
        "name": country.name,
        "code": getattr(country, "code", None),
    }


def _flatten_dict(d: dict, parent_key: str = "", sep: str = ".") -> dict:
    """Flatten nested dictionaries into a single-level dict with joined keys.

    Example: {"contact": {"phone": "123"}} -> {"contact.phone": "123"}
    """
    items = {}
    for k, v in (d or {}).items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            # attempt to collapse simple FK objects to their name if present
            if "name" in v and len(v.keys()) <= 4:
                items[new_key] = v.get("name")
            else:
                items.update(_flatten_dict(v, new_key, sep=sep))
        else:
            items[new_key] = v
    return items


def _excel_safe_value(v):
    """Convert a Python value into something openpyxl can write into a cell.

    - UUIDs -> str
    - dicts -> collapse to name if present else JSON
    - lists/tuples -> JSON
    - bytes -> utf-8 decode if possible else repr
    - Decimals -> float (best-effort)
    - Other non-primitive -> str()
    """
    if v is None:
        return ""
    if isinstance(v, uuid.UUID):
        return str(v)
    if isinstance(v, decimal.Decimal):
        try:
            return float(v)
        except Exception:
            return str(v)
    if isinstance(v, (list, tuple)):
        try:
            return json.dumps(v)
        except Exception:
            return str(v)
    if isinstance(v, dict):
        # prefer human-readable name when available
        if "name" in v:
            return v.get("name")
        try:
            return json.dumps(v)
        except Exception:
            return str(v)
    if isinstance(v, bytes):
        try:
            return v.decode("utf-8")
        except Exception:
            return repr(v)
    if isinstance(v, (str, int, float, bool)):
        return v
    # Fallback: string representation
    return str(v)


@jwt_login_required
@require_http_methods(["GET"])
def accounts_export_all_view(request):
    """Export all accounts for the user's company as an Excel file.

    This view fetches all records (no pagination), flattens nested fields and
    writes a header row with one column per flattened key. For FK objects
    that are represented as dicts (with `name`), the `name` value is used so
    the output is human-readable.
    """
    try:
        user = request.user
        queryset = _base_queryset(user)
        accounts = list(queryset)

        # Serialize and flatten each account
        rows = []
        for acc in accounts:
            data = _serialize(acc)
            flat = _flatten_dict(data)
            rows.append(flat)

        # Determine columns (union of keys)
        cols = []
        seen = set()
        for r in rows:
            for k in r.keys():
                if k not in seen:
                    cols.append(k)
                    seen.add(k)

        # Build workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Accounts"

        # Header row
        for c_idx, col in enumerate(cols, start=1):
            header = col.replace(".", " ").replace("_", " ").title()
            ws.cell(row=1, column=c_idx, value=header)

        # Data rows
        for r_idx, r in enumerate(rows, start=2):
            for c_idx, col in enumerate(cols, start=1):
                raw_val = r.get(col, "")
                val = _excel_safe_value(raw_val)
                ws.cell(row=r_idx, column=c_idx, value=val)

        # Auto-fit column widths (simple heuristic)
        from openpyxl.utils import get_column_letter

        for i, col in enumerate(cols, start=1):
            letter = get_column_letter(i)
            maxlen = len(col)
            for cell in ws[letter]:
                if cell.value is not None:
                    maxlen = max(maxlen, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(100, max(10, maxlen + 2))

        # Footer timestamp
        from datetime import datetime

        footer_row = len(rows) + 3
        ws.cell(row=footer_row, column=1, value="Exported On")
        ws.cell(row=footer_row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

        filename = "accounts_data.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    except Exception as exc:
        # Return JSON error to the frontend to help debugging rather than HTML
        import traceback

        tb = traceback.format_exc()
        return JsonResponse({"detail": str(exc), "trace": tb}, status=500)


@jwt_login_required
@require_http_methods(["GET"])
def subaccounts_export_all_view(request):
    """Export all SubAccount records for the user's company as an Excel file."""
    try:
        user = request.user
        queryset = SubAccount.objects.select_related("account", "account__spark_account_group").order_by("sub_account_name")
        if user.is_authenticated and getattr(user, "company_id", None):
            queryset = queryset.filter(created_by__company_id=user.company_id)
        else:
            queryset = queryset.none()

        rows = []
        for obj in queryset:
            data = SubAccountSerializer(obj).data
            # collapse account_detail/name where possible
            if isinstance(data.get("account_detail"), dict):
                # prefer human-readable account name in the main `account` column
                acct_name = data.get("account_detail").get("name")
                if acct_name:
                    data["account"] = acct_name
                # also provide an explicit flattened key for downstream consumers
                data["account_detail.name"] = acct_name
            rows.append(_flatten_dict(data))

        cols = []
        seen = set()
        for r in rows:
            for k in r.keys():
                if k not in seen:
                    cols.append(k)
                    seen.add(k)

        wb = Workbook()
        ws = wb.active
        ws.title = "SubAccounts"

        for c_idx, col in enumerate(cols, start=1):
            header = col.replace(".", " ").replace("_", " ").title()
            ws.cell(row=1, column=c_idx, value=header)

        for r_idx, r in enumerate(rows, start=2):
            for c_idx, col in enumerate(cols, start=1):
                raw_val = r.get(col, "")
                val = _excel_safe_value(raw_val)
                ws.cell(row=r_idx, column=c_idx, value=val)

        from openpyxl.utils import get_column_letter

        for i, col in enumerate(cols, start=1):
            letter = get_column_letter(i)
            maxlen = len(col)
            for cell in ws[letter]:
                if cell.value is not None:
                    maxlen = max(maxlen, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(100, max(10, maxlen + 2))

        from datetime import datetime

        footer_row = len(rows) + 3
        ws.cell(row=footer_row, column=1, value="Exported On")
        ws.cell(row=footer_row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

        filename = "subaccounts_data.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response
    except Exception as exc:
        import traceback

        tb = traceback.format_exc()
        return JsonResponse({"detail": str(exc), "trace": tb}, status=500)


@api_view(["GET", "POST"])
@authentication_classes([JWTAuthentication, CsrfExemptSessionAuthentication])
@permission_classes([IsAuthenticated])
def accounts_collection_view(request):
    user = request.user
    queryset = _base_queryset(user)

    if request.method == "GET":
        group_code = request.GET.get("group")
        if group_code:
            queryset = queryset.filter(
                Q(group_code=group_code) | Q(ledger_role__iexact=group_code)
            )

        status = request.GET.get("status")
        if status:
            queryset = queryset.filter(status=status)

        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)

        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(
                Q(account_name__icontains=search)
            )

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [_serialize(account) for account in queryset[start:end]]

        payload = {
            "count": total,
            "next": None,
            "previous": None,
            "results": results,
        }
        return Response(payload, status=200)

    payload = _parse_json(request.body)
    if payload is None:
        return Response({"detail": "Invalid JSON"}, status=400)

    payload = _normalise_account_payload(payload)

    # Pass company into serializer context so validate() can check uniqueness
    serializer = AccountSerializer(
        data=payload, context={"company": getattr(user, "company", None)}
    )
    if serializer.is_valid():
        try:
            account = serializer.save(company=getattr(user, "company", None), created_by=user)
        except IntegrityError:
            # Defensive: convert DB integrity errors into a friendly 400
            return Response(
                {"errors": {"account_no": ["Account number already exists for this company"]}},
                status=400,
            )
        return Response(_serialize(account), status=201)

    return Response({"errors": serializer.errors}, status=400)


@api_view(["GET", "PATCH", "DELETE"])
@authentication_classes([JWTAuthentication, CsrfExemptSessionAuthentication])
@permission_classes([IsAuthenticated])
def account_detail_view(request, pk):
    user = request.user
    queryset = _base_queryset(user)
    try:
        account = queryset.get(pk=pk)
    except Account.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return Response(_serialize(account), status=200)

    if request.method == "DELETE":
        account.delete()
        return Response(status=204)

    payload = _parse_json(request.body)
    if payload is None:
        return Response({"detail": "Invalid JSON"}, status=400)

    payload = _normalise_account_payload(payload)

    # Pass company into serializer context so uniqueness validation works on update
    serializer = AccountSerializer(
        account, data=payload, partial=True, context={"company": getattr(user, "company", None)}
    )
    if serializer.is_valid():
        try:
            updated = serializer.save(company=getattr(user, "company", None))
        except IntegrityError:
            return Response(
                {"errors": {"account_no": ["Account number already exists for this company"]}},
                status=400,
            )
        return Response(_serialize(updated), status=200)

    return Response({"errors": serializer.errors}, status=400)


@api_view(["GET"])
@authentication_classes([JWTAuthentication, CsrfExemptSessionAuthentication])
@permission_classes([IsAuthenticated])
def account_balance_view(request, pk):
    """Return the opening balance for a single account."""
    user = request.user
    queryset = _base_queryset(user)
    try:
        account = queryset.get(pk=pk)
    except Account.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)

    ob = getattr(account, "opening_balance", None)
    amount = float(ob.amount) if ob and ob.amount is not None else 0.0
    drcr = (ob.amount_drcr if ob and ob.amount_drcr else "") or ""

    return Response(
        {
            "account_id": str(account.id),
            "account_name": account.account_name,
            "balance": amount,
            "balance_drcr": drcr,
            "balance_display": f"{amount:.2f} {drcr}".strip(),
        },
        status=200,
    )


@api_view(["GET"])
@authentication_classes([JWTAuthentication, CsrfExemptSessionAuthentication])
@permission_classes([IsAuthenticated])
def account_master_list_view(request):
    user = request.user
    company = getattr(user, "company", None)
    if not company:
        return Response(
            {"branches": [], "locations": [], "states": [], "cities": [], "countries": []}
        )

    branches = Branch.objects.filter(company=company).order_by("name")
    locations = LocationMaster.objects.filter(company=company).order_by("name")

    state_id = request.GET.get("state") or request.GET.get("state_id") or request.GET.get("stateId")
    city_filters = {}
    if state_id:
        city_filters["state_id"] = state_id

    states = StateMaster.objects.select_related("country").order_by("name")
    cities = CityMaster.objects.filter(**city_filters).order_by("name")

    # Hardcoded countries list - all 192 UN recognized countries
    hardcoded_countries = [
        {"id": "1", "name": "Afghanistan", "code": "AF"},
        {"id": "2", "name": "Albania", "code": "AL"},
        {"id": "3", "name": "Algeria", "code": "DZ"},
        {"id": "4", "name": "Andorra", "code": "AD"},
        {"id": "5", "name": "Angola", "code": "AO"},
        {"id": "6", "name": "Antigua and Barbuda", "code": "AG"},
        {"id": "7", "name": "Argentina", "code": "AR"},
        {"id": "8", "name": "Armenia", "code": "AM"},
        {"id": "9", "name": "Australia", "code": "AU"},
        {"id": "10", "name": "Austria", "code": "AT"},
        {"id": "11", "name": "Azerbaijan", "code": "AZ"},
        {"id": "12", "name": "Bahamas", "code": "BS"},
        {"id": "13", "name": "Bahrain", "code": "BH"},
        {"id": "14", "name": "Bangladesh", "code": "BD"},
        {"id": "15", "name": "Barbados", "code": "BB"},
        {"id": "16", "name": "Belarus", "code": "BY"},
        {"id": "17", "name": "Belgium", "code": "BE"},
        {"id": "18", "name": "Belize", "code": "BZ"},
        {"id": "19", "name": "Benin", "code": "BJ"},
        {"id": "20", "name": "Bhutan", "code": "BT"},
        {"id": "21", "name": "Bolivia", "code": "BO"},
        {"id": "22", "name": "Bosnia and Herzegovina", "code": "BA"},
        {"id": "23", "name": "Botswana", "code": "BW"},
        {"id": "24", "name": "Brazil", "code": "BR"},
        {"id": "25", "name": "Brunei", "code": "BN"},
        {"id": "26", "name": "Bulgaria", "code": "BG"},
        {"id": "27", "name": "Burkina Faso", "code": "BF"},
        {"id": "28", "name": "Burundi", "code": "BI"},
        {"id": "29", "name": "Cabo Verde", "code": "CV"},
        {"id": "30", "name": "Cambodia", "code": "KH"},
        {"id": "31", "name": "Cameroon", "code": "CM"},
        {"id": "32", "name": "Canada", "code": "CA"},
        {"id": "33", "name": "Central African Republic", "code": "CF"},
        {"id": "34", "name": "Chad", "code": "TD"},
        {"id": "35", "name": "Chile", "code": "CL"},
        {"id": "36", "name": "China", "code": "CN"},
        {"id": "37", "name": "Colombia", "code": "CO"},
        {"id": "38", "name": "Comoros", "code": "KM"},
        {"id": "39", "name": "Congo", "code": "CG"},
        {"id": "40", "name": "Costa Rica", "code": "CR"},
        {"id": "41", "name": "Croatia", "code": "HR"},
        {"id": "42", "name": "Cuba", "code": "CU"},
        {"id": "43", "name": "Cyprus", "code": "CY"},
        {"id": "44", "name": "Czech Republic", "code": "CZ"},
        {"id": "45", "name": "Democratic Republic of the Congo", "code": "CD"},
        {"id": "46", "name": "Denmark", "code": "DK"},
        {"id": "47", "name": "Djibouti", "code": "DJ"},
        {"id": "48", "name": "Dominica", "code": "DM"},
        {"id": "49", "name": "Dominican Republic", "code": "DO"},
        {"id": "50", "name": "Ecuador", "code": "EC"},
        {"id": "51", "name": "Egypt", "code": "EG"},
        {"id": "52", "name": "El Salvador", "code": "SV"},
        {"id": "53", "name": "Equatorial Guinea", "code": "GQ"},
        {"id": "54", "name": "Eritrea", "code": "ER"},
        {"id": "55", "name": "Estonia", "code": "EE"},
        {"id": "56", "name": "Eswatini", "code": "SZ"},
        {"id": "57", "name": "Ethiopia", "code": "ET"},
        {"id": "58", "name": "Fiji", "code": "FJ"},
        {"id": "59", "name": "Finland", "code": "FI"},
        {"id": "60", "name": "France", "code": "FR"},
        {"id": "61", "name": "Gabon", "code": "GA"},
        {"id": "62", "name": "Gambia", "code": "GM"},
        {"id": "63", "name": "Georgia", "code": "GE"},
        {"id": "64", "name": "Germany", "code": "DE"},
        {"id": "65", "name": "Ghana", "code": "GH"},
        {"id": "66", "name": "Greece", "code": "GR"},
        {"id": "67", "name": "Grenada", "code": "GD"},
        {"id": "68", "name": "Guatemala", "code": "GT"},
        {"id": "69", "name": "Guinea", "code": "GN"},
        {"id": "70", "name": "Guinea-Bissau", "code": "GW"},
        {"id": "71", "name": "Guyana", "code": "GY"},
        {"id": "72", "name": "Haiti", "code": "HT"},
        {"id": "73", "name": "Honduras", "code": "HN"},
        {"id": "74", "name": "Hungary", "code": "HU"},
        {"id": "75", "name": "Iceland", "code": "IS"},
        {"id": "76", "name": "India", "code": "IN"},
        {"id": "77", "name": "Indonesia", "code": "ID"},
        {"id": "78", "name": "Iran", "code": "IR"},
        {"id": "79", "name": "Iraq", "code": "IQ"},
        {"id": "80", "name": "Ireland", "code": "IE"},
        {"id": "81", "name": "Israel", "code": "IL"},
        {"id": "82", "name": "Italy", "code": "IT"},
        {"id": "83", "name": "Ivory Coast", "code": "CI"},
        {"id": "84", "name": "Jamaica", "code": "JM"},
        {"id": "85", "name": "Japan", "code": "JP"},
        {"id": "86", "name": "Jordan", "code": "JO"},
        {"id": "87", "name": "Kazakhstan", "code": "KZ"},
        {"id": "88", "name": "Kenya", "code": "KE"},
        {"id": "89", "name": "Kiribati", "code": "KI"},
        {"id": "90", "name": "Kuwait", "code": "KW"},
        {"id": "91", "name": "Kyrgyzstan", "code": "KG"},
        {"id": "92", "name": "Laos", "code": "LA"},
        {"id": "93", "name": "Latvia", "code": "LV"},
        {"id": "94", "name": "Lebanon", "code": "LB"},
        {"id": "95", "name": "Lesotho", "code": "LS"},
        {"id": "96", "name": "Liberia", "code": "LR"},
        {"id": "97", "name": "Libya", "code": "LY"},
        {"id": "98", "name": "Liechtenstein", "code": "LI"},
        {"id": "99", "name": "Lithuania", "code": "LT"},
        {"id": "100", "name": "Luxembourg", "code": "LU"},
        {"id": "101", "name": "Madagascar", "code": "MG"},
        {"id": "102", "name": "Malawi", "code": "MW"},
        {"id": "103", "name": "Malaysia", "code": "MY"},
        {"id": "104", "name": "Maldives", "code": "MV"},
        {"id": "105", "name": "Mali", "code": "ML"},
        {"id": "106", "name": "Malta", "code": "MT"},
        {"id": "107", "name": "Marshall Islands", "code": "MH"},
        {"id": "108", "name": "Mauritania", "code": "MR"},
        {"id": "109", "name": "Mauritius", "code": "MU"},
        {"id": "110", "name": "Mexico", "code": "MX"},
        {"id": "111", "name": "Micronesia", "code": "FM"},
        {"id": "112", "name": "Moldova", "code": "MD"},
        {"id": "113", "name": "Monaco", "code": "MC"},
        {"id": "114", "name": "Mongolia", "code": "MN"},
        {"id": "115", "name": "Montenegro", "code": "ME"},
        {"id": "116", "name": "Morocco", "code": "MA"},
        {"id": "117", "name": "Mozambique", "code": "MZ"},
        {"id": "118", "name": "Myanmar", "code": "MM"},
        {"id": "119", "name": "Namibia", "code": "NA"},
        {"id": "120", "name": "Nauru", "code": "NR"},
        {"id": "121", "name": "Nepal", "code": "NP"},
        {"id": "122", "name": "Netherlands", "code": "NL"},
        {"id": "123", "name": "New Zealand", "code": "NZ"},
        {"id": "124", "name": "Nicaragua", "code": "NI"},
        {"id": "125", "name": "Niger", "code": "NE"},
        {"id": "126", "name": "Nigeria", "code": "NG"},
        {"id": "127", "name": "North Korea", "code": "KP"},
        {"id": "128", "name": "North Macedonia", "code": "MK"},
        {"id": "129", "name": "Norway", "code": "NO"},
        {"id": "130", "name": "Oman", "code": "OM"},
        {"id": "131", "name": "Pakistan", "code": "PK"},
        {"id": "132", "name": "Palau", "code": "PW"},
        {"id": "133", "name": "Palestine", "code": "PS"},
        {"id": "134", "name": "Panama", "code": "PA"},
        {"id": "135", "name": "Papua New Guinea", "code": "PG"},
        {"id": "136", "name": "Paraguay", "code": "PY"},
        {"id": "137", "name": "Peru", "code": "PE"},
        {"id": "138", "name": "Philippines", "code": "PH"},
        {"id": "139", "name": "Poland", "code": "PL"},
        {"id": "140", "name": "Portugal", "code": "PT"},
        {"id": "141", "name": "Qatar", "code": "QA"},
        {"id": "142", "name": "Romania", "code": "RO"},
        {"id": "143", "name": "Russia", "code": "RU"},
        {"id": "144", "name": "Rwanda", "code": "RW"},
        {"id": "145", "name": "Saint Kitts and Nevis", "code": "KN"},
        {"id": "146", "name": "Saint Lucia", "code": "LC"},
        {"id": "147", "name": "Saint Vincent and the Grenadines", "code": "VC"},
        {"id": "148", "name": "Samoa", "code": "WS"},
        {"id": "149", "name": "San Marino", "code": "SM"},
        {"id": "150", "name": "Sao Tome and Principe", "code": "ST"},
        {"id": "151", "name": "Saudi Arabia", "code": "SA"},
        {"id": "152", "name": "Senegal", "code": "SN"},
        {"id": "153", "name": "Serbia", "code": "RS"},
        {"id": "154", "name": "Seychelles", "code": "SC"},
        {"id": "155", "name": "Sierra Leone", "code": "SL"},
        {"id": "156", "name": "Singapore", "code": "SG"},
        {"id": "157", "name": "Slovakia", "code": "SK"},
        {"id": "158", "name": "Slovenia", "code": "SI"},
        {"id": "159", "name": "Solomon Islands", "code": "SB"},
        {"id": "160", "name": "Somalia", "code": "SO"},
        {"id": "161", "name": "South Africa", "code": "ZA"},
        {"id": "162", "name": "South Korea", "code": "KR"},
        {"id": "163", "name": "South Sudan", "code": "SS"},
        {"id": "164", "name": "Spain", "code": "ES"},
        {"id": "165", "name": "Sri Lanka", "code": "LK"},
        {"id": "166", "name": "Sudan", "code": "SD"},
        {"id": "167", "name": "Suriname", "code": "SR"},
        {"id": "168", "name": "Sweden", "code": "SE"},
        {"id": "169", "name": "Switzerland", "code": "CH"},
        {"id": "170", "name": "Syria", "code": "SY"},
        {"id": "171", "name": "Taiwan", "code": "TW"},
        {"id": "172", "name": "Tajikistan", "code": "TJ"},
        {"id": "173", "name": "Tanzania", "code": "TZ"},
        {"id": "174", "name": "Thailand", "code": "TH"},
        {"id": "175", "name": "Timor-Leste", "code": "TL"},
        {"id": "176", "name": "Togo", "code": "TG"},
        {"id": "177", "name": "Tonga", "code": "TO"},
        {"id": "178", "name": "Trinidad and Tobago", "code": "TT"},
        {"id": "179", "name": "Tunisia", "code": "TN"},
        {"id": "180", "name": "Turkey", "code": "TR"},
        {"id": "181", "name": "Turkmenistan", "code": "TM"},
        {"id": "182", "name": "Tuvalu", "code": "TV"},
        {"id": "183", "name": "Uganda", "code": "UG"},
        {"id": "184", "name": "Ukraine", "code": "UA"},
        {"id": "185", "name": "United Arab Emirates", "code": "AE"},
        {"id": "186", "name": "United Kingdom", "code": "GB"},
        {"id": "187", "name": "United States", "code": "US"},
        {"id": "188", "name": "Uruguay", "code": "UY"},
        {"id": "189", "name": "Uzbekistan", "code": "UZ"},
        {"id": "190", "name": "Vanuatu", "code": "VU"},
        {"id": "191", "name": "Vatican City", "code": "VA"},
        {"id": "192", "name": "Venezuela", "code": "VE"},
        {"id": "193", "name": "Vietnam", "code": "VN"},
        {"id": "194", "name": "Yemen", "code": "YE"},
        {"id": "195", "name": "Zambia", "code": "ZM"},
        {"id": "196", "name": "Zimbabwe", "code": "ZW"},
    ]

    payload = {
        "branches": [_serialize_branch(b) for b in branches],
        "locations": [_serialize_location(loc) for loc in locations],
        "states": [_serialize_state(state) for state in states],
        "cities": [_serialize_city(city) for city in cities],
        "countries": hardcoded_countries,  # Use hardcoded list instead of database
    }

    return Response(payload, status=200)


@api_view(["GET"])
@authentication_classes([JWTAuthentication, CsrfExemptSessionAuthentication])
@permission_classes([IsAuthenticated])
def accounts_dropdown_view(request):
    """Return minimal list of accounts for dropdowns in frontend forms.

    Optional query params:
      ?type=contra  → only Cash and Bank accounts (account_name contains 'cash' or 'bank')
      ?group=CASH   → filter by group_code (existing behaviour)
    """
    user = request.user
    queryset = _base_queryset(user)

    # Contra vouchers only allow Cash / Bank accounts
    if request.GET.get("type") == "contra":
        queryset = queryset.filter(group_code__in=["CASH", "BANK"])

    accounts = queryset.order_by("account_name").values("id", "account_name")
    data = [{"id": str(a["id"]), "name": a["account_name"]} for a in accounts]
    return Response({"accounts": data}, status=200)


@api_view(["GET", "POST"])
@authentication_classes([JWTAuthentication, CsrfExemptSessionAuthentication])
@permission_classes([IsAuthenticated])
def subaccounts_collection_view(request):
    user = request.user
    queryset = SubAccount.objects.select_related("account", "account__spark_account_group").order_by("-created_at")
    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(created_by__company_id=user.company_id)
    else:
        queryset = queryset.none()

    if request.method == "GET":
        # Filter by account_id if provided
        account_id = request.GET.get("account_id")
        if account_id:
            # Guard against non-UUID values (e.g. account name sent by mistake)
            import uuid as _uuid
            try:
                _uuid.UUID(str(account_id))
                queryset = queryset.filter(account_id=account_id)
            except (ValueError, AttributeError):
                queryset = queryset.none()

        # allow client to filter by search term (sub account, item name or parent account name)
        search = request.GET.get("search")
        if search:
            queryset = queryset.filter(
                Q(sub_account_name__icontains=search)
                | Q(item_name__icontains=search)
                | Q(account__account_name__icontains=search)
            )

        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [SubAccountSerializer(obj).data for obj in queryset[start:end]]

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return Response(payload, status=200)

    payload = _parse_json(request.body)
    if payload is None:
        return Response({"detail": "Invalid JSON"}, status=400)

    # normalize account foreign key
    acc = payload.get("account")
    if isinstance(acc, dict):
        payload["account"] = acc.get("id") or acc.get("pk")

    serializer = SubAccountSerializer(data=payload)
    if serializer.is_valid():
        obj = serializer.save(created_by=user)
        return Response(SubAccountSerializer(obj).data, status=201)

    return Response({"errors": serializer.errors}, status=400)


@api_view(["GET", "PATCH", "DELETE"])
@authentication_classes([JWTAuthentication, CsrfExemptSessionAuthentication])
@permission_classes([IsAuthenticated])
def subaccount_detail_view(request, pk):
    user = request.user
    queryset = SubAccount.objects.select_related("account", "account__spark_account_group").order_by("-created_at")
    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(created_by__company_id=user.company_id)
    else:
        queryset = queryset.none()

    try:
        obj = queryset.get(pk=pk)
    except SubAccount.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return Response(SubAccountSerializer(obj).data, status=200)

    if request.method == "DELETE":
        obj.delete()
        return Response(status=204)

    payload = _parse_json(request.body)
    if payload is None:
        return Response({"detail": "Invalid JSON"}, status=400)

    acc = payload.get("account")
    if isinstance(acc, dict):
        payload["account"] = acc.get("id") or acc.get("pk")

    serializer = SubAccountSerializer(obj, data=payload, partial=True)
    if serializer.is_valid():
        updated = serializer.save()
        return Response(SubAccountSerializer(updated).data, status=200)

    return Response({"errors": serializer.errors}, status=400)


@jwt_login_required
@require_http_methods(["GET"])
def subaccount_export_view(request, pk):
    """Export a single SubAccount record as an Excel file (field | value)."""
    user = request.user
    queryset = SubAccount.objects.select_related("account", "account__spark_account_group").order_by("-created_at")
    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(created_by__company_id=user.company_id)
    else:
        queryset = queryset.none()

    try:
        obj = queryset.get(pk=pk)
    except SubAccount.DoesNotExist:
        return JsonResponse({"detail": "Not found"}, status=404)

    data = SubAccountSerializer(obj).data

    from openpyxl import Workbook
    from django.utils.text import slugify

    wb = Workbook()
    ws = wb.active
    ws.title = "SubAccount Details"
    row = 1

    def write(label, value):
        nonlocal row
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value if value is not None else "")
        row += 1

    write(
        "Account",
        (data.get("account_detail") or {}).get("name")
        if isinstance(data.get("account_detail"), dict)
        else data.get("account"),
    )
    write("Sub Account Name", data.get("sub_account_name"))
    write("Item Name", data.get("item_name"))
    write("Address", data.get("address"))
    write("Phone Number", data.get("phone_number"))
    write("Ring Size", data.get("ring_size"))
    write("Bangle Size", data.get("bangle_size"))
    write("Gender", data.get("gender"))

    # footer
    from datetime import datetime

    row += 1
    ws.cell(row=row, column=1, value="Exported On")
    ws.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    safe_name = slugify(data.get("sub_account_name") or "subaccount")[:30]
    filename = f"SubAccount_{safe_name}_Details.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@api_view(["GET"])
@authentication_classes([JWTAuthentication, CsrfExemptSessionAuthentication])
@permission_classes([IsAuthenticated])
def ac_group_master_list_view(request):
    """Return available A/C Group master options (editable in admin)."""
    masters = ACGroupMaster.objects.order_by("name")
    data = [ACGroupMasterSerializer(m).data for m in masters]
    return Response({"ac_groups": data}, status=200)


def _is_admin(user):
    """Check if user has admin privileges for Account Group management."""
    if not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    if getattr(user, "role", None) in ("SUPER_ADMIN", "ADMIN"):
        return True
    return False


@api_view(["GET", "POST"])
@authentication_classes([JWTAuthentication, CsrfExemptSessionAuthentication])
@permission_classes([IsAuthenticated])
def ac_groups_collection_view(request):
    user = request.user
    queryset = ACGroup.objects.select_related("ac_group").order_by("-created_at")
    # scope by user's company if possible via created_by relation
    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(created_by__company_id=user.company_id)
    else:
        queryset = queryset.none()

    if request.method == "GET":
        page_size = int(request.GET.get("page_size", 50))
        page = int(request.GET.get("page", 1)) if request.GET.get("page") else 1
        page = page if page > 0 else 1
        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = [ACGroupSerializer(obj).data for obj in queryset[start:end]]

        payload = {"count": total, "next": None, "previous": None, "results": results}
        return Response(payload, status=200)

    # POST - Only admin users can create Account Groups
    if not _is_admin(user):
        return Response({"detail": "Only admin users can create Account Groups."}, status=403)

    payload = _parse_json(request.body)
    if payload is None:
        return Response({"detail": "Invalid JSON"}, status=400)

    # accept ac_group as id or dict
    ac_group = payload.get("ac_group") or payload.get("ac_group_id")
    if isinstance(ac_group, dict):
        ac_group = ac_group.get("id") or ac_group.get("pk")
    payload["ac_group_id"] = ac_group

    serializer = ACGroupSerializer(data=payload)
    if serializer.is_valid():
        obj = serializer.save(created_by=user)
        return Response(ACGroupSerializer(obj).data, status=201)

    return Response({"errors": serializer.errors}, status=400)


@api_view(["GET", "PATCH", "DELETE"])
@authentication_classes([JWTAuthentication, CsrfExemptSessionAuthentication])
@permission_classes([IsAuthenticated])
def ac_group_detail_view(request, pk):
    user = request.user
    queryset = ACGroup.objects.select_related("ac_group").order_by("-created_at")
    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(created_by__company_id=user.company_id)
    else:
        queryset = queryset.none()

    try:
        obj = queryset.get(pk=pk)
    except ACGroup.DoesNotExist:
        return Response({"detail": "Not found"}, status=404)

    if request.method == "GET":
        return Response(ACGroupSerializer(obj).data, status=200)

    # DELETE and PATCH - Only admin users can modify Account Groups
    if not _is_admin(user):
        return Response({"detail": "Only admin users can modify Account Groups."}, status=403)

    if request.method == "DELETE":
        obj.delete()
        return Response(status=204)

    payload = _parse_json(request.body)
    if payload is None:
        return Response({"detail": "Invalid JSON"}, status=400)

    ac_group = payload.get("ac_group") or payload.get("ac_group_id")
    if isinstance(ac_group, dict):
        ac_group = ac_group.get("id") or ac_group.get("pk")
    if ac_group is not None:
        payload["ac_group_id"] = ac_group

    serializer = ACGroupSerializer(obj, data=payload, partial=True)
    if serializer.is_valid():
        updated = serializer.save()
        return Response(ACGroupSerializer(updated).data, status=200)

    return Response({"errors": serializer.errors}, status=400)


@api_view(["GET"])
@authentication_classes([JWTAuthentication, CsrfExemptSessionAuthentication])
@permission_classes([IsAuthenticated])
def account_transactions_list_view(request):
    """Return a unified list of all transactions linked to Account Masters.

    Aggregates: Orders, Sales, Payments, Journals, Receipts, Purchases,
    Approvals, PurReturns, SalesReturns, Receives, Repairs.

    Query params:
    - start_date / end_date: YYYY-MM-DD (filter by created_at)
    - account_id: filter to a specific account
    - type: filter by transaction type (order, sale, payment, journal, receipt, etc.)
    - search: search by account name or reference number
    - page / page_size: pagination
    """
    from datetime import datetime
    from vouchers.models import (
        Order, Sale, PaymentEntry, JournalEntry, Receipt,
        PurchaseM, PurchaseTagwiseM, PurchaseDiamond,
        ApprovalLooseM, ApprovalTagM, PurAndApprovalM,
        PurReturn, SalesReturn, Receive, Repair,
        PaymentVoucher, JournalVoucher,
    )

    user = request.user
    company_id = getattr(user, "company_id", None)
    if not company_id:
        return Response({"count": 0, "results": []}, status=200)

    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")
    account_id = request.GET.get("account_id")
    txn_type = request.GET.get("type")
    search = request.GET.get("search")
    page_size = int(request.GET.get("page_size", 50))
    page = max(int(request.GET.get("page", 1)), 1)

    start_date = None
    end_date = None
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        except ValueError:
            pass
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        except ValueError:
            pass

    def _apply_filters(qs, date_field="date", account_field="account"):
        """Filter queryset. date_field should be the transaction date (e.g. 'date').
        For models without a date field, pass 'created_at'."""
        if start_date:
            if date_field == "created_at":
                qs = qs.filter(created_at__date__gte=start_date)
            else:
                qs = qs.filter(**{f"{date_field}__gte": start_date})
        if end_date:
            if date_field == "created_at":
                qs = qs.filter(created_at__date__lte=end_date)
            else:
                qs = qs.filter(**{f"{date_field}__lte": end_date})
        if account_id:
            qs = qs.filter(**{f"{account_field}_id": account_id})
        if search:
            qs = qs.filter(
                Q(**{f"{account_field}__account_name__icontains": search})
                | Q(**{f"{account_field}__account_no__icontains": search})
            )
        return qs

    # Collect all transactions
    txn_configs = [
        {
            "key": "order",
            "model": Order,
            "qs": Order.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"),
            "date_field": "date",
            "account_field": "account",
            "ref_fn": lambda o: o.bill_no or "",
            "date_fn": lambda o: o.date.isoformat() if o.date else (o.created_at.strftime("%Y-%m-%d") if o.created_at else ""),
            "amount_fn": lambda o: float(o.gold_rate or 0),
            "extra_fn": lambda o: {"item_name": o.item_name or ""},
        },
        {
            "key": "sale",
            "model": Sale,
            "qs": Sale.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"),
            "date_field": "created_at",
            "account_field": "account",
            "ref_fn": lambda o: o.order_no or o.tag_no or "",
            "date_fn": lambda o: o.created_at.strftime("%Y-%m-%d") if o.created_at else "",
            "amount_fn": lambda o: float(o.rate or 0),
            "extra_fn": lambda o: {"item_name": o.item_name or ""},
        },
        {
            "key": "payment",
            "model": PaymentEntry,
            "qs": PaymentEntry.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"),
            "date_field": "date",
            "account_field": "account",
            "ref_fn": lambda o: str(o.id)[:8],
            "date_fn": lambda o: o.date.isoformat() if o.date else (o.created_at.strftime("%Y-%m-%d") if o.created_at else ""),
            "amount_fn": lambda o: float(o.dr or o.cr or o.total or 0),
            "extra_fn": lambda o: {"type": o.type or "", "narration": o.narration or ""},
        },
        {
            "key": "journal",
            "model": JournalEntry,
            "qs": JournalEntry.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"),
            "date_field": "date",
            "account_field": "account",
            "ref_fn": lambda o: str(o.id)[:8],
            "date_fn": lambda o: o.date.isoformat() if o.date else (o.created_at.strftime("%Y-%m-%d") if o.created_at else ""),
            "amount_fn": lambda o: float(o.dr or o.cr or o.total or 0),
            "extra_fn": lambda o: {"type": o.type or "", "narration": o.narration or ""},
        },
        {
            "key": "receipt",
            "model": Receipt,
            "qs": Receipt.objects.filter(company_id=company_id).select_related("party_name", "party_name__spark_account_group"),
            "date_field": "date",
            "account_field": "party_name",
            "ref_fn": lambda o: o.voucher_no or str(o.id)[:8],
            "date_fn": lambda o: o.date.isoformat() if o.date else (o.created_at.strftime("%Y-%m-%d") if o.created_at else ""),
            "amount_fn": lambda o: float(o.cr or o.dr or 0),
            "extra_fn": lambda o: {"narration": o.narration or ""},
        },
        {
            "key": "purchase",
            "model": PurchaseM,
            "qs": PurchaseM.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"),
            "date_field": "date",
            "account_field": "account",
            "ref_fn": lambda o: o.bill_no or o.order_no or "",
            "date_fn": lambda o: o.date.isoformat() if o.date else (o.created_at.strftime("%Y-%m-%d") if o.created_at else ""),
            "amount_fn": lambda o: float(o.rate or 0),
            "extra_fn": lambda o: {},
        },
        {
            "key": "purchase_tagwise",
            "model": PurchaseTagwiseM,
            "qs": PurchaseTagwiseM.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"),
            "date_field": "date",
            "account_field": "account",
            "ref_fn": lambda o: o.bill_no or o.order_no or "",
            "date_fn": lambda o: o.date.isoformat() if o.date else (o.created_at.strftime("%Y-%m-%d") if o.created_at else ""),
            "amount_fn": lambda o: float(o.rate or 0),
            "extra_fn": lambda o: {},
        },
        {
            "key": "purchase_diamond",
            "model": PurchaseDiamond,
            "qs": PurchaseDiamond.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"),
            "date_field": "created_at",
            "account_field": "account",
            "ref_fn": lambda o: o.batch or "",
            "date_fn": lambda o: o.created_at.strftime("%Y-%m-%d") if o.created_at else "",
            "amount_fn": lambda o: float(o.ex_rate or 0),
            "extra_fn": lambda o: {},
        },
        {
            "key": "approval_loose",
            "model": ApprovalLooseM,
            "qs": ApprovalLooseM.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"),
            "date_field": "created_at",
            "account_field": "account",
            "ref_fn": lambda o: o.order_number or "",
            "date_fn": lambda o: o.created_at.strftime("%Y-%m-%d") if o.created_at else "",
            "amount_fn": lambda o: float(o.rate or 0),
            "extra_fn": lambda o: {},
        },
        {
            "key": "approval_tag",
            "model": ApprovalTagM,
            "qs": ApprovalTagM.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"),
            "date_field": "created_at",
            "account_field": "account",
            "ref_fn": lambda o: o.order_number or "",
            "date_fn": lambda o: o.created_at.strftime("%Y-%m-%d") if o.created_at else "",
            "amount_fn": lambda o: float(o.rate or 0),
            "extra_fn": lambda o: {},
        },
        {
            "key": "pur_approval",
            "model": PurAndApprovalM,
            "qs": PurAndApprovalM.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"),
            "date_field": "created_at",
            "account_field": "account",
            "ref_fn": lambda o: o.order_no or o.tag_no or "",
            "date_fn": lambda o: o.created_at.strftime("%Y-%m-%d") if o.created_at else "",
            "amount_fn": lambda o: float(o.rate or 0),
            "extra_fn": lambda o: {},
        },
        {
            "key": "pur_return",
            "model": PurReturn,
            "qs": PurReturn.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"),
            "date_field": "date",
            "account_field": "account",
            "ref_fn": lambda o: o.order_no or o.tag_no or "",
            "date_fn": lambda o: o.date.isoformat() if o.date else (o.created_at.strftime("%Y-%m-%d") if o.created_at else ""),
            "amount_fn": lambda o: float(o.rate or 0),
            "extra_fn": lambda o: {},
        },
        {
            "key": "sales_return",
            "model": SalesReturn,
            "qs": SalesReturn.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"),
            "date_field": "date",
            "account_field": "account",
            "ref_fn": lambda o: o.order_no or o.tag_no or "",
            "date_fn": lambda o: o.date.isoformat() if o.date else (o.created_at.strftime("%Y-%m-%d") if o.created_at else ""),
            "amount_fn": lambda o: float(o.rate or 0),
            "extra_fn": lambda o: {},
        },
        {
            "key": "receive",
            "model": Receive,
            "qs": Receive.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"),
            "date_field": "date",
            "account_field": "account",
            "ref_fn": lambda o: o.tag_no or "",
            "date_fn": lambda o: o.date.isoformat() if o.date else (o.created_at.strftime("%Y-%m-%d") if o.created_at else ""),
            "amount_fn": lambda o: float(o.total or o.rate or 0),
            "extra_fn": lambda o: {},
        },
        {
            "key": "repair",
            "model": Repair,
            "qs": Repair.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"),
            "date_field": "date",
            "account_field": "account",
            "ref_fn": lambda o: o.tag_no or "",
            "date_fn": lambda o: o.date.isoformat() if o.date else (o.created_at.strftime("%Y-%m-%d") if o.created_at else ""),
            "amount_fn": lambda o: float(o.total or o.rate or 0),
            "extra_fn": lambda o: {"type": o.type or ""},
        },
        {
            "key": "payment_voucher",
            "model": PaymentVoucher,
            "qs": PaymentVoucher.objects.filter(company_id=company_id).select_related("party_name", "party_name__spark_account_group"),
            "date_field": "date",
            "account_field": "party_name",
            "ref_fn": lambda o: o.voucher_no or str(o.id)[:8],
            "date_fn": lambda o: o.date.isoformat() if o.date else (o.created_at.strftime("%Y-%m-%d") if o.created_at else ""),
            "amount_fn": lambda o: float(o.dr or 0),
            "extra_fn": lambda o: {"narration": o.narration or ""},
        },
        {
            "key": "journal_voucher",
            "model": JournalVoucher,
            "qs": JournalVoucher.objects.filter(company_id=company_id).select_related("party_name", "party_name__spark_account_group"),
            "date_field": "date",
            "account_field": "party_name",
            "ref_fn": lambda o: o.voucher_no or str(o.id)[:8],
            "date_fn": lambda o: o.date.isoformat() if o.date else (o.created_at.strftime("%Y-%m-%d") if o.created_at else ""),
            "amount_fn": lambda o: float(o.dr or 0),
            "extra_fn": lambda o: {"narration": o.narration or ""},
        },
    ]

    # Filter by type if requested
    if txn_type:
        txn_configs = [c for c in txn_configs if c["key"] == txn_type]

    all_txns = []
    for cfg in txn_configs:
        qs = _apply_filters(cfg["qs"], date_field=cfg.get("date_field", "date"), account_field=cfg["account_field"])
        for obj in qs:
            acct = getattr(obj, cfg["account_field"], None)
            tax = getattr(acct, "tax", None) if acct else None
            opening = getattr(acct, "opening_balance", None) if acct else None
            all_txns.append({
                "id": str(obj.id),
                "transaction_type": cfg["key"],
                "reference_no": cfg["ref_fn"](obj),
                "date": cfg["date_fn"](obj),
                "account_id": str(acct.id) if acct else None,
                "account_name": acct.account_name if acct else "N/A",
                "account_no": acct.account_no if acct else "",
                "account_group_name": acct.spark_account_group.name if acct and acct.spark_account_group else "",
                "tally_parent_group": acct.tally_parent_group if acct else "",
                "financial_statement": acct.financial_statement if acct else "",
                "normal_balance": acct.normal_balance if acct else "",
                "gst_registration_type": acct.gst_registration_type if acct else "",
                "gstin": tax.gstin if tax else "",
                "pan": tax.pan if tax else "",
                "bill_wise_required": acct.bill_wise_required if acct else "",
                "cost_centre_required": acct.cost_centre_required if acct else "",
                "ledger_role": acct.ledger_role if acct else "",
                "notes": acct.remarks if acct else "",
                "opening_balance": float(opening.amount) if opening else "",
                "amount": cfg["amount_fn"](obj),
                "created_at": obj.created_at.isoformat() if obj.created_at else "",
                **cfg["extra_fn"](obj),
            })

    # Sort by created_at descending
    all_txns.sort(key=lambda t: t.get("created_at", ""), reverse=True)

    total = len(all_txns)
    start = (page - 1) * page_size
    end = start + page_size
    results = all_txns[start:end]

    return Response({
        "count": total,
        "results": results,
    }, status=200)


@api_view(["GET"])
@authentication_classes([JWTAuthentication, CsrfExemptSessionAuthentication])
@permission_classes([IsAuthenticated])
def account_transactions_tally_export_view(request):
    """Export Account Master transactions in Tally-compatible Excel format.

    Query params (all optional):
    - start_date (YYYY-MM-DD) - filter from this date onwards
    - end_date   (YYYY-MM-DD) - filter up to this date
    - account_id             - filter by specific account
    If no dates are provided, all transactions are exported.
    """
    from datetime import datetime
    from openpyxl import Workbook
    from vouchers.models import (
        Order, Sale, PaymentEntry, JournalEntry, Receipt,
        PurchaseM, PurchaseTagwiseM, PurchaseDiamond,
        ApprovalLooseM, ApprovalTagM, PurAndApprovalM,
        PurReturn, SalesReturn, Receive, Repair,
        PaymentVoucher, JournalVoucher,
    )

    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")

    start_date = None
    end_date = None

    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        except ValueError:
            return Response({"detail": "Invalid start_date format. Use YYYY-MM-DD"}, status=400)

    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        except ValueError:
            return Response({"detail": "Invalid end_date format. Use YYYY-MM-DD"}, status=400)

    if start_date and end_date and start_date > end_date:
        return Response({"detail": "start_date cannot be after end_date"}, status=400)

    user = request.user
    company_id = getattr(user, "company_id", None)
    if not company_id:
        return Response({"detail": "No company associated"}, status=400)

    account_id = request.GET.get("account_id")
    tally_parent_group = request.GET.get("tally_parent_group")

    def _filter(qs, account_field="account", date_field="date"):
        """Optionally filter by date range, account, and tally parent group."""
        if start_date:
            if date_field == "created_at":
                qs = qs.filter(created_at__date__gte=start_date)
            else:
                qs = qs.filter(**{f"{date_field}__gte": start_date})
        if end_date:
            if date_field == "created_at":
                qs = qs.filter(created_at__date__lte=end_date)
            else:
                qs = qs.filter(**{f"{date_field}__lte": end_date})
        if account_id:
            qs = qs.filter(**{f"{account_field}_id": account_id})
        if tally_parent_group:
            qs = qs.filter(**{f"{account_field}__tally_parent_group": tally_parent_group})
        return qs

    # Helper to get FK display name safely
    def _fk_name(obj, field):
        rel = getattr(obj, field, None)
        if rel is None:
            return ""
        return getattr(rel, "name", "") or str(rel)

    def _acct_fields(acct):
        """Return dict of Account Master fields for the linked account."""
        tax = getattr(acct, "tax", None) if acct else None
        opening = getattr(acct, "opening_balance", None) if acct else None
        if not acct:
            return {
                "Export Action": "",
                "Tally Ledger Name": "",
                "Tally Parent Group": "",
                "Opening Balance": "",
                "Dr / Cr": "",
                "GST Registration Type": "",
                "GSTIN": "",
                "PAN": "",
                "Bill-wise Required?": "",
                "Cost Centre Required?": "",
                "Spark Account Code": "",
                "Spark Account Group": "",
                "Ledger Role": "",
                "Notes": "",
            }
        return {
            "Export Action": "Create" if acct.export_to_tally == "YES" else "Ignore",
            "Tally Ledger Name": acct.tally_ledger_name_override or acct.account_name or "",
            "Tally Parent Group": acct.tally_parent_group or "",
            "Opening Balance": float(opening.amount) if opening else 0,
            "Dr / Cr": opening.amount_drcr if opening else acct.normal_balance or "",
            "GST Registration Type": acct.gst_registration_type or "",
            "GSTIN": tax.gstin if tax else "",
            "PAN": tax.pan if tax else "",
            "Bill-wise Required?": acct.bill_wise_required or "NO",
            "Cost Centre Required?": acct.cost_centre_required or "NO",
            "Spark Account Code": acct.account_no or "",
            "Spark Account Group": acct.spark_account_group.name if acct.spark_account_group else "",
            "Ledger Role": acct.ledger_role or "",
            "Notes": acct.remarks or "",
        }

    def _date_str(d):
        return d.isoformat() if d else ""

    def _ts_str(dt):
        return dt.strftime("%Y-%m-%d %H:%M:%S") if dt else ""

    def _dec(v):
        return float(v) if v else ""

    rows = []

    def _row(txn_type, acct):
        """Build a clean transaction row with Tally-relevant account fields."""
        acct_data = _acct_fields(acct) if acct else _acct_fields(None)
        return acct_data

    # --- Orders ---
    for o in _filter(Order.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"), date_field="date"):
        rows.append(_row("Order", o.account))

    # --- Sales ---
    for o in _filter(Sale.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"), date_field="created_at"):
        rows.append(_row("Sale", o.account))

    # --- PaymentEntry ---
    for o in _filter(PaymentEntry.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"), date_field="date"):
        rows.append(_row("Payment", o.account))

    # --- JournalEntry ---
    for o in _filter(JournalEntry.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"), date_field="date"):
        rows.append(_row("Journal", o.account))

    # --- Receipt ---
    for o in _filter(Receipt.objects.filter(company_id=company_id).select_related("party_name", "party_name__spark_account_group"), "party_name", date_field="date"):
        rows.append(_row("Receipt", o.party_name))

    # --- PurchaseM ---
    for o in _filter(PurchaseM.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"), date_field="date"):
        rows.append(_row("Purchase", o.account))

    # --- PurchaseTagwiseM ---
    for o in _filter(PurchaseTagwiseM.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"), date_field="date"):
        rows.append(_row("Purchase Tagwise", o.account))

    # --- PurchaseDiamond ---
    for o in _filter(PurchaseDiamond.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"), date_field="created_at"):
        rows.append(_row("Purchase Diamond", o.account))

    # --- ApprovalLooseM ---
    for o in _filter(ApprovalLooseM.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"), date_field="created_at"):
        rows.append(_row("Approval Loose", o.account))

    # --- ApprovalTagM ---
    for o in _filter(ApprovalTagM.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"), date_field="created_at"):
        rows.append(_row("Approval Tag", o.account))

    # --- PurAndApprovalM ---
    for o in _filter(PurAndApprovalM.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"), date_field="created_at"):
        rows.append(_row("Pur & Approval", o.account))

    # --- PurReturn ---
    for o in _filter(PurReturn.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"), date_field="date"):
        rows.append(_row("Pur Return", o.account))

    # --- SalesReturn ---
    for o in _filter(SalesReturn.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"), date_field="date"):
        rows.append(_row("Sales Return", o.account))

    # --- Receive ---
    for o in _filter(Receive.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"), date_field="date"):
        rows.append(_row("Receive", o.account))

    # --- Repair ---
    for o in _filter(Repair.objects.filter(company_id=company_id).select_related("account", "account__spark_account_group"), date_field="date"):
        rows.append(_row("Repair", o.account))

    # --- PaymentVoucher ---
    for o in _filter(PaymentVoucher.objects.filter(company_id=company_id).select_related("party_name", "party_name__spark_account_group"), "party_name", date_field="date"):
        rows.append(_row("Payment Voucher", o.party_name))

    # --- JournalVoucher ---
    for o in _filter(JournalVoucher.objects.filter(company_id=company_id).select_related("party_name", "party_name__spark_account_group"), "party_name", date_field="date"):
        rows.append(_row("Journal Voucher", o.party_name))

    # Remove duplicates by account to get unique accounts
    unique_accounts = {}
    for row in rows:
        ledger_name = row.get("Tally Ledger Name", "")
        if ledger_name and ledger_name not in unique_accounts:
            unique_accounts[ledger_name] = row
    rows = list(unique_accounts.values())

    # Headers with only the 14 required fields
    headers = [
        "Export Action",
        "Tally Ledger Name",
        "Tally Parent Group",
        "Opening Balance",
        "Dr / Cr",
        "GST Registration Type",
        "GSTIN",
        "PAN",
        "Bill-wise Required?",
        "Cost Centre Required?",
        "Spark Account Code",
        "Spark Account Group",
        "Ledger Role",
        "Notes",
    ]

    # Build Excel
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Header colour palette — one colour per column, cycling if more columns than colours
    HEADER_COLOURS = [
        "1F4E79",  # dark blue       — Export Action
        "2E75B6",  # medium blue     — Tally Ledger Name
        "2E75B6",  # medium blue     — Tally Parent Group
        "375623",  # dark green      — Opening Balance
        "375623",  # dark green      — Dr / Cr
        "7B2C2C",  # dark red        — GST Registration Type
        "7B2C2C",  # dark red        — GSTIN
        "7B2C2C",  # dark red        — PAN
        "7030A0",  # purple          — Bill-wise Required?
        "7030A0",  # purple          — Cost Centre Required?
        "C55A11",  # burnt orange    — Spark Account Code
        "C55A11",  # burnt orange    — Spark Account Group
        "404040",  # dark grey       — Ledger Role
        "404040",  # dark grey       — Notes
    ]

    thin_border = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
    )
    data_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Account Transactions"

    # ── Header row ────────────────────────────────────────────────────────────
    for col_idx, header in enumerate(headers, start=1):
        colour = HEADER_COLOURS[(col_idx - 1) % len(HEADER_COLOURS)]
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = PatternFill(fill_type="solid", fgColor=colour)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[1].height = 30  # taller header row
    ws.freeze_panes = "A2"            # freeze header so it stays visible when scrolling

    # ── Data rows ─────────────────────────────────────────────────────────────
    # Alternate row shading for readability
    EVEN_FILL = PatternFill(fill_type="solid", fgColor="DCE6F1")  # light blue
    ODD_FILL  = PatternFill(fill_type="solid", fgColor="FFFFFF")  # white

    for row_idx, row_data in enumerate(rows, start=2):
        row_fill = EVEN_FILL if row_idx % 2 == 0 else ODD_FILL
        for col_idx, header in enumerate(headers, start=1):
            val = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center")
            cell.border = data_border

    # ── Auto column widths ────────────────────────────────────────────────────
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row_idx in range(2, len(rows) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # ── Footer ────────────────────────────────────────────────────────────────
    footer_row = len(rows) + 3
    footer_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
    footer_font = Font(italic=True, color="595959", size=9)

    footer_data = [
        (1, "Exported On"),
        (2, datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        (3, f"Date Range: {start_date or 'All'} to {end_date or 'All'}"),
        (4, f"Total Records: {len(rows)}"),
    ]
    for col, val in footer_data:
        cell = ws.cell(row=footer_row, column=col, value=val)
        cell.fill = footer_fill
        cell.font = footer_font

    filename = f"account_transactions_tally_{start_date or 'all'}_to_{end_date or 'all'}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@api_view(["GET"])
@authentication_classes([JWTAuthentication, CsrfExemptSessionAuthentication])
@permission_classes([IsAuthenticated])
def accounts_tally_export_view(request):
    """Export Account Masters in Tally ledger format.

    Query params:
    - start_date: YYYY-MM-DD
    - end_date: YYYY-MM-DD
    """
    from datetime import datetime
    from openpyxl import Workbook

    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")

    if not start_date_str or not end_date_str:
        return Response({"detail": "start_date and end_date are required (YYYY-MM-DD)"}, status=400)

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
    except ValueError:
        return Response({"detail": "Invalid date format. Use YYYY-MM-DD"}, status=400)

    if start_date > end_date:
        return Response({"detail": "start_date cannot be after end_date"}, status=400)

    account_id = request.GET.get("account_id")

    user = request.user
    queryset = _base_queryset(user).filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date
    )

    if account_id:
        queryset = queryset.filter(id=account_id)

    queryset = queryset.order_by("account_no")

    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    headers = [
        "Export Action",
        "Tally Ledger Name",
        "Tally Parent Group",
        "Opening Balance",
        "Dr / Cr",
        "GST Registration Type",
        "GSTIN",
        "PAN",
        "Bill-wise Required?",
        "Cost Centre Required?",
        "Spark Account Code",
        "Spark Account Group",
        "Ledger Role",
        "Notes",
    ]

    HEADER_COLOURS = [
        "1F4E79", "2E75B6", "2E75B6", "375623", "375623",
        "7B2C2C", "7B2C2C", "7B2C2C", "7030A0", "7030A0",
        "C55A11", "C55A11", "404040", "404040",
    ]
    thin_border = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
    )
    data_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    EVEN_FILL = PatternFill(fill_type="solid", fgColor="DCE6F1")
    ODD_FILL  = PatternFill(fill_type="solid", fgColor="FFFFFF")

    wb = Workbook()
    ws = wb.active
    ws.title = "Tally Ledger Export"

    for col_idx, header in enumerate(headers, start=1):
        colour = HEADER_COLOURS[(col_idx - 1) % len(HEADER_COLOURS)]
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = PatternFill(fill_type="solid", fgColor=colour)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    for row_idx, account in enumerate(queryset, start=2):
        tax = getattr(account, "tax", None)
        opening = getattr(account, "opening_balance", None)
        row_fill = EVEN_FILL if row_idx % 2 == 0 else ODD_FILL

        values = [
            "Create" if account.export_to_tally == "YES" else "Ignore",
            account.tally_ledger_name_override or account.account_name or "",
            account.tally_parent_group or "",
            float(opening.amount) if opening else 0,
            opening.amount_drcr if opening else account.normal_balance or "",
            account.gst_registration_type or "",
            tax.gstin if tax else "",
            tax.pan if tax else "",
            account.bill_wise_required or "NO",
            account.cost_centre_required or "NO",
            account.account_no or "",
            account.spark_account_group.name if account.spark_account_group else "",
            account.ledger_role or "",
            account.remarks or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center")
            cell.border = data_border

    # Auto column widths
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row_idx in range(2, queryset.count() + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    footer_row = queryset.count() + 3
    footer_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
    footer_font = Font(italic=True, color="595959", size=9)
    for col, val in [
        (1, "Exported On"),
        (2, datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        (3, f"Date Range: {start_date} to {end_date}"),
        (4, f"Total Ledgers: {queryset.count()}"),
    ]:
        cell = ws.cell(row=footer_row, column=col, value=val)
        cell.fill = footer_fill
        cell.font = footer_font

    filename = f"tally_ledger_export_{start_date}_to_{end_date}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@api_view(["GET"])
@authentication_classes([JWTAuthentication, CsrfExemptSessionAuthentication])
@permission_classes([IsAuthenticated])
def ac_groups_tally_export_view(request):
    """Export account groups in Excel format with date range filter."""
    from datetime import datetime
    from openpyxl import Workbook
    
    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")
    
    start_date = None
    end_date = None
    
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        except ValueError:
            return Response({"detail": "Invalid start_date format. Use YYYY-MM-DD"}, status=400)
            
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        except ValueError:
            return Response({"detail": "Invalid end_date format. Use YYYY-MM-DD"}, status=400)
            
    if start_date and end_date and start_date > end_date:
        return Response({"detail": "start_date cannot be after end_date"}, status=400)
    
    group_id = request.GET.get("group_id")
    
    user = request.user
    queryset = ACGroup.objects.select_related("ac_group")
    
    if start_date:
        queryset = queryset.filter(created_at__date__gte=start_date)
    if end_date:
        queryset = queryset.filter(created_at__date__lte=end_date)
    
    if group_id:
        queryset = queryset.filter(ac_group_id=group_id)
        
    queryset = queryset.order_by("-created_at")
    
    if user.is_authenticated and getattr(user, "company_id", None):
        queryset = queryset.filter(created_by__company_id=user.company_id)
    else:
        queryset = queryset.none()
    
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    headers = [
        "Spark Account Group",
        "Tally Parent Group",
        "Financial Statement",
        "Universal Nature",
        "Normal Balance",
        "Status",
        "Incl. in Sale",
        "Incl. in Pur",
        "Incl. in Out",
        "Incl. in I/R",
        "Address Req",
        "Restrict Credit Facility",
        "Created At",
    ]

    AC_GROUP_COLOURS = [
        "1F4E79", "2E75B6", "375623", "7030A0", "C55A11",
        "7B2C2C", "404040", "404040", "404040", "404040",
        "404040", "404040", "595959",
    ]
    thin_border = Border(
        left=Side(style="thin", color="FFFFFF"),
        right=Side(style="thin", color="FFFFFF"),
    )
    data_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    EVEN_FILL = PatternFill(fill_type="solid", fgColor="DCE6F1")
    ODD_FILL  = PatternFill(fill_type="solid", fgColor="FFFFFF")

    wb = Workbook()
    ws = wb.active
    ws.title = "AC Groups Export"

    for col_idx, header in enumerate(headers, start=1):
        colour = AC_GROUP_COLOURS[(col_idx - 1) % len(AC_GROUP_COLOURS)]
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = PatternFill(fill_type="solid", fgColor=colour)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    rows_list = list(queryset)
    for row_idx, group in enumerate(rows_list, start=2):
        row_fill = EVEN_FILL if row_idx % 2 == 0 else ODD_FILL
        values = [
            group.ac_group.name if group.ac_group else "",
            group.tally_parent_group,
            group.financial_statement,
            group.universal_nature,
            group.normal_balance,
            group.status,
            group.incl_in_sale,
            group.incl_in_pur,
            group.incl_in_out,
            group.incl_in_ir,
            group.address_req,
            group.restrict_credit_facility,
            group.created_at.strftime("%Y-%m-%d %H:%M:%S") if group.created_at else "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center")
            cell.border = data_border

    # Auto column widths
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row_idx in range(2, len(rows_list) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    footer_row = len(rows_list) + 3
    footer_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
    footer_font = Font(italic=True, color="595959", size=9)
    for col, val in [
        (1, "Exported On"),
        (2, datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        (3, f"Date Range: {start_date} to {end_date}"),
    ]:
        cell = ws.cell(row=footer_row, column=col, value=val)
        cell.fill = footer_fill
        cell.font = footer_font
    
    filename = f"ac_groups_export_{start_date}_to_{end_date}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
